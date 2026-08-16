from __future__ import annotations

import io
import unittest
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font

from processors import store_report
from processors.common.excel import run_with_output_rollback

FONT = Font(name="微软雅黑", size=12)


LITERAL_DETAIL_ROW_LABELS = {
    3: "海尔系冰箱",
    4: "海尔系洗衣机",
    5: "美的系冰箱",
    6: "美的系洗衣机",
    7: "西门子冰箱",
    8: "西门子洗衣机",
    9: "博世冰箱",
    10: "博世洗衣机",
    11: "美菱冰箱",
    12: "美菱洗衣机",
    13: "小鸭洗衣机",
    14: "海信",
    15: "创维",
    16: "TCL",
    17: "海尔",
    18: "华为",
    19: "格力",
    20: "美的",
    21: "海尔",
    22: "海信",
    23: "奥克斯",
    24: "科龙",
    25: "TCL",
    26: "老板",
    27: "方太",
    28: "AO史密斯",
    29: "海尔",
    30: "美的",
    31: "万家乐",
    32: "数码",
}

LITERAL_BRAND_GROUP_ROW_LABELS = {
    40: "海尔系",
    41: "美的系",
    42: "格力",
    43: "博西",
    44: "海信系",
    45: "创维",
    46: "TCL",
}

LITERAL_STRUCTURE_CELLS: dict[str, str] = {
    "A2": "品类",
    "B2": "序号",
    "C2": "品牌",
    "D2": "26年发生额",
    "E2": "上传额",
    "F2": "上传率",
    "G2": "回款额",
    "H2": "回款率",
    **{f"C{row}": label for row, label in LITERAL_DETAIL_ROW_LABELS.items()},
    "A33": "费用总计",
    "C38": "表2                主要品牌国补上传及回款情况",
    "C39": "品牌",
    "D39": "26年国补发生额",
    "E39": "26年国补回款额",
    "F39": "回款率",
    **{f"C{row}": label for row, label in LITERAL_BRAND_GROUP_ROW_LABELS.items()},
    "C47": "合计",
    "C49": "表3",
    "D49": "审核中",
    "E49": "未上传",
    "C50": "家电",
    "C51": "数码",
    "C52": "合计",
    "A53": "模板版本：2026-V4",
}

LITERAL_EXPECTED_MERGED_RANGES: frozenset[str] = frozenset({
    "A1:H1",
    "A3:A13",
    "A14:A18",
    "A19:A25",
    "A26:A31",
    "A33:C33",
    "C38:F38",
})

LITERAL_ROW_RULES: tuple[store_report.RowRule, ...] = (
    store_report.RowRule(3, ("冰箱",), ("海尔", "卡萨帝"), ("冰箱",), ("海尔", "卡萨帝"), fill_digital=False),
    store_report.RowRule(4, ("洗衣机",), ("海尔", "卡萨帝"), ("洗衣机",), ("海尔", "卡萨帝"), fill_digital=False),
    store_report.RowRule(5, ("冰箱",), ("美的", "COLMO", "东芝"), ("冰箱",), ("美的系", "COLMO", "东芝JX"), fill_digital=False),
    store_report.RowRule(6, ("洗衣机",), ("美的", "小天鹅", "COLMO"), ("洗衣机",), ("美的系", "小天鹅", "COLMO"), fill_digital=False),
    store_report.RowRule(7, ("冰箱",), ("西门子",), ("冰箱",), ("西门子",), fill_digital=False),
    store_report.RowRule(8, ("洗衣机",), ("西门子",), ("洗衣机",), ("西门子",), fill_digital=False),
    store_report.RowRule(9, ("冰箱",), ("博世",), ("冰箱",), ("博世",), fill_digital=False),
    store_report.RowRule(10, ("洗衣机",), ("博世",), ("洗衣机",), ("博世",), fill_digital=False),
    store_report.RowRule(11, ("冰箱",), ("美菱",), ("冰箱",), ("美菱",), fill_digital=False),
    store_report.RowRule(12, ("洗衣机",), ("美菱",), ("洗衣机",), ("美菱",), fill_digital=False),
    store_report.RowRule(13, ("洗衣机",), ("小鸭",), ("洗衣机",), ("小鸭",), fill_digital=False),
    store_report.RowRule(14, ("国产彩电",), ("海信",), ("电视",), ("海信",), fill_digital=False),
    store_report.RowRule(15, ("国产彩电",), ("创维",), ("电视",), ("创维",), fill_digital=False),
    store_report.RowRule(16, ("国产彩电",), ("TCL",), ("电视",), ("TCL",), fill_digital=False),
    store_report.RowRule(17, ("国产彩电",), ("海尔", "卡萨帝"), ("电视",), ("海尔", "卡萨帝"), fill_digital=False),
    store_report.RowRule(18, ("国产彩电",), ("华为", "华为（终端）"), ("电视",), ("华为", "华为（终端）"), fill_digital=False),
    store_report.RowRule(19, ("空调",), ("格力",), ("空调",), ("格力",), fill_digital=False),
    store_report.RowRule(20, ("空调",), ("美的",), ("空调",), ("美的",), fill_digital=False),
    store_report.RowRule(21, ("空调",), ("海尔", "卡萨帝"), ("空调",), ("海尔", "卡萨帝"), fill_digital=False),
    store_report.RowRule(22, ("空调",), ("海信",), ("空调",), ("海信",), fill_digital=False),
    store_report.RowRule(23, ("空调",), ("奥克斯",), ("空调",), ("奥克斯",), fill_digital=False),
    store_report.RowRule(24, ("空调",), ("科龙",), ("空调",), ("科龙",), fill_digital=False),
    store_report.RowRule(25, ("空调",), ("TCL",), ("空调",), ("TCL",), fill_digital=False),
    store_report.RowRule(26, ("厨卫",), ("老板",), ("厨卫",), ("老板",), fill_digital=False),
    store_report.RowRule(27, ("厨卫", "冰箱"), ("方太",), ("厨卫", "冰箱"), ("方太",), fill_digital=False),
    store_report.RowRule(28, ("厨卫",), ("AO史密斯", "A.O.史密斯"), ("厨卫",), ("AO史密斯", "A.O.史密斯"), fill_digital=False),
    store_report.RowRule(29, ("厨卫",), ("海尔", "卡萨帝"), ("厨卫",), ("海尔", "卡萨帝"), fill_digital=False),
    store_report.RowRule(30, ("厨卫",), ("美的", "COLMO"), ("厨卫",), ("美的系", "美的", "COLMO"), fill_digital=False),
    store_report.RowRule(31, ("厨卫",), ("万家乐",), ("厨卫",), ("万家乐",), fill_digital=False),
    store_report.RowRule(32, (), (), ("数码",), (), fill_digital=True),
)

LITERAL_BRAND_GROUP_RULES: tuple[store_report.BrandGroupRule, ...] = (
    store_report.BrandGroupRule(
        40,
        "海尔系",
        (
            store_report.BrandGroupCategory("冰箱", "冰箱", ("海尔", "卡萨帝")),
            store_report.BrandGroupCategory("洗衣机", "洗衣机", ("海尔", "卡萨帝")),
            store_report.BrandGroupCategory("国产彩电", "电视", ("海尔", "卡萨帝")),
            store_report.BrandGroupCategory("空调", "空调", ("海尔", "卡萨帝")),
            store_report.BrandGroupCategory("厨卫", "厨卫", ("海尔", "卡萨帝")),
        ),
    ),
    store_report.BrandGroupRule(
        41,
        "美的系",
        (
            store_report.BrandGroupCategory("冰箱", "冰箱", ("美的", "COLMO", "东芝")),
            store_report.BrandGroupCategory("洗衣机", "洗衣机", ("美的", "小天鹅", "COLMO")),
            store_report.BrandGroupCategory("空调", "空调", ("美的",)),
            store_report.BrandGroupCategory("厨卫", "厨卫", ("美的", "COLMO")),
        ),
    ),
    store_report.BrandGroupRule(
        42,
        "格力",
        (store_report.BrandGroupCategory("空调", "空调", ("格力",)),),
    ),
    store_report.BrandGroupRule(
        43,
        "博西",
        (
            store_report.BrandGroupCategory("冰箱", "冰箱", ("西门子", "博世")),
            store_report.BrandGroupCategory("洗衣机", "洗衣机", ("西门子", "博世")),
        ),
    ),
    store_report.BrandGroupRule(
        44,
        "海信系",
        (
            store_report.BrandGroupCategory("国产彩电", "电视", ("海信",)),
            store_report.BrandGroupCategory("空调", "空调", ("海信", "科龙")),
        ),
    ),
    store_report.BrandGroupRule(
        45,
        "创维",
        (store_report.BrandGroupCategory("国产彩电", "电视", ("创维",)),),
    ),
    store_report.BrandGroupRule(
        46,
        "TCL",
        (
            store_report.BrandGroupCategory("国产彩电", "电视", ("TCL",)),
            store_report.BrandGroupCategory("空调", "空调", ("TCL",)),
        ),
    ),
)


LITERAL_TEMPLATE_HEADER_PREFIX = (
    "表1                  2026年（益庄店 ）门店国补上传及回款情况表_"
)


def _fill_columns(sheet) -> None:
    for column_index in range(1, store_report.EXPECTED_COLUMN_COUNT + 1):
        sheet.cell(row=1, column=column_index, value="")


def _build_minimal_template() -> Workbook:
    """A template with just enough structure to pass validate_template."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "益庄"
    for column_index in range(1, store_report.EXPECTED_COLUMN_COUNT + 1):
        sheet.cell(
            row=1,
            column=column_index,
            value=f"{LITERAL_TEMPLATE_HEADER_PREFIX}2026-08-15 18:06:27"
            if column_index == 1
            else "",
        )
    for coordinate, value in LITERAL_STRUCTURE_CELLS.items():
        sheet[coordinate] = value
    for range_ in LITERAL_EXPECTED_MERGED_RANGES:
        sheet.merge_cells(range_)
    for expected, row in enumerate(range(3, 33), start=1):
        sheet[f"B{row}"] = expected
    sheet.row_dimensions[53].hidden = True
    return workbook


def _upload_detail_row(
    *,
    document: str = "0001",
    category: str = "冰箱",
    brand: str = "海尔",
    reference: str = "12345678901N",
    subsidy: object = 100,
    remark: str = "已上传",
) -> list[object]:
    """One 家电-明细总表 row shaped like the real output (openpyxl 写回后
    calamine 读出的是 datetime，测试里直接给 date 对象)。"""
    header = store_report.coupon_appliance.COUPON_OUTPUT_HEADER
    values = {
        "单据号": document,
        "单据日期": date(2026, 1, 1),
        "商品名称": "测试商品",
        "品牌": brand,
        "财务大类": category,
        "明细摘要": reference,
        header[6]: subsidy,
        "备注": remark,
        "详细情况": "",
        "回款情况": "",
    }
    return [values.get(column) for column in header]


def _write_upload_detail_sheet(workbook, rows) -> None:
    """Append 家电-明细总表 to a fixture workbook."""
    sheet = workbook.create_sheet(store_report.coupon_appliance.DETAILS_SHEET_NAME)
    sheet.append(store_report.coupon_appliance.COUPON_OUTPUT_HEADER)
    for row in rows:
        sheet.append(row)


def _payment_detail_row(
    *,
    raw_category: str = "A02-电冰箱",
    brand: str = "海尔",
    reference: str = "12345678901N",
    subsidy: object = 100,
) -> list[object]:
    """One 回款明细「家电明细」row shaped like the real output."""
    header = (
        store_report.PAYMENT_PROFILES["家电"].detail_headers
        + store_report.PAYMENT_DERIVED_HEADERS
    )
    mapped = store_report.PAYMENT_CATEGORY_MAP.get(raw_category)
    values = {
        "拨付批次": "batch",
        "交易时间": "2026-01-01 10:00:00",
        "交易参考号": reference,
        "商户编号": "89813015722APT1",
        "编码品类": raw_category,
        "商品名称": "测试商品",
        "补贴金额": subsidy,
        "财务大类": mapped,
        "品牌": brand,
    }
    return [values.get(column) for column in header]


def _write_payment_detail_sheet(workbook, rows) -> None:
    """Append 回款明细「家电明细」 to a fixture workbook."""
    sheet = workbook.create_sheet(
        store_report.PAYMENT_PROFILES["家电"].detail_sheet_name
    )
    sheet.append(
        store_report.PAYMENT_PROFILES["家电"].detail_headers
        + store_report.PAYMENT_DERIVED_HEADERS
    )
    for row in rows:
        sheet.append(row)


class ReportRatioTests(unittest.TestCase):
    def test_to_decimal_quantizes_float_accumulation(self) -> None:
        """Excel 浮点尾差在读取时统一量化为两位小数：0.1+0.2 必须等于 0.30，
        而不是 0.30000000000000004 一路带进精确相等校验。"""
        self.assertEqual(
            store_report.to_decimal(0.1) + store_report.to_decimal(0.2),
            Decimal("0.30"),
        )
        self.assertEqual(
            store_report.to_decimal(2280155.70000000000101),
            Decimal("2280155.70"),
        )
        self.assertEqual(
            store_report.to_decimal(779458.0500000000004),
            Decimal("779458.05"),
        )
        self.assertEqual(store_report.to_decimal(None), Decimal("0"))
        self.assertEqual(store_report.to_decimal(""), Decimal("0"))

    def test_to_count_accepts_integers_only(self) -> None:
        """数量不能用金额的两位小数量化：1.004 量化后是 1.00 会漏过
        非整数检查，必须用未量化的 Decimal 原样校验。"""
        self.assertEqual(store_report.to_count(1), 1)
        self.assertEqual(store_report.to_count(1.0), 1)
        self.assertEqual(store_report.to_count("3"), 3)
        for value in (1.004, 1.005, 1.5, "1.004", "1.5"):
            with self.assertRaisesRegex(ValueError, "整数"):
                store_report.to_count(value)
        # NaN 与 Infinity 不是有限数，必须拒绝而不是转成整数。
        for value in (float("nan"), float("inf"), float("-inf")):
            with self.assertRaisesRegex(ValueError, "整数"):
                store_report.to_count(value)
        # 非数值文本（Decimal 抛 InvalidOperation）统一为数量错误，不泄漏底层异常。
        for value in ("abc", "1.2.3", "1,000"):
            with self.assertRaisesRegex(ValueError, "整数"):
                store_report.to_count(value)
        self.assertEqual(store_report.to_count(None), 0)
        self.assertEqual(store_report.to_count(""), 0)

    def test_current_timestamp_uses_shanghai_timezone(self) -> None:
        timestamp = store_report.current_timestamp()

        self.assertEqual(timestamp.tzinfo.key, "Asia/Shanghai")
        self.assertEqual(timestamp.utcoffset(), timedelta(hours=8))

    def test_upload_data_reads_embedded_digital_summary(self) -> None:
        workbook = Workbook()
        household_sheet = workbook.active
        household_sheet.title = store_report.UPLOAD_SHEET_NAME
        household_sheet.append(store_report.UPLOAD_HEADER)
        # 数据汇总的品牌行是明细聚合的校验基准（守恒校验要求逐键一致）。
        household_sheet.append(("冰箱", "海尔", "已上传", 1, 100))
        household_sheet.append((None, None, "未上传", 1, 50))
        household_sheet.append(("家电", None, "已上传", 2, 100))
        household_sheet.append((None, None, "未上传", 1, 50))
        household_sheet.append((None, None, "合计", 3, 150))
        household_sheet.append(("数码", None, "已上传", 4, 200))
        household_sheet.append((None, None, "未上传", 1, 40))
        household_sheet.append((None, None, "合计", 5, 240))
        _write_upload_detail_sheet(
            workbook,
            [
                _upload_detail_row(subsidy=100, remark="已上传"),
                _upload_detail_row(subsidy=50, remark="未上传"),
            ],
        )

        with TemporaryDirectory() as directory:
            upload_file = Path(directory) / "审核明细.xlsx"
            workbook.save(upload_file)
            (
                upload_data,
                digital_totals,
                project_metrics,
                corrections,
                reviews,
            ) = store_report.load_upload_data(upload_file)

        self.assertEqual(
            upload_data[("冰箱", "海尔")],
            {"已上传": Decimal("100"), "未上传": Decimal("50")},
        )
        self.assertNotIn(("家电", "海尔"), upload_data)
        self.assertEqual(corrections, [])
        self.assertEqual(reviews, [])
        self.assertEqual(
            digital_totals,
            {"发生额": Decimal("240"), "上传额": Decimal("200")},
        )
        self.assertEqual(
            project_metrics,
            {
                "家电": {"已上传": 2, "未上传": 1},
                "数码": {"已上传": 4, "未上传": 1},
            },
        )

    def test_payment_data_sums_known_digital_categories(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.PAYMENT_SHEET_NAME
        sheet.append(("财务大类", "品牌", "补贴金额合计", "补贴金额计数"))
        sheet.append(("冰箱", "海尔", 100, 1))
        sheet.append(("合计", None, 100, 1))
        sheet.append((None, None, None, None))
        sheet.append(("手机", "OPPO", 30, 1))
        sheet.append((None, "华为", 50, 1))
        sheet.append(("平板", "华为", 20, 1))
        sheet.append(("合计", None, 100, 3))
        sheet.append(("合计", None, 200, 4))
        _write_payment_detail_sheet(workbook, [_payment_detail_row()])

        with TemporaryDirectory() as directory:
            payment_file = Path(directory) / "回款明细.xlsx"
            workbook.save(payment_file)
            (
                payment_data,
                digital_amount,
                project_metrics,
                payment_index,
                ambiguous_refs,
            ) = store_report.load_payment_data(payment_file)

        self.assertEqual(payment_data, {("冰箱", "海尔"): Decimal("100")})
        self.assertEqual(
            payment_index["12345678901N"],
            store_report.PaymentDetailRecord(
                category="冰箱",
                raw_category="A02-电冰箱",
                brand="海尔",
                subsidy=Decimal("100"),
            ),
        )
        self.assertEqual(ambiguous_refs, {})
        self.assertEqual(digital_amount, Decimal("100"))
        self.assertEqual(
            project_metrics,
            {
                "家电": 1,
                "数码": 3,
            },
        )

    def test_payment_data_rejects_an_unconfigured_category(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.PAYMENT_SHEET_NAME
        sheet.append(("财务大类", "品牌", "补贴金额合计", "补贴金额计数"))
        sheet.append(("电脑", "联想", 100, 1))

        with TemporaryDirectory() as directory:
            payment_file = Path(directory) / "回款明细.xlsx"
            workbook.save(payment_file)
            with self.assertRaisesRegex(ValueError, "电脑"):
                store_report.load_payment_data(payment_file)

    def test_household_row_ratio_uses_occurred_amount(self) -> None:
        sheet = Workbook().active
        rule = store_report.RowRule(7, ("冰箱",), ("西门子",), ("冰箱",), ("西门子",))
        upload_data = {("冰箱", "西门子"): {"已上传": Decimal("100"), "未上传": Decimal("50")}}
        payment_data = {("冰箱", "西门子"): Decimal("40")}

        store_report.write_row(
            sheet, rule, upload_data, payment_data,
            {"发生额": Decimal("0"), "上传额": Decimal("0")}, Decimal("0"), FONT, {},
        )

        self.assertAlmostEqual(sheet["H7"].value, 40 / 150)
        self.assertAlmostEqual(sheet["F7"].value, 100 / 150)

    def test_fotile_row_combines_refrigerator_and_kitchen_data(self) -> None:
        """第 27 行（厨卫/方太）同时汇总冰箱和厨卫两侧方太金额，其余行不重复计入。"""
        sheet = Workbook().active
        fotile_rule = next(
            rule for rule in store_report.ROW_RULES if rule.row == 27
        )
        upload_data = {
            ("冰箱", "方太"): {"已上传": Decimal("1500"), "未上传": Decimal("0")},
            ("厨卫", "方太"): {"已上传": Decimal("800"), "未上传": Decimal("0")},
        }
        payment_data = {
            ("冰箱", "方太"): Decimal("1200"),
            ("厨卫", "方太"): Decimal("600"),
        }

        store_report.write_row(
            sheet, fotile_rule, upload_data, payment_data,
            {"发生额": Decimal("0"), "上传额": Decimal("0")}, Decimal("0"), FONT, {},
        )

        self.assertEqual(sheet["D27"].value, 2300)
        self.assertEqual(sheet["E27"].value, 2300)
        self.assertEqual(sheet["F27"].value, 1.0)
        self.assertEqual(sheet["G27"].value, 1800)
        self.assertAlmostEqual(sheet["H27"].value, 1800 / 2300)

    def test_digital_row_fills_amount_and_both_ratios(self) -> None:
        sheet = Workbook().active
        rule = store_report.RowRule(32, (), (), ("数码",), (), fill_digital=True)
        digital_upload = {"发生额": Decimal("1000"), "上传额": Decimal("800")}

        store_report.write_row(sheet, rule, {}, {}, digital_upload, Decimal("300"), FONT, {})

        self.assertEqual(sheet["D32"].value, 1000)
        self.assertEqual(sheet["E32"].value, 800)
        self.assertAlmostEqual(sheet["F32"].value, 0.8)
        self.assertEqual(sheet["G32"].value, 300)
        self.assertAlmostEqual(sheet["H32"].value, 0.3)
        for coordinate in ("D32", "E32", "G32"):
            self.assertEqual(
                sheet[coordinate].number_format,
                store_report.DATA_NUMBER_FORMAT,
            )
        for coordinate in ("F32", "H32"):
            self.assertEqual(
                sheet[coordinate].number_format,
                store_report.PERCENT_NUMBER_FORMAT,
            )

    def test_update_totals_computes_ratio_of_totals(self) -> None:
        sheet = Workbook().active
        for row in store_report.DETAIL_ROWS:
            sheet[f"D{row}"] = None
            sheet[f"E{row}"] = None
            sheet[f"G{row}"] = None
        sheet["D3"] = 100
        sheet["E3"] = 50
        sheet["G3"] = 20
        sheet["D4"] = 100
        sheet["E4"] = 50
        sheet["G4"] = 30

        store_report.update_totals(sheet, FONT, {})

        self.assertEqual(sheet["D33"].value, 200)
        self.assertEqual(sheet["E33"].value, 100)
        self.assertAlmostEqual(sheet["F33"].value, 0.5)
        self.assertEqual(sheet["G33"].value, 50)
        self.assertAlmostEqual(sheet["H33"].value, 0.25)
        for coordinate in ("D33", "E33", "G33"):
            self.assertEqual(
                sheet[coordinate].number_format,
                store_report.DATA_NUMBER_FORMAT,
            )
        for coordinate in ("F33", "H33"):
            self.assertEqual(
                sheet[coordinate].number_format,
                store_report.PERCENT_NUMBER_FORMAT,
            )

    def test_table3_subtracts_payment_from_uploaded_metrics(self) -> None:
        sheet = Workbook().active
        upload_metrics = {
            "家电": {
                "已上传": 10,
                "未上传": 3,
            },
            "数码": {
                "已上传": 8,
                "未上传": 2,
            },
        }
        payment_metrics = {
            "家电": 4,
            "数码": 3,
        }

        store_report.write_table3(sheet, upload_metrics, payment_metrics, FONT, {})

        self.assertEqual(
            [sheet[cell].value for cell in ("D50", "E50")],
            [6, 3],
        )
        self.assertEqual(
            [sheet[cell].value for cell in ("D51", "E51")],
            [5, 2],
        )
        self.assertEqual(
            [sheet[cell].value for cell in ("D52", "E52")],
            [11, 5],
        )
        for row in (50, 51, 52):
            for column in ("D", "E"):
                self.assertEqual(
                    sheet[f"{column}{row}"].number_format,
                    store_report.DATA_NUMBER_FORMAT,
                )


class SourceHeaderValidationTests(unittest.TestCase):
    def test_load_upload_data_rejects_a_wrong_header(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.UPLOAD_SHEET_NAME
        # The pre-rename header (备注 / 2026家电国补（计入收入）合计) is rejected:
        # a stale 审核明细.xlsx must fail loudly instead of being misread, and
        # rerunning the 审核明细 mode regenerates it with the new header.
        sheet.append(("财务大类", "品牌", "备注", "数量", "2026家电国补（计入收入）合计"))

        with TemporaryDirectory() as directory:
            upload_file = Path(directory) / "审核明细.xlsx"
            workbook.save(upload_file)
            with self.assertRaisesRegex(ValueError, "表头"):
                store_report.load_upload_data(upload_file)

    def test_load_payment_data_rejects_a_wrong_header(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.PAYMENT_SHEET_NAME
        sheet.append(("品类", "品牌", "金额", "笔数"))  # renamed columns

        with TemporaryDirectory() as directory:
            payment_file = Path(directory) / "回款明细.xlsx"
            workbook.save(payment_file)
            with self.assertRaisesRegex(ValueError, "表头"):
                store_report.load_payment_data(payment_file)

    def test_upload_header_literal_is_pinned(self) -> None:
        """The 审核明细 数据汇总 header is a cross-module contract: a rename
        must break this test rather than silently re-parse store data."""
        self.assertEqual(
            store_report.UPLOAD_HEADER,
            ("财务大类", "品牌", "上传状态", "数量", "2026国补金额"),
        )

    def test_load_upload_data_parses_statuses_with_the_literal_header(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.UPLOAD_SHEET_NAME
        sheet.append(("财务大类", "品牌", "上传状态", "数量", "2026国补金额"))
        sheet.append(("家电", None, "已上传", 2, 100))
        sheet.append((None, None, "未上传", 1, 50))
        sheet.append((None, None, "合计", 3, 150))
        sheet.append(("数码", None, "已上传", 4, 200))
        sheet.append((None, None, "未上传", 1, 40))
        sheet.append((None, None, "合计", 5, 240))
        _write_upload_detail_sheet(workbook, [])

        with TemporaryDirectory() as directory:
            upload_file = Path(directory) / "审核明细.xlsx"
            workbook.save(upload_file)
            (
                _amounts,
                digital_totals,
                project_metrics,
                _corrections,
                _reviews,
            ) = store_report.load_upload_data(upload_file)

        self.assertEqual(
            digital_totals,
            {"发生额": Decimal("240"), "上传额": Decimal("200")},
        )
        self.assertEqual(project_metrics["家电"]["已上传"], 2)
        self.assertEqual(project_metrics["家电"]["未上传"], 1)
        self.assertEqual(project_metrics["数码"]["已上传"], 4)
        self.assertEqual(project_metrics["数码"]["未上传"], 1)

    def test_load_upload_data_parses_six_column_header_with_returned_column(self) -> None:
        """新版六列审核明细（带 F 列“退回”）能被正常读取，且结果与原五列版本完全一致。"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.UPLOAD_SHEET_NAME
        sheet.append(("财务大类", "品牌", "上传状态", "数量", "2026国补金额", "退回"))
        sheet.append(("家电", None, "已上传", 2, 100, 1))
        sheet.append((None, None, "未上传", 1, 50, 0))
        sheet.append((None, None, "合计", 3, 150, 1))
        sheet.append(("数码", None, "已上传", 4, 200, 2))
        sheet.append((None, None, "未上传", 1, 40, 0))
        sheet.append((None, None, "合计", 5, 240, 2))
        _write_upload_detail_sheet(workbook, [])

        with TemporaryDirectory() as directory:
            upload_file = Path(directory) / "审核明细.xlsx"
            workbook.save(upload_file)
            (
                _amounts,
                digital_totals,
                project_metrics,
                _corrections,
                _reviews,
            ) = store_report.load_upload_data(upload_file)

        self.assertEqual(
            digital_totals,
            {"发生额": Decimal("240"), "上传额": Decimal("200")},
        )
        self.assertEqual(project_metrics["家电"]["已上传"], 2)
        self.assertEqual(project_metrics["家电"]["未上传"], 1)
        self.assertEqual(project_metrics["数码"]["已上传"], 4)
        self.assertEqual(project_metrics["数码"]["未上传"], 1)

    def test_load_upload_data_rejects_non_integer_total_count(self) -> None:
        """审核明细项目「合计」行的数量为非整数（如 abc 或 1.5）时拒绝。"""
        for invalid_count in ("abc", 1.5):
            with self.subTest(invalid_count=invalid_count):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = store_report.UPLOAD_SHEET_NAME
                sheet.append(("财务大类", "品牌", "上传状态", "数量", "2026国补金额"))
                sheet.append(("家电", None, "已上传", 2, 100))
                sheet.append((None, None, "未上传", 1, 50))
                sheet.append((None, None, "合计", invalid_count, 150))
                _write_upload_detail_sheet(workbook, [])

                with TemporaryDirectory() as directory:
                    upload_file = Path(directory) / "审核明细.xlsx"
                    workbook.save(upload_file)
                    with self.assertRaisesRegex(ValueError, "数量应为整数"):
                        store_report.load_upload_data(upload_file)


class RuleCoverageTests(unittest.TestCase):
    def test_accepts_when_every_group_is_claimed_by_a_rule(self) -> None:
        upload_data = {("冰箱", "海尔"): {"已上传": Decimal("100"), "未上传": Decimal("0")}}
        payment_data = {("冰箱", "海尔"): Decimal("80")}

        store_report.validate_rule_coverage(upload_data, payment_data)

    def test_accepts_both_fotile_keys(self) -> None:
        """方太冰箱和厨卫统一归入第 27 行规则覆盖。"""
        store_report.validate_rule_coverage(
            {
                ("冰箱", "方太"): {"已上传": Decimal("1500")},
                ("厨卫", "方太"): {"已上传": Decimal("800")},
            },
            {
                ("冰箱", "方太"): Decimal("1500"),
                ("厨卫", "方太"): Decimal("800"),
            },
        )

    def test_ignores_a_zero_amount_unclaimed_group(self) -> None:
        upload_data = {("平板电脑", "未知品牌"): {"已上传": Decimal("0"), "未上传": Decimal("0")}}
        payment_data = {("平板电脑", "未知品牌"): Decimal("0")}

        store_report.validate_rule_coverage(upload_data, payment_data)

    def test_rejects_an_unclaimed_nonzero_upload_group(self) -> None:
        """Simulates a brand newly appearing upstream that ROW_RULES hasn't been updated for."""
        upload_data = {("冰箱", "新品牌"): {"已上传": Decimal("500"), "未上传": Decimal("0")}}
        payment_data: dict[tuple[str, str], Decimal] = {}

        with self.assertRaisesRegex(ValueError, "新品牌"):
            store_report.validate_rule_coverage(upload_data, payment_data)

    def test_rejects_an_unclaimed_nonzero_payment_group(self) -> None:
        upload_data: dict[tuple[str, str], dict[str, Decimal]] = {}
        payment_data = {("冰箱", "新品牌"): Decimal("300")}

        with self.assertRaisesRegex(ValueError, "新品牌"):
            store_report.validate_rule_coverage(upload_data, payment_data)

    def test_real_row_rules_have_no_internal_conflicts(self) -> None:
        """Regression test: RowRule(20, ...) lists both 华为 and 华为（终端）, which
        normalize to the same brand — that must be tolerated as redundancy, not
        flagged as two different rules claiming the same key."""
        store_report.validate_rule_coverage({}, {})


class UploadCategoryCorrectionTests(unittest.TestCase):
    """审核侧品类纠正：以回款明细编码品类为权威，按交易参考号匹配。

    方案要点：A02-电冰箱 的方太 BCD 在审核侧被标成厨卫，经参考号关联到
    回款侧后纠正为冰箱，并入厨卫/方太汇总行（第 27 行）；真正的厨卫方太与未回款的记录
    保留原品类；参考号重复、品牌/金额不符、编码品类未配置时不自动纠正。
    """

    FOTILE_REFERENCE = "17914133741N"

    def _load(self, detail_rows, payment_rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.UPLOAD_SHEET_NAME
        sheet.append(store_report.UPLOAD_HEADER)
        # 数据汇总的品牌行必须与明细聚合逐键一致（守恒校验）。
        header = store_report.coupon_appliance.COUPON_OUTPUT_HEADER
        for row in detail_rows:
            category = row[header.index("财务大类")]
            brand = row[header.index("品牌")]
            remark = row[header.index("备注")]
            subsidy = row[header.index(header[6])]
            if category == "数码" or remark not in ("已上传", "未上传"):
                continue
            sheet.append((category, brand, remark, 1, subsidy))
        sheet.append(("家电", None, "已上传", 0, 0))
        sheet.append((None, None, "未上传", 0, 0))
        sheet.append((None, None, "合计", 0, 0))
        sheet.append(("数码", None, "已上传", 0, 0))
        sheet.append((None, None, "未上传", 0, 0))
        sheet.append((None, None, "合计", 0, 0))
        _write_upload_detail_sheet(workbook, detail_rows)
        payment_workbook = Workbook()
        payment_sheet = payment_workbook.active
        payment_sheet.title = store_report.PAYMENT_SHEET_NAME
        payment_sheet.append(("财务大类", "品牌", "补贴金额合计", "补贴金额计数"))
        _write_payment_detail_sheet(payment_workbook, payment_rows)
        with TemporaryDirectory() as directory:
            upload_file = Path(directory) / "审核明细.xlsx"
            payment_file = Path(directory) / "回款明细.xlsx"
            workbook.save(upload_file)
            payment_workbook.save(payment_file)
            (
                _payment_data,
                _digital_amount,
                _payment_metrics,
                payment_index,
                ambiguous_refs,
            ) = store_report.load_payment_data(payment_file)
            return store_report.load_upload_data(
                upload_file, payment_index, ambiguous_refs
            )

    def test_corrects_kitchen_fotile_to_refrigerator_by_reference(self) -> None:
        amounts, _totals, _metrics, corrections, reviews = self._load(
            [
                _upload_detail_row(
                    document="ZHLT000259",
                    category="厨卫",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
            [
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
        )

        self.assertEqual(
            amounts,
            {
                ("冰箱", "方太"): {
                    "已上传": Decimal("1500"),
                    "未上传": Decimal("0"),
                }
            },
        )
        self.assertEqual(len(corrections), 1)
        correction = corrections[0]
        self.assertEqual(
            (
                correction.document_number,
                correction.reference,
                correction.brand,
                correction.original_category,
                correction.corrected_category,
                correction.raw_payment_category,
            ),
            ("ZHLT000259", self.FOTILE_REFERENCE, "方太", "厨卫", "冰箱", "A02-电冰箱"),
        )
        self.assertEqual(reviews, [])

    def test_unmatched_upload_keeps_original_category(self) -> None:
        amounts, _totals, _metrics, corrections, reviews = self._load(
            [
                _upload_detail_row(
                    category="厨卫",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
            [],
        )

        self.assertEqual(
            amounts,
            {("厨卫", "方太"): {"已上传": Decimal("1500"), "未上传": Decimal("0")}},
        )
        self.assertEqual(corrections, [])
        self.assertEqual(reviews, [])

    def test_blank_reference_keeps_original_category(self) -> None:
        amounts, _totals, _metrics, corrections, reviews = self._load(
            [
                _upload_detail_row(
                    category="厨卫",
                    brand="方太",
                    reference="",
                    subsidy=1500,
                )
            ],
            [
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
        )

        self.assertEqual(
            amounts,
            {("厨卫", "方太"): {"已上传": Decimal("1500"), "未上传": Decimal("0")}},
        )
        self.assertEqual(corrections, [])
        self.assertEqual(reviews, [])

    def test_brand_mismatch_is_not_corrected_and_flagged(self) -> None:
        amounts, _totals, _metrics, corrections, reviews = self._load(
            [
                _upload_detail_row(
                    category="厨卫",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
            [
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="海尔",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
        )

        self.assertEqual(
            amounts,
            {("厨卫", "方太"): {"已上传": Decimal("1500"), "未上传": Decimal("0")}},
        )
        self.assertEqual(corrections, [])
        self.assertEqual(len(reviews), 1)
        self.assertIn("品牌不一致", reviews[0][0])

    def test_amount_mismatch_is_not_corrected_and_flagged(self) -> None:
        amounts, _totals, _metrics, corrections, reviews = self._load(
            [
                _upload_detail_row(
                    category="厨卫",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
            [
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1499.99,
                )
            ],
        )

        self.assertEqual(
            amounts,
            {("厨卫", "方太"): {"已上传": Decimal("1500"), "未上传": Decimal("0")}},
        )
        self.assertEqual(corrections, [])
        self.assertEqual(len(reviews), 1)
        self.assertIn("补贴金额不一致", reviews[0][0])

    def test_system_category_difference_is_not_corrected_silently(self) -> None:
        """国产彩电 ↔ 电视 是两套口径的体系差异，不是错标：不纠正也不提示。"""
        amounts, _totals, _metrics, corrections, reviews = self._load(
            [
                _upload_detail_row(
                    category="国产彩电",
                    brand="海信",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=100,
                )
            ],
            [
                _payment_detail_row(
                    raw_category="A01-电视机",
                    brand="海信",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=100,
                )
            ],
        )

        self.assertEqual(
            amounts,
            {("国产彩电", "海信"): {"已上传": Decimal("100"), "未上传": Decimal("0")}},
        )
        self.assertEqual(corrections, [])
        self.assertEqual(reviews, [])

    def test_multiple_payment_records_flagged_when_category_differs(self) -> None:
        """同一参考号的正负冲正对（品类一致）→ 无法唯一匹配；审核侧品类不同
        时列入人工核对而不是猜测。"""
        amounts, _totals, _metrics, corrections, reviews = self._load(
            [
                _upload_detail_row(
                    category="厨卫",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
            [
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                ),
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=-1500,
                ),
            ],
        )

        self.assertEqual(
            amounts,
            {("厨卫", "方太"): {"已上传": Decimal("1500"), "未上传": Decimal("0")}},
        )
        self.assertEqual(corrections, [])
        self.assertEqual(len(reviews), 1)
        self.assertIn("多条回款记录", reviews[0][0])

    def test_multiple_distinct_categories_are_flagged_with_candidates(self) -> None:
        """同一参考号对应 A02-电冰箱、A04-空调 两个品类：无法唯一确定，
        必须保留原品类并列出全部候选，绝不自动纠正。"""
        amounts, _totals, _metrics, corrections, reviews = self._load(
            [
                _upload_detail_row(
                    category="厨卫",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
            [
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                ),
                _payment_detail_row(
                    raw_category="A04-空调",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                ),
            ],
        )

        self.assertEqual(
            amounts,
            {("厨卫", "方太"): {"已上传": Decimal("1500"), "未上传": Decimal("0")}},
        )
        self.assertEqual(corrections, [])
        self.assertEqual(len(reviews), 1)
        title, details = reviews[0]
        self.assertIn("参考号对应多个编码品类", title)
        self.assertIn("A02-电冰箱｜方太", details[1])
        self.assertIn("A04-空调｜方太", details[1])

    def test_multiple_payment_records_same_category_are_silent(self) -> None:
        """冲正对但审核侧品类与回款侧相同：无需纠正也无需核对。"""
        _amounts, _totals, _metrics, corrections, reviews = self._load(
            [
                _upload_detail_row(
                    category="冰箱",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
            [
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                ),
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=-1500,
                ),
            ],
        )

        self.assertEqual(corrections, [])
        self.assertEqual(reviews, [])

    def test_unmapped_payment_category_is_flagged(self) -> None:
        amounts, _totals, _metrics, corrections, reviews = self._load(
            [
                _upload_detail_row(
                    category="厨卫",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
            [
                _payment_detail_row(
                    raw_category="A99-未知品类",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
        )

        self.assertEqual(
            amounts,
            {("厨卫", "方太"): {"已上传": Decimal("1500"), "未上传": Decimal("0")}},
        )
        self.assertEqual(corrections, [])
        self.assertEqual(len(reviews), 1)
        self.assertIn("编码品类未配置", reviews[0][0])

    def test_category_map_mismatch_raises(self) -> None:
        """编码品类映射结果与明细财务大类不一致是口径漂移，直接报错不猜测。"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.PAYMENT_SHEET_NAME
        sheet.append(("财务大类", "品牌", "补贴金额合计", "补贴金额计数"))
        _write_payment_detail_sheet(
            workbook,
            [
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
        )
        detail_header = (
            store_report.PAYMENT_PROFILES["家电"].detail_headers
            + store_report.PAYMENT_DERIVED_HEADERS
        )
        sheet2 = workbook[store_report.PAYMENT_PROFILES["家电"].detail_sheet_name]
        sheet2.cell(row=2, column=detail_header.index("财务大类") + 1, value="厨卫")

        with TemporaryDirectory() as directory:
            payment_file = Path(directory) / "回款明细.xlsx"
            workbook.save(payment_file)
            with self.assertRaisesRegex(ValueError, "不一致"):
                store_report.load_payment_data(payment_file)

    def test_summary_brand_rows_must_match_detail_aggregation(self) -> None:
        """数据汇总品牌行与明细聚合不一致（如明细状态改名、行被跳过）→ 报错。"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.UPLOAD_SHEET_NAME
        sheet.append(store_report.UPLOAD_HEADER)
        sheet.append(("冰箱", "海尔", "已上传", 1, 100))
        sheet.append(("家电", None, "已上传", 0, 0))
        sheet.append((None, None, "未上传", 0, 0))
        sheet.append((None, None, "合计", 0, 0))
        sheet.append(("数码", None, "已上传", 0, 0))
        sheet.append((None, None, "未上传", 0, 0))
        sheet.append((None, None, "合计", 0, 0))
        # 明细里是 150，与数据汇总的 100 不一致。
        _write_upload_detail_sheet(
            workbook,
            [_upload_detail_row(subsidy=150, remark="已上传")],
        )

        with TemporaryDirectory() as directory:
            upload_file = Path(directory) / "审核明细.xlsx"
            workbook.save(upload_file)
            with self.assertRaisesRegex(
                ValueError, "品牌金额不一致，共 1 项.*冰箱/海尔/已上传：明细 150.00，汇总 100.00"
            ):
                store_report.load_upload_data(upload_file)

    def test_correction_moves_only_the_matching_row(self) -> None:
        """同一 (原品类, 品牌) 下部分行纠正、部分行保留，金额互不串扰。"""
        amounts, _totals, _metrics, corrections, reviews = self._load(
            [
                _upload_detail_row(
                    document="A1",
                    category="厨卫",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                ),
                _upload_detail_row(
                    document="A2",
                    category="厨卫",
                    brand="方太",
                    reference="",  # 无参考号 → 保留厨卫
                    subsidy=800,
                ),
            ],
            [
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="方太",
                    reference=self.FOTILE_REFERENCE,
                    subsidy=1500,
                )
            ],
        )

        self.assertEqual(
            amounts,
            {
                ("冰箱", "方太"): {
                    "已上传": Decimal("1500"),
                    "未上传": Decimal("0"),
                },
                ("厨卫", "方太"): {
                    "已上传": Decimal("800"),
                    "未上传": Decimal("0"),
                },
            },
        )
        self.assertEqual(len(corrections), 1)
        self.assertEqual(reviews, [])


class BrandGroupTests(unittest.TestCase):
    def test_haier_group_sums_across_categories(self) -> None:
        rule = next(rule for rule in store_report.BRAND_GROUP_RULES if rule.name == "海尔系")
        upload_data = {
            ("冰箱", "海尔"): {"已上传": Decimal("100"), "未上传": Decimal("20")},
            ("国产彩电", "卡萨帝"): {"已上传": Decimal("50"), "未上传": Decimal("0")},
        }
        payment_data = {
            ("冰箱", "海尔"): Decimal("60"),
            ("电视", "卡萨帝"): Decimal("30"),
        }

        occurred, paid = store_report.sum_brand_group(upload_data, payment_data, rule.categories)

        self.assertEqual(occurred, Decimal("170"))
        self.assertEqual(paid, Decimal("90"))

    def test_midea_group_sums_across_categories(self) -> None:
        rule = next(rule for rule in store_report.BRAND_GROUP_RULES if rule.name == "美的系")
        upload_data = {
            ("冰箱", "美的"): {"已上传": Decimal("40"), "未上传": Decimal("10")},
            ("空调", "美的"): {"已上传": Decimal("200"), "未上传": Decimal("50")},
        }
        payment_data = {
            ("冰箱", "美的"): Decimal("30"),
            ("空调", "美的"): Decimal("150"),
        }

        occurred, paid = store_report.sum_brand_group(upload_data, payment_data, rule.categories)

        self.assertEqual(occurred, Decimal("300"))
        self.assertEqual(paid, Decimal("180"))

    def test_upload_category_differs_from_payment_category_for_tv_brands(self) -> None:
        """Regression test: TV rows are 国产彩电 in 审核明细 but 电视 in 回款明细."""
        rule = next(rule for rule in store_report.BRAND_GROUP_RULES if rule.name == "创维")
        upload_data = {("国产彩电", "创维"): {"已上传": Decimal("70"), "未上传": Decimal("10")}}
        payment_data = {("电视", "创维"): Decimal("50")}

        occurred, paid = store_report.sum_brand_group(upload_data, payment_data, rule.categories)

        self.assertEqual(occurred, Decimal("80"))
        self.assertEqual(paid, Decimal("50"))

    def test_brand_group_writing_and_totals_with_distinct_amounts(self) -> None:
        """使用互不相等的发生额、上传额、回款额验证表 2 的 D/E/F 写入和第 47 行合计。"""
        sheet = Workbook().active
        upload_data = {
            ("冰箱", "海尔"): {"已上传": Decimal("100"), "未上传": Decimal("20")},
            ("国产彩电", "卡萨帝"): {"已上传": Decimal("50"), "未上传": Decimal("10")},
            ("空调", "美的"): {"已上传": Decimal("200"), "未上传": Decimal("30")},
            ("空调", "格力"): {"已上传": Decimal("80"), "未上传": Decimal("15")},
            ("冰箱", "西门子"): {"已上传": Decimal("70"), "未上传": Decimal("10")},
            ("国产彩电", "海信"): {"已上传": Decimal("90"), "未上传": Decimal("10")},
            ("国产彩电", "创维"): {"已上传": Decimal("40"), "未上传": Decimal("5")},
            ("空调", "TCL"): {"已上传": Decimal("60"), "未上传": Decimal("10")},
        }
        payment_data = {
            ("冰箱", "海尔"): Decimal("60"),
            ("电视", "卡萨帝"): Decimal("30"),
            ("空调", "美的"): Decimal("115"),
            ("空调", "格力"): Decimal("38"),
            ("冰箱", "西门子"): Decimal("32"),
            ("电视", "海信"): Decimal("50"),
            ("电视", "创维"): Decimal("18"),
            ("空调", "TCL"): Decimal("35"),
        }

        expected_cells: dict[str, object] = {}
        for rule in store_report.BRAND_GROUP_RULES:
            store_report.write_brand_group_row(
                sheet, rule, upload_data, payment_data, FONT, expected_cells
            )

        store_report.update_brand_group_totals(sheet, FONT, expected_cells)

        # 40: 海尔系 发生 180, 回款 90, 比例 0.5
        self.assertEqual(sheet["D40"].value, 180)
        self.assertEqual(sheet["E40"].value, 90)
        self.assertAlmostEqual(sheet["F40"].value, 0.5)

        # 41: 美的系 发生 230, 回款 115, 比例 0.5
        self.assertEqual(sheet["D41"].value, 230)
        self.assertEqual(sheet["E41"].value, 115)
        self.assertAlmostEqual(sheet["F41"].value, 0.5)

        # 42: 格力 发生 95, 回款 38, 比例 0.4
        self.assertEqual(sheet["D42"].value, 95)
        self.assertEqual(sheet["E42"].value, 38)
        self.assertAlmostEqual(sheet["F42"].value, 0.4)

        # 43: 博西 发生 80, 回款 32, 比例 0.4
        self.assertEqual(sheet["D43"].value, 80)
        self.assertEqual(sheet["E43"].value, 32)
        self.assertAlmostEqual(sheet["F43"].value, 0.4)

        # 44: 海信系 发生 100, 回款 50, 比例 0.5
        self.assertEqual(sheet["D44"].value, 100)
        self.assertEqual(sheet["E44"].value, 50)
        self.assertAlmostEqual(sheet["F44"].value, 0.5)

        # 45: 创维 发生 45, 回款 18, 比例 0.4
        self.assertEqual(sheet["D45"].value, 45)
        self.assertEqual(sheet["E45"].value, 18)
        self.assertAlmostEqual(sheet["F45"].value, 0.4)

        # 46: TCL 发生 70, 回款 35, 比例 0.5
        self.assertEqual(sheet["D46"].value, 70)
        self.assertEqual(sheet["E46"].value, 35)
        self.assertAlmostEqual(sheet["F46"].value, 0.5)

        # 47: 合计 D47=800, E47=378, F47=378/800=0.4725
        total_occurred = 180 + 230 + 95 + 80 + 100 + 45 + 70
        total_paid = 90 + 115 + 38 + 32 + 50 + 18 + 35
        self.assertEqual(sheet["D47"].value, total_occurred)
        self.assertEqual(sheet["E47"].value, total_paid)
        self.assertAlmostEqual(sheet["F47"].value, total_paid / total_occurred)


class TemplateValidationTests(unittest.TestCase):
    def test_literal_structure_contract_matches_production(self) -> None:
        """独立字面量契约必须与生产代码的 TEMPLATE_STRUCTURE_CELLS 完全一致。"""
        self.assertEqual(
            store_report.TEMPLATE_STRUCTURE_CELLS,
            LITERAL_STRUCTURE_CELLS,
        )

    def test_literal_header_prefix_matches_production(self) -> None:
        """A1 标题前缀独立契约必须与生产代码的 TEMPLATE_HEADER_PREFIX 完全一致。"""
        self.assertEqual(
            store_report.TEMPLATE_HEADER_PREFIX,
            LITERAL_TEMPLATE_HEADER_PREFIX,
        )

    def test_literal_row_rules_contract_matches_production(self) -> None:
        """明细行与品牌组完整规则契约必须与独立字面量契约完全全等。"""
        self.assertEqual(store_report.ROW_RULES, LITERAL_ROW_RULES)
        self.assertEqual(store_report.BRAND_GROUP_RULES, LITERAL_BRAND_GROUP_RULES)
        self.assertEqual(
            store_report.EXPECTED_MERGED_RANGES,
            LITERAL_EXPECTED_MERGED_RANGES,
        )

    def test_rejects_template_with_wrong_store_name(self) -> None:
        """A1 标题中门店名称错误（非益庄店）必须被拒绝。"""
        workbook = _build_minimal_template()
        workbook.active["A1"] = "表1                  2026年（其他店 ）门店国补上传及回款情况表_2026-08-15 18:06:27"
        with self.assertRaisesRegex(ValueError, "A1 标题前缀不符合预期"):
            store_report.validate_template(workbook)

    def test_rejects_template_with_wrong_header_prefix(self) -> None:
        """A1 标题前缀任意缺失或错误必须被拒绝。"""
        workbook = _build_minimal_template()
        workbook.active["A1"] = "2026年门店国补上传及回款情况表_2026-08-15 18:06:27"
        with self.assertRaisesRegex(ValueError, "A1 标题前缀不符合预期"):
            store_report.validate_template(workbook)

    def test_rejects_template_with_invalid_timestamp(self) -> None:
        """A1 标题时间戳为非法日期（如 2026-99-99 99:99:99）必须被拒绝。"""
        workbook = _build_minimal_template()
        workbook.active["A1"] = f"{store_report.TEMPLATE_HEADER_PREFIX}2026-99-99 99:99:99"
        with self.assertRaisesRegex(ValueError, "A1 标题时间戳无效"):
            store_report.validate_template(workbook)

    def test_rejects_template_with_non_standard_timestamp_formats(self) -> None:
        """非标准格式的时间戳（如单数月份/时间、多余空格、制表符、首尾空格、非法闰月日期）必须被严格拒绝。"""
        invalid_suffixes = [
            "2026-8-5 1:2:3",
            "2026-08-05  01:02:03",
            "2026-08-05\t01:02:03",
            " 2026-08-05 01:02:03",
            "2026-08-05 01:02:03 ",
            "2026-02-30 01:02:03",
        ]
        for suffix in invalid_suffixes:
            with self.subTest(suffix=suffix):
                workbook = _build_minimal_template()
                workbook.active["A1"] = f"{store_report.TEMPLATE_HEADER_PREFIX}{suffix}"
                with self.assertRaisesRegex(ValueError, "A1 标题时间戳无效"):
                    store_report.validate_template(workbook)

    def test_rejects_template_with_missing_timestamp(self) -> None:
        """A1 标题缺少时间戳必须被拒绝。"""
        workbook = _build_minimal_template()
        workbook.active["A1"] = store_report.TEMPLATE_HEADER_PREFIX
        with self.assertRaisesRegex(ValueError, "A1 标题时间戳无效"):
            store_report.validate_template(workbook)

    def test_all_brand_cells_mutation_rejected_by_template_validation(self) -> None:
        """参数化测试：C3:C32 与 C40:C46 中任何一个品牌标签被篡改都会被拒绝。"""
        brand_cells = [f"C{r}" for r in range(3, 33)] + [f"C{r}" for r in range(40, 47)]
        for cell in brand_cells:
            with self.subTest(cell=cell):
                workbook = _build_minimal_template()
                workbook.active[cell] = "错误品牌"
                with self.assertRaisesRegex(ValueError, cell):
                    store_report.validate_template(workbook)

    def test_accepts_a_correctly_structured_template(self) -> None:
        store_report.validate_template(_build_minimal_template())

    def test_rejects_a_blank_workbook(self) -> None:
        with self.assertRaises(ValueError):
            store_report.validate_template(Workbook())

    def test_rejects_a_legacy_template_without_version_marker(self) -> None:
        """旧模板（含手工修改版）没有版本标记，必须给出更换提示而不是
        静默产出报表。"""
        workbook = _build_minimal_template()
        workbook.active[store_report.TEMPLATE_VERSION_CELL] = None

        with self.assertRaisesRegex(ValueError, "请更换为新版模板"):
            store_report.validate_template(workbook)

    def test_rejects_a_legacy_template_with_old_row_layout(self) -> None:
        """旧版模板第 14 行是方太冰箱：若第 14 行被改成方太冰箱则结构校验拒绝（新版为海信）。"""
        workbook = _build_minimal_template()
        workbook.active["C14"] = "方太冰箱"

        with self.assertRaisesRegex(ValueError, "请更换为新版模板"):
            store_report.validate_template(workbook)

    def test_rejects_a_template_with_a_moved_label(self) -> None:
        workbook = _build_minimal_template()
        workbook.active["C40"] = "格力系"  # 海尔系 moved or renamed

        with self.assertRaisesRegex(ValueError, "C40"):
            store_report.validate_template(workbook)

    def test_rejects_a_template_missing_a_key_merge(self) -> None:
        workbook = _build_minimal_template()
        workbook.active.unmerge_cells("A3:A13")

        with self.assertRaisesRegex(ValueError, "合并区域"):
            store_report.validate_template(workbook)

    def test_rejects_a_template_with_broken_sequence_numbers(self) -> None:
        workbook = _build_minimal_template()
        workbook.active["B14"] = 99

        with self.assertRaisesRegex(ValueError, "序号"):
            store_report.validate_template(workbook)

    def test_rejects_a_template_with_visible_version_row(self) -> None:
        workbook = _build_minimal_template()
        workbook.active.row_dimensions[
            store_report.TEMPLATE_VERSION_ROW
        ].hidden = False

        with self.assertRaisesRegex(ValueError, "隐藏"):
            store_report.validate_template(workbook)

    def test_accepts_a_template_with_extra_merges(self) -> None:
        """多余合并允许（模板可能有其他区域合并），关键合并缺失才拒绝。"""
        workbook = _build_minimal_template()
        workbook.active.merge_cells("G36:H36")

        store_report.validate_template(workbook)

    def test_row_rules_cover_every_template_detail_row(self) -> None:
        """行号契约：ROW_RULES 必须恰好覆盖模板明细区 3..TOTAL_ROW，无跳空。"""
        rule_rows = {rule.row for rule in store_report.ROW_RULES}
        self.assertEqual(
            rule_rows,
            set(range(3, store_report.TOTAL_ROW)),
        )
        self.assertEqual(
            sorted(rule_rows),
            list(range(3, store_report.TOTAL_ROW)),
        )

    def test_brand_group_rules_cover_template_rows(self) -> None:
        group_rows = {rule.row for rule in store_report.BRAND_GROUP_RULES}
        self.assertEqual(
            group_rows,
            set(range(40, store_report.BRAND_GROUP_TOTAL_ROW)),
        )

    def test_layout_constants_match_structure_cells(self) -> None:
        """表 1 总计、表 2 合计、表 3 行号与模板关键标签互相锁定。"""
        self.assertEqual(store_report.TOTAL_ROW, 33)
        self.assertEqual(store_report.TEMPLATE_STRUCTURE_CELLS["A33"], "费用总计")
        self.assertEqual(store_report.BRAND_GROUP_TOTAL_ROW, 47)
        self.assertEqual(store_report.TEMPLATE_STRUCTURE_CELLS["C47"], "合计")
        self.assertEqual(
            store_report.TABLE3_PROJECT_ROWS,
            {"家电": 50, "数码": 51},
        )
        self.assertEqual(store_report.TABLE3_TOTAL_ROW, 52)
        self.assertEqual(store_report.TEMPLATE_STRUCTURE_CELLS["C50"], "家电")
        self.assertEqual(store_report.TEMPLATE_STRUCTURE_CELLS["C51"], "数码")
        self.assertEqual(store_report.TEMPLATE_STRUCTURE_CELLS["C52"], "合计")

    def test_fotile_rule_covers_both_categories(self) -> None:
        """方太规则在第 27 行同时覆盖厨卫与冰箱，两侧品类一致。"""
        fotile = next(rule for rule in store_report.ROW_RULES if rule.row == 27)
        self.assertEqual(
            fotile,
            store_report.RowRule(27, ("厨卫", "冰箱"), ("方太",), ("厨卫", "冰箱"), ("方太",)),
        )
        self.assertEqual(store_report.TEMPLATE_STRUCTURE_CELLS["C27"], "方太")
        # 白名单必须覆盖纠正目标：冰箱在审核侧有行规则，电视（对应国产彩电）不在。
        self.assertIn("冰箱", store_report.UPLOAD_CATEGORIES)
        self.assertIn("厨卫", store_report.UPLOAD_CATEGORIES)
        self.assertNotIn("电视", store_report.UPLOAD_CATEGORIES)

    def test_summary_header_compatibility_with_store_report(self) -> None:
        """审核明细的 SUMMARY_HEADER 扩展为 8 列后，前 5 列仍与 SUMMARY_CORE_HEADER 完全一致。"""
        from processors.coupon_report import SUMMARY_CORE_HEADER, SUMMARY_HEADER

        self.assertEqual(SUMMARY_HEADER[:5], SUMMARY_CORE_HEADER)
        self.assertEqual(store_report.UPLOAD_HEADER, SUMMARY_CORE_HEADER)

    def test_rejects_extra_sheets(self) -> None:
        workbook = _build_minimal_template()
        workbook.create_sheet("附表")

        with self.assertRaisesRegex(ValueError, "工作表数量"):
            store_report.validate_template(workbook)

    def test_rejects_wrong_column_count(self) -> None:
        workbook = _build_minimal_template()
        workbook.active["I1"] = "多出来的列"

        with self.assertRaisesRegex(ValueError, "列数"):
            store_report.validate_template(workbook)

    def test_rejects_too_few_rows(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.EXPECTED_SHEET_TITLE
        for column_index in range(1, store_report.EXPECTED_COLUMN_COUNT + 1):
            sheet.cell(row=1, column=column_index, value="")
        for coordinate, value in store_report.TEMPLATE_STRUCTURE_CELLS.items():
            if int("".join(filter(str.isdigit, coordinate))) <= 33:
                sheet[coordinate] = value

        with self.assertRaisesRegex(ValueError, "行数"):
            store_report.validate_template(workbook)

    def test_rejects_the_wrong_sheet_title(self) -> None:
        workbook = _build_minimal_template()
        workbook.active.title = "其他门店"

        with self.assertRaisesRegex(ValueError, "工作表名称"):
            store_report.validate_template(workbook)


class TemplateResolutionTests(unittest.TestCase):
    def test_resolves_the_sole_matching_template(self) -> None:
        with TemporaryDirectory() as directory:
            data_dir = Path(directory)
            template = data_dir / "2026年门店国补上传及回款情况表 （益庄店）.xlsx"
            Workbook().save(template)
            store_report.configure_data_dir(data_dir)

            self.assertEqual(store_report.resolve_template_file(), template)

    def test_raises_when_no_template_is_found(self) -> None:
        with TemporaryDirectory() as directory:
            store_report.configure_data_dir(Path(directory))

            with self.assertRaises(FileNotFoundError):
                store_report.resolve_template_file()


class ValidateOutputTests(unittest.TestCase):
    def test_accepts_a_workbook_matching_its_expected_cells_and_totals(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "益庄"
        _fill_columns(sheet)
        sheet["A1"] = "表1                  2026年（益庄店 ）门店国补上传及回款情况表_2026-07-28 00:00:00"
        sheet["D3"] = 100
        sheet["D33"] = 100
        sheet["D40"] = 100
        sheet["D47"] = 100
        sheet["D50"] = 1
        sheet["D51"] = 2
        sheet["D52"] = 3
        sheet["E50"] = 1
        sheet["E51"] = 2
        sheet["E52"] = 3

        with TemporaryDirectory() as directory:
            path = Path(directory) / "report.xlsx"
            workbook.save(path)
            store_report.validate_output(path, {"D3": 100.0}, "益庄")

    def test_rejects_a_workbook_whose_total_disagrees_with_its_detail_rows(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "益庄"
        _fill_columns(sheet)
        sheet["A1"] = "表1                  2026年（益庄店 ）门店国补上传及回款情况表_2026-07-28 00:00:00"
        sheet["D3"] = 100
        sheet["D33"] = 999

        with TemporaryDirectory() as directory:
            path = Path(directory) / "report.xlsx"
            workbook.save(path)
            with self.assertRaises(ValueError):
                store_report.validate_output(path, {}, "益庄")

    def test_rejects_a_workbook_whose_written_cell_does_not_match_expectation(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "益庄"
        _fill_columns(sheet)
        sheet["A1"] = "表1                  2026年（益庄店 ）门店国补上传及回款情况表_2026-07-28 00:00:00"
        sheet["D3"] = 100

        with TemporaryDirectory() as directory:
            path = Path(directory) / "report.xlsx"
            workbook.save(path)
            with self.assertRaisesRegex(ValueError, "D3"):
                store_report.validate_output(path, {"D3": 999.0}, "益庄")

    def test_rejects_a_workbook_missing_the_update_timestamp(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "益庄"
        _fill_columns(sheet)
        sheet["A1"] = "无时间戳标题"

        with TemporaryDirectory() as directory:
            path = Path(directory) / "report.xlsx"
            workbook.save(path)
            with self.assertRaisesRegex(ValueError, "A1 标题前缀不符合预期"):
                store_report.validate_output(path, {}, "益庄")

    def test_rejects_a_workbook_with_invalid_timestamp(self) -> None:
        invalid_suffixes = [
            "2026-99-99 99:99:99",
            "2026-8-5 1:2:3",
            "2026-08-05  01:02:03",
            "2026-08-05\t01:02:03",
            " 2026-08-05 01:02:03",
            "2026-08-05 01:02:03 ",
            "2026-02-30 01:02:03",
        ]
        for suffix in invalid_suffixes:
            with self.subTest(suffix=suffix):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "益庄"
                _fill_columns(sheet)
                sheet["A1"] = f"{store_report.TEMPLATE_HEADER_PREFIX}{suffix}"

                with TemporaryDirectory() as directory:
                    path = Path(directory) / "report.xlsx"
                    workbook.save(path)
                    with self.assertRaisesRegex(ValueError, "A1 标题时间戳无效"):
                        store_report.validate_output(path, {}, "益庄")

    def test_rejects_a_workbook_with_the_wrong_sheet_title(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "其他门店"
        _fill_columns(sheet)
        sheet["A1"] = "表1                  2026年（益庄店 ）门店国补上传及回款情况表_2026-07-28 00:00:00"

        with TemporaryDirectory() as directory:
            path = Path(directory) / "report.xlsx"
            workbook.save(path)
            with self.assertRaisesRegex(ValueError, "工作表名称"):
                store_report.validate_output(path, {}, "益庄")


class UpdateHeaderTests(unittest.TestCase):
    def test_update_header_writes_exact_title_and_updates_expected_cells(self) -> None:
        """update_header 传入固定 datetime，精确断言 A1 严格等于 TEMPLATE_HEADER_PREFIX + 格式化时间戳。"""
        workbook = _build_minimal_template()
        sheet = workbook.active
        fixed_time = datetime(2026, 8, 16, 10, 20, 30, tzinfo=store_report.SHANGHAI_TZ)
        expected_cells: dict[str, object] = {}

        store_report.update_header(sheet, fixed_time, expected_cells)

        expected_title = f"{store_report.TEMPLATE_HEADER_PREFIX}2026-08-16 10:20:30"
        self.assertEqual(sheet["A1"].value, expected_title)
        self.assertEqual(expected_cells["A1"], expected_title)


class ProcessStoreReportIntegrationTests(unittest.TestCase):
    def _write_upload_file(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.UPLOAD_SHEET_NAME
        sheet.append(store_report.UPLOAD_HEADER)
        # 数据汇总的品牌行是明细聚合的守恒校验基准。
        sheet.append(("冰箱", "海尔", "已上传", 1, 100))
        sheet.append((None, None, "未上传", 1, 20))
        sheet.append(("厨卫", "方太", "已上传", 1, 1500))
        sheet.append(("家电", None, "已上传", 2, 1600))
        sheet.append((None, None, "未上传", 1, 20))
        sheet.append((None, None, "合计", 3, 1620))
        sheet.append(("数码", None, "已上传", 1, 40))
        sheet.append((None, None, "未上传", 1, 10))
        sheet.append((None, None, "合计", 2, 50))
        # 品牌行的权威来源是家电-明细总表（行级纠正）；其中方太 BCD 在
        # 审核侧被标成厨卫，集成链路里靠回款明细的参考号纠正为冰箱。
        _write_upload_detail_sheet(
            workbook,
            [
                _upload_detail_row(
                    document="ZHLT000001",
                    category="冰箱",
                    brand="海尔",
                    reference="12345678901N",
                    subsidy=100,
                    remark="已上传",
                ),
                _upload_detail_row(
                    document="ZHLT000002",
                    category="冰箱",
                    brand="海尔",
                    reference="12345678902N",
                    subsidy=20,
                    remark="未上传",
                ),
                _upload_detail_row(
                    document="ZHLT000259",
                    category="厨卫",
                    brand="方太",
                    reference="17914133741N",
                    subsidy=1500,
                    remark="已上传",
                ),
            ],
        )
        workbook.save(path)

    def _write_payment_file(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = store_report.PAYMENT_SHEET_NAME
        sheet.append(("财务大类", "品牌", "补贴金额合计", "补贴金额计数"))
        sheet.append(("冰箱", "海尔", 60, 1))
        sheet.append(("冰箱", "方太", 1500, 1))
        sheet.append(("合计", None, 1560, 2))
        sheet.append((None, None, None, None))
        sheet.append(("手机", "OPPO", 15, 1))
        _write_payment_detail_sheet(
            workbook,
            [
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="海尔",
                    reference="12345678901N",
                    subsidy=60,
                ),
                _payment_detail_row(
                    raw_category="A02-电冰箱",
                    brand="方太",
                    reference="17914133741N",
                    subsidy=1500,
                ),
            ],
        )
        workbook.save(path)

    def test_runs_end_to_end_and_produces_a_validated_report(self) -> None:
        with TemporaryDirectory() as directory:
            base = Path(directory)
            data_dir = base / "data"
            data_dir.mkdir()
            upload_file = base / "审核明细.xlsx"
            payment_file = base / "回款明细.xlsx"
            output_file = base / "output.xlsx"
            template_file = data_dir / "2026年门店国补上传及回款情况表 （益庄店）.xlsx"

            self._write_upload_file(upload_file)
            self._write_payment_file(payment_file)
            _build_minimal_template().save(template_file)
            store_report.configure_data_dir(data_dir)

            reporter_stream = io.StringIO()
            reporter = store_report.ConsoleReporter(stream=reporter_stream)
            with (
                patch.object(store_report, "UPLOAD_FILE", upload_file),
                patch.object(store_report, "PAYMENT_FILE", payment_file),
                patch.object(store_report, "OUTPUT_FILE", output_file),
                patch.object(
                    store_report,
                    "resolve_font",
                    return_value=("Maple Mono NF CN", Path("unused.ttf")),
                ),
            ):
                store_report.process_store_report(reporter)
            reporter.finish(success=True, succeeded=1, total=1)
            reporter_text = reporter_stream.getvalue()

            self.assertTrue(output_file.exists())
            result = load_workbook(output_file, data_only=True)
            sheet = result[result.sheetnames[0]]
            self.assertEqual(sheet["D3"].value, 120)
            self.assertEqual(sheet["E3"].value, 100)
            self.assertEqual(sheet["G3"].value, 60)
            # 审核侧厨卫/方太经参考号纠正为冰箱，并入第 27 行（厨卫/方太）。
            self.assertEqual(sheet["D27"].value, 1500)
            self.assertEqual(sheet["E27"].value, 1500)
            self.assertEqual(sheet["G27"].value, 1500)
            self.assertEqual(sheet["H27"].value, 1)
            self.assertEqual(sheet["D32"].value, 50)
            self.assertEqual(sheet["E32"].value, 40)
            self.assertEqual(sheet["G32"].value, 15)
            self.assertEqual(sheet["D33"].value, 1670)
            self.assertEqual(sheet["E33"].value, 1640)
            self.assertEqual(sheet["G33"].value, 1575)
            # 表 2 品牌汇总：海尔系 发生 120, 回款 60, 回款率 0.5；其余组为空；第 47 行合计。
            self.assertEqual(sheet["D40"].value, 120)
            self.assertEqual(sheet["E40"].value, 60)
            self.assertAlmostEqual(sheet["F40"].value, 0.5)
            for row in range(41, 47):
                self.assertIsNone(sheet[f"D{row}"].value)
                self.assertIsNone(sheet[f"E{row}"].value)
                self.assertIsNone(sheet[f"F{row}"].value)
            self.assertEqual(sheet["D47"].value, 120)
            self.assertEqual(sheet["E47"].value, 60)
            self.assertAlmostEqual(sheet["F47"].value, 0.5)
            self.assertIsNone(sheet["D50"].value)
            self.assertEqual(sheet["E50"].value, 1)
            self.assertIsNone(sheet["D51"].value)
            self.assertEqual(sheet["E51"].value, 1)
            self.assertIsNone(sheet["D52"].value)
            self.assertEqual(sheet["E52"].value, 2)
            ts_str = str(sheet["A1"].value)[len(store_report.TEMPLATE_HEADER_PREFIX):]
            parsed = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
            self.assertEqual(
                sheet["A1"].value,
                f"{store_report.TEMPLATE_HEADER_PREFIX}{parsed.strftime('%Y-%m-%d %H:%M:%S')}",
            )
            self.assertIn("审核明细品类纠正：1 条", reporter_text)
            # 合并区域内部的 MergedCell 不参与渲染，openpyxl 保存时会丢弃
            # 对其样式的写入（保持模板原样），所以只断言普通单元格的字体。
            output_fonts = {
                cell.font.name
                for row in sheet.iter_rows()
                for cell in row
                if not isinstance(cell, MergedCell)
            }
            self.assertEqual(output_fonts, {"Maple Mono NF CN"})
            leftover = [entry.name for entry in output_file.parent.iterdir() if entry.name.startswith(".")]
            self.assertEqual(leftover, [])


class RollbackTests(unittest.TestCase):
    def test_prior_report_is_restored_when_store_report_fails(self) -> None:
        with TemporaryDirectory() as directory:
            output_file = Path(directory) / "report.xlsx"
            other_output = Path(directory) / "other.xlsx"
            output_file.write_bytes(b"old report")
            other_output.write_bytes(b"old other")

            def fail_after_writing_other() -> None:
                other_output.write_bytes(b"new other")
                output_file.write_bytes(b"corrupted partial report")
                raise ValueError("门店报表处理失败")

            with self.assertRaisesRegex(ValueError, "门店报表处理失败"):
                run_with_output_rollback((other_output, output_file), fail_after_writing_other)

            self.assertEqual(output_file.read_bytes(), b"old report")
            self.assertEqual(other_output.read_bytes(), b"old other")

    def test_no_half_written_output_remains_when_nothing_existed_before(self) -> None:
        with TemporaryDirectory() as directory:
            output_dir = Path(directory)
            output_file = output_dir / "report.xlsx"

            def fail_after_creating_output() -> None:
                output_file.write_bytes(b"half-written report")
                raise ValueError("门店报表处理失败")

            with self.assertRaisesRegex(ValueError, "门店报表处理失败"):
                run_with_output_rollback((output_file,), fail_after_creating_output)

            self.assertFalse(output_file.exists())
            self.assertEqual(list(output_dir.iterdir()), [])


if __name__ == "__main__":
    unittest.main()
