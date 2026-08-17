import io
import signal
import sys
import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook

import main as app_main
from processors import coupon_report, store_report
from processors.common.console import ConsoleReporter
from processors.common.excel import OutputCleanupError, StaleFileCleanup
from processors.coupons import appliance as coupon_appliance
from processors.coupons import digital as coupons_digital
from tests.test_store_report import (
    LITERAL_EXPECTED_MERGED_RANGES,
    LITERAL_STRUCTURE_CELLS,
)


class RequireLinuxTest(unittest.TestCase):
    def test_accepts_linux(self) -> None:
        with patch.object(sys, "platform", "linux"):
            app_main.require_linux()

    def test_rejects_windows(self) -> None:
        with patch.object(sys, "platform", "win32"):
            with self.assertRaises(SystemExit) as raised:
                app_main.require_linux()
        self.assertEqual(raised.exception.code, 2)

    def test_rejects_macos(self) -> None:
        with patch.object(sys, "platform", "darwin"):
            with self.assertRaises(SystemExit) as raised:
                app_main.require_linux()
        self.assertEqual(raised.exception.code, 2)


class ParseCliArgsTest(unittest.TestCase):
    def test_empty_argv_is_interactive(self) -> None:
        args = app_main.parse_cli_args([])
        self.assertFalse(args.all)
        self.assertIsNone(args.mode)

    def test_all_flag(self) -> None:
        args = app_main.parse_cli_args(["--all"])
        self.assertTrue(args.all)
        self.assertIsNone(args.mode)

    def test_mode_flag(self) -> None:
        args = app_main.parse_cli_args(["--mode", "3"])
        self.assertFalse(args.all)
        self.assertEqual(args.mode, 3)

    def test_all_and_mode_are_mutually_exclusive(self) -> None:
        with self.assertRaises(SystemExit) as raised:
            app_main.parse_cli_args(["--all", "--mode", "1"])
        self.assertEqual(raised.exception.code, 2)


class SelectionHelpersTest(unittest.TestCase):
    def _processors(self):
        return (
            ("已上传数据（家电+数码）", "已上传数据", lambda r: None),
            ("收款单统计", "收款单统计", lambda r: None),
        )

    def test_selection_for_mode_in_range(self) -> None:
        selection = app_main.selection_for_mode(self._processors(), 2)
        self.assertFalse(selection.is_all)
        self.assertEqual(selection.step_label, "收款单统计")

    def test_selection_for_mode_out_of_range(self) -> None:
        with self.assertRaises(SystemExit) as raised:
            app_main.selection_for_mode(self._processors(), 6)
        self.assertEqual(raised.exception.code, 2)

    def test_selection_for_all(self) -> None:
        selection = app_main.selection_for_all(self._processors())
        self.assertTrue(selection.is_all)
        self.assertEqual(selection.step_label, "全部模式")


class InstanceLockTest(unittest.TestCase):
    def test_second_acquire_fails_with_exit_3(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            lock_path = Path(directory) / "test.lock"
            first = app_main.acquire_instance_lock(lock_path)
            try:
                with self.assertRaises(SystemExit) as raised:
                    app_main.acquire_instance_lock(lock_path)
                self.assertEqual(raised.exception.code, 3)
            finally:
                first.close()

    def test_lock_released_after_close_allows_reacquire(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            lock_path = Path(directory) / "test.lock"
            first = app_main.acquire_instance_lock(lock_path)
            first.close()
            second = app_main.acquire_instance_lock(lock_path)
            try:
                self.assertIsNotNone(second)
            finally:
                second.close()


class SigtermHandlerTest(unittest.TestCase):
    def test_sigterm_handler_raises_keyboard_interrupt(self) -> None:
        with self.assertRaises(KeyboardInterrupt):
            app_main._sigterm_handler(signal.SIGTERM, None)

    def test_install_registers_handler(self) -> None:
        previous = signal.getsignal(signal.SIGTERM)
        try:
            app_main.install_sigterm_handler()
            self.assertIs(signal.getsignal(signal.SIGTERM), app_main._sigterm_handler)
        finally:
            signal.signal(signal.SIGTERM, previous)


class CombinedOutputRollbackTest(unittest.TestCase):
    def test_submitted_outputs_are_restored_when_second_project_fails(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory)
            appliance_output = output_dir / "家电_已上传.xlsx"
            digital_output = output_dir / "数码_已上传.xlsx"
            appliance_output.write_bytes(b"old appliance")
            digital_output.write_bytes(b"old digital")

            def fake_process(profile_name: str, reporter) -> None:
                if profile_name == "家电":
                    appliance_output.write_bytes(b"new appliance")
                else:
                    digital_output.write_bytes(b"new digital")
                    raise ValueError("digital failed")

            with (
                patch.object(
                    app_main.submitted,
                    "OUTPUT_FILES",
                    (appliance_output, digital_output),
                ),
                patch.object(
                    app_main.submitted,
                    "process_submitted_files",
                    fake_process,
                ),
            ):
                with self.assertRaisesRegex(ValueError, "digital failed"):
                    app_main.submitted.process_all(
                        ConsoleReporter(stream=io.StringIO())
                    )

            self.assertEqual(appliance_output.read_bytes(), b"old appliance")
            self.assertEqual(digital_output.read_bytes(), b"old digital")


class AllOutputFilesTest(unittest.TestCase):
    def test_includes_the_store_report_output(self) -> None:
        self.assertIn(app_main.store_report.OUTPUT_FILE, app_main.all_output_files())


class ProcessorOrderTest(unittest.TestCase):
    def test_store_report_runs_after_its_two_upstream_processing_modes(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            data_dir = Path(directory)
            app_main.submitted.configure_data_dir(data_dir)
            app_main.receipts.configure_data_dir(data_dir)
            app_main.coupon_sources.configure_data_dir(data_dir)
            app_main.payment.configure_data_dir(data_dir)
            app_main.store_report.configure_data_dir(data_dir)

            labels = [label for label, _, _ in app_main.build_processors()]
            coupon_report_index = labels.index("审核明细（销售用券情况统计）")
            store_report_index = labels.index("门店国补上传及回款情况表")

            self.assertLess(labels.index("回款明细（家电+数码）"), coupon_report_index)
            self.assertLess(coupon_report_index, store_report_index)


class AllModeRollbackTest(unittest.TestCase):
    def test_every_real_output_target_is_restored_when_store_report_fails_last(self) -> None:
        """Exercises the actual all_output_files() rollback set — not just the two
        files a given step happens to touch — the way `all` mode really runs."""
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory)
            paths = {
                "large_appliances": output_dir / "家电_已上传.xlsx",
                "digital": output_dir / "数码_已上传.xlsx",
                "receipts": output_dir / "收款单统计.xlsx",
                "coupon_report": output_dir / "审核明细.xlsx",
                "payment": output_dir / "回款明细.xlsx",
                "store_report": output_dir / "门店报表.xlsx",
            }
            for key, path in paths.items():
                path.write_bytes(f"old {key}".encode())

            def write_ok(key: str):
                return lambda reporter: paths[key].write_bytes(
                    f"new {key}".encode()
                )

            def fail_store_report(reporter) -> None:
                paths["store_report"].write_bytes(b"corrupted partial store report")
                raise ValueError("store report failed")

            processors = (
                ("已上传数据（家电+数码）", "已上传数据", write_ok("large_appliances")),
                ("审核明细（销售用券情况统计）", "审核明细", write_ok("coupon_report")),
                ("回款明细（家电+数码）", "回款明细", write_ok("payment")),
                ("门店国补上传及回款情况表", "门店报表", fail_store_report),
            )

            with (
                patch.object(
                    app_main.submitted,
                    "OUTPUT_FILES",
                    (paths["large_appliances"], paths["digital"]),
                ),
                patch.object(app_main.receipts, "OUTPUT_FILE", paths["receipts"]),
                patch.object(coupon_report, "OUTPUT_FILE", paths["coupon_report"]),
                patch.object(app_main.payment, "OUTPUT_FILE", paths["payment"]),
                patch.object(app_main.store_report, "OUTPUT_FILE", paths["store_report"]),
            ):
                with self.assertRaisesRegex(ValueError, "store report failed"):
                    app_main.process_all(
                        processors,
                        ConsoleReporter(stream=io.StringIO()),
                    )

            for key, path in paths.items():
                self.assertEqual(path.read_bytes(), f"old {key}".encode())


class MainErrorHandlingTest(unittest.TestCase):
    def test_configuration_failure_uses_normal_error_report(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            data_dir = Path(directory)
            with (
                patch.object(app_main, "resolve_data_dir", return_value=data_dir),
                patch.object(
                    app_main.coupon_sources,
                    "configure_data_dir",
                    side_effect=ValueError("bad config"),
                ),
                patch.object(
                    app_main.ConsoleReporter, "error"
                ) as error,
            ):
                # argv=[] avoids picking up pytest's own flags.
                self.assertEqual(app_main.main([]), 1)

            error.assert_called_once()
            self.assertEqual(str(error.call_args.args[1]), "bad config")

    def test_cli_all_skips_interactive_menu(self) -> None:
        ran = {"all": False}

        class AllRun:
            is_all = True
            step_label = "全部模式"

            def run(self, reporter) -> None:
                ran["all"] = True

        with (
            patch.object(app_main, "resolve_data_dir", return_value=Path("data")),
            patch.object(
                app_main,
                "remove_stale_temporary_files",
                return_value=StaleFileCleanup((), ()),
            ),
            patch.object(app_main, "resolve_selection", return_value=AllRun()),
            patch.object(app_main, "acquire_instance_lock", return_value=MagicMock()),
            patch.object(app_main, "choose_data_processor") as choose,
        ):
            self.assertEqual(app_main.main(["--all"]), 0)
        choose.assert_not_called()
        self.assertTrue(ran["all"])


class LockBeforeCleanupTest(unittest.TestCase):
    def test_lock_failure_skips_stale_file_cleanup(self) -> None:
        class Selection:
            is_all = False
            step_label = "测试模式"

            def run(self, reporter) -> None:
                raise AssertionError("锁失败后不应开始处理")

        with (
            patch.object(app_main, "resolve_data_dir", return_value=Path("data")),
            patch.object(app_main.submitted, "configure_data_dir"),
            patch.object(app_main.receipts, "configure_data_dir"),
            patch.object(app_main.coupon_sources, "configure_data_dir"),
            patch.object(app_main.payment, "configure_data_dir"),
            patch.object(app_main.store_report, "configure_data_dir"),
            patch.object(
                app_main,
                "resolve_selection",
                return_value=Selection(),
            ),
            patch.object(
                app_main,
                "acquire_instance_lock",
                side_effect=SystemExit(3),
            ),
            patch.object(app_main, "remove_stale_temporary_files") as cleanup,
        ):
            with self.assertRaises(SystemExit) as raised:
                app_main.main(["--mode", "1"])

        self.assertEqual(raised.exception.code, 3)
        cleanup.assert_not_called()


class TransactionLifecycleTest(unittest.TestCase):
    """process_all's console lifecycle: success only after commit, failures
    and cancellations flush concerns, and a post-commit cleanup failure is
    never reported as a rollback."""

    def _processors(self, *behaviors):
        return tuple(
            (f"模式{index}", f"步骤{index}", behavior)
            for index, behavior in enumerate(behaviors, start=1)
        )

    def test_batch_success_finishes_after_commit(self) -> None:
        reporter = ConsoleReporter(
            stream=io.StringIO(),
            error_stream=io.StringIO(),
        )
        processors = self._processors(
            lambda r: None,
            lambda r: None,
        )
        with patch.object(
            app_main,
            "run_with_output_rollback",
            side_effect=lambda paths, operation: operation(),
        ):
            app_main.process_all(processors, reporter)
        text = reporter.stream.getvalue()
        self.assertIn("处理完成：2/2 步骤成功", text)
        self.assertIn("已提交输出", text)
        self.assertEqual(reporter.error_stream.getvalue(), "")

    def test_batch_failure_reports_rollback_and_no_output_list(self) -> None:
        reporter = ConsoleReporter(
            stream=io.StringIO(),
            error_stream=io.StringIO(),
        )
        processors = self._processors(
            lambda r: None,
            lambda r: (_ for _ in ()).throw(ValueError("坏了")),
        )
        with patch.object(
            app_main,
            "run_with_output_rollback",
            side_effect=lambda paths, operation: operation(),
        ):
            with self.assertRaisesRegex(ValueError, "坏了"):
                app_main.process_all(processors, reporter)
        self.assertIn("处理失败：1/2 步骤成功", reporter.error_stream.getvalue())
        self.assertIn(
            "本次输出已回滚，原有输出文件保持不变",
            reporter.error_stream.getvalue(),
        )
        self.assertNotIn("已提交输出", reporter.stream.getvalue())
        self.assertNotIn("处理完成", reporter.stream.getvalue())

    def test_batch_cancel_reports_cancelled_summary(self) -> None:
        reporter = ConsoleReporter(
            stream=io.StringIO(),
            error_stream=io.StringIO(),
        )
        processors = self._processors(lambda r: (_ for _ in ()).throw(KeyboardInterrupt()))
        with patch.object(
            app_main,
            "run_with_output_rollback",
            side_effect=lambda paths, operation: operation(),
        ):
            with self.assertRaises(KeyboardInterrupt):
                app_main.process_all(processors, reporter)
        self.assertIn("处理已取消：0/1 步骤已完成", reporter.error_stream.getvalue())
        self.assertIn(
            "本次输出已回滚，原有输出文件保持不变",
            reporter.error_stream.getvalue(),
        )

    def test_commit_cleanup_failure_is_not_reported_as_rollback(self) -> None:
        """Every step committed, then backup cleanup raised: the summary must
        say the outputs were committed, not rolled back."""
        reporter = ConsoleReporter(
            stream=io.StringIO(),
            error_stream=io.StringIO(),
        )
        processors = self._processors(lambda r: None)

        def cleanup_fails(paths, operation):
            operation()  # steps all succeed
            raise OutputCleanupError("输出已提交，备份清理失败")

        with patch.object(
            app_main,
            "run_with_output_rollback",
            side_effect=cleanup_fails,
        ):
            with self.assertRaisesRegex(OutputCleanupError, "备份清理失败"):
                app_main.process_all(processors, reporter)
        text = reporter.error_stream.getvalue()
        self.assertIn("输出已提交，未回滚", text)
        self.assertIn("备份清理失败", text)
        self.assertNotIn("本次输出已回滚", text)
        # No success summary on the main stream either.
        self.assertNotIn("已提交输出", reporter.stream.getvalue())
        self.assertNotIn("处理完成：", reporter.stream.getvalue())

    def test_nested_cleanup_failure_in_all_mode_reports_rollback(self) -> None:
        """Mode 1 has its own inner rollback transaction. Its post-commit
        cleanup failure reaches process_all as OutputCleanupError, but the
        outer transaction rolled every output back — the summary must say
        rolled back, with no second [失败] and no committed claim."""
        reporter = ConsoleReporter(
            stream=io.StringIO(),
            error_stream=io.StringIO(),
        )

        def inner_cleanup_fails(reporter) -> None:
            raise OutputCleanupError("内层清理失败")

        processors = self._processors(inner_cleanup_fails)
        with patch.object(
            app_main,
            "run_with_output_rollback",
            side_effect=lambda paths, operation: operation(),
        ):
            with self.assertRaisesRegex(OutputCleanupError, "内层清理失败"):
                app_main.process_all(processors, reporter)
        text = reporter.error_stream.getvalue()
        # Only the step-level [失败] (with the rollback remedy) prints.
        self.assertEqual(text.count("[失败]"), 1)
        self.assertIn("处理失败：0/1 步骤成功", text)
        self.assertIn("本次输出已回滚，原有输出文件保持不变", text)
        self.assertNotIn("输出已提交，未回滚", text)
        self.assertNotIn("已提交输出", reporter.stream.getvalue())

    def test_single_mode_cleanup_failure_does_not_claim_rollback(self) -> None:
        """Mode 1's own rollback transaction can hit a post-commit cleanup
        failure; the single-mode handler must not claim the files were
        preserved or rolled back."""

        class CleanupFailRun:
            is_all = False
            step_label = "已上传数据"

            def run(self, reporter) -> None:
                raise OutputCleanupError("输出已提交，备份清理失败")

        reporter = ConsoleReporter(
            stream=io.StringIO(),
            error_stream=io.StringIO(),
        )
        with (
            patch.object(app_main, "ConsoleReporter", return_value=reporter),
            patch.object(app_main, "resolve_data_dir", return_value=Path("data")),
            patch.object(
                app_main,
                "choose_data_processor",
                return_value=CleanupFailRun(),
            ),
            patch.object(
                app_main,
                "remove_stale_temporary_files",
                return_value=StaleFileCleanup((), ()),
            ),
            patch.object(app_main, "acquire_instance_lock", return_value=MagicMock()),
        ):
            self.assertEqual(app_main.main([]), 1)
        text = reporter.error_stream.getvalue()
        self.assertIn("输出已提交，未回滚", text)
        self.assertIn("备份清理失败", text)
        self.assertNotIn("本次输出已回滚", text)
        self.assertNotIn("现有输出文件保持不变", text)


class UploadDataDebugTest(unittest.TestCase):
    def test_debug_enabled_when_value_is_one(self) -> None:
        def failing_processor(reporter):
            raise ValueError("boom")

        processors = [("1", "失败步骤", failing_processor)]
        reporter = ConsoleReporter(
            stream=io.StringIO(), error_stream=io.StringIO()
        )

        with patch.dict("os.environ", {"UPLOAD_DATA_DEBUG": "1"}):
            with patch("traceback.print_exc") as mock_print:
                with patch.object(
                    app_main,
                    "run_with_output_rollback",
                    side_effect=lambda paths, op: op(),
                ):
                    with self.assertRaises(ValueError):
                        app_main.process_all(processors, reporter)
                mock_print.assert_called_once()

    def test_debug_disabled_when_unset_or_empty_or_zero(self) -> None:
        def failing_processor(reporter):
            raise ValueError("boom")

        processors = [("1", "失败步骤", failing_processor)]

        for env_val in (None, "", "0", "true", "yes"):
            with self.subTest(env_val=env_val):
                env = {} if env_val is None else {"UPLOAD_DATA_DEBUG": env_val}
                reporter = ConsoleReporter(
                    stream=io.StringIO(), error_stream=io.StringIO()
                )
                with patch.dict("os.environ", env, clear=True):
                    with patch("traceback.print_exc") as mock_print:
                        with patch.object(
                            app_main,
                            "run_with_output_rollback",
                            side_effect=lambda paths, op: op(),
                        ):
                            with self.assertRaises(ValueError):
                                app_main.process_all(processors, reporter)
                        mock_print.assert_not_called()


class EndToEndAllModeTest(unittest.TestCase):
    """完整 --all 端到端：5 个模式用最小真实源文件跑通。

    数据链覆盖方太场景：审核侧 ZHLT000259 被标为厨卫/方太，回款
    明细编码品类 A02-电冰箱 表明它是冰箱，门店报表因豁免规则保持厨卫，
    第 27 行自动汇总厨卫和冰箱两侧方太。
    """

    FOTILE_REFERENCE = "17914133741N"
    HAIER_REFERENCE = "12345678901N"
    DIGITAL_REFERENCE = "12345678902N"

    def _write_mer_source(self, directory: Path, merchant: str, rows) -> Path:
        """MER_<商户编号>_*.xlsx：标题行 + 24 列表头（D/E/F/G/I/J 语义）+ 数据。"""
        path = directory / f"MER_{merchant}_20260801000000_yjhx.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["报表标题"])
        header = [f"列{i}" for i in range(24)]
        header[3] = "订单号"
        header[4] = "交易日期"
        header[5] = "交易金额"
        header[6] = "检索参考号"
        header[8] = "状态"
        header[9] = "描述"
        sheet.append(header)
        for item in rows:
            if len(item) == 5:
                order, reference, amount, status, description = item
            else:
                order, reference, amount = item
                status, description = "审核通过", "说明"
            row = ["v"] * 24
            row[3] = order
            row[4] = "2026-01-01"
            row[5] = amount
            row[6] = reference
            row[8] = status
            row[9] = description
            sheet.append(row)
        workbook.save(path)
        return path

    def _write_receipts_source(self, path: Path) -> None:
        """60 列收款单统计导出：标题行 + 表头 + 数据 + 合计。"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["收款单统计"])
        header = [f"列{i}" for i in range(60)]
        header[0] = "单据号"
        header[1] = "日期"
        header[2] = "原票号"
        header[3] = "商品名称"
        header[4] = "销售类别"
        sheet.append(header)
        row = [""] * 60
        row[0] = "REC-1"
        row[1] = date(2026, 1, 1)
        row[2] = ""
        row[3] = "海尔冰箱"
        row[4] = "零售"
        sheet.append(row)
        total = [""] * 60
        total[0] = "合计"
        sheet.append(total)
        workbook.save(path)

    def _write_payment_source(
        self, path: Path, profile, rows, merchant: str
    ) -> None:
        workbook = Workbook()
        detail = workbook.active
        detail.append(profile.detail_headers)
        for reference, raw_category, product, subsidy in rows:
            row = [None] * len(profile.detail_headers)
            row[profile.detail_headers.index("拨付批次")] = "batch"
            row[profile.detail_headers.index("交易时间")] = "2026-01-01 10:00:00"
            row[profile.detail_headers.index("交易参考号")] = reference
            row[profile.detail_headers.index("商户编号")] = merchant
            row[profile.detail_headers.index("销售金额")] = 1000
            row[profile.detail_headers.index("补贴金额")] = subsidy
            row[profile.detail_headers.index("编码品类")] = raw_category
            row[profile.detail_headers.index("商品名称")] = product
            detail.append(row)
        workbook.save(path)

    def _write_coupon_source(self, path: Path) -> None:
        """销售用券情况统计：标题 + 表头行（第 26/27 列为补贴列）+ 数据 + 合计。"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["销售用券情况统计"])
        header = [""] * 27
        header[2] = "单据号"
        header[3] = "单据日期"
        header[5] = "商品名称"
        header[7] = "品牌"
        header[14] = "财务大类"
        header[17] = "明细摘要"
        header[25] = coupon_appliance.COUPON_SUBSIDY_HEADER
        header[26] = coupons_digital.COUPON_SUBSIDY_HEADER
        sheet.append(header)
        sheet.append(
            [
                None,
                None,
                "0001",
                date(2026, 1, 1),
                None,
                "海尔冰箱",
                None,
                "海尔",
                *([None] * 6),
                "冰箱",
                *([None] * 2),
                self.HAIER_REFERENCE,
                *([None] * 7),
                150,
                None,
            ]
        )
        sheet.append(
            [
                None,
                None,
                "ZHLT000259",
                date(2026, 1, 1),
                None,
                "方太冰箱",
                None,
                "方太",
                *([None] * 6),
                "厨卫",
                *([None] * 2),
                self.FOTILE_REFERENCE,
                *([None] * 7),
                1500,
                None,
            ]
        )
        sheet.append(
            [
                None,
                None,
                "0003",
                date(2026, 1, 2),
                None,
                "华为手机",
                None,
                "华为",
                *([None] * 6),
                "数码",
                *([None] * 2),
                self.DIGITAL_REFERENCE,
                *([None] * 7),
                None,
                75,
            ]
        )
        total = [None] * 27
        total[0] = "合计"
        total[25] = 1650
        sheet.append(total)
        workbook.save(path)

    def _write_supplement_file(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(("参考号", "单据号", "单据日期"))
        workbook.save(path)

    def _write_template(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "益庄"
        for column_index in range(1, store_report.EXPECTED_COLUMN_COUNT + 1):
            sheet.cell(
                row=1,
                column=column_index,
                value="表1                  2026年（益庄店 ）门店国补上传及回款情况表_2026-08-15 18:06:27"
                if column_index == 1
                else "",
            )
        for coordinate, value in LITERAL_STRUCTURE_CELLS.items():
            sheet[coordinate] = value
        for range_ in LITERAL_EXPECTED_MERGED_RANGES:
            sheet.merge_cells(range_)
        for expected, row in enumerate(range(3, 33), start=1):
            sheet[f"B{row}"] = expected
        sheet.row_dimensions[53].hidden = True
        workbook.save(path)

    @staticmethod
    def _measurement_font() -> Path | None:
        """A real font file for the patched resolve_font: PIL must open it."""
        candidates = (
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf"),
            Path.home() / ".local/share/fonts/MapleMono-NF-CN/MapleMono-NF-CN-Regular.ttf",
        )
        return next((path for path in candidates if path.exists()), None)

    def _patches(self, data_dir: Path, output_dir: Path, font_path: Path):
        output_files = {
            "appliance": output_dir / "家电_已上传.xlsx",
            "digital": output_dir / "数码_已上传.xlsx",
            "receipts": output_dir / "收款单统计.xlsx",
            "payment": output_dir / "回款明细.xlsx",
            "coupon": output_dir / "审核明细.xlsx",
            "store": output_dir / "2026年门店国补上传及回款情况表（益庄店）.xlsx",
        }
        submitted_profiles = dict(app_main.submitted.PROFILES)
        submitted_profiles["家电"] = app_main.submitted.SubmittedProfile(
            output_files["appliance"], Decimal("0.15"), Decimal("1500")
        )
        submitted_profiles["数码"] = app_main.submitted.SubmittedProfile(
            output_files["digital"], Decimal("0.15"), Decimal("500")
        )
        return (
            output_files,
            [
                patch.object(app_main, "resolve_data_dir", return_value=data_dir),
                patch.object(
                    app_main,
                    "remove_stale_temporary_files",
                    return_value=StaleFileCleanup((), ()),
                ),
                patch.object(
                    app_main, "acquire_instance_lock", return_value=MagicMock()
                ),
                patch.object(
                    app_main.submitted,
                    "OUTPUT_FILES",
                    (output_files["appliance"], output_files["digital"]),
                ),
                patch.object(app_main.submitted, "PROFILES", submitted_profiles),
                patch.object(
                    app_main.receipts, "OUTPUT_FILE", output_files["receipts"]
                ),
                patch.object(
                    app_main.payment, "OUTPUT_FILE", output_files["payment"]
                ),
                patch.object(
                    coupon_report, "OUTPUT_FILE", output_files["coupon"]
                ),
                patch.object(
                    coupon_report, "PAYMENT_FILE", output_files["payment"]
                ),
                patch.object(
                    coupon_appliance,
                    "COUPON_REMARK_SOURCE_FILE",
                    output_files["receipts"],
                ),
                patch.object(
                    coupon_appliance,
                    "COUPON_UPLOADED_SOURCE_FILE",
                    output_files["appliance"],
                ),
                patch.object(
                    coupons_digital,
                    "COUPON_REMARK_SOURCE_FILE",
                    output_files["receipts"],
                ),
                patch.object(
                    coupons_digital,
                    "COUPON_UPLOADED_SOURCE_FILE",
                    output_files["digital"],
                ),
                patch.object(store_report, "UPLOAD_FILE", output_files["coupon"]),
                patch.object(store_report, "PAYMENT_FILE", output_files["payment"]),
                patch.object(store_report, "OUTPUT_FILE", output_files["store"]),
                *(
                    patch.object(
                        module,
                        "resolve_font",
                        return_value=("Maple Mono NF CN", font_path),
                    )
                    for module in (
                        app_main.submitted,
                        app_main.receipts,
                        app_main.payment,
                        coupon_report,
                        store_report,
                    )
                ),
            ],
        )

    def test_full_all_mode_runs_end_to_end(self) -> None:
        font_path = self._measurement_font()
        if font_path is None:
            self.skipTest("无可用字体文件（端到端测试需要真实字体测宽）")
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            data_dir = base / "data"
            data_dir.mkdir()
            output_dir = base / "output"
            output_dir.mkdir()

            self._write_mer_source(
                data_dir,
                "89813015722APT1",
                [
                    (
                        "ORDER-1",
                        self.HAIER_REFERENCE,
                        1000,
                        "核销失败",
                        "商品已核销",
                    ),
                    (
                        "ORDER-2",
                        self.FOTILE_REFERENCE,
                        10000,
                        "审核通过",
                        "说明",
                    ),
                ],
            )
            self._write_mer_source(
                data_dir,
                "89813014812B06R",
                [
                    (
                        "ORDER-3",
                        self.DIGITAL_REFERENCE,
                        500,
                        "审核失败",
                        "SN编码错误",
                    )
                ],
            )
            self._write_receipts_source(data_dir / "收款单统计.xlsx")
            appliance_profile = app_main.payment.PROFILES["家电"]
            digital_profile = app_main.payment.PROFILES["数码"]
            self._write_payment_source(
                data_dir / "以旧换新补贴明细.xlsx",
                appliance_profile,
                [
                    (self.HAIER_REFERENCE, "A02-电冰箱", "海尔冰箱", 150),
                    (self.FOTILE_REFERENCE, "A02-电冰箱", "方太冰箱", 1500),
                ],
                "89813015722APT1",
            )
            self._write_payment_source(
                data_dir / "数码补贴明细.xlsx",
                digital_profile,
                [(self.DIGITAL_REFERENCE, "B01-手机", "华为手机", 75)],
                "89813014812B06R",
            )
            self._write_coupon_source(data_dir / "销售用券情况统计.xlsx")
            self._write_supplement_file(
                data_dir / "新建 Microsoft Excel 工作表.xlsx"
            )
            self._write_template(
                data_dir / "2026年门店国补上传及回款情况表 （益庄店）.xlsx"
            )

            output_files, patches = self._patches(data_dir, output_dir, font_path)
            from contextlib import ExitStack

            with ExitStack() as stack:
                for patch_ in patches:
                    stack.enter_context(patch_)
                self.assertEqual(app_main.main(["--all"]), 0)

            for name, path in output_files.items():
                self.assertTrue(path.exists(), f"缺少输出：{name}")

            # 审核明细：验证数据汇总表 G/H 列回款统计正确
            coupon_result = load_workbook(output_files["coupon"], data_only=True)
            summary_sheet = coupon_result[coupon_report.SUMMARY_SHEET_NAME]
            self.assertEqual(summary_sheet.max_column, 8)
            self.assertEqual(summary_sheet["F1"].value, "退回")
            self.assertEqual(summary_sheet["G1"].value, "回款数量")
            self.assertEqual(summary_sheet["H1"].value, "回款金额")
            summary_data = {}
            current_category = None
            current_brand = None
            for row in summary_sheet.iter_rows(min_row=2, values_only=True):
                if not any(v is not None for v in row):
                    continue
                category, brand, status, count, amount, returned, p_cnt, p_amt = row[:8]
                if category is not None:
                    current_category = category
                    current_brand = brand
                elif brand is not None:
                    current_brand = brand
                summary_data[(current_category, current_brand, status)] = (
                    count,
                    amount,
                    returned,
                    p_cnt,
                    p_amt,
                )
            self.assertEqual(summary_data[("冰箱", "海尔", "已上传")][2], 1)
            self.assertEqual(summary_data[("冰箱", "海尔", "已上传")][3], 1)
            self.assertEqual(summary_data[("冰箱", "海尔", "已上传")][4], 150)
            self.assertIsNone(summary_data[("厨卫", "方太", "已上传")][2])
            self.assertEqual(summary_data[("厨卫", "方太", "已上传")][3], 1)
            self.assertEqual(summary_data[("厨卫", "方太", "已上传")][4], 1500)
            self.assertEqual(summary_data[("家电", None, "已上传")][2], 1)
            self.assertEqual(summary_data[("家电", None, "已上传")][3], 2)
            self.assertEqual(summary_data[("家电", None, "已上传")][4], 1650)
            self.assertIsNone(summary_data[("家电", None, "未上传")][2])
            self.assertIsNone(summary_data[("家电", None, "未上传")][3])
            self.assertIsNone(summary_data[("家电", None, "未上传")][4])
            self.assertEqual(summary_data[("家电", None, "合计")][2], 1)
            self.assertIsNone(summary_data[("家电", None, "合计")][3])
            self.assertIsNone(summary_data[("家电", None, "合计")][4])
            self.assertEqual(summary_data[("数码", None, "已上传")][2], 1)
            self.assertEqual(summary_data[("数码", None, "已上传")][3], 1)
            self.assertEqual(summary_data[("数码", None, "已上传")][4], 75)
            self.assertIsNone(summary_data[("数码", None, "未上传")][2])
            self.assertIsNone(summary_data[("数码", None, "未上传")][3])
            self.assertIsNone(summary_data[("数码", None, "未上传")][4])
            self.assertEqual(summary_data[("数码", None, "合计")][2], 1)
            self.assertIsNone(summary_data[("数码", None, "合计")][3])
            self.assertIsNone(summary_data[("数码", None, "合计")][4])

            # 门店报表：方太冰箱与厨卫统一汇总至第 27 行。
            result = load_workbook(output_files["store"], data_only=True)
            sheet = result[result.sheetnames[0]]
            self.assertEqual(sheet.max_column, 8)
            self.assertEqual(sheet["D3"].value, 150)
            self.assertEqual(sheet["E3"].value, 150)
            self.assertEqual(sheet["G3"].value, 150)
            self.assertEqual(sheet["D27"].value, 1500)
            self.assertEqual(sheet["E27"].value, 1500)
            self.assertEqual(sheet["G27"].value, 1500)
            self.assertEqual(sheet["D32"].value, 75)
            self.assertEqual(sheet["E32"].value, 75)
            self.assertEqual(sheet["G32"].value, 75)
            # 表 1 总计：
            self.assertEqual(sheet["D33"].value, 1725)
            self.assertEqual(sheet["E33"].value, 1725)
            self.assertEqual(sheet["G33"].value, 1725)
            self.assertEqual(sheet["F33"].value, 1)
            self.assertEqual(sheet["H33"].value, 1)
            # 表 2 品牌汇总：海尔系 发生 150, 上传 150, 回款 150, 上传率 1, 回款率 1；其余为空；第 47 行合计。
            self.assertEqual(sheet["D40"].value, 150)
            self.assertEqual(sheet["E40"].value, 150)
            self.assertEqual(sheet["F40"].value, 150)
            self.assertEqual(sheet["G40"].value, 1)
            self.assertEqual(sheet["H40"].value, 1)
            for row in range(41, 47):
                self.assertIsNone(sheet[f"D{row}"].value)
                self.assertIsNone(sheet[f"E{row}"].value)
                self.assertIsNone(sheet[f"F{row}"].value)
                self.assertIsNone(sheet[f"G{row}"].value)
                self.assertIsNone(sheet[f"H{row}"].value)
            self.assertEqual(sheet["D47"].value, 150)
            self.assertEqual(sheet["E47"].value, 150)
            self.assertEqual(sheet["F47"].value, 150)
            self.assertEqual(sheet["G47"].value, 1)
            self.assertEqual(sheet["H47"].value, 1)


if __name__ == "__main__":
    unittest.main()
