import io
import tempfile
import unittest
from collections import Counter
from contextlib import redirect_stdout
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from xlsxwriter import Workbook as XlsxWorkbook

from processors.common.excel import load_measurement_font, resolve_font
from processors.coupons import appliance, sources, xlsx_output
from processors.coupons import digital as coupons_digital


def summary_row(
    module: object,
    *,
    category: str,
    brand: str,
    remark: str,
    detail: str,
    subsidy: object,
) -> list[object]:
    header = module.COUPON_OUTPUT_HEADER
    values = {
        "财务大类": category,
        "品牌": brand,
        "备注": remark,
        "详细情况": detail,
        module.COUPON_OUTPUT_HEADER[6]: subsidy,
    }
    return [values.get(column) for column in header]


class CouponSummaryTest(unittest.TestCase):
    def test_digital_uploaded_subsidy_stats_come_from_uploaded_summary(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "已上传.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Summary"
            sheet.append(["检索参考号", "状态", "描述", "补贴金额"])
            sheet.append(["12345678901N", "已完成", "匹配成功", 10.1])
            sheet.append(["12345678902N", "已完成", "匹配成功", None])
            sheet.append(["12345678903N", "已完成", "匹配成功", 20])
            workbook.save(source)
            workbook.close()

            _lookup, subsidy_count, subsidy_total = sources.load_uploaded_summary(
                source
            )
            self.assertEqual((subsidy_count, subsidy_total), (2, Decimal("30.1")))

    def test_digital_summary_subtracts_uploaded_stats_from_coupon_totals(
        self,
    ) -> None:
        rows = [
            list(coupons_digital.COUPON_OUTPUT_HEADER),
            summary_row(
                coupons_digital,
                category="数码",
                brand="A",
                remark="已上传",
                detail="审核通过",
                subsidy="10.10",
            ),
            summary_row(
                coupons_digital,
                category="数码",
                brand="B",
                remark="未上传",
                detail="",
                subsidy=20,
            ),
            summary_row(
                coupons_digital,
                category="数码",
                brand="C",
                remark="已上传",
                detail="备注不影响汇总口径",
                subsidy=0,
            ),
            summary_row(
                coupons_digital,
                category="数码",
                brand="D",
                remark="未上传",
                detail="负数按负一计数",
                subsidy=-5,
            ),
        ]

        result = coupons_digital.build_coupon_summary(
            rows,
            uploaded_count=1,
            uploaded_subsidy_total=Decimal("20.10"),
        )

        self.assertEqual(
            result,
            [
                ("已上传", 1, 20.1),
                ("未上传", 0, 5.0),
                ("合计", 1, 25.1),
            ],
        )

    def test_large_appliance_summaries_have_independent_expectations(self) -> None:
        rows = [
            list(appliance.COUPON_OUTPUT_HEADER),
            summary_row(
                appliance,
                category="冰箱",
                brand="海尔",
                remark="已上传",
                detail="状态：审核通过",
                subsidy="10.11",
            ),
            summary_row(
                appliance,
                category="冰箱",
                brand="海尔",
                remark="已上传",
                detail="状态：待审核",
                subsidy="20.20",
            ),
            summary_row(
                appliance,
                category="空调",
                brand="格力",
                remark="未上传",
                detail="",
                subsidy=None,
            ),
            summary_row(
                appliance,
                category="底部",
                brand="排除",
                remark="已上传",
                detail="审核通过",
                subsidy=999,
            ),
        ]

        summary, zero_subsidy_count = (
            appliance.build_coupon_summary(
                rows,
                excluded_bottom_rows=1,
                uploaded_subsidy_count=1,
                uploaded_subsidy_total=Decimal("20.20"),
            )
        )

        self.assertEqual(
            summary,
            [
                ("冰箱", "海尔", "已上传", 2, 30.31),
                ("空调", "格力", "未上传", 1, 0.0),
                # 已上传 is measured from the 已上传 workbook, 合计 from this
                # coupon file's own 国补 column, and 未上传 is the difference.
                # 财务大类=家电 with no 品牌 so the block matches the 数码 one
                # appended after it in 审核明细.
                ("家电", None, "已上传", 1, 20.20),
                ("家电", None, "未上传", 1, 10.11),
                ("家电", None, "合计", 2, 30.31),
            ],
        )
        self.assertEqual(zero_subsidy_count, 0)

    def test_reversal_counts_as_minus_one_and_zero_is_flagged(self) -> None:
        """A return cancels its original, so it counts -1, not +1.

        A zero 国补 cannot happen legitimately; it is counted as 0 and
        reported so the operator can go fix the source row.
        """
        rows = [
            list(appliance.COUPON_OUTPUT_HEADER),
            summary_row(
                appliance,
                category="冰箱",
                brand="海尔",
                remark="",
                detail="",
                subsidy="100.00",
            ),
            summary_row(
                appliance,
                category="冰箱",
                brand="海尔",
                remark="",
                detail="",
                subsidy="-100.00",
            ),
            summary_row(
                appliance,
                category="冰箱",
                brand="海尔",
                remark="",
                detail="",
                subsidy=0,
            ),
        ]

        summary, zero_subsidy_count = (
            appliance.build_coupon_summary(
                rows,
                excluded_bottom_rows=0,
                uploaded_subsidy_count=0,
                uploaded_subsidy_total=Decimal("0"),
            )
        )

        self.assertEqual(zero_subsidy_count, 1)
        self.assertEqual(summary[-1], ("家电", None, "合计", 0, 0.0))

    def test_invalid_subsidy_is_rejected(self) -> None:
        for module in (coupons_digital, appliance):
            with self.subTest(module=module.__name__):
                rows = [
                    list(module.COUPON_OUTPUT_HEADER),
                    summary_row(
                        module,
                        category="冰箱",
                        brand="海尔",
                        remark="已上传",
                        detail="",
                        subsidy="非数字",
                    ),
                ]

                with self.assertRaisesRegex(ValueError, "国补金额无效"):
                    if module is coupons_digital:
                        module.build_coupon_summary(
                            rows,
                            uploaded_count=0,
                            uploaded_subsidy_total=Decimal("0"),
                        )
                    else:
                        module.build_coupon_summary(
                            rows,
                            excluded_bottom_rows=0,
                            uploaded_subsidy_count=0,
                            uploaded_subsidy_total=Decimal("0"),
                        )

    def test_group_sheet_inherits_payment_status_from_main_detail(self) -> None:
        row = summary_row(
            appliance,
            category="冰箱",
            brand="海尔",
            remark="已上传",
            detail="审核通过：同意",
            subsidy=100,
        )
        row[appliance.COUPON_OUTPUT_HEADER.index("回款情况")] = "已回款"
        groups = appliance.build_coupon_group_sheets(
            [list(appliance.COUPON_OUTPUT_HEADER), row],
            excluded_bottom_rows=0,
        )

        grouped_row = appliance.select_coupon_group_columns(groups[0][3][0][0])
        payment_index = appliance.COUPON_GROUP_HEADER.index("回款情况")
        self.assertEqual(grouped_row[payment_index], "已回款")


class SummarySheetLayoutTest(unittest.TestCase):
    """Cover the sheet builder itself, not just the row computation.

    Removing the 备注汇总 panel once left a dangling reference to its row
    numbers behind; nothing failed until the real workbook was built, because
    no test called the builder.
    """

    def build_computation(self, **overrides) -> object:
        rows = [
            list(appliance.COUPON_OUTPUT_HEADER),
            summary_row(
                appliance,
                category="冰箱",
                brand="海尔",
                remark="已上传",
                detail="状态：审核通过",
                subsidy="10.11",
            ),
        ]
        summary, zero = appliance.build_coupon_summary(
            rows,
            excluded_bottom_rows=0,
            uploaded_subsidy_count=1,
            uploaded_subsidy_total=Decimal("10.11"),
        )
        fields = dict(
            rows=rows,
            data_row_count=1,
            matched_count=0,
            matched_subsidy_total=Decimal("0"),
            remark_lookup={},
            detail_lookup={},
            reference_universe=set(),
            payment_references=frozenset(),
            payment_match_count=0,
            reference_supplement_count=0,
            ambiguous_reference_supplement_count=0,
            reference_supplement_matches=Counter(),
            corrected_count=0,
            unresolved_count=0,
            correction_collision_count=0,
            reference_decisions=[],
            final_unresolved_reference_count=0,
            uploaded_count=0,
            unmatched_count=0,
            excluded_category_row_count=0,
            uploaded_subsidy_count=1,
            uploaded_subsidy_total=Decimal("10.11"),
            zero_subsidy_count=zero,
            source_total=None,
            computed_total=Decimal("10.11"),
            summary_rows=summary,
            group_sheets=[],
        )
        fields.update(overrides)
        return appliance.CouponComputation(**fields)

    def build_sheet(self):
        """Write 数据汇总 with the XlsxWriter writer and read it back.

        The assertions below are about the finished sheet — its rows, its
        merged ranges — so they are unchanged from when an openpyxl builder
        produced it in memory. Only how the sheet comes into being differs.
        """
        computation = self.build_computation()
        rows = computation.summary_rows
        blocks = appliance.project_summary_blocks(rows)
        brand_rows_end = blocks[0][0] if blocks else len(rows)
        font_name, font_path = resolve_font()
        measurement_font = load_measurement_font(font_path)

        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        path = Path(directory.name) / "summary.xlsx"
        with XlsxWorkbook(str(path)) as workbook:
            formats = xlsx_output.CouponFormatCache(
                workbook, font_name, appliance.COUPON_MATCH_FILL_COLOR
            )
            xlsx_output.write_summary_sheet(
                workbook,
                appliance.SUMMARY_SHEET_NAME,
                appliance.COUPON_SUMMARY_HEADER,
                rows,
                formats,
                measurement_font,
                group_merges=(
                    appliance.coupon_summary_group_merges(rows, brand_rows_end)
                    if brand_rows_end
                    else []
                ),
                project_merges=appliance.coupon_summary_project_merges(blocks),
            )
        return load_workbook(path), computation

    def test_summary_sheet_drops_the_side_panels(self) -> None:
        """数据汇总 is the 财务大类/品牌/上传状态 表 and nothing else.

        The 备注汇总 and 审核通过明细 panels that once sat beside it were
        removed; both left dangling row-number references behind when they
        went, which nothing caught until the real workbook was built.
        """
        workbook, _ = self.build_sheet()
        try:
            sheet = workbook[appliance.SUMMARY_SHEET_NAME]
            values = {
                str(cell.value)
                for row in sheet.iter_rows()
                for cell in row
                if cell.value is not None
            }
            self.assertNotIn("审核通过明细", values)
            self.assertNotIn("备注汇总", values)
            self.assertEqual(
                sheet.max_column,
                len(appliance.COUPON_SUMMARY_HEADER),
            )
        finally:
            workbook.close()

    def test_summary_sheet_ends_with_the_three_tail_rows(self) -> None:
        """已上传/未上传/合计 sit in 上传状态, under a single merged 家电 cell.

        The status used to live in 财务大类, which read as three more
        categories alongside 冰箱/空调 and did not line up with the 数码
        block appended below it in 审核明细.
        """
        self.assertEqual(
            appliance.COUPON_SUMMARY_HEADER,
            ("财务大类", "品牌", "上传状态", "数量", "2026国补金额"),
        )
        workbook, computation = self.build_sheet()
        try:
            sheet = workbook[appliance.SUMMARY_SHEET_NAME]
            tail_start = 1 + len(computation.summary_rows) - 3
            status_column = appliance.COUPON_SUMMARY_HEADER.index(
                "上传状态"
            ) + 1
            labels = [
                sheet.cell(tail_start + offset + 1, status_column).value
                for offset in range(3)
            ]
            self.assertEqual(labels, ["已上传", "未上传", "合计"])
            categories = [
                sheet.cell(tail_start + offset + 1, 1).value
                for offset in range(3)
            ]
            # Merged vertically, so only the first row keeps the label.
            self.assertEqual(
                categories,
                [appliance.COUPON_SUMMARY_PROJECT_LABEL, None, None],
            )
        finally:
            workbook.close()

    def test_project_block_spans_both_label_columns(self) -> None:
        """财务大类 and 品牌 become one cell over each project's block.

        The block is not a brand breakdown, so an empty 品牌 cell beside the
        家电 label read as a missing value rather than an inapplicable one.
        """
        workbook, computation = self.build_sheet()
        try:
            sheet = workbook[appliance.SUMMARY_SHEET_NAME]
            tail_start = 1 + len(computation.summary_rows) - 3
            merges = {str(r) for r in sheet.merged_cells.ranges}
            self.assertIn(
                f"A{tail_start + 1}:B{tail_start + 3}",
                merges,
            )
        finally:
            workbook.close()


class ProjectSummaryBlocksTest(unittest.TestCase):
    """The blocks are found by 品牌 being absent, not by counting three rows
    from the bottom, so any number of projects can be appended."""

    def test_blocks_are_the_brandless_runs_split_per_project(self) -> None:
        rows = [
            ("冰箱", "海尔", "已上传", 1, 1.0),
            ("空调", "格力", "未上传", 1, 2.0),
            ("家电", None, "已上传", 1, 1.0),
            ("家电", None, "未上传", 1, 2.0),
            ("家电", None, "合计", 2, 3.0),
            ("数码", None, "已上传", 1, 4.0),
            ("数码", None, "合计", 1, 4.0),
        ]

        self.assertEqual(
            appliance.project_summary_blocks(rows),
            [(2, 4), (5, 6)],
        )

    def test_brand_only_rows_produce_no_blocks(self) -> None:
        rows = [
            ("冰箱", "海尔", "已上传", 1, 1.0),
            ("空调", "格力", "未上传", 1, 2.0),
        ]

        self.assertEqual(appliance.project_summary_blocks(rows), [])


class SubsidyCorrectionWarningTests(unittest.TestCase):
    """The attribution warnings are formatted by one pure function and
    printed by the flow — the reader no longer prints while reading."""

    def test_format_names_every_field(self) -> None:
        from processors.coupon_report import format_subsidy_correction_warning

        correction = sources.SubsidyCorrection(
            row_number=5346,
            document_number="ZG2J000016",
            financial_category="数码",
            amount=Decimal("314.85"),
            from_header=sources.COUPON_FAMILY_SUBSIDY_HEADER,
            to_header=sources.COUPON_DIGITAL_SUBSIDY_HEADER,
        )
        message = format_subsidy_correction_warning(
            correction,
            "销售用券情况统计.xlsx",
        )
        self.assertIn("销售用券情况统计.xlsx", message)
        self.assertIn("第 5346 行", message)
        self.assertIn("ZG2J000016", message)
        self.assertIn("'数码'", message)
        self.assertIn("314.85", message)
        self.assertIn("从“2026家电国补（计入收入）”", message)
        self.assertIn("调整到“2026数码国补（计入收入）”", message)

    def test_process_coupon_sales_prints_one_warning_per_correction(self) -> None:
        """Flow-level check: the warnings appear exactly once per recorded
        correction, in source row order, no matter what the reader did."""
        from contextlib import ExitStack
        from unittest.mock import patch

        from processors import coupon_report

        export = sources.CouponExport(
            appliance_rows=[list(sources.APPLIANCE_PROFILE.output_header)],
            digital_rows=[list(sources.DIGITAL_PROFILE.output_header)],
            source_total=None,
            subsidy_corrections=(
                sources.SubsidyCorrection(
                    5346,
                    "ZG2J000016",
                    "数码",
                    Decimal("314.85"),
                    sources.COUPON_FAMILY_SUBSIDY_HEADER,
                    sources.COUPON_DIGITAL_SUBSIDY_HEADER,
                ),
                sources.SubsidyCorrection(
                    5347,
                    "ZG2J000017",
                    "冰箱",
                    Decimal("100.00"),
                    sources.COUPON_DIGITAL_SUBSIDY_HEADER,
                    sources.COUPON_FAMILY_SUBSIDY_HEADER,
                ),
            ),
        )
        with ExitStack() as stack:
            stack.enter_context(
                patch.object(
                    coupon_report.sources,
                    "COUPON_SOURCE_FILE",
                    Path("销售用券情况统计.xlsx"),
                    create=True,
                )
            )
            # COUPON_REFERENCE_SUPPLEMENT_FILE only exists after
            # configure_data_dir; without this patch the test fails whenever
            # it runs before any test that configured the module globals.
            stack.enter_context(
                patch.object(
                    coupon_report.sources,
                    "COUPON_REFERENCE_SUPPLEMENT_FILE",
                    Path("参考号补充.xlsx"),
                    create=True,
                )
            )
            stack.enter_context(
                patch.object(
                    coupon_report.sources,
                    "read_coupon_export",
                    return_value=export,
                )
            )
            stack.enter_context(
                patch.object(
                    coupon_report.sources,
                    "load_payment_reference_locations",
                    return_value={"家电": {}, "数码": {}},
                )
            )
            stack.enter_context(
                patch.object(
                    coupon_report,
                    "load_coupon_remark_lookup",
                    return_value={},
                )
            )
            stack.enter_context(
                patch.object(
                    appliance,
                    "load_uploaded_summary",
                    return_value=({}, 0, Decimal("0")),
                )
            )
            stack.enter_context(
                patch.object(
                    appliance,
                    "load_coupon_reference_supplement",
                    return_value={},
                )
            )
            stack.enter_context(
                patch.object(
                    coupons_digital,
                    "load_uploaded_summary",
                    return_value=({}, 0, Decimal("0")),
                )
            )
            stack.enter_context(
                patch.object(coupon_report, "CalamineWorkbook")
            )
            stack.enter_context(
                patch.object(coupon_report, "write_xlsx_atomically")
            )
            output = io.StringIO()
            with redirect_stdout(output):
                coupon_report.process_coupon_sales()

        warning_lines = [
            line
            for line in output.getvalue().splitlines()
            if line.startswith("警告：")
        ]
        self.assertEqual(len(warning_lines), 2)
        self.assertIn("第 5346 行单据 ZG2J000016", warning_lines[0])
        self.assertIn("314.85", warning_lines[0])
        self.assertIn("第 5347 行单据 ZG2J000017", warning_lines[1])


class SheetHeaderContractTests(unittest.TestCase):
    """数据汇总 speaks 上传状态; the detail sheets keep calling it 备注."""

    def test_summary_uses_upload_status_but_details_keep_remark(self) -> None:
        self.assertEqual(
            appliance.COUPON_SUMMARY_HEADER[2],
            "上传状态",
        )
        self.assertNotIn("上传状态", appliance.COUPON_OUTPUT_HEADER)
        self.assertNotIn("上传状态", coupons_digital.COUPON_OUTPUT_HEADER)
        self.assertIn("备注", appliance.COUPON_OUTPUT_HEADER)
        self.assertIn("备注", coupons_digital.COUPON_OUTPUT_HEADER)
        self.assertEqual(appliance.DETAILS_SHEET_NAME, "家电-明细总表")
        self.assertEqual(coupons_digital.DETAILS_SHEET_NAME, "数码-明细总表")


class SourceTotalGapTest(unittest.TestCase):
    """The coupon export's own 合计 row can disagree with its detail rows.

    Seen in production: a return recorded after the detail rows were written
    left the 合计 row 1,290.00 lower. The program keeps the detail-row total
    (so 数据汇总 matches 家电-明细总表) and reports the gap instead.
    """

    def report(self, source_total, computed_total) -> str:
        from processors.coupon_report import report_source_total_gap

        output = io.StringIO()
        with redirect_stdout(output):
            report_source_total_gap("家电", source_total, computed_total)
        return output.getvalue()

    def test_gap_is_reported_with_both_totals_and_the_difference(self) -> None:
        message = self.report(Decimal("2866331.40"), Decimal("2867621.40"))

        self.assertIn("2,866,331.40", message)
        self.assertIn("2,867,621.40", message)
        self.assertIn("1,290.00", message)

    def test_matching_totals_report_nothing(self) -> None:
        self.assertEqual(
            self.report(Decimal("2867621.40"), Decimal("2867621.40")),
            "",
        )

    def test_unparsable_source_total_reports_nothing(self) -> None:
        self.assertEqual(self.report(None, Decimal("2867621.40")), "")


if __name__ == "__main__":
    unittest.main()
