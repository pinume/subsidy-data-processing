"""Contracts for the XlsxWriter audit output and its calamine validator.

Style tests read the file with openpyxl; runtime validation tests exercise the
calamine path that checks every value and merge. The format rules were read off
the original openpyxl output (数据汇总's currency column is located by header
name, and its header differs per project).
"""

import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from openpyxl.styles import Font
from xlsxwriter import Workbook

from processors import coupon_report
from processors.common.excel import load_measurement_font, resolve_font
from processors.coupons import appliance, digital
from processors.coupons.xlsx_output import (
    CouponFormatCache,
    FormatKey,
    write_detail_sheet,
    write_summary_sheet,
)

FILL_COLOR = "FFC7CE"
DETAIL_HEADER = ("单据号", "单据日期", "商品名称", "备注", "详细情况")
SUMMARY_HEADER = ("财务大类", "品牌", "上传状态", "数量", "2026国补金额", "退回")


class WriterContractTest(unittest.TestCase):
    def setUp(self) -> None:
        self._directory = tempfile.TemporaryDirectory()
        self.path = Path(self._directory.name) / "audit.xlsx"
        self.addCleanup(self._directory.cleanup)
        font_name, font_path = resolve_font()
        self.font_name = font_name
        self.measurement_font = load_measurement_font(font_path)

    def _read(self, name: str):
        return load_workbook(self.path)[name]

    def test_detail_sheet_formats_and_trailing_pink_block(self) -> None:
        rows = [[f"ZH{i:04d}", None, "海尔冰箱", "", ""] for i in range(5)]
        with Workbook(str(self.path)) as workbook:
            formats = CouponFormatCache(workbook, self.font_name, FILL_COLOR)
            write_detail_sheet(
                workbook,
                "明细",
                DETAIL_HEADER,
                rows,
                formats,
                self.measurement_font,
                left_aligned_headers=("商品名称", "详细情况"),
                matched_count=2,
            )
        sheet = self._read("明细")

        self.assertEqual(sheet["A2"].number_format, "@")
        # 单据日期 is empty here, so it keeps the general format — the openpyxl
        # writer guarded the date format on the cell having a value.
        self.assertEqual(sheet["B2"].number_format, "General")
        self.assertEqual(sheet["C2"].alignment.horizontal, "left")
        self.assertEqual(sheet["E2"].alignment.horizontal, "left")
        self.assertEqual(sheet["D2"].alignment.horizontal, "center")

        def pink(row: int) -> bool:
            cell = sheet.cell(row, 1)
            return (
                cell.fill.patternType == "solid"
                and str(cell.fill.fgColor.rgb)[-6:] == FILL_COLOR
            )

        # 5 data rows in sheet rows 2..6; the last 2 are the matched block.
        self.assertEqual([pink(r) for r in range(2, 7)], [False, False, False, True, True])

    def test_detail_sheet_with_no_matches_has_no_pink(self) -> None:
        rows = [["ZH0001", None, "海尔冰箱", "", ""]]
        with Workbook(str(self.path)) as workbook:
            formats = CouponFormatCache(workbook, self.font_name, FILL_COLOR)
            write_detail_sheet(
                workbook,
                "明细",
                DETAIL_HEADER,
                rows,
                formats,
                self.measurement_font,
                left_aligned_headers=(),
                matched_count=0,
            )
        self.assertIsNone(self._read("明细")["A2"].fill.patternType)

    def test_summary_sheet_borders_currency_and_merges(self) -> None:
        rows = [
            ("空调", "格力", "已上传", 1, 1.5, None),
            ("空调", "格力", "未上传", 2, 2.5, None),
            ("冰箱", "海尔", "已上传", 3, 3.5, 1),
        ]
        with Workbook(str(self.path)) as workbook:
            formats = CouponFormatCache(workbook, self.font_name, FILL_COLOR)
            write_summary_sheet(
                workbook,
                "数据汇总",
                SUMMARY_HEADER,
                rows,
                formats,
                self.measurement_font,
                group_merges=[(2, 3, 1), (2, 3, 2)],
                project_merges=[],
            )
        sheet = self._read("数据汇总")

        self.assertEqual(sheet.max_column, 6)
        # The currency column is E (2026国补金额), and F (退回) is General.
        self.assertEqual(sheet["E2"].number_format, "0.00")
        self.assertEqual(sheet["E4"].number_format, "0.00")
        self.assertEqual(sheet["F2"].number_format, "General")
        self.assertEqual(sheet["A2"].number_format, "General")
        # Borders cover the header row and every data row across A:F.
        for coordinate in ("A1", "F1", "A4", "F4"):
            with self.subTest(coordinate=coordinate):
                self.assertEqual(sheet[coordinate].border.left.style, "thin")
        self.assertEqual(
            {str(r) for r in sheet.merged_cells.ranges}, {"A2:A3", "B2:B3"}
        )
        self.assertEqual(sheet["A2"].value, "空调")

    def test_format_cache_reuses_one_format_per_distinct_key(self) -> None:
        with Workbook(str(self.path)) as workbook:
            formats = CouponFormatCache(workbook, self.font_name, FILL_COLOR)
            first = formats.get(FormatKey(align="left", pink=True))
            second = formats.get(FormatKey(align="left", pink=True))
            third = formats.get(FormatKey(align="left"))
            self.assertIs(first, second)
            self.assertIsNot(first, third)
            workbook.add_worksheet("x")

    def test_width_measurer_is_shared_across_workbook_sheets(self) -> None:
        with Workbook(str(self.path)) as workbook:
            formats = CouponFormatCache(workbook, self.font_name, FILL_COLOR)
            first = formats.get_width_measurer(self.measurement_font)
            second = formats.get_width_measurer(self.measurement_font)

            self.assertIs(first, second)
            workbook.add_worksheet("x")


class CouponOutputValidationTest(unittest.TestCase):
    def setUp(self) -> None:
        self._directory = tempfile.TemporaryDirectory()
        self.path = Path(self._directory.name) / "audit.xlsx"
        self.addCleanup(self._directory.cleanup)
        self.appliance_computation = SimpleNamespace(
            summary_rows=[
                ("空调", "格力", "已上传", 1, 1.5, 1),
                ("空调", "格力", "未上传", 1, 2.5, None),
            ],
            rows=[list(appliance.COUPON_OUTPUT_HEADER)],
            matched_count=0,
            group_sheets=[],
        )
        self.digital_computation = SimpleNamespace(
            rows=[list(digital.COUPON_OUTPUT_HEADER)],
            matched_count=0,
        )
        self.extra_summary_rows = []
        self.decisions = []
        coupon_report.write_coupon_workbook(
            self.path,
            self.appliance_computation,
            self.digital_computation,
            self.extra_summary_rows,
            self.decisions,
        )

    def validate(self) -> None:
        coupon_report.validate_merged_coupon_output(
            self.path,
            self.appliance_computation,
            self.digital_computation,
            self.extra_summary_rows,
            self.decisions,
        )

    def test_calamine_validation_accepts_the_written_contract(self) -> None:
        self.validate()

    def test_calamine_validation_rejects_a_changed_value(self) -> None:
        workbook = load_workbook(self.path)
        workbook[appliance.DETAILS_SHEET_NAME]["A1"] = "错误表头"
        workbook.save(self.path)
        workbook.close()

        with self.assertRaisesRegex(RuntimeError, "输出校验失败"):
            self.validate()

    def test_calamine_validation_rejects_a_changed_returned_cell(self) -> None:
        workbook = load_workbook(self.path)
        workbook[appliance.SUMMARY_SHEET_NAME]["F2"] = 999
        workbook.save(self.path)
        workbook.close()

        with self.assertRaisesRegex(RuntimeError, "输出校验失败"):
            self.validate()

    def test_calamine_validation_rejects_zero_in_empty_returned_cell(self) -> None:
        """F 列原本为空（无退回，如未上传行 F3）时，若被写成数值 0，校验必须失败。"""
        workbook = load_workbook(self.path)
        workbook[appliance.SUMMARY_SHEET_NAME]["F3"] = 0
        workbook.save(self.path)
        workbook.close()

        with self.assertRaisesRegex(RuntimeError, "输出校验失败"):
            self.validate()

    def test_calamine_validation_rejects_boolean_true_in_returned_cell(self) -> None:
        """F 列原本为整数 1（如 F2）时，若被写成布尔值 True（对应 Excel TRUE），校验必须失败。"""
        workbook = load_workbook(self.path)
        workbook[appliance.SUMMARY_SHEET_NAME]["F2"] = True
        workbook.save(self.path)
        workbook.close()

        with self.assertRaisesRegex(RuntimeError, "输出校验失败"):
            self.validate()

    def test_calamine_validation_rejects_a_missing_merge(self) -> None:
        workbook = load_workbook(self.path)
        workbook[appliance.SUMMARY_SHEET_NAME].unmerge_cells("A2:A3")
        workbook.save(self.path)
        workbook.close()

        with self.assertRaisesRegex(RuntimeError, "合并范围校验失败"):
            self.validate()

    def test_runtime_validation_does_not_treat_font_as_business_data(self) -> None:
        workbook = load_workbook(self.path)
        workbook[appliance.DETAILS_SHEET_NAME]["A1"].font = Font(bold=False)
        workbook.save(self.path)
        workbook.close()

        self.validate()
