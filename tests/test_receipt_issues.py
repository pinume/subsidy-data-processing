import io
import tempfile
import unittest
from contextlib import redirect_stdout
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from processors import receipts

HEADER = list(receipts.RECEIPTS_SOURCE_HEADER)
SOURCE_NAME = "收款单统计.xlsx"


def pad(row):
    """Fixture rows name the leading five fields; widen them to the header."""
    return [*row, *[None] * (len(HEADER) - len(row))]


def prepare(kept_rows):
    return receipts.prepare_receipt_data(
        [pad(row) for row in kept_rows], source_name=SOURCE_NAME
    )


class PrepareReceiptDataIssuesTest(unittest.TestCase):
    def test_issues_cover_duplicate_missing_and_invalid_original(self) -> None:
        kept_rows = [
            HEADER,
            ["ZH0001", "2026-01-24", "", "海尔冰箱"],
            ["ZH0001", "2026-01-24", "", "美的空调"],
            ["ZH0002", None, "", "格力空调"],
            ["ZH0003", "2026-01-25", "notvalid", "小米手机"],
        ]

        _output_rows, stats, issues = prepare(kept_rows)

        # A duplicate match key is expected here (multi-line suite sales
        # share one 单据号/日期) and is never reported as an issue — see
        # test_duplicate_match_keys_are_not_reported_as_issues. Only the
        # genuine problems below reach 问题明细, numbered by their row in the
        # output sheet (row 2 is ZH0001, row 3 the ZH0001 duplicate, row 4
        # ZH0002, row 5 ZH0003).
        self.assertEqual(
            issues,
            [
                (
                    "缺少匹配键",
                    "4",
                    "",
                    "日期或单据号为空，无法生成匹配键",
                ),
                (
                    "原票号格式异常",
                    "5",
                    "notvalid",
                    "原票号应为6位日期加单据号",
                ),
                (
                    "原票号未匹配",
                    "5",
                    "notvalid",
                    "未找到日期与单据号组合键相同的原单",
                ),
            ],
        )
        self.assertEqual(stats["重复匹配键数量"], 1)
        self.assertEqual(stats["缺少匹配键数量"], 1)
        self.assertEqual(stats["原票号格式异常数量"], 1)

    def test_duplicate_match_keys_are_not_reported_as_issues(self) -> None:
        """Same 单据号/日期 with two line items is a normal suite sale (e.g.
        烟机+灶具 sold together), not a data problem — it still counts
        towards 重复匹配键数量 for diagnostics, but is neither highlighted
        nor written to 问题明细."""
        kept_rows = [
            HEADER,
            ["ZH0001", "2026-01-24", "", "老板-欧式烟机-K1L"],
            ["ZH0001", "2026-01-24", "", "老板-嵌入式灶-9B5-B1"],
        ]

        _output_rows, stats, issues = prepare(kept_rows)

        self.assertEqual(issues, [])
        self.assertEqual(stats["重复匹配键数量"], 1)

    def test_beiguo_rows_are_excluded_and_recorded_not_issued(self) -> None:
        """北国 rows keep being dropped — from the output, from remark
        matching, and from 问题明细 — while their source locations are
        recorded for the console warning."""
        kept_rows = [
            HEADER,
            ["ZH0001", "2026-01-24", "", "海尔冰箱"],
            ["ZH0002", "2026-01-25", "", "北国电器"],
            ["ZH0003", "2026-01-26", "", "美的空调"],
            ["ZH0004", "2026-01-27", "", "北国商城电视"],
        ]

        output_rows, stats, issues = prepare(kept_rows)

        self.assertEqual([row[0] for row in output_rows], ["ZH0001", "ZH0003"])
        self.assertEqual(stats["删除北国商品行数"], 2)
        self.assertEqual(stats["备注总数"], 0)
        self.assertEqual(issues, [])
        self.assertEqual(
            stats["北国剔除明细"],
            (
                receipts.ExcludedProductRecord(4, "ZH0002", "北国电器"),
                receipts.ExcludedProductRecord(6, "ZH0004", "北国商城电视"),
            ),
        )

    def test_beiguo_warning_names_rows_and_caps_examples(self) -> None:
        """The warning shows up to 10 examples and reports the remainder."""
        kept_rows = [
            HEADER,
            *[
                ["ZH%04d" % index, "2026-01-24", "", "北国%d" % index]
                for index in range(12)
            ],
        ]
        output = io.StringIO()
        with patch.object(
            receipts, "read_receipt_rows", return_value=kept_rows
        ), patch.object(
            receipts,
            "RECEIPTS_SOURCE_FILE",
            Path("收款单统计.xlsx"),
            create=True,
        ), patch.object(
            receipts, "write_xlsx_atomically"
        ), redirect_stdout(output):
            receipts.process_receipts()

        text = output.getvalue()
        self.assertIn(
            "[收款单] 警告：商品名称含“北国”的 12 行已按业务规则剔除",
            text,
        )
        self.assertIn("源第 3 行，单据号 ZH0000，商品名称 北国0", text)
        self.assertIn("源第 12 行，单据号 ZH0009，商品名称 北国9", text)
        self.assertIn("其余 2 行未展开", text)
        self.assertNotIn("北国10", text)  # beyond the 10-example cap


class UnremarkedSaleCategoryTest(unittest.TestCase):
    """零售补差 / 同型号换货 are left out of the remark rules entirely.

    They are booked as a pair sharing one 单据号/日期 whose 销售金额 cancel —
    the original was replaced, not returned — so neither the pair nor the
    original it names is a return to report.
    """

    def rows(self, category: str) -> list[list[object]]:
        return [
            HEADER,
            ["ZFP3000067", "2026-01-07", "", "美的空调", "正常销售"],
            ["0233000588", "2026-01-23", "260107ZFP3000067", "美的空调", category],
            ["0233000588", "2026-01-23", "260107ZFP3000067", "美的空调", category],
        ]

    def test_neither_the_pair_nor_its_original_is_remarked(self) -> None:
        for category in ("零售补差", "同型号换货"):
            with self.subTest(category):
                output_rows, stats, issues = prepare(self.rows(category))

                self.assertEqual([row[-1] for row in output_rows], [None] * 3)
                self.assertEqual(stats["备注总数"], 0)
                self.assertEqual(issues, [])
                # Every one of these rows carries an 原票号, so the tallies
                # would claim them as 退单 while their 备注 stayed blank.
                self.assertEqual(
                    (
                        stats["仅退单数量"],
                        stats["仅原单数量"],
                        stats["退单及原单数量"],
                    ),
                    (0, 0, 0),
                )

    def test_later_return_traces_through_unremarked_bridge(self) -> None:
        """A return through either bridge pairs with the subsidy-bearing original."""
        for category in ("零售补差", "同型号换货"):
            with self.subTest(category):
                kept_rows = [
                    *self.rows(category),
                    # 原票号 names the 2026-01-23 bridge: yymmdd + 单据号.
                    [
                        "0233000901",
                        "2026-02-01",
                        "2601230233000588",
                        "美的空调",
                        "退货",
                    ],
                ]

                output_rows, stats, _issues = prepare(kept_rows)

                self.assertEqual(
                    {(row[0], row[-1]) for row in output_rows},
                    {
                        ("ZFP3000067", receipts.RECEIPTS_REMARK_ORIGINAL),
                        ("0233000588", None),
                        ("0233000901", receipts.RECEIPTS_REMARK_RETURN),
                    },
                )
                self.assertEqual(stats["备注总数"], 2)
                self.assertEqual(
                    (
                        stats["仅退单数量"],
                        stats["仅原单数量"],
                        stats["退单及原单数量"],
                    ),
                    (1, 1, 0),
                )

    def test_reference_bridge_is_followed_transitively(self) -> None:
        kept_rows = [
            HEADER,
            ["ZH3X000025", "2026-04-26", "", "小鸭洗衣机", "正常销售"],
            [
                "0233000049",
                "2026-05-03",
                "260426ZH3X000025",
                "小鸭洗衣机",
                "同型号换货",
            ],
            [
                "0233000077",
                "2026-05-05",
                "2605030233000049",
                "小鸭洗衣机",
                "退货",
            ],
        ]

        output_rows, stats, issues = prepare(kept_rows)

        self.assertEqual(
            {(row[0], row[-1]) for row in output_rows},
            {
                ("ZH3X000025", receipts.RECEIPTS_REMARK_ORIGINAL),
                ("0233000049", None),
                ("0233000077", receipts.RECEIPTS_REMARK_RETURN),
            },
        )
        self.assertEqual(stats["备注总数"], 2)
        self.assertEqual(issues, [])

    def test_an_ordinary_return_of_the_same_original_still_reports(self) -> None:
        """The exclusion is by 销售类别, not by 原票号: a real 退货 row naming
        the same original must still produce the 退单/原单 pairing."""
        kept_rows = [
            *self.rows("零售补差"),
            ["0233000900", "2026-01-24", "260107ZFP3000067", "美的空调", "退货"],
        ]

        output_rows, stats, _issues = prepare(kept_rows)

        self.assertEqual(
            {(row[0], row[-1]) for row in output_rows},
            {
                ("ZFP3000067", receipts.RECEIPTS_REMARK_ORIGINAL),
                ("0233000588", None),
                ("0233000900", receipts.RECEIPTS_REMARK_RETURN),
            },
        )
        self.assertEqual(stats["备注总数"], 2)

class PrepareReceiptDataOriginalInvoiceTest(unittest.TestCase):
    def test_original_invoice_from_before_the_file_is_not_reported(self) -> None:
        """收款单统计 only ever covers one year's receipts, so an 原票号
        pointing at an earlier year can never resolve to a match_key in this
        file by construction — that is not a data problem either."""
        kept_rows = [
            HEADER,
            ["ZH0001", "2026-01-24", "250101ZH9999", "海尔冰箱"],
        ]

        output_rows, stats, issues = prepare(kept_rows)

        self.assertEqual(
            output_rows[0][-1],
            receipts.RECEIPTS_REMARK_RETURN,
        )
        self.assertEqual(issues, [])
        self.assertEqual(stats["未匹配原票号数量"], 0)

    def test_original_invoice_from_the_same_year_is_still_reported(self) -> None:
        kept_rows = [
            HEADER,
            ["ZH0001", "2026-01-24", "260101ZH9999", "海尔冰箱"],
        ]

        _output_rows, stats, issues = prepare(kept_rows)

        self.assertEqual(
            issues,
            [
                (
                    "原票号未匹配",
                    "2",
                    "260101ZH9999",
                    "未找到日期与单据号组合键相同的原单",
                )
            ],
        )
        self.assertEqual(stats["未匹配原票号数量"], 1)

    def test_no_issues_yields_empty_list(self) -> None:
        kept_rows = [
            HEADER,
            ["ZH0001", "2026-01-24", "", "海尔冰箱"],
        ]

        _output_rows, _stats, issues = prepare(kept_rows)

        self.assertEqual(issues, [])

    def test_invalid_date_error_names_source_file_and_real_row(self) -> None:
        kept_rows = [
            HEADER,
            ["ZH0001", "2026-01-24", "", "海尔冰箱"],
            ["ZH0002", "2026.1.25", "", "美的空调"],
        ]

        with self.assertRaises(ValueError) as caught:
            prepare(kept_rows)

        message = str(caught.exception)
        self.assertIn(SOURCE_NAME, message)
        self.assertIn("第 4 行", message)
        self.assertIn("2026.1.25", message)


class BlankAndTotalRowTest(unittest.TestCase):
    def test_fully_blank_row_is_skipped(self) -> None:
        kept_rows = [
            HEADER,
            ["ZH0001", "2026-01-24", "", "海尔冰箱"],
            [None, None, None, None],
            ["", "  ", "", ""],
        ]

        output_rows, stats, issues = prepare(kept_rows)

        self.assertEqual(len(output_rows), 1)
        self.assertEqual(stats["总数据量"], 1)
        self.assertEqual(stats["跳过空白行数"], 2)
        self.assertEqual(issues, [])

    def test_row_with_any_value_is_kept(self) -> None:
        kept_rows = [
            HEADER,
            [None, None, None, "海尔冰箱"],
        ]

        output_rows, stats, issues = prepare(kept_rows)

        self.assertEqual(len(output_rows), 1)
        self.assertEqual(stats["跳过空白行数"], 0)
        # Still incomplete data, so it must keep reporting itself.
        self.assertEqual([issue[0] for issue in issues], ["缺少匹配键"])

    def test_numeric_zero_is_not_blank(self) -> None:
        kept_rows = [
            HEADER,
            [0, None, None, None],
        ]

        output_rows, stats, _issues = prepare(kept_rows)

        self.assertEqual(len(output_rows), 1)
        self.assertEqual(stats["跳过空白行数"], 0)

    def test_total_row_is_skipped_in_either_column(self) -> None:
        for total_row in (
            ["合计", "合计", "", ""],
            ["合计", None, "", ""],
            [None, "合计", "", ""],
        ):
            with self.subTest(total_row=total_row):
                kept_rows = [
                    HEADER,
                    ["ZH0001", "2026-01-24", "", "海尔冰箱"],
                    total_row,
                ]

                output_rows, stats, _issues = prepare(kept_rows)

                self.assertEqual(len(output_rows), 1)
                self.assertEqual(stats["跳过合计行数"], 1)

    def test_合计_inside_text_fields_is_not_a_total_row(self) -> None:
        kept_rows = [
            HEADER,
            ["ZH0001", "2026-01-24", "", "合计套装"],
        ]

        output_rows, stats, _issues = prepare(kept_rows)

        self.assertEqual(len(output_rows), 1)
        self.assertEqual(stats["跳过合计行数"], 0)

    def test_skipped_rows_shift_the_reported_output_row(self) -> None:
        kept_rows = [
            HEADER,
            ["ZH0001", "2026-01-24", "", "海尔冰箱"],
            [None, None, None, None],
            ["合计", "合计", "", ""],
            # Source Excel row 6, but the blank and 合计 rows above never
            # reach the output sheet — this is only its second data row,
            # so 问题明细 must report 3 (an operator's coordinate in
            # Sheet1), not 6 (its position in the raw import).
            ["ZH0002", None, "", "格力空调"],
        ]

        _output_rows, _stats, issues = prepare(kept_rows)

        self.assertEqual(
            issues,
            [("缺少匹配键", "3", "", "日期或单据号为空，无法生成匹配键")],
        )


class ReceiptRemarkStatsTest(unittest.TestCase):
    def test_ordinary_remarks_still_counted(self) -> None:
        kept_rows = [
            HEADER,
            ["ZH0001", "2026-01-24", "260101ZH9999", "海尔冰箱"],
        ]

        _output_rows, stats, _issues = prepare(kept_rows)

        self.assertEqual(stats["仅退单数量"], 1)
        self.assertEqual(stats["备注总数"], 1)


class ReceiptOutputSortTest(unittest.TestCase):
    def test_remark_groups_follow_explicit_business_order(self) -> None:
        remarks = [
            receipts.RECEIPTS_REMARK_BOTH,
            receipts.RECEIPTS_REMARK_RETURN,
            receipts.RECEIPTS_REMARK_ORIGINAL,
            None,
        ]

        self.assertEqual(
            sorted(
                remarks,
                key=lambda remark: receipts.receipt_output_sort_key(
                    remark,
                    date(2026, 1, 1),
                    "ZH0001",
                    "商品",
                ),
            ),
            [
                None,
                receipts.RECEIPTS_REMARK_ORIGINAL,
                receipts.RECEIPTS_REMARK_RETURN,
                receipts.RECEIPTS_REMARK_BOTH,
            ],
        )

    def test_product_name_breaks_ties(self) -> None:
        """商品名称 is still the last sort key even though it is not written.

        Two rows alike in remark, date and 单据号 are indistinguishable in the
        output, so this is the only place the tie-break can be asserted.
        """
        key = receipts.receipt_output_sort_key
        self.assertLess(
            key(None, date(2026, 1, 1), "ZH0001", "A商品"),
            key(None, date(2026, 1, 1), "ZH0001", "Z商品"),
        )

    def test_rows_are_sorted_by_remark_date_document_and_product(self) -> None:
        output_rows, _stats, issues = prepare(
            [
                HEADER,
                ["ZH0002", "2026-01-02", "", "B商品"],
                ["ZH0004", "2026-01-01", "250101OLD4", "D退货"],
                ["ZH0001", "2026-01-01", "", "Z商品"],
                ["ZH0001", "2026-01-01", "", "A商品"],
                ["ZH0003", "2026-01-01", "250101OLD3", "C退货"],
            ]
        )

        self.assertEqual(issues, [])
        self.assertEqual(
            [(row[0], row[-1]) for row in output_rows],
            [
                # 商品名称 is no longer written out, so the two ZH0001 lines are
                # indistinguishable here; their A商品/Z商品 ordering is asserted
                # on the sort key itself in test_product_name_breaks_ties.
                ("ZH0001", None),
                ("ZH0001", None),
                ("ZH0002", None),
                ("ZH0003", receipts.RECEIPTS_REMARK_RETURN),
                ("ZH0004", receipts.RECEIPTS_REMARK_RETURN),
            ],
        )

    def test_issue_row_number_uses_the_sorted_output_position(self) -> None:
        output_rows, _stats, issues = prepare(
            [
                HEADER,
                ["ZH0009", None, "", "缺少日期"],
                ["ZH0001", "2026-01-01", "", "正常商品"],
            ]
        )

        self.assertEqual([row[0] for row in output_rows], ["ZH0001", "ZH0009"])
        self.assertEqual(
            issues,
            [("缺少匹配键", "3", "", "日期或单据号为空，无法生成匹配键")],
        )

    def test_missing_date_sorts_last_inside_a_remark_group(self) -> None:
        output_rows, _stats, issues = prepare(
            [
                HEADER,
                ["ZH0009", None, "250101OLD9", "缺少日期退货"],
                ["ZH0001", "2026-01-01", "250101OLD1", "正常日期退货"],
            ]
        )

        self.assertEqual([row[0] for row in output_rows], ["ZH0001", "ZH0009"])
        self.assertEqual(
            [row[-1] for row in output_rows],
            [receipts.RECEIPTS_REMARK_RETURN] * 2,
        )
        self.assertEqual(
            issues,
            [("缺少匹配键", "3", "", "日期或单据号为空，无法生成匹配键")],
        )

    def test_validation_rejects_unsorted_output(self) -> None:
        output_rows, _stats, issues = prepare(
            [
                HEADER,
                ["ZH0002", "2026-01-02", "", "B商品"],
                ["ZH0001", "2026-01-01", "", "A商品"],
            ]
        )
        output_rows.reverse()

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "收款单统计.xlsx"
            receipts._write_receipts_workbook(path, output_rows, issues)

            with self.assertRaisesRegex(RuntimeError, "收款单排序校验失败"):
                receipts.validate_receipts_output(path, output_rows, issues)


class ReceiptRemarkFillTest(unittest.TestCase):
    @staticmethod
    def _is_pink(cell) -> bool:
        return (
            cell.fill.patternType == "solid"
            and str(cell.fill.fgColor.rgb)[-6:]
            == receipts.RECEIPTS_REMARK_FILL_COLOR
        )

    def _write_and_read(self, kept_rows):
        output_rows, _stats, issues = prepare(kept_rows)
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        path = Path(directory.name) / "收款单统计.xlsx"
        receipts._write_receipts_workbook(path, output_rows, issues)
        workbook = load_workbook(path)
        self.addCleanup(workbook.close)
        receipts.validate_receipts_output(path, output_rows, issues)
        return output_rows, issues, workbook

    def test_validation_rejects_a_tampered_data_row(self) -> None:
        """The remark can no longer be re-derived from 原票号/摘要 here, so the
        read-back's remaining job is proving the writer wrote what was built."""
        output_rows, _stats, issues = prepare(
            [
                HEADER,
                ["ZH0001", "2026-01-24", "250101OLD1", "海尔冰箱"],
            ]
        )
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "收款单统计.xlsx"
            receipts._write_receipts_workbook(path, output_rows, issues)
            tampered = [list(output_rows[0])]
            tampered[0][0] = "ZH9999"

            with self.assertRaisesRegex(RuntimeError, "收款单数据行与生成结果不一致"):
                receipts.validate_receipts_output(path, tampered, issues)

    def test_suite_sale_duplicate_without_return_is_not_filled(self) -> None:
        output_rows, issues, workbook = self._write_and_read(
            [
                HEADER,
                ["ZH0001", "2026-01-24", "", "老板-欧式烟机-K1L"],
                ["ZH0001", "2026-01-24", "", "老板-嵌入式灶-9B5-B1"],
            ]
        )

        self.assertEqual([row[-1] for row in output_rows], [None, None])
        self.assertEqual(issues, [])
        self.assertEqual(workbook.sheetnames, ["Sheet1"])
        for row in workbook["Sheet1"].iter_rows(min_row=2):
            self.assertFalse(any(self._is_pink(cell) for cell in row))

    def test_return_referencing_suite_sale_marks_every_related_row(self) -> None:
        output_rows, issues, workbook = self._write_and_read(
            [
                HEADER,
                ["ZH0001", "2026-01-24", "", "老板-欧式烟机-K1L"],
                ["ZH0001", "2026-01-24", "", "老板-嵌入式灶-9B5-B1"],
                [
                    "ZH0002",
                    "2026-01-25",
                    "260124ZH0001",
                    "",
                    "老板烟灶套装退货",
                ],
            ]
        )

        self.assertEqual(
            [row[-1] for row in output_rows],
            [
                receipts.RECEIPTS_REMARK_ORIGINAL,
                receipts.RECEIPTS_REMARK_ORIGINAL,
                receipts.RECEIPTS_REMARK_RETURN,
            ],
        )
        self.assertEqual(issues, [])
        for row in workbook["Sheet1"].iter_rows(min_row=2):
            self.assertTrue(all(self._is_pink(cell) for cell in row))

    def test_prior_year_original_invoice_is_return_and_is_filled(self) -> None:
        output_rows, issues, workbook = self._write_and_read(
            [
                HEADER,
                [
                    "ZH0001",
                    "2026-01-24",
                    "250101ZH9999",
                    "",
                    "海尔冰箱",
                ],
            ]
        )

        self.assertEqual(
            output_rows[0][-1],
            receipts.RECEIPTS_REMARK_RETURN,
        )
        self.assertEqual(issues, [])
        self.assertEqual(workbook.sheetnames, ["Sheet1"])
        self.assertTrue(
            all(
                self._is_pink(cell)
                for cell in workbook["Sheet1"][2]
            )
        )


class IssuesSheetRoundTripTest(unittest.TestCase):
    def test_issues_sheet_is_written_and_validated(self) -> None:
        issues = [
            ("缺少匹配键", "5", "", "日期或单据号为空，无法生成匹配键"),
        ]
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "收款单统计.xlsx"
            receipts._write_receipts_workbook(
                path,
                [],
                issues,
            )

            workbook = load_workbook(path)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["Sheet1", receipts.ISSUES_SHEET_NAME],
                )
                issues_sheet = workbook[receipts.ISSUES_SHEET_NAME]
                self.assertEqual(
                    tuple(cell.value for cell in issues_sheet[1]),
                    receipts.ISSUES_HEADER,
                )
                self.assertEqual(
                    [
                        row
                        for row in issues_sheet.iter_rows(
                            min_row=2,
                            values_only=True,
                        )
                    ],
                    [("缺少匹配键", "5", None, "日期或单据号为空，无法生成匹配键")],
                )
            finally:
                workbook.close()

            receipts.validate_receipts_output(path, [], issues)

    def test_validation_rejects_mismatched_issues(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "收款单统计.xlsx"
            receipts._write_receipts_workbook(
                path,
                [],
                [("缺少匹配键", "5", "", "错误说明")],
            )

            with self.assertRaisesRegex(RuntimeError, "问题明细工作表内容校验失败"):
                receipts.validate_receipts_output(
                    path,
                    [],
                    [("缺少匹配键", "5", "", "说明")],
                )


if __name__ == "__main__":
    unittest.main()
