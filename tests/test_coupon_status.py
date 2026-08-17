import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from processors.coupons import appliance, matching
from processors.coupons import digital as coupons_digital
from processors.coupons.sources import (
    load_payment_reference_locations,
    load_uploaded_summary,
    validate_payment_reference_subset,
)
from processors.coupons.validation import (
    validate_payment_statuses,
    validate_row_statuses_and_matched_subsidy,
)

HEADERS = (coupons_digital.COUPON_OUTPUT_HEADER, appliance.COUPON_OUTPUT_HEADER)


def coupon_row(
    header: tuple[str, ...],
    reference: str,
    *,
    remark: str = "",
    detail: str = "",
) -> list[object]:
    values = {
        "明细摘要": reference,
        "备注": remark,
        "详细情况": detail,
    }
    return [values.get(column) for column in header]


class CouponStatusLookupTest(unittest.TestCase):
    def test_uploaded_lookup_normalizes_reference_and_builds_detail(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "uploaded.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Summary"
            sheet.append(["检索参考号", "状态", "描述", "补贴金额"])
            sheet.append([" 12345678901n ", " 已完成 ", "匹配成功 ", 1])
            workbook.save(source)
            workbook.close()

            self.assertEqual(
                load_uploaded_summary(source)[0],
                {"12345678901N": "已完成：匹配成功"},
            )

    def test_uploaded_lookup_rejects_malformed_reference(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "uploaded.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Summary"
            sheet.append(["检索参考号", "状态", "描述", "补贴金额"])
            sheet.append(["12345678901A", "已完成", "匹配成功", 1])
            workbook.save(source)
            workbook.close()

            with self.assertRaisesRegex(
                ValueError,
                "11位数字后跟大写字母 N",
            ):
                load_uploaded_summary(source)

    def test_uploaded_lookup_rejects_conflicting_duplicate(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "uploaded.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Summary"
            sheet.append(["检索参考号", "状态", "描述", "补贴金额"])
            sheet.append(["12345678901N", "已完成", "结果一", 1])
            sheet.append(["12345678901N", "已完成", "结果二", 1])
            workbook.save(source)
            workbook.close()

            with self.assertRaisesRegex(
                ValueError,
                "检索参考号存在冲突",
            ):
                load_uploaded_summary(source)


class CouponStatusFillTest(unittest.TestCase):
    def test_upload_status_is_filled_before_payment_matching(self) -> None:
        reference = "12345678901N"
        for header in HEADERS:
            with self.subTest(header=header):
                row = coupon_row(header, "12345678901n")
                rows = [list(header), row]
                lookup = {reference: "已完成：匹配成功"}

                upload_counts = matching.fill_upload_statuses(
                    rows,
                    lookup,
                )
                paid_count = matching.fill_payment_statuses(rows, {reference})

                remark_index = header.index("备注")
                detail_index = header.index("详细情况")
                payment_index = header.index("回款情况")
                self.assertEqual(upload_counts, (1, 0))
                self.assertEqual(paid_count, 1)
                self.assertEqual(row[remark_index], "已上传")
                self.assertEqual(row[detail_index], "已完成：匹配成功")
                self.assertEqual(row[payment_index], "已回款")

    def test_reference_outside_submitted_data_is_marked_unsubmitted(self) -> None:
        """Without unsubmitted data, anything not submitted is 未上传."""
        for header in HEADERS:
            with self.subTest(header=header):
                row = coupon_row(header, "99999999999X")
                rows = [list(header), row]

                upload_counts = matching.fill_upload_statuses(
                    rows,
                    {},
                )
                paid_count = matching.fill_payment_statuses(rows, set())

                remark_index = header.index("备注")
                payment_index = header.index("回款情况")
                self.assertEqual(upload_counts, (0, 1))
                self.assertEqual(paid_count, 0)
                self.assertEqual(row[remark_index], "未上传")
                self.assertIsNone(row[payment_index])

    def test_payment_reference_is_ignored_until_remark_is_uploaded(self) -> None:
        reference = "12345678901N"
        for header in HEADERS:
            with self.subTest(header=header):
                row = coupon_row(header, reference)
                rows = [list(header), row]

                paid_count = matching.fill_payment_statuses(rows, {reference})

                self.assertEqual(paid_count, 0)
                self.assertEqual(row[header.index("备注")], "")
                self.assertIsNone(row[header.index("回款情况")])

    def test_only_exact_uploaded_remark_is_eligible_for_payment(self) -> None:
        reference = "12345678901N"
        header = appliance.COUPON_OUTPUT_HEADER
        rows = [
            list(header),
            coupon_row(header, reference, remark="已上传"),
            coupon_row(header, reference, remark="未上传"),
            coupon_row(header, reference, remark="退换货/倒票（退单）"),
            coupon_row(header, reference, remark=" 已上传 "),
        ]

        paid_count = matching.fill_payment_statuses(rows, {reference})

        payment_index = header.index("回款情况")
        self.assertEqual(paid_count, 1)
        self.assertEqual(rows[1][payment_index], "已回款")
        for row in rows[2:]:
            self.assertIsNone(row[payment_index])

    def test_excluded_bottom_rows_are_left_alone(self) -> None:
        reference = "12345678901N"
        header = appliance.COUPON_OUTPUT_HEADER
        uploaded_row = coupon_row(
            header,
            reference,
            remark="已上传",
            detail="已完成：匹配成功",
        )
        bottom_row = coupon_row(header, "99999999999X")
        rows = [list(header), uploaded_row, bottom_row]

        upload_counts = matching.fill_upload_statuses(
            rows,
            {reference: "已完成：匹配成功"},
            excluded_bottom_rows=1,
        )
        paid_count = matching.fill_payment_statuses(
            rows,
            {reference},
            excluded_bottom_rows=1,
        )

        remark_index = header.index("备注")
        payment_index = header.index("回款情况")
        self.assertEqual(upload_counts, (1, 0))
        self.assertEqual(paid_count, 1)
        self.assertEqual(uploaded_row[remark_index], "已上传")
        self.assertEqual(uploaded_row[payment_index], "已回款")
        self.assertEqual(bottom_row[remark_index], "")
        self.assertIsNone(bottom_row[payment_index])

    def test_validation_rejects_an_incorrect_payment_status(self) -> None:
        reference = "12345678901N"
        header = appliance.COUPON_OUTPUT_HEADER
        row = coupon_row(header, reference)
        row[header.index("回款情况")] = "已回款"
        rows = [list(header), row]

        with self.assertRaisesRegex(RuntimeError, "回款情况匹配校验失败"):
            validate_payment_statuses(
                rows,
                header,
                {reference},
                excluded_bottom_rows=0,
                expected_paid_rows=1,
            )


class PaymentReferenceSourceTest(unittest.TestCase):
    def test_loads_appliance_and_digital_references_separately(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "回款明细.xlsx"
            workbook = Workbook()
            appliance_sheet = workbook.active
            appliance_sheet.title = "家电明细"
            appliance_sheet.append(["交易参考号"])
            appliance_sheet.append(["12345678901n"])
            digital_sheet = workbook.create_sheet("数码明细")
            digital_sheet.append(["交易参考号"])
            digital_sheet.append(["12345678902N"])
            workbook.save(source)
            workbook.close()

            result = load_payment_reference_locations(source)

        self.assertEqual(set(result["家电"]), {"12345678901N"})
        self.assertEqual(set(result["数码"]), {"12345678902N"})

    def test_requires_both_detail_sheets_and_reference_headers(self) -> None:
        cases = (
            (False, True, "缺少 数码明细 工作表"),
            (True, False, "家电明细 缺少字段：交易参考号"),
        )
        for include_digital, include_header, expected in cases:
            with self.subTest(expected=expected), tempfile.TemporaryDirectory() as directory:
                source = Path(directory) / "回款明细.xlsx"
                workbook = Workbook()
                appliance_sheet = workbook.active
                appliance_sheet.title = "家电明细"
                appliance_sheet.append(
                    ["交易参考号"] if include_header else ["订单号"]
                )
                if include_digital:
                    digital_sheet = workbook.create_sheet("数码明细")
                    digital_sheet.append(["交易参考号"])
                workbook.save(source)
                workbook.close()

                with self.assertRaisesRegex(ValueError, expected):
                    load_payment_reference_locations(source)

    def test_rejects_payment_reference_outside_submitted_data(self) -> None:
        locations = {"12345678901N": "回款明细.xlsx 的 家电明细 第 2 行"}

        with self.assertRaisesRegex(
            ValueError,
            "家电回款参考号子集校验失败.*第 2 行",
        ):
            validate_payment_reference_subset("家电", locations, set())


class RowStatusesAndMatchedSubsidyTest(unittest.TestCase):
    def test_accumulates_matched_partition_and_tracks_supplement_matches(self) -> None:
        header = appliance.COUPON_OUTPUT_HEADER
        subsidy_header = appliance.COUPON_SUBSIDY_HEADER
        doc_idx = header.index("单据号")
        date_idx = header.index("单据日期")
        summary_idx = header.index("明细摘要")
        subsidy_idx = header.index(subsidy_header)
        remark_idx = header.index("备注")
        detail_idx = header.index("详细情况")

        # Row 1: Regular row matching reference supplement
        row1 = [None] * len(header)
        row1[doc_idx] = "DOC001"
        row1[date_idx] = date(2026, 1, 1)
        row1[summary_idx] = "REF001"
        row1[subsidy_idx] = Decimal("100.00")
        row1[remark_idx] = "已上传"
        row1[detail_idx] = "已完成：匹配成功"

        # Row 2: Matched partition row (退换货)
        row2 = [None] * len(header)
        row2[doc_idx] = "DOC002"
        row2[date_idx] = date(2026, 1, 2)
        row2[summary_idx] = "REF002"
        row2[subsidy_idx] = Decimal("-50.00")
        row2[remark_idx] = "退货备注"
        row2[detail_idx] = ""

        # Row 3: Matched partition row with positive subsidy
        row3 = [None] * len(header)
        row3[doc_idx] = "DOC003"
        row3[date_idx] = date(2026, 1, 3)
        row3[summary_idx] = "REF003"
        row3[subsidy_idx] = Decimal("20.00")
        row3[remark_idx] = "换货备注"
        row3[detail_idx] = ""

        rows = [list(header), row1, row2, row3]

        remark_lookup = {
            ("DOC001", date(2026, 1, 1)): "原始备注",
            ("DOC002", date(2026, 1, 2)): "退货备注",
            ("DOC003", date(2026, 1, 3)): "换货备注",
        }
        detail_lookup = {"REF001": "已完成：匹配成功"}
        reference_universe = {"REF001"}
        expected_supplements = {("DOC001", date(2026, 1, 1), "REF001"): 1}

        total = validate_row_statuses_and_matched_subsidy(
            rows,
            header=header,
            subsidy_header=subsidy_header,
            remark_lookup=remark_lookup,
            detail_lookup=detail_lookup,
            reference_universe=reference_universe,
            expected_matched_rows=2,
            expected_reference_supplement_matches=expected_supplements,
        )

        self.assertEqual(total, Decimal("-30.00"))

    def test_rejects_insufficient_supplement_matches(self) -> None:
        header = appliance.COUPON_OUTPUT_HEADER
        subsidy_header = appliance.COUPON_SUBSIDY_HEADER
        doc_idx = header.index("单据号")
        date_idx = header.index("单据日期")
        summary_idx = header.index("明细摘要")
        subsidy_idx = header.index(subsidy_header)
        remark_idx = header.index("备注")
        detail_idx = header.index("详细情况")

        # Row 1: Regular row matching REF001
        row1 = [None] * len(header)
        row1[doc_idx] = "DOC001"
        row1[date_idx] = date(2026, 1, 1)
        row1[summary_idx] = "REF001"
        row1[subsidy_idx] = Decimal("100.00")
        row1[remark_idx] = "已上传"
        row1[detail_idx] = "已完成：匹配成功"

        rows = [list(header), row1]

        remark_lookup = {("DOC001", date(2026, 1, 1)): "原始备注"}
        detail_lookup = {"REF001": "已完成：匹配成功"}
        reference_universe = {"REF001"}
        # Expecting 2 matches but only 1 provided in rows
        expected_supplements = {("DOC001", date(2026, 1, 1), "REF001"): 2}

        with self.assertRaisesRegex(
            RuntimeError, "销售用券补充参考号逐行匹配结果校验失败"
        ):
            validate_row_statuses_and_matched_subsidy(
                rows,
                header=header,
                subsidy_header=subsidy_header,
                remark_lookup=remark_lookup,
                detail_lookup=detail_lookup,
                reference_universe=reference_universe,
                expected_matched_rows=0,
                expected_reference_supplement_matches=expected_supplements,
            )

    def test_regular_partition_remarks_unuploaded_or_fallback_to_receipt_remark(self) -> None:
        header = coupons_digital.COUPON_OUTPUT_HEADER
        subsidy_header = coupons_digital.COUPON_SUBSIDY_HEADER
        doc_idx = header.index("单据号")
        date_idx = header.index("单据日期")
        summary_idx = header.index("明细摘要")
        subsidy_idx = header.index(subsidy_header)
        remark_idx = header.index("备注")
        detail_idx = header.index("详细情况")

        # Row 1: Not in reference universe -> 未上传
        row1 = [None] * len(header)
        row1[doc_idx] = "DOC001"
        row1[date_idx] = date(2026, 1, 1)
        row1[summary_idx] = "REF_NOT_IN_UNIVERSE"
        row1[subsidy_idx] = Decimal("100.00")
        row1[remark_idx] = "未上传"
        row1[detail_idx] = ""

        # Row 2: In universe but not uploaded -> keeps receipt remark
        row2 = [None] * len(header)
        row2[doc_idx] = "DOC002"
        row2[date_idx] = date(2026, 1, 2)
        row2[summary_idx] = "REF_IN_UNIVERSE"
        row2[subsidy_idx] = Decimal("200.00")
        row2[remark_idx] = "回单备注"
        row2[detail_idx] = ""

        rows = [list(header), row1, row2]

        remark_lookup = {
            ("DOC001", date(2026, 1, 1)): "原始备注",
            ("DOC002", date(2026, 1, 2)): "回单备注",
        }
        detail_lookup = {}
        reference_universe = {"REF_IN_UNIVERSE"}

        total = validate_row_statuses_and_matched_subsidy(
            rows,
            header=header,
            subsidy_header=subsidy_header,
            remark_lookup=remark_lookup,
            detail_lookup=detail_lookup,
            reference_universe=reference_universe,
            expected_matched_rows=0,
        )

        self.assertEqual(total, Decimal("0"))


if __name__ == "__main__":
    unittest.main()
