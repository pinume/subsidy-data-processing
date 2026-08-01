import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from xlsxwriter import Workbook as XlsxWorkbook

from processors.common.dates import (
    normalize_coupon_date,
    normalize_document_number,
    normalize_receipt_date,
)
from processors.common.excel import (
    load_measurement_font,
    pixels_to_excel_width,
    remove_stale_temporary_files,
    resolve_font,
    run_with_output_rollback,
    save_workbook_atomically,
    write_formatted_sheet,
    write_xlsx_atomically,
)


class RemoveStaleTemporaryFilesTest(unittest.TestCase):
    def test_removes_only_dot_prefixed_files(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory)
            keep = (
                "回款明细.xlsx",
                "家电_已上传.xlsx",
                "~$回款明细.xlsx",
            )
            remove = (
                ".收款单统计-evx66gk1.xlsx",
                ".2026年数码补贴明细.xlsx.working.xlsx",
            )
            for name in (*keep, *remove):
                (output_dir / name).write_bytes(b"x")
            (output_dir / ".子目录").mkdir()

            removed = remove_stale_temporary_files(output_dir, minimum_age_seconds=0)

            self.assertEqual(sorted(removed), sorted(remove))
            self.assertEqual(
                sorted(path.name for path in output_dir.iterdir()),
                sorted([*keep, ".子目录"]),
            )

    def test_missing_output_directory_is_not_an_error(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            self.assertEqual(
                remove_stale_temporary_files(Path(directory) / "output"),
                [],
            )

    def test_leaves_a_freshly_written_file_alone(self) -> None:
        # A second instance's startup cleanup must not delete a first
        # instance's temporary file while save_workbook_atomically is still
        # writing to it — that would turn a harmless leftover-file feature
        # into a way for one run to corrupt another's in-flight save.
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory)
            in_flight = output_dir / ".收款单统计-abc123.xlsx"
            in_flight.write_bytes(b"x")

            removed = remove_stale_temporary_files(
                output_dir, minimum_age_seconds=180
            )

            self.assertEqual(removed, [])
            self.assertTrue(in_flight.exists())


SOURCE_NAME = "收款单统计.xlsx"


class DateHelpersTest(unittest.TestCase):
    def test_normalize_receipt_date_accepts_supported_formats(self) -> None:
        expected = date(2026, 7, 6)
        for value in ("2026-07-06", "2026/07/06", "20260706"):
            with self.subTest(value=value):
                self.assertEqual(
                    normalize_receipt_date(value, 3, SOURCE_NAME), expected
                )
        self.assertEqual(
            normalize_receipt_date(datetime(2026, 7, 6), 3, SOURCE_NAME), expected
        )
        self.assertEqual(
            normalize_receipt_date(date(2026, 7, 6), 3, SOURCE_NAME), expected
        )

    def test_normalize_receipt_date_treats_blank_as_missing(self) -> None:
        self.assertIsNone(normalize_receipt_date(None, 3, SOURCE_NAME))
        self.assertIsNone(normalize_receipt_date("", 3, SOURCE_NAME))

    def test_normalize_receipt_date_error_names_file_row_and_value(self) -> None:
        with self.assertRaises(ValueError) as caught:
            normalize_receipt_date("2026.7.6", 128, SOURCE_NAME)

        message = str(caught.exception)
        self.assertIn(SOURCE_NAME, message)
        self.assertIn("128", message)
        self.assertIn("2026.7.6", message)
        self.assertIn("YYYY-MM-DD", message)

    def test_normalize_coupon_date_rejects_invalid_value(self) -> None:
        with self.assertRaisesRegex(ValueError, "row 9"):
            normalize_coupon_date("not-a-date", 9)

    def test_normalize_document_number_only_removes_prefix(self) -> None:
        self.assertEqual(normalize_document_number("收款123收款"), "123收款")


class AtomicSaveTest(unittest.TestCase):
    def test_replaces_output_only_after_validation(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "result.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "ok"

            save_workbook_atomically(
                workbook,
                output,
                lambda path: self.assertTrue(path.exists()),
            )

            saved = load_workbook(output, read_only=True)
            try:
                self.assertEqual(saved.active["A1"].value, "ok")
            finally:
                saved.close()

    def test_validation_failure_does_not_replace_existing_output(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "result.xlsx"
            existing = Workbook()
            existing.active["A1"] = "old"
            existing.save(output)
            existing.close()

            replacement = Workbook()
            replacement.active["A1"] = "new"

            def reject(_path: Path) -> None:
                raise RuntimeError("invalid")

            with self.assertRaisesRegex(RuntimeError, "invalid"):
                save_workbook_atomically(replacement, output, reject)

            saved = load_workbook(output, read_only=True)
            try:
                self.assertEqual(saved.active["A1"].value, "old")
            finally:
                saved.close()

    def test_path_writer_replaces_output_only_after_validation(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "result.xlsx"

            write_xlsx_atomically(
                output,
                lambda path: path.write_bytes(b"valid xlsx placeholder"),
                lambda path: self.assertEqual(
                    path.read_bytes(),
                    b"valid xlsx placeholder",
                ),
            )

            self.assertEqual(output.read_bytes(), b"valid xlsx placeholder")

    def test_path_writer_validation_failure_preserves_existing_output(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "result.xlsx"
            output.write_bytes(b"old")

            def reject(_path: Path) -> None:
                raise RuntimeError("invalid")

            with self.assertRaisesRegex(RuntimeError, "invalid"):
                write_xlsx_atomically(
                    output,
                    lambda path: path.write_bytes(b"new"),
                    reject,
                )

            self.assertEqual(output.read_bytes(), b"old")
            self.assertEqual([path.name for path in output.parent.iterdir()], [output.name])


class OutputRollbackTest(unittest.TestCase):
    def test_restores_existing_outputs_and_removes_new_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory)
            existing = output_dir / "existing.xlsx"
            new = output_dir / "new.xlsx"
            existing.write_bytes(b"old")

            def fail_after_writes() -> None:
                existing.write_bytes(b"new")
                new.write_bytes(b"new")
                raise ValueError("later step failed")

            with self.assertRaisesRegex(ValueError, "later step failed"):
                run_with_output_rollback((existing, new), fail_after_writes)

            self.assertEqual(existing.read_bytes(), b"old")
            self.assertFalse(new.exists())
            self.assertEqual(
                [path.name for path in output_dir.iterdir()],
                ["existing.xlsx"],
            )

    def test_keeps_successful_outputs_and_removes_backups(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory)
            first = output_dir / "first.xlsx"
            second = output_dir / "second.xlsx"
            first.write_bytes(b"old")

            def succeed() -> str:
                first.write_bytes(b"new")
                second.write_bytes(b"new")
                return "done"

            self.assertEqual(
                run_with_output_rollback((first, second), succeed),
                "done",
            )
            self.assertEqual(first.read_bytes(), b"new")
            self.assertEqual(second.read_bytes(), b"new")
            self.assertEqual(
                sorted(path.name for path in output_dir.iterdir()),
                ["first.xlsx", "second.xlsx"],
            )

    def test_failed_backup_copy_does_not_leave_temporary_file(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory)
            existing = output_dir / "existing.xlsx"
            existing.write_bytes(b"old")

            with patch(
                "processors.common.excel.shutil.copy2",
                side_effect=OSError("disk full"),
            ):
                with self.assertRaisesRegex(OSError, "disk full"):
                    run_with_output_rollback((existing,), lambda: None)

            self.assertEqual(existing.read_bytes(), b"old")
            self.assertEqual(
                [path.name for path in output_dir.iterdir()],
                ["existing.xlsx"],
            )


class WriteFormattedSheetTest(unittest.TestCase):
    def test_native_dates_get_the_formats_openpyxl_applied_on_its_own(self) -> None:
        """XlsxWriter formats nothing unless told to; openpyxl stamped a date
        format onto the cell as soon as a date was assigned. Without this a
        native date would render as its serial number (2026-01-02 as 46024).
        Both writers must produce the same-looking cell."""
        font_name, font_path = resolve_font()
        measurement_font = load_measurement_font(font_path)
        values = [date(2026, 1, 2), datetime(2026, 1, 2, 13, 5), "2026-01-02"]

        with tempfile.TemporaryDirectory() as directory:
            openpyxl_path = Path(directory) / "openpyxl.xlsx"
            reference = Workbook()
            reference.active.append(["日期", "时间", "文本"])
            reference.active.append(values)
            reference.save(openpyxl_path)
            reference.close()

            xlsx_path = Path(directory) / "xlsxwriter.xlsx"
            with XlsxWorkbook(str(xlsx_path)) as workbook:
                write_formatted_sheet(
                    workbook,
                    "Sheet1",
                    ["日期", "时间", "文本"],
                    [values],
                    font_name,
                    measurement_font,
                )

            expected = load_workbook(openpyxl_path).active
            actual = load_workbook(xlsx_path)["Sheet1"]
            for column in (1, 2, 3):
                with self.subTest(column=column):
                    self.assertEqual(
                        actual.cell(2, column).number_format,
                        expected.cell(2, column).number_format,
                    )
                    self.assertEqual(
                        actual.cell(2, column).value,
                        expected.cell(2, column).value,
                    )


class ExcelHelpersTest(unittest.TestCase):
    def test_excel_column_width_is_capped_at_format_limit(self) -> None:
        self.assertEqual(pixels_to_excel_width(100_000), 255)

    def test_resolve_font_uses_first_existing_candidate(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            missing = Path(directory) / "missing.ttf"
            existing = Path(directory) / "existing.ttf"
            existing.touch()

            with patch(
                "processors.common.excel.FONT_CANDIDATES",
                (
                    ("Primary", (missing,), ()),
                    ("Fallback", (existing,), ()),
                ),
            ), patch.dict("os.environ", {}, clear=True):
                self.assertEqual(resolve_font(), ("Fallback", existing))

    def test_resolve_font_searches_font_roots_by_file_name(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            font_root = Path(directory) / "fonts"
            font_path = font_root / "nested" / "candidate.ttf"
            font_path.parent.mkdir(parents=True)
            font_path.touch()

            with patch(
                "processors.common.excel.FONT_CANDIDATES",
                (("Candidate", (), ("candidate.ttf",)),),
            ), patch(
                "processors.common.excel.FONT_SEARCH_ROOTS",
                (font_root,),
            ), patch.dict("os.environ", {}, clear=True):
                self.assertEqual(resolve_font(), ("Candidate", font_path))

    def test_resolve_font_uses_configured_font_path(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            font_path = Path(directory) / "custom.ttf"
            font_path.touch()

            with patch.dict(
                "os.environ",
                {
                    "UPLOAD_DATA_FONT_PATH": str(font_path),
                    "UPLOAD_DATA_FONT_NAME": "Custom Font",
                },
                clear=True,
            ):
                self.assertEqual(resolve_font(), ("Custom Font", font_path))

if __name__ == "__main__":
    unittest.main()
