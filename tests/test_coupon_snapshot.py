"""The snapshot reader must agree with openpyxl on everything it replaces.

Each assertion below is paired against openpyxl reading the same file, because
the snapshot exists only to be a faster substitute for that read. A difference
between them is a bug in the substitute, whichever value looks more plausible.
"""

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from xlsxwriter import Workbook

from processors.coupons.snapshot import (
    a1_range,
    column_letter,
    read_workbook_snapshot,
)


def write_fixture(path: Path) -> None:
    """A workbook shaped like the audit output: merges, freeze, filter, widths."""
    with Workbook(str(path)) as workbook:
        merged = workbook.add_worksheet("数据汇总")
        centered = workbook.add_format({"align": "center"})
        merged.write_row(0, 0, ["财务大类", "品牌", "补贴金额"])
        for row in range(1, 7):
            merged.write_row(row, 0, ["空调", f"品牌{row}", row * 1.5])
        merged.merge_range(1, 0, 3, 0, "空调", centered)
        merged.merge_range(4, 0, 6, 0, "冰箱", centered)
        merged.freeze_panes(1, 0)
        merged.set_column_pixels(0, 0, 120)
        merged.set_column_pixels(2, 2, 200)

        plain = workbook.add_worksheet("明细")
        plain.write_row(0, 0, ["单据号", "日期"])
        plain.write_row(1, 0, ["ZH0001", "2026-01-24"])
        plain.autofilter(0, 0, 1, 1)

        workbook.add_worksheet("空表")


class ColumnLetterTest(unittest.TestCase):
    def test_matches_openpyxl_for_the_boundaries_that_carry(self) -> None:
        from openpyxl.utils import get_column_letter

        for index in (0, 1, 25, 26, 27, 51, 52, 701, 702):
            with self.subTest(index=index):
                self.assertEqual(column_letter(index), get_column_letter(index + 1))

    def test_a1_range_is_one_based_and_inclusive(self) -> None:
        self.assertEqual(a1_range((0, 0), (0, 0)), "A1:A1")
        self.assertEqual(a1_range((1, 0), (3, 0)), "A2:A4")
        self.assertEqual(a1_range((0, 2), (5, 27)), "C1:AB6")


class SnapshotAgreesWithOpenpyxlTest(unittest.TestCase):
    def setUp(self) -> None:
        self._directory = tempfile.TemporaryDirectory()
        self.path = Path(self._directory.name) / "audit.xlsx"
        write_fixture(self.path)
        self.snapshot = read_workbook_snapshot(self.path)
        self.workbook = load_workbook(self.path)
        self.addCleanup(self.workbook.close)
        self.addCleanup(self._directory.cleanup)

    def test_sheet_names_keep_workbook_order(self) -> None:
        self.assertEqual(self.snapshot.sheet_names, tuple(self.workbook.sheetnames))

    def test_merged_ranges_match(self) -> None:
        for name in self.snapshot.sheet_names:
            with self.subTest(sheet=name):
                self.assertEqual(
                    self.snapshot.merged_ranges[name],
                    {str(r) for r in self.workbook[name].merged_cells.ranges},
                )

    def test_freeze_panes_match(self) -> None:
        for name in self.snapshot.sheet_names:
            with self.subTest(sheet=name):
                self.assertEqual(
                    self.snapshot.metadata[name].freeze_panes,
                    self.workbook[name].freeze_panes or None,
                )

    def test_autofilter_ref_matches(self) -> None:
        for name in self.snapshot.sheet_names:
            with self.subTest(sheet=name):
                self.assertEqual(
                    self.snapshot.metadata[name].autofilter_ref,
                    self.workbook[name].auto_filter.ref or None,
                )

    def test_column_widths_match(self) -> None:
        for name in self.snapshot.sheet_names:
            expected = {}
            for dimension in self.workbook[name].column_dimensions.values():
                if not dimension.width:
                    continue
                first = dimension.min or 1
                for index in range(first, (dimension.max or first) + 1):
                    expected[index] = dimension.width
            with self.subTest(sheet=name):
                self.assertEqual(
                    self.snapshot.metadata[name].column_widths, expected
                )

    def test_values_match(self) -> None:
        for name in self.snapshot.sheet_names:
            expected = [
                tuple(row)
                for row in self.workbook[name].iter_rows(values_only=True)
            ]
            with self.subTest(sheet=name):
                self.assertEqual(self.snapshot.values[name], expected)

    def test_a_sheet_with_no_cells_yields_no_rows(self) -> None:
        """calamine's iter_rows() panics on an empty sheet; it must not reach it."""
        self.assertEqual(self.snapshot.values["空表"], [])
        self.assertEqual(self.snapshot.merged_ranges["空表"], frozenset())
