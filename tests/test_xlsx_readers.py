"""Regression coverage for the two hot-path .xlsx readers introduced when
receipts and coupon sources moved off xlrd.

Both processors/receipts.py:read_receipt_rows and
processors/coupons/sources.py:read_coupon_export used to
call sheet.cell(row, column) once per row on a read_only worksheet, which
silently degrades to O(n^2) on a real 10000+ row export (each random access
re-parses the sheet XML from the start). Nothing caught that until a real
file was run through main.py. These tests pin the correct output on small
synthetic fixtures shaped like the real exports, and assert iter_rows() is
called exactly once per read so a future edit can't reintroduce the same
random-access pattern without a test failing.
"""

import tempfile
import unittest
from contextlib import ExitStack
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import Workbook, load_workbook
from python_calamine import CalamineWorkbook

from processors import receipts
from processors.coupons import appliance, digital, sources
from processors.receipts import read_receipt_rows


class _CountingSheetWrapper:
    """Wraps a worksheet, counting iter_rows() calls made through it."""

    def __init__(self, sheet):
        self._sheet = sheet
        self.iter_rows_calls = 0

    def iter_rows(self, *args, **kwargs):
        self.iter_rows_calls += 1
        return self._sheet.iter_rows(*args, **kwargs)

    def __getattr__(self, name):
        return getattr(self._sheet, name)


class _CountingCalamineSheetWrapper:
    """Wraps a calamine sheet, counting iter_rows() calls made through it."""

    def __init__(self, sheet):
        self._sheet = sheet
        self.iter_rows_calls = 0

    @property
    def start(self):
        return self._sheet.start

    def iter_rows(self):
        self.iter_rows_calls += 1
        return self._sheet.iter_rows()


class _CountingCalamineWorkbookWrapper:
    """Wraps a calamine workbook so get_sheet_by_index(0) returns a
    call-counting sheet."""

    def __init__(self, workbook):
        self._workbook = workbook
        self.sheet_wrapper = _CountingCalamineSheetWrapper(
            workbook.get_sheet_by_index(0)
        )

    def get_sheet_by_index(self, index):
        return self.sheet_wrapper

    def close(self):
        self._workbook.close()


class _CountingWorkbookWrapper:
    """Wraps a workbook so worksheets[0] returns a call-counting sheet."""

    def __init__(self, workbook):
        self._workbook = workbook
        self.sheet_wrapper = _CountingSheetWrapper(workbook.worksheets[0])

    @property
    def worksheets(self):
        return [self.sheet_wrapper]

    @property
    def sheetnames(self):
        return self._workbook.sheetnames

    def __getitem__(self, name):
        if name == self.sheet_wrapper.title:
            return self.sheet_wrapper
        return self._workbook[name]

    def close(self):
        self._workbook.close()


def _pad(row: list[object]) -> list[object]:
    """Widen a fixture row to the full source header, leaving 销售类别 blank."""
    return [*row, *[None] * (len(receipts.RECEIPTS_SOURCE_HEADER) - len(row))]


def _write_receipt_source(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["收款单统计"])
    sheet.append(list(receipts.RECEIPTS_SOURCE_HEADER))
    for row in rows:
        sheet.append(_pad(row))
    workbook.save(path)
    workbook.close()


class ReadReceiptRowsTest(unittest.TestCase):
    def test_合计行与空白行保留在原位置(self) -> None:
        """Filtering is prepare_receipt_data's job, not this function's.

        Dropping rows here would renumber every row after them, and the
        reported Excel row number of a problem row is derived from position
        in this list — so the rows stay, blanks and 合计 included.
        """
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "收款单统计.xlsx"
            _write_receipt_source(
                source,
                [
                    ["收款ZH0001", "2026-01-24", "", "", "海尔冰箱"],
                    [None, None, None, None, None],
                    ["收款ZH0002", "2026-01-25", "", "", "美的空调"],
                    ["合计", "合计", "", "", ""],
                ],
            )

            rows = read_receipt_rows(source)

            self.assertEqual(
                rows,
                [
                    list(receipts.RECEIPTS_SOURCE_HEADER),
                    _pad(["收款ZH0001", "2026-01-24", None, None, "海尔冰箱"]),
                    _pad([None, None, None, None, None]),
                    _pad(["收款ZH0002", "2026-01-25", None, None, "美的空调"]),
                    _pad(["合计", "合计", None, None, None]),
                ],
            )

    def test_missing_header_row_raises(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "收款单统计.xlsx"
            workbook = Workbook()
            workbook.active.append(["收款单统计"])
            workbook.save(source)
            workbook.close()

            with self.assertRaises(ValueError):
                read_receipt_rows(source)

    def test_reads_via_exactly_one_iter_rows_pass(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "收款单统计.xlsx"
            _write_receipt_source(
                source,
                [
                    ["收款ZH0001", "2026-01-24", "", "", "海尔冰箱"],
                    ["合计", "合计", "", "", ""],
                ],
            )
            workbook = CalamineWorkbook.from_path(str(source))
            wrapped = _CountingCalamineWorkbookWrapper(workbook)
            # read_receipt_rows opens its own workbook internally (and closes
            # it in its finally block) — count via monkeypatching
            # CalamineWorkbook.from_path to hand back the wrapper.
            import processors.receipts as receipts_module

            with patch.object(
                receipts_module.CalamineWorkbook,
                "from_path",
                return_value=wrapped,
            ):
                read_receipt_rows(source)

            self.assertEqual(wrapped.sheet_wrapper.iter_rows_calls, 1)


class ReceiptOutputPerformanceTest(unittest.TestCase):
    def _build_output_rows(self, row_count: int = 100):
        receipt_date = datetime(2026, 1, 24)
        return [
            [f"ZH{index:04d}", receipt_date, None]
            for index in range(row_count)
        ]

    def test_remark_fill_and_number_formats_are_written(self) -> None:
        rows = self._build_output_rows()
        for index in range(0, 100, 2):
            rows[index][-1] = receipts.RECEIPTS_REMARK_RETURN
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "收款单统计.xlsx"
            receipts._write_receipts_workbook(
                path,
                rows,
                [],
            )
            workbook = load_workbook(path)
            try:
                sheet = workbook["Sheet1"]
                self.assertEqual(sheet["A2"].number_format, "@")
                self.assertEqual(sheet["B2"].number_format, "yyyymmdd")
                self.assertEqual(
                    sheet["A2"].fill.fgColor.rgb[-6:],
                    receipts.RECEIPTS_REMARK_FILL_COLOR[-6:],
                )
                self.assertIsNone(sheet["A3"].fill.fill_type)
            finally:
                workbook.close()

    def test_saved_output_validation_reads_rows_once(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "收款单统计.xlsx"
            receipts._write_receipts_workbook(
                path,
                self._build_output_rows(row_count=1),
                [],
            )

            loaded = load_workbook(path, read_only=True, data_only=True)
            wrapped = _CountingWorkbookWrapper(loaded)
            with patch.object(receipts, "load_workbook", return_value=wrapped):
                receipts.validate_receipts_output(
                    path, self._build_output_rows(row_count=1), []
                )

            self.assertEqual(wrapped.sheet_wrapper.iter_rows_calls, 1)


def _write_coupon_source(path: Path, data_rows: list[list[object]]) -> None:
    """Build a minimal 28-column export matching the real column layout at
    the positions read_coupon_export actually uses (title/header/合计 rows,
    columns 3/4/6/8/15/18/26/27); every other column is left blank."""
    header = [None] * 28
    header[2] = "单据号"
    header[3] = "单据日期"
    header[5] = "商品名称"
    header[7] = "品牌"
    header[14] = "财务大类"
    header[17] = "明细摘要"
    header[25] = "2026家电国补（计入收入）"
    header[26] = "2026数码国补（计入收入）"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["销售用券情况统计"])
    sheet.append(header)
    for row in data_rows:
        sheet.append(row)
    total_row = [None] * 28
    total_row[0] = "合计"
    total_row[25] = "￥123.40"
    sheet.append(total_row)
    workbook.save(path)
    workbook.close()


def _coupon_row(
    *,
    document: str,
    day: str,
    product: str,
    brand: str,
    category: str,
    summary: str,
    family_subsidy: object = None,
    digital_subsidy: object = None,
) -> list[object]:
    row = [None] * 28
    row[2] = document
    row[3] = day
    row[5] = product
    row[7] = brand
    row[14] = category
    row[17] = summary
    row[25] = family_subsidy
    row[26] = digital_subsidy
    return row


class ReadCouponExportTest(unittest.TestCase):
    def test_classifies_by_which_subsidy_column_is_populated(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "销售用券情况统计.xlsx"
            _write_coupon_source(
                source,
                [
                    _coupon_row(
                        document="收款001",
                        day="2026-01-24",
                        product="海尔冰箱",
                        brand="海尔",
                        category="冰箱",
                        summary="ref1",
                        family_subsidy=100,
                    ),
                    _coupon_row(
                        document="收款002",
                        day="2026-01-25",
                        product="小米手机",
                        brand="小米",
                        category="手机",
                        summary="ref2",
                        digital_subsidy=50,
                    ),
                ],
            )

            export = sources.read_coupon_export(source)

            self.assertEqual(len(export.appliance_rows) - 1, 1)
            self.assertEqual(export.appliance_rows[1][0], "001")
            self.assertEqual(export.appliance_rows[1][1], date(2026, 1, 24))

            self.assertEqual(len(export.digital_rows) - 1, 1)
            self.assertEqual(export.digital_rows[1][0], "002")

    def test_合计行不会被当作数据行(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "销售用券情况统计.xlsx"
            _write_coupon_source(
                source,
                [
                    _coupon_row(
                        document="收款001",
                        day="2026-01-24",
                        product="海尔冰箱",
                        brand="海尔",
                        category="冰箱",
                        summary="ref1",
                        family_subsidy=100,
                    ),
                ],
            )

            export = sources.read_coupon_export(source)

            self.assertEqual(len(export.appliance_rows), 2)  # header + one data row, no 合计

    def test_missing_合计_row_raises(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "销售用券情况统计.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["销售用券情况统计"])
            sheet.append([None] * 28)
            workbook.save(source)
            workbook.close()

            with self.assertRaises(ValueError):
                sources.read_coupon_export(source)

    def test_reads_via_exactly_one_iter_rows_pass(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "销售用券情况统计.xlsx"
            _write_coupon_source(
                source,
                [
                    _coupon_row(
                        document="收款001",
                        day="2026-01-24",
                        product="海尔冰箱",
                        brand="海尔",
                        category="冰箱",
                        summary="ref1",
                        family_subsidy=100,
                    ),
                ],
            )
            workbook = CalamineWorkbook.from_path(str(source))
            wrapped = _CountingCalamineWorkbookWrapper(workbook)
            try:
                sources.read_coupon_export(source, wrapped)
            finally:
                wrapped.close()

            self.assertEqual(wrapped.sheet_wrapper.iter_rows_calls, 1)

    def test_parses_currency_formatted_total(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "销售用券情况统计.xlsx"
            _write_coupon_source(source, [])

            export = sources.read_coupon_export(source)

            self.assertEqual(str(export.source_total), "123.40")

    def test_zero_total_is_not_treated_as_missing(self) -> None:
        """A source total of exactly 0 must parse to Decimal("0"), not None
        — `value or ""` previously treated 0 as falsy and silently dropped
        the source-vs-detail gap warning for a legitimately zero total."""
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "销售用券情况统计.xlsx"
            header = [None] * 28
            header[2] = "单据号"
            header[3] = "单据日期"
            header[5] = "商品名称"
            header[7] = "品牌"
            header[14] = "财务大类"
            header[17] = "明细摘要"
            header[25] = "2026家电国补（计入收入）"
            header[26] = "2026数码国补（计入收入）"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["销售用券情况统计"])
            sheet.append(header)
            total_row = [None] * 28
            total_row[0] = "合计"
            total_row[25] = 0
            sheet.append(total_row)
            workbook.save(source)
            workbook.close()

            export = sources.read_coupon_export(source)

            self.assertIsNotNone(export.source_total)
            self.assertEqual(export.source_total, 0)

    def test_moves_wrong_subsidy_columns_by_financial_category_and_records_corrections(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "销售用券情况统计.xlsx"
            _write_coupon_source(
                source,
                [
                    _coupon_row(
                        document="收款001",
                        day="2026-01-24",
                        product="海尔冰箱",
                        brand="海尔",
                        category="冰箱",
                        summary="ref1",
                        digital_subsidy=100,
                    ),
                    _coupon_row(
                        document="收款002",
                        day="2026-01-25",
                        product="数码商品",
                        brand="品牌",
                        category="新业务类",
                        summary="ref2",
                        family_subsidy=50,
                    ),
                ],
            )

            export = sources.read_coupon_export(source, remark_lookup={})

            self.assertEqual(len(export.appliance_rows), 2)
            self.assertEqual(export.appliance_rows[1][0], "001")
            self.assertEqual(export.appliance_rows[1][6], 100)
            self.assertEqual(len(export.digital_rows), 2)
            self.assertEqual(export.digital_rows[1][0], "002")
            self.assertEqual(export.digital_rows[1][6], 50)
            self.assertEqual(str(export.source_total), "173.40")

            self.assertEqual(len(export.subsidy_corrections), 2)
            first, second = export.subsidy_corrections
            self.assertEqual(first.row_number, 3)
            self.assertEqual(first.document_number, "001")
            self.assertEqual(first.financial_category, "冰箱")
            self.assertEqual(first.amount, Decimal("100"))
            self.assertEqual(
                first.from_header,
                sources.COUPON_DIGITAL_SUBSIDY_HEADER,
            )
            self.assertEqual(
                first.to_header,
                sources.COUPON_FAMILY_SUBSIDY_HEADER,
            )
            self.assertEqual(second.row_number, 4)
            self.assertEqual(second.document_number, "002")
            self.assertEqual(second.financial_category, "新业务类")
            self.assertEqual(second.amount, Decimal("50"))
            self.assertEqual(
                second.from_header,
                sources.COUPON_FAMILY_SUBSIDY_HEADER,
            )
            self.assertEqual(
                second.to_header,
                sources.COUPON_DIGITAL_SUBSIDY_HEADER,
            )

    def test_return_remark_prevents_automatic_subsidy_adjustment(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "销售用券情况统计.xlsx"
            _write_coupon_source(
                source,
                [
                    _coupon_row(
                        document="收款003",
                        day="2026-01-26",
                        product="海尔冰箱",
                        brand="海尔",
                        category="冰箱",
                        summary="ref3",
                        digital_subsidy=100,
                    ),
                ],
            )
            remark_lookup = {
                ("003", date(2026, 1, 26)): "退换货/倒票（退单）"
            }

            export = sources.read_coupon_export(
                source,
                remark_lookup=remark_lookup,
            )

            self.assertEqual(len(export.appliance_rows), 1)
            self.assertEqual(len(export.digital_rows), 2)
            self.assertEqual(export.digital_rows[1][0], "003")
            self.assertEqual(export.digital_rows[1][6], 100)
            self.assertEqual(export.subsidy_corrections, ())

    def test_invalid_date_error_names_source_file_and_row(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "销售用券情况统计.xlsx"
            _write_coupon_source(
                source,
                [
                    _coupon_row(
                        document="收款001",
                        day="2026.1.24",
                        product="海尔冰箱",
                        brand="海尔",
                        category="冰箱",
                        summary="ref1",
                        family_subsidy=100,
                    ),
                ],
            )

            with self.assertRaises(ValueError) as caught:
                sources.read_coupon_export(source)

            message = str(caught.exception)
            self.assertIn(source.name, message)
            self.assertIn("第 3 行", message)
            self.assertIn("2026.1.24", message)

    def test_rejects_excel_error_in_subsidy_column(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "销售用券情况统计.xlsx"
            _write_coupon_source(
                source,
                [
                    _coupon_row(
                        document="收款001",
                        day="2026-01-24",
                        product="海尔冰箱",
                        brand="海尔",
                        category="冰箱",
                        summary="ref1",
                        family_subsidy="#DIV/0!",
                    ),
                ],
            )

            with self.assertRaisesRegex(ValueError, "Excel 错误值.*#DIV/0!"):
                sources.read_coupon_export(source)


class CouponComputationSingleReadTests(unittest.TestCase):
    """compute_coupon_data's no-rows fallback must read the export once.

    The 合计 used to be fetched by a second full-sheet read; both now come
    from the same CouponExport, so one call to read_coupon_export must be
    the entire read cost of the fallback path.
    """

    def _patched_fallbacks(self):
        export = sources.CouponExport(
            appliance_rows=[list(sources.APPLIANCE_PROFILE.output_header)],
            digital_rows=[list(sources.DIGITAL_PROFILE.output_header)],
            source_total=Decimal("123.40"),
            subsidy_corrections=(),
        )
        stack = ExitStack()
        stack.enter_context(
            patch.object(
                sources,
                "COUPON_SOURCE_FILE",
                Path("销售用券情况统计.xlsx"),
                create=True,
            )
        )
        # COUPON_REFERENCE_SUPPLEMENT_FILE only exists after
        # configure_data_dir; patch it explicitly so this test is
        # order-independent (it passed in the full suite only while another
        # test happened to have configured the module globals first).
        stack.enter_context(
            patch.object(
                sources,
                "COUPON_REFERENCE_SUPPLEMENT_FILE",
                Path("参考号补充.xlsx"),
                create=True,
            )
        )
        counting = Mock(wraps=sources.read_coupon_export, return_value=export)
        stack.enter_context(
            patch.object(sources, "read_coupon_export", counting)
        )
        stack.enter_context(
            patch.object(
                appliance,
                "load_uploaded_summary",
                return_value=({}, 0, Decimal("0")),
            )
        )
        stack.enter_context(
            patch.object(
                appliance,
                "load_coupon_reference_supplement",
                return_value={},
            )
        )
        stack.enter_context(
            patch.object(
                digital,
                "load_uploaded_summary",
                return_value=({}, 0, Decimal("0")),
            )
        )
        return stack, counting, export

    def test_appliance_fallback_reads_the_export_exactly_once(self) -> None:
        stack, counting, export = self._patched_fallbacks()
        with stack:
            computation = appliance.compute_coupon_data()

        counting.assert_called_once()
        self.assertEqual(computation.rows, export.appliance_rows)
        self.assertEqual(computation.source_total, Decimal("123.40"))

    def test_digital_fallback_reads_the_export_exactly_once(self) -> None:
        stack, counting, export = self._patched_fallbacks()
        with stack:
            computation = digital.compute_coupon_data()

        counting.assert_called_once()
        self.assertEqual(computation.rows, export.digital_rows)


if __name__ == "__main__":
    unittest.main()
