import io
import tempfile
import unittest
from contextlib import redirect_stdout
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from processors import submitted
from processors.common.config import load_merchants, submitted_file_marker
from processors.common.excel import (
    load_measurement_font,
    measurement_text,
    resolve_font,
    width_measurer,
    widths_are_additive,
)
from processors.submitted import STATUS_ORDER


def text_pixel_width(value: object, font) -> float:
    """Test oracle mirroring width_measurer's uncached, unoptimized path."""
    text = measurement_text(value)
    return font.getlength(text) if text else 0


SOURCE_COLUMN_COUNT = 24
# Column letters as they appear in a real MER_*.xlsx export, so a fixture that
# passes here is shaped like the file the operator actually feeds in.
# add_subsidy_column reads the amount from the third kept column, which is F.
AMOUNT_COLUMN = "F"
REFERENCE_COLUMN = "G"
STATUS_COLUMN = "I"
DESCRIPTION_COLUMN = "J"


def build_header(**overrides: str) -> tuple[str, ...]:
    """Build a source header row, naming fields by their source column letter."""
    header = [
        chr(ord("A") + index) * 2 for index in range(SOURCE_COLUMN_COUNT)
    ]
    names = {
        AMOUNT_COLUMN: "交易金额",
        STATUS_COLUMN: "状态",
        REFERENCE_COLUMN: "检索参考号",
        DESCRIPTION_COLUMN: "描述",
    }
    names.update(overrides)
    for letter, name in names.items():
        header[column_index_from_string(letter) - 1] = name
    return tuple(header)


SUBMITTED_HEADER = build_header()


def submitted_row(
    header: tuple[str, ...],
    *,
    reference: object = "12345678901A",
) -> list[object]:
    row: list[object] = ["v"] * len(header)
    row[column_index_from_string(AMOUNT_COLUMN) - 1] = 1000
    row[column_index_from_string(STATUS_COLUMN) - 1] = "审核通过"
    row[column_index_from_string(REFERENCE_COLUMN) - 1] = reference
    row[column_index_from_string(DESCRIPTION_COLUMN) - 1] = "说明"
    return row


def write_submitted_source(
    path: Path,
    header: tuple[str, ...],
    rows: list[list[object]] | None = None,
) -> None:
    """Write a source file shaped like the upload export: title row, then header."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["报表标题"])
    sheet.append(list(header))
    for row in rows if rows is not None else [submitted_row(header)]:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


class SubsidyCapTest(unittest.TestCase):
    """The two projects share the 15% rate but cap at different amounts.

    Household appliances cap at 1500, digital at 500. They were once both
    written as 500, which silently understated 43% of the appliance rows, so
    the caps are pinned here rather than left to a shared default.
    """

    def subsidy_for(self, profile_name: str, amount: object) -> object:
        row = submitted.add_subsidy_column(
            [None, None, amount], profile_name=profile_name
        )
        return row[3]

    def test_rate_is_15_percent_below_either_cap(self) -> None:
        for profile_name in ("家电", "数码"):
            with self.subTest(profile_name=profile_name):
                self.assertEqual(self.subsidy_for(profile_name, 1000), 150.00)

    def test_caps_differ_between_projects(self) -> None:
        self.assertEqual(submitted.PROFILES["家电"].subsidy_cap, Decimal("1500"))
        self.assertEqual(submitted.PROFILES["数码"].subsidy_cap, Decimal("500"))

    def test_appliance_cap_applies_at_1500_not_500(self) -> None:
        # 6000 * 0.15 = 900: above digital's cap, still below the appliance one.
        self.assertEqual(self.subsidy_for("家电", 6000), 900.00)
        self.assertEqual(self.subsidy_for("数码", 6000), 500.00)
        # 20000 * 0.15 = 3000: both capped, but at different values.
        self.assertEqual(self.subsidy_for("家电", 20000), 1500.00)
        self.assertEqual(self.subsidy_for("数码", 20000), 500.00)

    def test_blank_amount_yields_no_subsidy(self) -> None:
        for profile_name in ("家电", "数码"):
            with self.subTest(profile_name=profile_name):
                self.assertIsNone(self.subsidy_for(profile_name, None))
                self.assertIsNone(self.subsidy_for(profile_name, ""))


class WidthMeasurerTest(unittest.TestCase):
    """The cached measurer must stay numerically identical to the direct call."""

    def test_matches_text_pixel_width_for_every_value_kind(self) -> None:
        _, font_path = resolve_font()
        font = load_measurement_font(font_path)
        measure = width_measurer(font)

        for value in (
            None,
            "",
            "ABC",
            "商品名称示例",
            0,
            0.0,
            123.456,
            True,
            date(2026, 1, 24),
            "重复值",
            "重复值",
        ):
            with self.subTest(value=value):
                self.assertEqual(measure(value), text_pixel_width(value, font))

    def test_repeated_values_cost_nothing_on_the_installed_font(self) -> None:
        """Whichever path the installed font takes, repeats must be free."""
        _, font_path = resolve_font()
        counter = CallCountingFont(load_measurement_font(font_path))
        measure = width_measurer(counter)

        measure("同一个值")
        after_first = counter.calls
        for _ in range(49):
            measure("同一个值")

        self.assertEqual(counter.calls, after_first)


class CallCountingFont:
    """Wraps a font object and counts getlength calls."""

    def __init__(self, font) -> None:
        self._font = font
        self.calls = 0

    def getlength(self, text: str) -> float:
        self.calls += 1
        return self._font.getlength(text)


class FakeMonospaceFont:
    """Every glyph advances the same width, so widths sum exactly."""

    def __init__(self, width: float = 10.0) -> None:
        self.width = width
        self.calls = 0

    def getlength(self, text: str) -> float:
        self.calls += 1
        return self.width * len(text)


class FakeProportionalFont:
    """Kerns the pairs real proportional fonts kern, so widths do not sum."""

    KERNED_PAIRS = {"AV": -2.0, "To": -1.5, "Ta": -1.0}

    def __init__(self) -> None:
        self.calls = 0

    def getlength(self, text: str) -> float:
        self.calls += 1
        total = sum(7.0 if character.isascii() else 14.0 for character in text)
        for pair, adjustment in self.KERNED_PAIRS.items():
            total += adjustment * (len(text) - len(text.replace(pair, "")) ) / len(pair)
        return total


class WidthMeasurerFontPathsTest(unittest.TestCase):
    """Both measuring strategies, on fonts the test itself controls.

    Using fake fonts rather than whatever is installed keeps the assertions
    about call counts meaningful on any machine.
    """

    SAMPLE = "同一个值"

    def test_additive_font_takes_the_per_character_path(self) -> None:
        self.assertTrue(widths_are_additive(FakeMonospaceFont()))

    def test_kerning_font_is_rejected_by_detection(self) -> None:
        self.assertFalse(widths_are_additive(FakeProportionalFont()))

    def test_additive_font_measures_each_distinct_character_once(self) -> None:
        font = FakeMonospaceFont()
        measure = width_measurer(font)
        font.calls = 0

        measure(self.SAMPLE)
        self.assertLessEqual(font.calls, len(set(self.SAMPLE)))
        after_first = font.calls

        for _ in range(49):
            measure(self.SAMPLE)
        self.assertEqual(font.calls, after_first)

    def test_additive_font_result_equals_the_direct_measurement(self) -> None:
        font = FakeMonospaceFont()
        measure = width_measurer(font)
        for value in ("SN2026ABCD", "商品名称示例", "中文A1，混排", "0123456789"):
            with self.subTest(value=value):
                self.assertEqual(measure(value), font.getlength(value))

    def test_kerning_font_keeps_whole_string_measurement(self) -> None:
        font = FakeProportionalFont()
        measure = width_measurer(font)
        font.calls = 0

        for _ in range(50):
            measure(self.SAMPLE)
        self.assertEqual(font.calls, 1)

        # A kerned string must keep the font's own answer, not a summed one.
        self.assertEqual(measure("AV"), font.getlength("AV"))
        self.assertNotEqual(
            font.getlength("AV"),
            font.getlength("A") + font.getlength("V"),
        )


class SubmittedFileMarkerTest(unittest.TestCase):
    """The marker is derived from the merchant id, never configured twice.

    The export is named MER_<商户编号>_<导出时间>_yjhx.xlsx, so a hardcoded
    marker could drift out of sync with config/merchants.yaml and send a run
    looking for files that no longer exist.
    """

    def test_marker_is_the_merchant_id_with_the_exporter_prefix(self) -> None:
        merchants = load_merchants()
        for data_type in ("家电", "数码"):
            with self.subTest(data_type=data_type):
                self.assertEqual(
                    submitted_file_marker(data_type),
                    f"MER_{merchants[data_type]}",
                )

    def test_each_profile_uses_its_own_data_type(self) -> None:
        self.assertEqual(submitted.PROFILES["数码"].data_type, "数码")
        self.assertEqual(submitted.PROFILES["家电"].data_type, "家电")
        with tempfile.TemporaryDirectory() as directory:
            data_dir = Path(directory)
            submitted.configure_data_dir(data_dir)
            self.assertEqual(
                submitted.SUBMITTED_FILE_MARKERS["数码"],
                submitted_file_marker("数码"),
            )
            self.assertEqual(
                submitted.SUBMITTED_FILE_MARKERS["家电"],
                submitted_file_marker("家电"),
            )

    def test_unknown_data_type_is_reported(self) -> None:
        with self.assertRaisesRegex(ValueError, "缺少家具的商户编号"):
            submitted_file_marker("家具")


class SubmittedHeaderValidationTest(unittest.TestCase):
    """A wrong export format must name the file and the missing fields."""

    def run_build(self, profile_name: str, header: tuple[str, ...]):
        with tempfile.TemporaryDirectory() as directory:
            data_dir = Path(directory)
            write_submitted_source(
                data_dir / f"{submitted_file_marker(profile_name)}_export.xlsx",
                header,
            )
            submitted.configure_data_dir(data_dir)
            return submitted.build_report(profile_name)

    def test_missing_required_fields_are_reported_by_name(self) -> None:
        for profile_name in ("家电", "数码"):
            with self.subTest(profile_name=profile_name):
                header = build_header(
                    **{
                        STATUS_COLUMN: "占位一",
                        REFERENCE_COLUMN: "占位二",
                        DESCRIPTION_COLUMN: "占位三",
                    }
                )
                with self.assertRaises(ValueError) as raised:
                    self.run_build(profile_name, header)

                message = str(raised.exception)
                self.assertIn("export.xlsx", message)
                self.assertIn("检索参考号", message)
                self.assertIn("状态", message)
                self.assertIn("描述", message)

    def test_valid_header_is_accepted(self) -> None:
        for profile_name in ("家电", "数码"):
            with self.subTest(profile_name=profile_name):
                report = self.run_build(
                    profile_name,
                    SUBMITTED_HEADER,
                )
                self.assertEqual(report.file_count, 1)
                self.assertEqual(report.data_row_count, 1)

    def test_workbook_without_sheets_is_deleted_reported_and_skipped(self) -> None:
        profile_name = "数码"
        with tempfile.TemporaryDirectory() as directory:
            data_dir = Path(directory)
            marker = submitted_file_marker(profile_name)
            valid_path = data_dir / f"{marker}_valid.xlsx"
            invalid_path = data_dir / f"{marker}_invalid.xlsx"
            write_submitted_source(valid_path, SUBMITTED_HEADER)
            invalid_path.touch()
            submitted.configure_data_dir(data_dir)

            real_from_path = submitted.CalamineWorkbook.from_path

            class EmptyWorkbook:
                sheet_names: list[str] = []

                def close(self) -> None:
                    pass

            def open_source(path: str):
                if Path(path) == invalid_path:
                    return EmptyWorkbook()
                return real_from_path(path)

            with (
                patch.object(
                    submitted.CalamineWorkbook,
                    "from_path",
                    side_effect=open_source,
                ),
                patch("builtins.print") as mocked_print,
            ):
                report = submitted.build_report(profile_name)

            self.assertFalse(invalid_path.exists())
            self.assertTrue(valid_path.exists())
            self.assertEqual(report.file_count, 1)
            mocked_print.assert_called_once_with(
                f"已删除无效导出文件（没有工作表）：{invalid_path}"
            )

    def test_output_carries_only_the_kept_columns(self) -> None:
        """详细地址/tel/发票金额/图片1/S/N码 were dropped from both projects.

        The placeholder header names are the doubled column letter, so a
        column that sneaks back in is identifiable by where it came from.
        """
        for profile_name in ("家电", "数码"):
            with self.subTest(profile_name=profile_name):
                header = self.run_build(profile_name, SUBMITTED_HEADER).header

                self.assertEqual(len(header), 7)
                self.assertEqual(header[3], "补贴金额")
                for dropped in ("QQ", "SS", "UU", "WW", "XX"):
                    self.assertNotIn(dropped, header)
                for required in submitted.REQUIRED_SUBMITTED_HEADERS:
                    self.assertIn(required, header)


class SubmittedRowValidationTest(unittest.TestCase):
    def build_from_rows(
        self,
        rows: list[list[object]],
        *,
        profile_name: str = "家电",
    ) -> submitted.SubmittedReport:
        with tempfile.TemporaryDirectory() as directory:
            data_dir = Path(directory)
            write_submitted_source(
                data_dir / f"{submitted_file_marker(profile_name)}_export.xlsx",
                SUBMITTED_HEADER,
                rows,
            )
            submitted.configure_data_dir(data_dir)
            return submitted.build_report(profile_name)

    def test_unknown_statuses_are_counted_not_dropped(self) -> None:
        """核销成功 and 待同步 appear in real exports but have no status sheet.

        Such rows still belong in Summary, so they must be reported rather
        than silently missing from every status tab.
        """
        rows = []
        for status, count in (("核销成功", 2), ("待同步", 1), ("", 1)):
            for index in range(count):
                row = submitted_row(
                    SUBMITTED_HEADER,
                    reference=f"1234567890{len(rows)}A",
                )
                row[column_index_from_string(STATUS_COLUMN) - 1] = status
                rows.append(row)
        rows.append(submitted_row(SUBMITTED_HEADER, reference="12345678999A"))

        report = self.build_from_rows(rows)

        self.assertEqual(report.data_row_count, 5)
        self.assertEqual(
            report.unknown_status_counts,
            {"核销成功": 2, "待同步": 1, "": 1},
        )
        self.assertEqual(
            report.unknown_status_records,
            (
                submitted.UnknownStatusRecord(
                    "MER_89813015722APT1_export.xlsx",
                    3,
                    "12345678900A",
                    "核销成功",
                ),
                submitted.UnknownStatusRecord(
                    "MER_89813015722APT1_export.xlsx",
                    4,
                    "12345678901A",
                    "核销成功",
                ),
                submitted.UnknownStatusRecord(
                    "MER_89813015722APT1_export.xlsx",
                    5,
                    "12345678902A",
                    "待同步",
                ),
                submitted.UnknownStatusRecord(
                    "MER_89813015722APT1_export.xlsx",
                    6,
                    "12345678903A",
                    "",
                ),
            ),
        )
        # The known-status row is the only one reaching a status sheet, but
        # every row is still carried in Summary.
        self.assertEqual(len(report.summary_rows), 5)
        self.assertEqual(
            sum(len(rows) for rows in report.status_rows.values()), 1
        )

    def test_no_unknown_statuses_reports_nothing(self) -> None:
        report = self.build_from_rows([submitted_row(SUBMITTED_HEADER)])

        self.assertEqual(report.unknown_status_counts, {})

    def test_unknown_status_warning_names_count_and_sources(self) -> None:
        rows = []
        for reference in ("12345678900A", "12345678901A", "12345678902A"):
            row = submitted_row(SUBMITTED_HEADER, reference=reference)
            row[column_index_from_string(STATUS_COLUMN) - 1] = "待同步"
            rows.append(row)
        with tempfile.TemporaryDirectory() as directory:
            data_dir = Path(directory)
            write_submitted_source(
                data_dir / f"{submitted_file_marker('数码')}_export.xlsx",
                SUBMITTED_HEADER,
                rows,
            )
            submitted.configure_data_dir(data_dir)
            output = io.StringIO()
            with patch.object(
                submitted, "write_xlsx_atomically"
            ), redirect_stdout(output):
                submitted.process_submitted_files("数码")

        text = output.getvalue()
        self.assertIn("[数码] 警告：3 行", text)
        self.assertIn("状态为 待同步×3", text)
        self.assertIn("未配置独立工作表", text)
        self.assertIn("数据已保留在 Summary，未被删除", text)
        # Title and header occupy source rows 1-2, so the first data row is 3.
        self.assertIn("源文件 MER_89813014812B06R_export.xlsx", text)
        self.assertIn("源行 3", text)
        self.assertIn("检索参考号 12345678900A", text)
        self.assertIn("状态 待同步", text)

    def test_invalid_reference_reports_source_location(self) -> None:
        row = submitted_row(SUBMITTED_HEADER, reference="not-valid")

        with self.assertRaisesRegex(
            ValueError,
            r"export\.xlsx 第 3 行检索参考号格式无效",
        ):
            self.build_from_rows([row])

    def test_blank_reference_is_allowed(self) -> None:
        row = submitted_row(SUBMITTED_HEADER, reference=None)

        report = self.build_from_rows([row])

        self.assertEqual(report.data_row_count, 1)

    def test_row_populated_only_outside_kept_columns_is_skipped(self) -> None:
        row: list[object] = [None] * len(SUBMITTED_HEADER)
        row[0] = "非保留列中的值"

        report = self.build_from_rows([row])

        self.assertEqual(report.data_row_count, 0)
        self.assertEqual(report.summary_rows, [])

    def test_repeated_reference_is_rejected_however_the_rows_differ(self) -> None:
        """检索参考号 is one per row in every real export, so a repeat is a
        duplicate no matter which other fields happen to differ."""
        variants = {
            "完全相同": lambda row: row,
            "商品明细不同": lambda row: self._with(row, "D", "另一条商品明细"),
            "描述不同": lambda row: self._with(row, DESCRIPTION_COLUMN, "另一种说明"),
        }
        for label, vary in variants.items():
            with self.subTest(label):
                first = submitted_row(SUBMITTED_HEADER)
                second = vary(submitted_row(SUBMITTED_HEADER))

                with self.assertRaisesRegex(
                    ValueError,
                    r"检索参考号重复：12345678901A.*第 3 行.*第 4 行",
                ):
                    self.build_from_rows([first, second])

    @staticmethod
    def _with(row: list[object], column: str, value: object) -> list[object]:
        row[column_index_from_string(column) - 1] = value
        return row

    def test_blank_references_do_not_collide(self) -> None:
        """A blank reference is tolerated, so several of them must not read as
        repeats of each other."""
        rows = [
            submitted_row(SUBMITTED_HEADER, reference=None),
            submitted_row(SUBMITTED_HEADER, reference=""),
        ]

        report = self.build_from_rows(rows)

        self.assertEqual(report.data_row_count, 2)

    def test_overlapping_exports_are_rejected_by_reference(self) -> None:
        """Two exports covering an overlapping period repeat whole records.

        The repeated rows are matched on 检索参考号 rather than on the row as a
        whole, so an incidental difference between the two exports — here a
        reformatted 交易金额 — cannot smuggle the duplicate past the check and
        double-count its subsidy.
        """
        profile_name = "家电"
        rows = {
            "first": submitted_row(SUBMITTED_HEADER),
            "second": self._with(submitted_row(SUBMITTED_HEADER), AMOUNT_COLUMN, 1000.0),
        }
        with tempfile.TemporaryDirectory() as directory:
            data_dir = Path(directory)
            marker = submitted_file_marker(profile_name)
            for suffix, row in rows.items():
                write_submitted_source(
                    data_dir / f"{marker}_{suffix}.xlsx",
                    SUBMITTED_HEADER,
                    [row],
                )
            submitted.configure_data_dir(data_dir)

            with self.assertRaisesRegex(
                ValueError,
                r"检索参考号重复.*first\.xlsx 第 3 行.*second\.xlsx 第 3 行",
            ):
                submitted.build_report(profile_name)


class ValidatorRejectsBadOutputTest(unittest.TestCase):
    """The atomic save only protects the output if the validators really reject.

    Each case writes a deliberately wrong workbook and asserts it is refused.
    """

    def build_valid_output(self, profile_name: str, path: Path) -> int:
        with tempfile.TemporaryDirectory() as directory:
            data_dir = Path(directory)
            write_submitted_source(
                data_dir / f"{submitted_file_marker(profile_name)}_export.xlsx",
                SUBMITTED_HEADER,
            )
            submitted.configure_data_dir(data_dir)
            report = submitted.build_report(profile_name)
            submitted.write_workbook(path, report)
            return report.data_row_count

    def test_accepts_the_workbook_it_just_built(self) -> None:
        for profile_name in ("家电", "数码"):
            with self.subTest(profile_name=profile_name):
                with tempfile.TemporaryDirectory() as directory:
                    output = Path(directory) / "已上传.xlsx"
                    data_rows = self.build_valid_output(profile_name, output)
                    submitted.validate_output(output, data_rows, profile_name)

                    workbook = load_workbook(output)
                    try:
                        sheet = workbook["Summary"]
                        self.assertEqual(sheet.freeze_panes, "A2")
                        # One column per kept source column, plus 补贴金额.
                        last_column = get_column_letter(
                            len(submitted.KEPT_SOURCE_COLUMNS) + 1
                        )
                        self.assertEqual(
                            sheet.auto_filter.ref, f"A1:{last_column}2"
                        )
                        self.assertEqual(sheet["A1"].fill.fgColor.rgb[-6:], "000000")
                        self.assertEqual(sheet["D2"].number_format, "0.00")
                    finally:
                        workbook.close()

    def test_rejects_wrong_row_count(self) -> None:
        for profile_name in ("家电", "数码"):
            with self.subTest(profile_name=profile_name):
                with tempfile.TemporaryDirectory() as directory:
                    output = Path(directory) / "已上传.xlsx"
                    data_rows = self.build_valid_output(profile_name, output)
                    with self.assertRaises(RuntimeError):
                        submitted.validate_output(
                            output, data_rows + 1, profile_name
                        )

    def test_rejects_missing_status_worksheet(self) -> None:
        for profile_name in ("家电", "数码"):
            with self.subTest(profile_name=profile_name):
                with tempfile.TemporaryDirectory() as directory:
                    output = Path(directory) / "已上传.xlsx"
                    data_rows = self.build_valid_output(profile_name, output)

                    workbook = load_workbook(output)
                    try:
                        del workbook[STATUS_ORDER[0]]
                        workbook.save(output)
                    finally:
                        workbook.close()

                    with self.assertRaises(RuntimeError):
                        submitted.validate_output(output, data_rows, profile_name)

    def test_rejects_invalid_reference_in_saved_summary(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "已上传.xlsx"
            data_rows = self.build_valid_output("家电", output)
            workbook = load_workbook(output)
            try:
                sheet = workbook["Summary"]
                header = [cell.value for cell in sheet[1]]
                reference_column = header.index("检索参考号") + 1
                sheet.cell(2, reference_column, "invalid")
                workbook.save(output)
            finally:
                workbook.close()

            with self.assertRaisesRegex(RuntimeError, "检索参考号格式无效"):
                submitted.validate_output(output, data_rows, "家电")

    def tamper_with_subsidy(self, sheet_name: str) -> None:
        """Overwrite one 补贴金额 cell and assert the validator refuses it."""
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "已上传.xlsx"
            data_rows = self.build_valid_output("家电", output)
            workbook = load_workbook(output)
            try:
                sheet = workbook[sheet_name]
                header = [cell.value for cell in sheet[1]]
                sheet.cell(2, header.index("补贴金额") + 1, 99999.99)
                workbook.save(output)
            finally:
                workbook.close()

            with self.assertRaises(RuntimeError):
                submitted.validate_output(output, data_rows, "家电")

    def test_rejects_wrong_subsidy_in_summary(self) -> None:
        self.tamper_with_subsidy("Summary")

    def test_rejects_wrong_subsidy_in_status_sheet(self) -> None:
        """Summary and the status sheets are written by separate calls, so a
        subsidy that is only wrong on the status sheet must still be caught."""
        self.tamper_with_subsidy("审核通过")


if __name__ == "__main__":
    unittest.main()
