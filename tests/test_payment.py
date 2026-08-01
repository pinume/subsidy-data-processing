from __future__ import annotations

import unittest
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from processors import payment


def _detail_row(profile, category: str, product: str, subsidy: str | None = None):
    row = [None] * len(profile.detail_headers)
    row[profile.detail_headers.index("拨付批次")] = "batch"
    row[profile.detail_headers.index("商户编号")] = "ABC123"
    row[profile.detail_headers.index("编码品类")] = category
    row[profile.detail_headers.index("商品名称")] = product
    if subsidy is not None:
        row[profile.detail_headers.index("补贴金额")] = subsidy
    return row


def _write_source(path: Path, profile, category: str, product: str, subsidy: str) -> None:
    workbook = Workbook()
    detail = workbook.active
    detail.append(profile.detail_headers)
    detail.append(_detail_row(profile, category, product, subsidy))
    workbook.save(path)


class DetailProcessingTests(unittest.TestCase):
    def test_actual_rows_include_data_after_long_blank_run(self) -> None:
        workbook = Workbook()
        source = workbook.active
        source.append(["first"])
        for _ in range(100):
            source.append([None])
        source.append(["last"])

        rows = list(payment._iter_actual_rows_with_numbers(source))

        self.assertEqual([row_number for row_number, _ in rows], [1, 102])
        self.assertEqual(rows[-1][1][0], "last")

    def test_source_headers_are_trimmed(self) -> None:
        profile = payment.PROFILES["家电"]
        workbook = Workbook()
        source = workbook.active
        source.append(tuple(f" {header} " for header in profile.detail_headers))
        source.append(_detail_row(profile, "A04-空调", "格力空调"))
        target = workbook.create_sheet("target")

        rows, unidentified = payment._collect_normalized_detail(
            payment._iter_actual_rows_with_numbers(source),
            source.title,
            profile=profile,
            merchant_id="ABC123",
        )
        target.append(profile.detail_headers + payment.DERIVED_HEADERS)
        for row in rows:
            target.append(row)

        self.assertEqual(len(rows), 1)
        self.assertEqual(unidentified, 0)

    def test_header_fields_after_column_twenty_are_supported(self) -> None:
        profile = payment.PROFILES["家电"]
        workbook = Workbook()
        source = workbook.active
        headers = ("辅助列",) + profile.detail_headers
        source.append(headers)
        source.append(["ignored"] + _detail_row(profile, "A04-空调", "格力空调"))
        target = workbook.create_sheet("target")

        rows, unidentified = payment._collect_normalized_detail(
            payment._iter_actual_rows_with_numbers(source),
            source.title,
            profile=profile,
            merchant_id="ABC123",
        )
        target.append(profile.detail_headers + payment.DERIVED_HEADERS)
        for row in rows:
            target.append(row)

        self.assertEqual(len(rows), 1)
        self.assertEqual(unidentified, 0)
        self.assertEqual(target.cell(2, len(profile.detail_headers) + 2).value, "格力")

    def test_summary_uses_decimal_and_signed_count(self) -> None:
        detail_rows = [
            ["财务大类", "品牌", "补贴金额"],
            ["空调", "格力", "0.10"],
            ["空调", "格力", "0.20"],
            ["空调", "格力", "-0.05"],
        ]

        rows, bold_rows, groups = payment._build_summary_rows(
            [("家电", [("明细", detail_rows)], payment.APPLIANCE_CATEGORY_MAP)]
        )

        self.assertEqual(groups, 1)
        rows = [tuple(row) for row in rows]
        self.assertEqual(
            rows,
            [
                ("财务大类", "品牌", "补贴金额合计", "补贴金额计数"),
                ("空调", "格力", 0.25, 1),
                ("合计", None, 0.25, 1),
                ("合计", None, 0.25, 1),
            ],
        )
        self.assertEqual(bold_rows, [3, 4])

    def test_large_read_only_detail_uses_streaming_iteration(self) -> None:
        with TemporaryDirectory(dir=".") as temporary_dir:
            path = Path(temporary_dir) / "detail.xlsx"
            workbook = Workbook()
            detail = workbook.active
            detail.title = "家电明细"
            detail.append(["财务大类", "品牌", "补贴金额"])
            for _ in range(3000):
                detail.append(["空调", "格力", 1])
            workbook.save(path)
            workbook.close()

            read_only_book = load_workbook(
                path,
                read_only=True,
                data_only=True,
            )
            try:
                read_only_detail = read_only_book["家电明细"]
                with patch.object(
                    type(read_only_detail),
                    "cell",
                    side_effect=AssertionError(
                        "read-only aggregation must not use random cell access"
                    ),
                ):
                    groups = payment._sum_detail_groups(
                        [
                            (
                                read_only_detail.title,
                                payment._worksheet_rows(read_only_detail),
                            )
                        ]
                    )
            finally:
                read_only_book.close()

            self.assertEqual(
                groups[("空调", "格力")],
                [Decimal("3000"), 3000],
            )


class WorkbookLoadingTests(unittest.TestCase):
    def test_process_sources_opens_formula_workbook_lazily_and_uncached(
        self,
    ) -> None:
        # The main source read is calamine, which has no formula view — a
        # blank subsidy cell still needs an openpyxl, data_only=False read of
        # the same file to check for an uncached formula, but only once that
        # blank cell is actually encountered.
        profile = payment.PROFILES["家电"]

        with TemporaryDirectory(dir=".") as temporary_dir:
            source = Path(temporary_dir).resolve() / "source.xlsx"
            workbook = Workbook()
            detail = workbook.active
            detail.append(profile.detail_headers)
            detail.append(_detail_row(profile, "A04-空调", "格力空调", None))
            workbook.save(source)

            calls = []
            original_load_workbook = load_workbook

            def recording_loader(path, **kwargs):
                calls.append((path, kwargs))
                return original_load_workbook(path, **kwargs)

            with patch.object(payment, "load_workbook", recording_loader):
                payment._process_sources([source], profile, "ABC123")

        self.assertEqual(len(calls), 1)
        self.assertTrue(calls[0][1]["read_only"])
        self.assertFalse(calls[0][1]["data_only"])

    def test_process_sources_rejects_uncached_formula_subsidy_amount(self) -> None:
        profile = payment.PROFILES["家电"]

        with TemporaryDirectory(dir=".") as temporary_dir:
            source = Path(temporary_dir).resolve() / "source.xlsx"
            workbook = Workbook()
            detail = workbook.active
            detail.append(profile.detail_headers)
            detail.append(
                _detail_row(profile, "A04-空调", "格力空调", "=0.1+0.2")
            )
            workbook.save(source)

            with self.assertRaisesRegex(ValueError, "公式但没有缓存计算结果"):
                payment._process_sources([source], profile, "ABC123")
            self.assertTrue(source.exists())

    def test_process_sources_rejects_excel_error_subsidy_amount(self) -> None:
        profile = payment.PROFILES["家电"]

        with TemporaryDirectory(dir=".") as temporary_dir:
            source = Path(temporary_dir).resolve() / "source.xlsx"
            workbook = Workbook()
            detail = workbook.active
            detail.append(profile.detail_headers)
            detail.append(
                _detail_row(profile, "A04-空调", "格力空调", "#DIV/0!")
            )
            workbook.save(source)

            with self.assertRaisesRegex(ValueError, "Excel 错误值.*#DIV/0!"):
                payment._process_sources([source], profile, "ABC123")
            self.assertTrue(source.exists())

    def test_process_sources_skips_other_merchants_without_error(self) -> None:
        """Another merchant's rows are excluded silently, never raised on.

        The source carries every merchant's sales, so this is the common case,
        not an anomaly — including when 商户编号 holds an Excel error value
        (the 数码 source has one). See the comment at the merchant filter.
        """
        profile = payment.PROFILES["家电"]

        with TemporaryDirectory(dir=".") as temporary_dir:
            source = Path(temporary_dir).resolve() / "source.xlsx"
            workbook = Workbook()
            detail = workbook.active
            detail.append(profile.detail_headers)
            detail.append(_detail_row(profile, "A04-空调", "格力空调", "10.00"))
            other = _detail_row(profile, "A04-空调", "格力空调", "10.00")
            other[profile.detail_headers.index("商户编号")] = "OTHER99"
            detail.append(other)
            workbook.save(source)

            section = payment._process_sources([source], profile, "ABC123")

            self.assertEqual(len(section.rows), 1)

    def test_process_sources_returns_final_rows_without_creating_a_workbook(
        self,
    ) -> None:
        """Brand inference and sorting must be done before anything is written.

        This used to be checked by counting the cells already created on the
        target worksheet when each step ran. _process_sources no longer builds
        a worksheet at all — it returns rows, and the writer runs afterwards —
        so the invariant is now that no workbook exists while it works, and
        that the rows it hands back are already inferred and sorted.
        """
        profile = payment.PROFILES["家电"]
        with TemporaryDirectory(dir=".") as temporary_dir:
            source = Path(temporary_dir).resolve() / "source.xlsx"
            _write_source(source, profile, "A04-空调", "格力空调", "10.00")
            observed = []
            original_infer = payment._infer_missing_brands
            original_sort = payment._sort_detail_rows

            def recording_infer(rows, headers):
                observed.append("infer")
                return original_infer(rows, headers)

            def recording_sort(rows, headers, category_map):
                observed.append("sort")
                return original_sort(rows, headers, category_map)

            def forbidden_workbook(*args, **kwargs):
                raise AssertionError(
                    "_process_sources must not create a workbook; the writer does"
                )

            with (
                patch.object(payment, "_infer_missing_brands", recording_infer),
                patch.object(payment, "_sort_detail_rows", recording_sort),
                patch.object(payment, "Workbook", forbidden_workbook),
            ):
                section = payment._process_sources([source], profile, "ABC123")

            self.assertEqual(observed, ["infer", "sort"])
            self.assertEqual(section.name, profile.detail_sheet_name)
            self.assertEqual(section.header, profile.detail_headers + payment.DERIVED_HEADERS)
            self.assertEqual(len(section.rows), 1)
            self.assertEqual(section.rows[0][-1], "格力")

    def _write_detail_workbook(self, path: Path, headers) -> None:
        workbook = Workbook()
        detail = workbook.active
        detail.title = "明细"
        detail.append(["报表标题"])
        detail.append(headers)
        workbook.save(path)

    def test_detect_profile_matches_appliance_and_digital_headers(self) -> None:
        with TemporaryDirectory(dir=".") as temporary_dir:
            temporary_path = Path(temporary_dir).resolve()
            for expected_name, headers in (
                ("家电", payment.APPLIANCE_DETAIL_HEADERS),
                ("数码", payment.DIGITAL_DETAIL_HEADERS),
            ):
                source = temporary_path / f"{expected_name}.xlsx"
                self._write_detail_workbook(source, headers)

                self.assertEqual(payment.detect_profile(source).name, expected_name)

    def test_detect_profile_accepts_header_aliases(self) -> None:
        aliased = {"交易时间": "交易完成时间", "销售企业": "销方名称", "发票号": "发票号码"}
        headers = [
            aliased.get(header, header) for header in payment.DIGITAL_DETAIL_HEADERS
        ]
        with TemporaryDirectory(dir=".") as temporary_dir:
            source = Path(temporary_dir).resolve() / "source.xlsx"
            self._write_detail_workbook(source, headers)

            self.assertEqual(payment.detect_profile(source).name, "数码")

    def test_header_normalization_preserves_canonical_names(self) -> None:
        canonical = ("交易时间", "商户编号", "销售金额", "发票号", "ID")

        self.assertEqual(payment._normalize_header_names(canonical), list(canonical))

    def test_detect_profile_prefers_the_source_filename(self) -> None:
        with TemporaryDirectory(dir=".") as temporary_dir:
            source = Path(temporary_dir).resolve() / "source.xlsx"
            self._write_detail_workbook(source, payment.APPLIANCE_DETAIL_HEADERS)

            self.assertEqual(
                payment.detect_profile(source, "2026年数码补贴明细.xlsx").name, "数码"
            )
            self.assertEqual(
                payment.detect_profile(source, "1_2026年以旧换新补贴明细.xlsx").name,
                "家电",
            )

    def test_detect_profile_falls_back_to_headers_for_other_filenames(self) -> None:
        with TemporaryDirectory(dir=".") as temporary_dir:
            source = Path(temporary_dir).resolve() / "source.xlsx"
            self._write_detail_workbook(source, payment.DIGITAL_DETAIL_HEADERS)

            self.assertEqual(payment.detect_profile(source, "导出结果.xlsx").name, "数码")

    def test_detect_profile_rejects_unrecognized_headers(self) -> None:
        headers = [
            header
            for header in payment.APPLIANCE_DETAIL_HEADERS
            if header != "补贴比例"
        ]
        with TemporaryDirectory(dir=".") as temporary_dir:
            source = Path(temporary_dir).resolve() / "source.xlsx"
            self._write_detail_workbook(source, headers)

            with self.assertRaisesRegex(ValueError, "无法识别数据类型"):
                payment.detect_profile(source)


class PaymentPipelineTests(unittest.TestCase):
    """End-to-end runs against a temporary data directory and output file."""

    def _run(self, temporary_path: Path, merchants: dict[str, str]) -> Path:
        output_dir = temporary_path / "output"
        output_dir.mkdir(exist_ok=True)
        output_file = output_dir / "回款明细.xlsx"

        def fake_merchant_id(data_type: str) -> str:
            merchant = merchants.get(data_type)
            if not merchant:
                raise ValueError(f"merchants.yaml 缺少{data_type}的商户编号")
            return merchant

        with (
            patch.object(payment, "OUTPUT_DIR", output_dir),
            patch.object(payment, "OUTPUT_FILE", output_file),
            patch.object(payment, "merchant_id", fake_merchant_id),
        ):
            payment.configure_data_dir(temporary_path / "data")
            payment.process_payment_files()
        return output_file

    def _prepare_data_dir(self, temporary_path: Path) -> Path:
        data_dir = temporary_path / "data"
        data_dir.mkdir()
        _write_source(
            data_dir / "1_2026年以旧换新补贴明细.xlsx",
            payment.PROFILES["家电"],
            "A04-空调",
            "格力空调",
            "10.00",
        )
        _write_source(
            data_dir / "2026年数码补贴明细.xlsx",
            payment.PROFILES["数码"],
            "B01-手机",
            "华为手机",
            "25.00",
        )
        return data_dir

    def test_both_data_types_share_one_output_workbook(self) -> None:
        with TemporaryDirectory(dir=".") as temporary_dir:
            temporary_path = Path(temporary_dir).resolve()
            self._prepare_data_dir(temporary_path)

            output_file = self._run(
                temporary_path, {"家电": "ABC123", "数码": "ABC123"}
            )

            result = load_workbook(output_file, data_only=True)
            try:
                self.assertEqual(
                    result.sheetnames, ["汇总", "家电明细", "数码明细"]
                )
                summary = result["汇总"]
                rows = [
                    tuple(summary.cell(row, column).value for column in range(1, 5))
                    for row in range(2, summary.max_row + 1)
                ]
                self.assertEqual(
                    rows,
                    [
                        ("空调", "格力", 10.0, 1),
                        ("合计", None, 10.0, 1),
                        (None, None, None, None),
                        ("手机", "华为", 25.0, 1),
                        ("合计", None, 25.0, 1),
                        ("合计", None, 35.0, 2),
                    ],
                )
                self.assertEqual(
                    Decimal(str(summary.cell(summary.max_row, 3).value)),
                    Decimal("35"),
                )
                self.assertTrue(
                    all(
                        "," not in cell.number_format
                        for sheet in result.worksheets
                        for row in sheet.iter_rows()
                        for cell in row
                    )
                )
                self.assertEqual(summary["C2"].number_format, "0.00")
            finally:
                result.close()

    def test_sources_are_kept_and_only_the_output_file_is_written(self) -> None:
        with TemporaryDirectory(dir=".") as temporary_dir:
            temporary_path = Path(temporary_dir).resolve()
            data_dir = self._prepare_data_dir(temporary_path)

            output_file = self._run(
                temporary_path, {"家电": "ABC123", "数码": "ABC123"}
            )

            self.assertEqual(
                sorted(path.name for path in data_dir.iterdir()),
                ["1_2026年以旧换新补贴明细.xlsx", "2026年数码补贴明细.xlsx"],
            )
            self.assertEqual(
                [path.name for path in output_file.parent.iterdir()],
                ["回款明细.xlsx"],
            )

    def test_missing_merchant_id_stops_before_writing_output(self) -> None:
        with TemporaryDirectory(dir=".") as temporary_dir:
            temporary_path = Path(temporary_dir).resolve()
            self._prepare_data_dir(temporary_path)

            with self.assertRaisesRegex(ValueError, "缺少数码的商户编号"):
                self._run(temporary_path, {"家电": "ABC123"})

            self.assertEqual(
                list((temporary_path / "output").iterdir()),
                [],
                "失败时不应留下工作副本或半成品输出",
            )

    def test_wrong_merchant_id_does_not_replace_existing_output(self) -> None:
        with TemporaryDirectory(dir=".") as temporary_dir:
            temporary_path = Path(temporary_dir).resolve()
            self._prepare_data_dir(temporary_path)
            output_dir = temporary_path / "output"
            output_dir.mkdir()
            output_file = output_dir / "回款明细.xlsx"
            existing = Workbook()
            existing.active["A1"] = "old"
            existing.save(output_file)
            existing.close()

            with self.assertRaisesRegex(ValueError, "未找到商户 WRONG 的数据"):
                self._run(
                    temporary_path,
                    {"家电": "WRONG", "数码": "WRONG"},
                )

            saved = load_workbook(output_file, read_only=True, data_only=True)
            try:
                self.assertEqual(saved.active["A1"].value, "old")
            finally:
                saved.close()

    def test_validator_rejects_summary_that_disagrees_with_details(self) -> None:
        with TemporaryDirectory(dir=".") as temporary_dir:
            temporary_path = Path(temporary_dir).resolve()
            self._prepare_data_dir(temporary_path)
            output_file = self._run(
                temporary_path,
                {"家电": "ABC123", "数码": "ABC123"},
            )

            saved = load_workbook(output_file)
            expected_summary = payment._summary_snapshot(
                payment._worksheet_rows(saved["汇总"])
            )
            saved["汇总"].cell(saved["汇总"].max_row, 3).value = 999
            saved.save(output_file)
            saved.close()

            expectations = {
                "家电明细": payment.PaymentOutputExpectation(
                    payment.PROFILES["家电"],
                    1,
                    "ABC123",
                ),
                "数码明细": payment.PaymentOutputExpectation(
                    payment.PROFILES["数码"],
                    1,
                    "ABC123",
                ),
            }
            with self.assertRaisesRegex(ValueError, "汇总.*内容校验失败"):
                payment.validate_output(
                    output_file,
                    expectations,
                    expected_summary,
                )

    def test_output_not_created_when_a_later_file_is_unrecognized(self) -> None:
        with TemporaryDirectory(dir=".") as temporary_dir:
            temporary_path = Path(temporary_dir).resolve()
            data_dir = temporary_path / "data"
            data_dir.mkdir()
            output_dir = temporary_path / "output"
            output_dir.mkdir()
            profile = payment.PROFILES["家电"]
            _write_source(
                data_dir / "A补贴明细.xlsx", profile, "A04-空调", "格力空调", "10.00"
            )
            unrecognized = Workbook()
            unrecognized.active.append(
                [header for header in profile.detail_headers if header != "补贴比例"]
            )
            unrecognized.save(data_dir / "B补贴明细.xlsx")

            with self.assertRaisesRegex(ValueError, "无法识别数据类型"):
                self._run(temporary_path, {"家电": "ABC123"})

            self.assertEqual(list(output_dir.iterdir()), [])

    def test_empty_data_directory_is_reported(self) -> None:
        with TemporaryDirectory(dir=".") as temporary_dir:
            temporary_path = Path(temporary_dir).resolve()
            (temporary_path / "data").mkdir()

            with self.assertRaisesRegex(FileNotFoundError, "回款原始数据文件"):
                self._run(temporary_path, {"家电": "ABC123", "数码": "ABC123"})


if __name__ == "__main__":
    unittest.main()
