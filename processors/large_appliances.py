
import re
from collections import Counter
from collections.abc import Callable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import column_index_from_string

from processors.common.config import load_brand_mapping
from processors.common.excel import (
    create_sheet_styles,
    format_sheet,
    load_measurement_font,
    pixels_to_excel_width,
    read_rows,
    resolve_font,
    save_workbook_atomically,
    width_measurer,
)
from processors.common.dates import (
    is_valid_original_invoice_number,
    normalize_coupon_date,
    normalize_document_number,
    normalize_receipt_date,
    normalize_receipt_identifier,
    receipt_match_key,
)
from processors.common.paths import resolve_existing_data_file


BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "output" / "large_appliances"
DATA_DIR: Path
INPUT_DIR: Path
OUTPUT_FILE = OUTPUT_DIR / "已上传.xlsx"
RECEIPTS_SOURCE_FILE: Path
RECEIPTS_OUTPUT_FILE = OUTPUT_DIR / "收款单统计.xlsx"
COUPON_SOURCE_FILE: Path
COUPON_REFERENCE_SUPPLEMENT_FILE: Path
COUPON_OUTPUT_FILE = OUTPUT_DIR / "销售用券情况统计.xlsx"
COUPON_REMARK_SOURCE_FILE = RECEIPTS_OUTPUT_FILE
COUPON_UPLOADED_SOURCE_FILE = OUTPUT_FILE


def configure_data_dir(data_dir: Path) -> None:
    global DATA_DIR
    global INPUT_DIR
    global RECEIPTS_SOURCE_FILE
    global COUPON_SOURCE_FILE
    global COUPON_REFERENCE_SUPPLEMENT_FILE

    DATA_DIR = data_dir
    INPUT_DIR = DATA_DIR / "submitted"
    RECEIPTS_SOURCE_FILE = resolve_existing_data_file(
        DATA_DIR,
        (
            Path("receipt_statistics") / "receipt_statistics.XLS",
            Path("收款单统计") / "收款单统计.XLS",
        ),
    )
    COUPON_SOURCE_FILE = resolve_existing_data_file(
        DATA_DIR,
        (
            Path("subsidy_coupon_statistics") / "subsidy_coupon_statistics.XLS",
            Path("销售用券情况统计") / "销售用券情况统计.XLS",
        ),
    )
    COUPON_REFERENCE_SUPPLEMENT_FILE = resolve_existing_data_file(
        DATA_DIR,
        (
            Path("reference_number_supplement")
            / "reference_number_supplement.xlsx",
            Path("参考号异常数据补充") / "参考号补充.xlsx",
        ),
    )

# Keep the required source columns in their original order.
KEPT_SOURCE_COLUMNS = ("D", "E", "F", "G", "I", "J", "Q", "S", "U", "W", "X")
KEPT_COLUMN_INDEXES = tuple(
    column_index_from_string(column) for column in KEPT_SOURCE_COLUMNS
)
# "补贴金额" is inserted by add_subsidy_column, so only source fields are listed.
REQUIRED_SUBMITTED_HEADERS = ("状态", "描述", "交易金额")
STATUS_ORDER = (
    "核销失败",
    "审核失败",
    "暂存",
    "同步(已上送)",
    "待审核",
    "审核通过",
)
STATUS_PRIORITY = {status: index for index, status in enumerate(STATUS_ORDER)}
RECEIPTS_SOURCE_HEADER = ("单据号", "日期", "原票号", "摘要", "商品名称")
RECEIPTS_OUTPUT_HEADER = (*RECEIPTS_SOURCE_HEADER, "备注")
RECEIPTS_REMARK_RETURN = "退换货/倒票（退单）"
RECEIPTS_REMARK_ORIGINAL = "退换货/倒票（原单）"
RECEIPTS_REMARK_BOTH = "退换货/倒票（退单及原单）"
RECEIPTS_REMARK_SAME_MODEL_REPLACEMENT = "已做同型号换货处理"
RECEIPTS_REMARK_SPECIAL = r"退换货\倒票"
RECEIPTS_SPECIAL_REMARK_KEYS = {
    "2605050233000077",
    "2605030233000049",
    "260426ZH3X000025",
}
RECEIPTS_REPORT_HEADER = ("异常类型", "源文件行", "匹配值", "说明")
RECEIPTS_ROW_HEIGHT = 20
RECEIPTS_DUPLICATE_FILL_COLOR = "FFC7CE"
RECEIPTS_EXCLUDED_PRODUCT_KEYWORD = "北国"
RECEIPTS_SAME_MODEL_REPLACEMENT_KEYWORD = "同型号换货"
COUPON_KEPT_SOURCE_COLUMNS = (3, 4, 6, 8, 15, 18, 26)
COUPON_SUBSIDY_HEADER = "2026家电国补（计入收入）"
COUPON_OUTPUT_HEADER = (
    "单据号",
    "单据日期",
    "商品名称",
    "品牌",
    "财务大类",
    "明细摘要",
    COUPON_SUBSIDY_HEADER,
    "备注",
    "详细情况",
)
COUPON_GROUP_HEADER = (
    "单据号",
    "单据日期",
    "商品名称",
    COUPON_SUBSIDY_HEADER,
    "备注",
    "详细情况",
)
COUPON_GROUP_COLUMN_INDEXES = tuple(
    COUPON_OUTPUT_HEADER.index(header) for header in COUPON_GROUP_HEADER
)
REFERENCE_REPORT_CORRECTED = "已自动纠正"
REFERENCE_REPORT_UNRESOLVED = "无唯一候选"
REFERENCE_REPORT_COLLISION = "目标冲突"
REFERENCE_REPORT_ORDER = {
    REFERENCE_REPORT_CORRECTED: 0,
    REFERENCE_REPORT_COLLISION: 1,
    REFERENCE_REPORT_UNRESOLVED: 2,
}
REFERENCE_REPORT_HEADER = ("处理结果", "输出行号", "原参考号", "说明")
REFERENCE_REPORT_SHEET = "Processing Report"
COUPON_MATCH_FILL_COLOR = "FFC7CE"
COUPON_BRAND_REPLACEMENTS = load_brand_mapping()
COUPON_SUMMARY_HEADER = (
    "财务大类",
    "品牌",
    "备注",
    "数量",
    f"{COUPON_SUBSIDY_HEADER}合计",
)
COUPON_APPROVED_SUMMARY_TITLE = "审核通过明细"
COUPON_APPROVED_SUMMARY_HEADER = (
    "财务大类",
    "品牌",
    "数量",
    f"{COUPON_SUBSIDY_HEADER}合计",
)
COUPON_REMARK_SUMMARY_TITLE = "备注汇总"
COUPON_REMARK_SUMMARY_HEADER = (
    "备注",
    "数量",
    f"{COUPON_SUBSIDY_HEADER}合计",
)
COUPON_REMARK_SORT_PRIORITY = {
    "未上传": 0,
    "已上传": 1,
}
COUPON_SUMMARY_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
INVALID_SHEET_TITLE_RE = re.compile(r"[\[\]:*?/\\]")
COUPON_REFERENCE_RE = re.compile(r"\d{11}[A-Z]")
COUPON_REFERENCE_SUPPLEMENT_HEADER = ("参考号", "单据号", "单据日期")


def as_currency(amount: Decimal) -> Decimal:
    """Round to cents before comparing.

    Source amounts arrive as floats, so Decimal(str(value)) carries binary
    noise such as 1234.5600000000001 into the running total. Totals are money
    and are only ever meaningful to two decimal places.
    """
    return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def select_columns(row: list[object]) -> list[object]:
    return [
        row[column - 1] if column <= len(row) else None
        for column in KEPT_COLUMN_INDEXES
    ]


def add_subsidy_column(
    row: list[object],
    *,
    is_header: bool = False,
    source_name: str | None = None,
    source_row: int | None = None,
) -> list[object]:
    result = list(row)
    if is_header:
        result.insert(3, "补贴金额")
        return result

    amount = result[2]
    if amount in (None, ""):
        subsidy = None
    else:
        try:
            calculated = Decimal(str(amount)) * Decimal("0.15")
            subsidy = float(
                min(calculated, Decimal("500")).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP,
                )
            )
        except (InvalidOperation, ValueError) as error:
            location = (
                f"{source_name} 第 {source_row} 行"
                if source_name is not None and source_row is not None
                else "数据"
            )
            raise ValueError(f"{location}的交易金额无效：{amount!r}") from error

    result.insert(3, subsidy)
    return result


def build_workbook() -> tuple[Workbook, int, int]:
    files = sorted(
        path
        for path in INPUT_DIR.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    if not files:
        raise FileNotFoundError(f"未在 {INPUT_DIR} 中找到 .xlsx 文件")

    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "Summary"

    expected_header: list[object] | None = None
    output_header: list[object] | None = None
    data_row_count = 0
    data_rows: list[list[object]] = []

    for path in files:
        rows = read_rows(path)
        title = next(rows, None)
        header = next(rows, None)
        if title is None or header is None:
            raise ValueError(f"{path.name} 缺少标题行或表头行")

        if expected_header is None:
            expected_header = header
            # Drop the source title row and use the selected headers as row 1.
            output_header = add_subsidy_column(
                select_columns(header),
                is_header=True,
            )
            # Reject a wrong export before parsing any rows, so the operator sees
            # the missing fields instead of a downstream value error.
            missing_headers = [
                required
                for required in REQUIRED_SUBMITTED_HEADERS
                if required not in output_header
            ]
            if missing_headers:
                raise ValueError(
                    f"{path.name} 不是已上传数据的导出格式，缺少字段："
                    f"{'、'.join(missing_headers)}；"
                    f"实际字段 {tuple(output_header)}"
                )
            sheet.append(output_header)
        elif header != expected_header:
            raise ValueError(f"{path.name} 的表头与第一个文件不一致")

        for source_row, row in enumerate(rows, start=3):
            if any(value not in (None, "") for value in row):
                data_rows.append(
                    add_subsidy_column(
                        select_columns(row),
                        source_name=path.name,
                        source_row=source_row,
                    )
                )
                data_row_count += 1

    if output_header is None:
        raise RuntimeError("未能生成输出表头")

    status_column_index = output_header.index("状态")
    data_rows.sort(
        key=lambda row: STATUS_PRIORITY.get(
            str(row[status_column_index]) if row[status_column_index] is not None else "",
            len(STATUS_ORDER),
        )
    )
    for row in data_rows:
        sheet.append(row)

    description_column_index = output_header.index("描述")
    rows_by_status: dict[str, list[list[object]]] = {
        status: [] for status in STATUS_ORDER
    }
    for row in data_rows:
        status = str(row[status_column_index] or "")
        if status in rows_by_status:
            rows_by_status[status].append(row)

    font_name, font_path = resolve_font()
    measurement_font = load_measurement_font(font_path)
    format_sheet(sheet, font_name, measurement_font)
    for status in STATUS_ORDER:
        status_sheet = workbook.create_sheet(title=status)
        status_sheet.append(output_header)
        status_rows = rows_by_status[status]
        status_rows.sort(
            key=lambda row: (
                row[description_column_index] not in (None, ""),
                str(row[description_column_index] or ""),
            ),
            reverse=True,
        )
        for row in status_rows:
            status_sheet.append(row)
        format_sheet(status_sheet, font_name, measurement_font)

    return workbook, len(files), data_row_count


def validate_output(path: Path, expected_data_rows: int) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        expected_sheet_names = ["Summary", *STATUS_ORDER]
        if workbook.sheetnames != expected_sheet_names:
            raise RuntimeError(
                f"工作表校验失败：预期 {expected_sheet_names}，"
                f"实际 {workbook.sheetnames}"
            )

        sheet = workbook["Summary"]
        actual_data_rows = max(sheet.max_row - 1, 0)
        if actual_data_rows != expected_data_rows:
            raise RuntimeError(
                f"输出校验失败：预期 {expected_data_rows} 条，实际 {actual_data_rows} 条"
            )
        expected_columns = len(KEPT_SOURCE_COLUMNS) + 1
        if sheet.max_column != expected_columns:
            raise RuntimeError(
                f"输出校验失败：预期 {expected_columns} 列，"
                f"实际 {sheet.max_column} 列"
            )

        header = tuple(cell.value for cell in next(sheet.iter_rows(max_row=1)))
        status_column = header.index("状态")
        description_column = header.index("描述")
        amount_column = header.index("交易金额")
        subsidy_column = header.index("补贴金额")

        status_total = sum(
            max(workbook[status].max_row - 1, 0)
            for status in STATUS_ORDER
        )
        known_status_total = sum(
            1
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[status_column] in STATUS_ORDER
        )
        if status_total != known_status_total:
            raise RuntimeError(
                f"状态工作表校验失败：预期共 {known_status_total} 条，"
                f"实际共 {status_total} 条"
            )

        for status in STATUS_ORDER:
            status_sheet = workbook[status]
            status_header = tuple(
                cell.value for cell in next(status_sheet.iter_rows(max_row=1))
            )
            if status_header != header:
                raise RuntimeError(f"{status}工作表的标题行与汇总表不一致")

            descriptions: list[str] = []
            blank_description_found = False
            for row in status_sheet.iter_rows(min_row=2, values_only=True):
                if row[status_column] != status:
                    raise RuntimeError(f"{status}工作表中存在其他状态的数据")

                description = row[description_column]
                if description in (None, ""):
                    blank_description_found = True
                else:
                    if blank_description_found:
                        raise RuntimeError(
                            f"{status}工作表的空白描述未全部排在末尾"
                        )
                    descriptions.append(str(description))

                amount = row[amount_column]
                subsidy = row[subsidy_column]
                if amount not in (None, ""):
                    expected_subsidy = min(
                        Decimal(str(amount)) * Decimal("0.15"),
                        Decimal("500"),
                    ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    if Decimal(str(subsidy)) != expected_subsidy:
                        raise RuntimeError(
                            f"{status}工作表存在补贴金额计算错误"
                        )

            if descriptions != sorted(descriptions, reverse=True):
                raise RuntimeError(f"{status}工作表的描述列未按降序排列")
    finally:
        workbook.close()


def process_submitted_files() -> None:
    workbook, file_count, data_row_count = build_workbook()
    save_workbook_atomically(
        workbook,
        OUTPUT_FILE,
        lambda path: validate_output(path, data_row_count),
    )

    print(f"Submitted data complete: merged {file_count} files, {data_row_count} rows")
    print(f"Output file: {OUTPUT_FILE}")


def read_receipt_rows(source: Path) -> list[list[object]]:
    """Read legacy XLS files in pure Python without invoking Excel."""
    source_workbook = xlrd.open_workbook(source)
    try:
        source_sheet = source_workbook.sheet_by_index(0)
        if source_sheet.nrows < 2:
            raise ValueError(f"{source.name} 缺少总标题行或字段标题行")
        source_headers = [
            str(source_sheet.cell_value(1, column_index)).strip()
            for column_index in range(source_sheet.ncols)
        ]
        missing_headers = [
            header for header in RECEIPTS_SOURCE_HEADER
            if header not in source_headers
        ]
        if missing_headers:
            raise ValueError(
                f"{source.name} 缺少必要字段：{tuple(missing_headers)}；"
                f"实际字段 {tuple(source_headers)}"
            )
        source_column_indexes = [
            source_headers.index(header) for header in RECEIPTS_SOURCE_HEADER
        ]

        rows: list[list[object]] = []
        for row_index in range(1, source_sheet.nrows):
            row: list[object] = []
            for column_index in source_column_indexes:
                cell = source_sheet.cell(row_index, column_index)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(
                        value,
                        source_workbook.datemode,
                    )
                elif cell.ctype == xlrd.XL_CELL_NUMBER and value == int(value):
                    value = int(value)
                row.append(value)
            if row_index > 1 and str(row[1]).strip() == "合计":
                continue
            rows.append(row)

        actual_header = tuple(
            str(value).strip() if value is not None else ""
            for value in rows[0]
        )
        if actual_header != RECEIPTS_SOURCE_HEADER:
            raise ValueError(
                f"{source.name} 字段标题不一致："
                f"预期 {RECEIPTS_SOURCE_HEADER}，实际 {actual_header}"
            )
        return rows
    finally:
        source_workbook.release_resources()


def receipt_remark(
    has_original: bool,
    is_referenced: bool,
    is_same_model_replacement: bool = False,
    is_special: bool = False,
) -> str | None:
    if is_special:
        return RECEIPTS_REMARK_SPECIAL
    if has_original and is_referenced:
        return RECEIPTS_REMARK_BOTH
    if is_same_model_replacement:
        return RECEIPTS_REMARK_SAME_MODEL_REPLACEMENT
    if has_original:
        return RECEIPTS_REMARK_RETURN
    if is_referenced:
        return RECEIPTS_REMARK_ORIGINAL
    return None


def prepare_receipt_data(kept_rows: list[list[object]]):
    records: list[dict[str, object]] = []
    key_rows: dict[str, list[int]] = {}
    original_invoice_numbers: set[str] = set()
    referenced_original_invoice_numbers: set[str] = set()
    same_model_replacement_original_invoice_numbers: set[str] = set()
    excluded_product_count = 0

    for source_row, row in enumerate(kept_rows[1:], start=3):
        product_name = normalize_receipt_identifier(row[4])
        if RECEIPTS_EXCLUDED_PRODUCT_KEYWORD in product_name:
            excluded_product_count += 1
            continue

        document_number = normalize_document_number(row[0])
        receipt_date = normalize_receipt_date(row[1], source_row=source_row)
        original_invoice_number = normalize_receipt_identifier(row[2])
        match_key = (
            receipt_match_key(receipt_date, document_number)
            if receipt_date is not None and document_number
            else ""
        )
        if match_key:
            key_rows.setdefault(match_key, []).append(source_row)
        if original_invoice_number:
            original_invoice_numbers.add(original_invoice_number)
            if RECEIPTS_SAME_MODEL_REPLACEMENT_KEYWORD in (
                normalize_receipt_identifier(row[3])
            ):
                same_model_replacement_original_invoice_numbers.add(
                    original_invoice_number
                )
            else:
                referenced_original_invoice_numbers.add(
                    original_invoice_number
                )
        records.append(
            {
                "source_row": source_row,
                "document_number": document_number,
                "receipt_date": receipt_date,
                "original_invoice_number": original_invoice_number,
                "summary": row[3],
                "product_name": product_name or None,
                "match_key": match_key,
            }
        )

    issues: list[tuple[str, str, str, str]] = []
    for match_key, source_rows in key_rows.items():
        if len(source_rows) > 1:
            issues.append(
                (
                    "重复匹配键",
                    "、".join(str(row) for row in source_rows),
                    match_key,
                    "多个数据行生成了相同的日期与单据号组合键",
                )
            )

    only_return_count = 0
    only_original_count = 0
    both_count = 0
    unmatched_original_count = 0
    invalid_original_count = 0
    missing_match_key_count = 0
    output_rows: list[list[object]] = []
    for record in records:
        source_row = int(record["source_row"])
        original_invoice_number = str(record["original_invoice_number"])
        match_key = str(record["match_key"])
        has_original = bool(original_invoice_number)
        is_referenced = bool(
            match_key and match_key in referenced_original_invoice_numbers
        )
        is_same_model_replacement = bool(
            match_key
            and match_key
            in same_model_replacement_original_invoice_numbers
            and not is_referenced
        )
        is_special = match_key in RECEIPTS_SPECIAL_REMARK_KEYS

        if not match_key:
            missing_match_key_count += 1
            issues.append(
                (
                    "缺少匹配键",
                    str(source_row),
                    "",
                    "日期或单据号为空，无法生成匹配键",
                )
            )
        if has_original and not is_valid_original_invoice_number(
            original_invoice_number
        ):
            invalid_original_count += 1
            issues.append(
                (
                    "原票号格式异常",
                    str(source_row),
                    original_invoice_number,
                    "原票号应为6位日期加单据号",
                )
            )
        if has_original and original_invoice_number not in key_rows:
            unmatched_original_count += 1
            issues.append(
                (
                    "原票号未匹配",
                    str(source_row),
                    original_invoice_number,
                    "未找到日期与单据号组合键相同的原单",
                )
            )

        if has_original and (is_referenced or is_same_model_replacement):
            both_count += 1
        elif has_original:
            only_return_count += 1
        elif is_referenced or is_same_model_replacement:
            only_original_count += 1

        output_rows.append(
            [
                record["document_number"],
                record["receipt_date"],
                original_invoice_number or None,
                record["summary"],
                record["product_name"],
                receipt_remark(
                    has_original,
                    is_referenced,
                    is_same_model_replacement,
                    is_special,
                ),
            ]
        )

    stats = {
        "总数据量": len(records),
        "删除北国商品行数": excluded_product_count,
        "仅退单数量": only_return_count,
        "仅原单数量": only_original_count,
        "退单及原单数量": both_count,
        "备注总数": only_return_count + only_original_count + both_count,
        "未匹配原票号数量": unmatched_original_count,
        "重复匹配键数量": sum(
            1 for source_rows in key_rows.values() if len(source_rows) > 1
        ),
        "原票号格式异常数量": invalid_original_count,
        "缺少匹配键数量": missing_match_key_count,
    }
    duplicate_match_keys = {
        match_key
        for match_key, source_rows in key_rows.items()
        if len(source_rows) > 1
    }
    return output_rows, stats, issues, duplicate_match_keys


def validate_receipts_output(
    path: Path,
    expected_data_rows: int,
    expected_stats: dict[str, int],
    expected_issues: list[tuple[str, str, str, str]],
) -> None:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        if workbook.sheetnames != ["Sheet1", "Processing Report"]:
            raise RuntimeError(
                f"收款单工作表校验失败：实际工作表为 {workbook.sheetnames}"
            )

        sheet = workbook["Sheet1"]
        if sheet.max_row - 1 != expected_data_rows:
            raise RuntimeError(
                f"收款单行数校验失败：预期 {expected_data_rows} 条，"
                f"实际 {sheet.max_row - 1} 条"
            )
        if sheet.max_column != len(RECEIPTS_OUTPUT_HEADER):
            raise RuntimeError(
                f"收款单列数校验失败：预期 {len(RECEIPTS_OUTPUT_HEADER)} 列，"
                f"实际 {sheet.max_column} 列"
            )

        header = tuple(cell.value for cell in sheet[1])
        if header != RECEIPTS_OUTPUT_HEADER:
            raise RuntimeError(
                f"收款单表头校验失败：预期 {RECEIPTS_OUTPUT_HEADER}，"
                f"实际 {header}"
            )

        referenced_original_invoice_numbers = {
            normalize_receipt_identifier(sheet.cell(row_number, 3).value)
            for row_number in range(2, sheet.max_row + 1)
            if (
                sheet.cell(row_number, 3).value not in (None, "")
                and normalize_receipt_identifier(
                    sheet.cell(row_number, 4).value
                )
                .find(RECEIPTS_SAME_MODEL_REPLACEMENT_KEYWORD)
                == -1
            )
        }
        same_model_replacement_original_invoice_numbers = {
            normalize_receipt_identifier(sheet.cell(row_number, 3).value)
            for row_number in range(2, sheet.max_row + 1)
            if (
                sheet.cell(row_number, 3).value not in (None, "")
                and normalize_receipt_identifier(
                    sheet.cell(row_number, 4).value
                )
                .find(RECEIPTS_SAME_MODEL_REPLACEMENT_KEYWORD)
                != -1
            )
        }
        match_key_counts: dict[str, int] = {}
        for row_number in range(2, sheet.max_row + 1):
            document_number = normalize_receipt_identifier(
                sheet.cell(row_number, 1).value
            )
            date_value = sheet.cell(row_number, 2).value
            if date_value is None or not document_number:
                continue
            match_key = receipt_match_key(
                date_value.date()
                if isinstance(date_value, datetime)
                else date_value,
                document_number,
            )
            match_key_counts[match_key] = match_key_counts.get(match_key, 0) + 1

        for row_number in range(2, sheet.max_row + 1):
            document_number = sheet.cell(row_number, 1).value
            if document_number is not None and (
                not isinstance(document_number, str)
                or document_number.startswith("收款")
            ):
                raise RuntimeError(
                    f"收款单第 {row_number} 行的单据号格式不正确"
                )

            date_cell = sheet.cell(row_number, 2)
            if date_cell.value is not None:
                if not isinstance(date_cell.value, (date, datetime)):
                    raise RuntimeError(
                        f"收款单第 {row_number} 行的日期不是有效日期"
                    )
                if date_cell.number_format != "yyyymmdd":
                    raise RuntimeError(
                        f"收款单第 {row_number} 行的日期格式不正确"
                    )

            original_invoice_number = normalize_receipt_identifier(
                sheet.cell(row_number, 3).value
            )
            match_key = (
                receipt_match_key(
                    date_cell.value.date()
                    if isinstance(date_cell.value, datetime)
                    else date_cell.value,
                    normalize_receipt_identifier(document_number),
                )
                if date_cell.value is not None and document_number is not None
                else ""
            )
            expected_remark = receipt_remark(
                bool(original_invoice_number),
                bool(
                    match_key
                    and match_key in referenced_original_invoice_numbers
                ),
                bool(
                    match_key
                    and match_key
                    in same_model_replacement_original_invoice_numbers
                    and match_key
                    not in referenced_original_invoice_numbers
                ),
                match_key in RECEIPTS_SPECIAL_REMARK_KEYS,
            )
            if sheet.cell(row_number, 6).value != expected_remark:
                raise RuntimeError(
                    f"收款单第 {row_number} 行的备注校验失败"
                )

            is_duplicate = bool(match_key and match_key_counts[match_key] > 1)
            for cell in sheet[row_number]:
                fill_color = cell.fill.fgColor.rgb
                is_pink = (
                    cell.fill.fill_type == "solid"
                    and fill_color is not None
                    and fill_color[-6:] == RECEIPTS_DUPLICATE_FILL_COLOR[-6:]
                )
                if is_pink != is_duplicate:
                    raise RuntimeError(
                        f"收款单第 {row_number} 行的重复匹配键标记不正确"
                    )

        report_sheet = workbook["Processing Report"]
        actual_stats = {
            report_sheet.cell(row_number, 1).value: report_sheet.cell(
                row_number,
                2,
            ).value
            for row_number in range(2, 2 + len(expected_stats))
        }
        if actual_stats != expected_stats:
            raise RuntimeError(
                f"处理报告统计校验失败：预期 {expected_stats}，"
                f"实际 {actual_stats}"
            )
        issue_header_row = len(expected_stats) + 3
        issue_header = tuple(
            report_sheet.cell(issue_header_row, column).value
            for column in range(1, len(RECEIPTS_REPORT_HEADER) + 1)
        )
        if issue_header != RECEIPTS_REPORT_HEADER:
            raise RuntimeError("处理报告异常明细表头校验失败")
        actual_issues = [
            tuple(
                (
                    report_sheet.cell(row_number, column).value
                    if report_sheet.cell(row_number, column).value is not None
                    else ""
                )
                for column in range(1, len(RECEIPTS_REPORT_HEADER) + 1)
            )
            for row_number in range(
                issue_header_row + 1,
                issue_header_row + 1 + len(expected_issues),
            )
        ]
        if actual_issues != expected_issues:
            raise RuntimeError("处理报告异常明细校验失败")
    finally:
        workbook.close()


def process_receipts() -> None:
    if not RECEIPTS_SOURCE_FILE.exists():
        raise FileNotFoundError(f"未找到源文件：{RECEIPTS_SOURCE_FILE}")

    kept_rows = read_receipt_rows(RECEIPTS_SOURCE_FILE)
    output_rows, stats, issues, duplicate_match_keys = prepare_receipt_data(
        kept_rows
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(RECEIPTS_OUTPUT_HEADER)
    for row in output_rows:
        sheet.append(row)

    font_name, font_path = resolve_font()
    measurement_font = load_measurement_font(font_path)
    normal_font, header_font, header_fill, centered = create_sheet_styles(
        font_name
    )

    measure = width_measurer(measurement_font)
    maximum_widths = [0.0] * sheet.max_column
    duplicate_fill = PatternFill(
        "solid",
        fgColor=RECEIPTS_DUPLICATE_FILL_COLOR,
    )
    for row_number, row in enumerate(sheet.iter_rows(), start=1):
        sheet.row_dimensions[row_number].height = RECEIPTS_ROW_HEIGHT
        row_match_key = ""
        if row_number >= 2:
            document_number = normalize_receipt_identifier(row[0].value)
            receipt_date = row[1].value
            if receipt_date is not None and document_number:
                row_match_key = receipt_match_key(
                    receipt_date.date()
                    if isinstance(receipt_date, datetime)
                    else receipt_date,
                    document_number,
                )
        for cell in row:
            cell.font = header_font if row_number == 1 else normal_font
            cell.alignment = centered
            if row_number == 1:
                cell.fill = header_fill
            elif row_match_key in duplicate_match_keys:
                cell.fill = duplicate_fill

            if row_number >= 2 and cell.column == 1:
                cell.number_format = "@"
            elif row_number >= 2 and cell.column == 2:
                if cell.value not in (None, ""):
                    cell.number_format = "yyyymmdd"

            width = measure(cell.value)
            if width > maximum_widths[cell.column - 1]:
                maximum_widths[cell.column - 1] = width

    sheet.freeze_panes = "A2"
    for column_index, maximum_pixels in enumerate(maximum_widths, start=1):
        column_letter = sheet.cell(1, column_index).column_letter
        sheet.column_dimensions[column_letter].width = min(
            pixels_to_excel_width(maximum_pixels),
            255,
        )

    report_sheet = workbook.create_sheet("Processing Report")
    report_sheet.append(("统计项目", "数量"))
    for label, value in stats.items():
        report_sheet.append((label, value))
    report_sheet.append(())
    report_sheet.append(RECEIPTS_REPORT_HEADER)
    for issue in issues:
        report_sheet.append(issue)
    format_sheet(report_sheet, font_name, measurement_font)

    row_count = len(output_rows)
    save_workbook_atomically(
        workbook,
        RECEIPTS_OUTPUT_FILE,
        lambda path: validate_receipts_output(
            path,
            row_count,
            stats,
            issues,
        ),
    )
    print(f"Receipt statistics complete: {row_count} rows")
    print(
        f"Remarks: {stats['备注总数']}; "
        f"unmatched original invoices: {stats['未匹配原票号数量']}; "
        f"duplicate match keys: {stats['重复匹配键数量']}"
    )
    print(f"Output file: {RECEIPTS_OUTPUT_FILE}")


def read_coupon_rows(source: Path) -> list[list[object]]:
    source_workbook = xlrd.open_workbook(source)
    try:
        source_sheet = source_workbook.sheet_by_index(0)
        if source_sheet.nrows < 3:
            raise ValueError(f"{source.name} 缺少标题行、字段标题行或合计行")
        if source_sheet.ncols < max(COUPON_KEPT_SOURCE_COLUMNS):
            raise ValueError(
                f"{source.name} 列数不足：至少需要 "
                f"{max(COUPON_KEPT_SOURCE_COLUMNS)} 列"
            )
        if str(source_sheet.cell_value(source_sheet.nrows - 1, 0)).strip() != "合计":
            raise ValueError(f"{source.name} 最后一行不是合计行")

        source_header = tuple(
            source_sheet.cell_value(1, column - 1)
            for column in COUPON_KEPT_SOURCE_COLUMNS
        )
        expected_source_header = COUPON_OUTPUT_HEADER[
            :len(COUPON_KEPT_SOURCE_COLUMNS)
        ]
        if source_header != expected_source_header:
            raise ValueError(
                f"{source.name} 保留列字段标题不符合要求："
                f"预期为 {expected_source_header}，实际为 {source_header}。"
                f"如果实际字段包含“2026数码国补（计入收入）”，"
                f"请选择 digital 数据类型，或更换为大家电销售用券文件。"
            )

        rows: list[list[object]] = [list(COUPON_OUTPUT_HEADER)]
        for row_index in range(2, source_sheet.nrows - 1):
            row: list[object] = []
            for column_index in COUPON_KEPT_SOURCE_COLUMNS:
                cell = source_sheet.cell(row_index, column_index - 1)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(
                        value,
                        source_workbook.datemode,
                    )
                elif (
                    cell.ctype == xlrd.XL_CELL_NUMBER
                    and value == int(value)
                ):
                    value = int(value)
                row.append(value)
            if any(value not in (None, "") for value in row):
                document_number = (
                    ""
                    if row[0] is None
                    else str(row[0]).replace("收款", "")
                )
                document_date = normalize_coupon_date(row[1], row_index + 1)
                brand = str(row[3] or "").strip()
                row[3] = COUPON_BRAND_REPLACEMENTS.get(brand, row[3])
                rows.append(
                    [document_number, document_date, *row[2:], None, None]
                )
        return rows
    finally:
        source_workbook.release_resources()


def load_coupon_remark_lookup(source: Path) -> dict[tuple[str, date], str]:
    if not source.exists():
        raise FileNotFoundError(f"未找到备注匹配文件：{source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if "Sheet1" not in workbook.sheetnames:
            raise ValueError(f"{source.name} 缺少 Sheet1 工作表")
        sheet = workbook["Sheet1"]
        header = [cell.value for cell in sheet[1]]
        required_headers = ("单据号", "日期", "备注")
        missing_headers = [
            required_header
            for required_header in required_headers
            if required_header not in header
        ]
        if missing_headers:
            raise ValueError(
                f"{source.name} 缺少字段：{'、'.join(missing_headers)}"
            )

        document_index = header.index("单据号")
        date_index = header.index("日期")
        remark_index = header.index("备注")
        lookup: dict[tuple[str, date], str] = {}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            document_number = normalize_document_number(row[document_index])
            remark = str(row[remark_index] or "").strip()
            if not document_number or not remark:
                continue
            receipt_date = normalize_coupon_date(
                row[date_index],
                row_number,
            )
            key = (document_number, receipt_date)
            existing_remark = lookup.get(key)
            if existing_remark is not None and existing_remark != remark:
                raise ValueError(
                    f"{source.name} 第 {row_number} 行组合键存在冲突备注："
                    f"{document_number} + {receipt_date:%Y-%m-%d}"
                )
            lookup[key] = remark
        return lookup
    finally:
        workbook.close()


def fill_coupon_remarks(
    rows: list[list[object]],
    remark_lookup: dict[tuple[str, date], str],
) -> tuple[int, Decimal, int]:
    matched_rows: list[list[object]] = []
    unmatched_rows: list[list[object]] = []
    subsidy_index = COUPON_OUTPUT_HEADER.index(COUPON_SUBSIDY_HEADER)
    matched_subsidy_total = Decimal("0")
    remark_index = COUPON_OUTPUT_HEADER.index("备注")
    receipt_remark_count = 0
    for row in rows[1:]:
        key = (normalize_document_number(row[0]), row[1])
        remark = remark_lookup.get(key, "")
        row[remark_index] = remark
        if remark:
            receipt_remark_count += 1
        if remark and remark != RECEIPTS_REMARK_SAME_MODEL_REPLACEMENT:
            matched_rows.append(row)
            subsidy = row[subsidy_index]
            if subsidy not in (None, ""):
                try:
                    matched_subsidy_total += Decimal(str(subsidy))
                except InvalidOperation as error:
                    raise ValueError(
                        f"组合键 {key} 的2026家电国补金额无效：{subsidy!r}"
                    ) from error
        else:
            unmatched_rows.append(row)
    rows[1:] = [*unmatched_rows, *matched_rows]
    return len(matched_rows), matched_subsidy_total, receipt_remark_count


def normalize_coupon_reference_supplement_header(value: object) -> str:
    text = str(value or "").strip()
    if text in COUPON_REFERENCE_SUPPLEMENT_HEADER:
        return text
    try:
        return text.encode("cp949").decode("gb18030")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return text


def load_coupon_reference_supplement(
    source: Path,
) -> dict[tuple[str, date], frozenset[str]]:
    if not source.exists():
        print(f"Optional reference supplement file not found; skipping: {source}")
        return {}

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header = tuple(
            normalize_coupon_reference_supplement_header(cell.value)
            for cell in sheet[1]
        )
        if header != COUPON_REFERENCE_SUPPLEMENT_HEADER:
            raise ValueError(
                f"{source.name} 字段标题不符合要求：实际为 {header}"
            )

        references_by_key: dict[tuple[str, date], set[str]] = {}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if all(value in (None, "") for value in row):
                continue
            reference = normalize_receipt_identifier(row[0]).upper()
            document_number = normalize_document_number(row[1])
            if not COUPON_REFERENCE_RE.fullmatch(reference):
                raise ValueError(
                    f"{source.name} 第 {row_number} 行参考号格式无效："
                    f"{row[0]!r}"
                )
            if not document_number:
                raise ValueError(
                    f"{source.name} 第 {row_number} 行单据号为空"
                )
            document_date = normalize_coupon_date(row[2], row_number)
            key = (document_number, document_date)
            references_by_key.setdefault(key, set()).add(reference)
        return {
            key: frozenset(references)
            for key, references in references_by_key.items()
        }
    finally:
        workbook.close()


def fill_coupon_reference_supplement(
    rows: list[list[object]],
    reference_lookup: dict[tuple[str, date], frozenset[str]],
    reference_universe: set[str],
    excluded_bottom_rows: int,
) -> tuple[int, int, set[int], Counter[tuple[str, date, str]]]:
    summary_index = COUPON_OUTPUT_HEADER.index("明细摘要")
    included_rows = (
        rows[1:-excluded_bottom_rows]
        if excluded_bottom_rows > 0
        else rows[1:]
    )
    matched_count = 0
    ambiguous_count = 0
    matched_row_ids: set[int] = set()
    matched_values: Counter[tuple[str, date, str]] = Counter()
    for row in included_rows:
        current_reference = normalize_receipt_identifier(
            row[summary_index]
        ).upper()
        if current_reference in reference_universe:
            continue
        key = (normalize_document_number(row[0]), row[1])
        references = reference_lookup.get(key)
        if references is None:
            continue
        if len(references) == 1:
            reference = next(iter(references))
        elif current_reference in references:
            reference = current_reference
        else:
            ambiguous_count += 1
            continue
        row[summary_index] = reference
        matched_count += 1
        matched_row_ids.add(id(row))
        matched_values[(key[0], key[1], reference)] += 1
    return (
        matched_count,
        ambiguous_count,
        matched_row_ids,
        matched_values,
    )


def load_uploaded_detail_lookup(source: Path) -> dict[str, str]:
    if not source.exists():
        raise FileNotFoundError(f"未找到已上传匹配文件：{source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if "Summary" not in workbook.sheetnames:
            raise ValueError(f"{source.name} 缺少“汇总”工作表")
        sheet = workbook["Summary"]
        header = [cell.value for cell in sheet[1]]
        required_headers = ("检索参考号", "状态", "描述")
        missing_headers = [
            required_header
            for required_header in required_headers
            if required_header not in header
        ]
        if missing_headers:
            raise ValueError(
                f"{source.name} 缺少字段：{'、'.join(missing_headers)}"
            )

        reference_index = header.index("检索参考号")
        status_index = header.index("状态")
        description_index = header.index("描述")
        lookup: dict[str, str] = {}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            reference = normalize_receipt_identifier(
                row[reference_index]
            ).upper()
            if not reference:
                continue
            status = str(row[status_index] or "").strip()
            description = str(row[description_index] or "").strip()
            detail = f"{status}：{description}"
            existing_detail = lookup.get(reference)
            if existing_detail is not None and existing_detail != detail:
                raise ValueError(
                    f"{source.name} 第 {row_number} 行检索参考号存在冲突："
                    f"{reference}"
                )
            lookup[reference] = detail
        return lookup
    finally:
        workbook.close()


def fill_uploaded_details(
    rows: list[list[object]],
    detail_lookup: dict[str, str],
    excluded_bottom_rows: int,
) -> int:
    summary_index = COUPON_OUTPUT_HEADER.index("明细摘要")
    remark_index = COUPON_OUTPUT_HEADER.index("备注")
    detail_index = COUPON_OUTPUT_HEADER.index("详细情况")
    matched_count = 0
    included_rows = (
        rows[1:-excluded_bottom_rows]
        if excluded_bottom_rows > 0
        else rows[1:]
    )
    for row in included_rows:
        reference = normalize_receipt_identifier(
            row[summary_index]
        ).upper()
        detail = detail_lookup.get(reference, "")
        row[detail_index] = detail
        if detail:
            row[remark_index] = "已上传"
            matched_count += 1
    return matched_count


def reference_correction_candidates(
    raw_reference: str,
    reference_universe: set[str],
) -> set[str]:
    candidates: set[str] = set()
    upper_reference = raw_reference.upper()
    compact = re.sub(r"\s+", "", upper_reference)
    cleaned = re.sub(r"[^0-9A-Z]", "", upper_reference)

    for token in re.findall(
        r"(?<!\d)(\d{11}[A-Z])(?![A-Z0-9])",
        upper_reference,
    ):
        if token in reference_universe:
            candidates.add(token)
    if cleaned in reference_universe:
        candidates.add(cleaned)
    if re.fullmatch(r"\d{11}", compact):
        candidates.update(
            reference
            for reference in reference_universe
            if reference[:11] == compact
        )
    if len(compact) == 11:
        for reference in reference_universe:
            if len(reference) == 12 and any(
                reference[:index] + reference[index + 1:] == compact
                for index in range(12)
            ):
                candidates.add(reference)
    elif len(compact) == 13:
        for index in range(13):
            candidate = compact[:index] + compact[index + 1:]
            if candidate in reference_universe:
                candidates.add(candidate)
    elif len(compact) == 12:
        for reference in reference_universe:
            if len(reference) == 12 and sum(
                left != right
                for left, right in zip(compact, reference)
            ) == 1:
                candidates.add(reference)
    return candidates


def correct_coupon_references(
    rows: list[list[object]],
    reference_universe: set[str],
    excluded_bottom_rows: int,
    protected_row_ids: set[int] | None = None,
) -> tuple[int, int, int, list[tuple[str, str, str, str]]]:
    """Correct references and record every decision for the processing report.

    The universe is built from submitted data only, so an operator has to be
    able to review each applied correction, not just the counts.
    """
    summary_index = COUPON_OUTPUT_HEADER.index("明细摘要")
    included_end = len(rows) - excluded_bottom_rows
    included_rows = rows[1:included_end]
    existing_counts = Counter(
        normalize_receipt_identifier(row[summary_index]).upper()
        for row in included_rows
        if normalize_receipt_identifier(row[summary_index])
    )
    proposed: dict[int, str] = {}
    target_counts: Counter[str] = Counter()
    unresolved_count = 0
    decisions: list[tuple[str, str, str, str]] = []

    for row_index, row in enumerate(included_rows, start=1):
        if protected_row_ids is not None and id(row) in protected_row_ids:
            continue
        raw_reference = normalize_receipt_identifier(
            row[summary_index]
        ).upper()
        if not raw_reference or raw_reference in reference_universe:
            continue
        candidates = reference_correction_candidates(
            raw_reference,
            reference_universe,
        )
        if len(candidates) != 1:
            unresolved_count += 1
            decisions.append(
                (
                    REFERENCE_REPORT_UNRESOLVED,
                    str(row_index + 1),
                    raw_reference,
                    f"候选数量 {len(candidates)}，"
                    f"未在已上传数据中找到唯一匹配，保留原值",
                )
            )
            continue
        target = next(iter(candidates))
        proposed[row_index] = target
        target_counts[target] += 1

    corrected_count = 0
    collision_count = 0
    for row_index, target in proposed.items():
        raw_reference = normalize_receipt_identifier(
            rows[row_index][summary_index]
        ).upper()
        if existing_counts[target] > 0 or target_counts[target] > 1:
            collision_count += 1
            decisions.append(
                (
                    REFERENCE_REPORT_COLLISION,
                    str(row_index + 1),
                    raw_reference,
                    f"目标参考号 {target} 已被其他行占用，未纠正",
                )
            )
            continue
        rows[row_index][summary_index] = target
        corrected_count += 1
        decisions.append(
            (
                REFERENCE_REPORT_CORRECTED,
                str(row_index + 1),
                raw_reference,
                f"已自动纠正为 {target}，请人工复核",
            )
        )

    decisions.sort(key=lambda decision: REFERENCE_REPORT_ORDER[decision[0]])
    return corrected_count, unresolved_count, collision_count, decisions


def fill_unmatched_remarks(
    rows: list[list[object]],
    reference_universe: set[str],
    excluded_bottom_rows: int,
) -> int:
    summary_index = COUPON_OUTPUT_HEADER.index("明细摘要")
    remark_index = COUPON_OUTPUT_HEADER.index("备注")
    unmatched_count = 0
    included_rows = (
        rows[1:-excluded_bottom_rows]
        if excluded_bottom_rows > 0
        else rows[1:]
    )
    for row in included_rows:
        reference = normalize_receipt_identifier(
            row[summary_index]
        ).upper()
        if reference not in reference_universe:
            row[remark_index] = "未上传"
            unmatched_count += 1
    return unmatched_count


def coupon_text_sort_value(value: object) -> str:
    return str(value or "").strip()


def coupon_date_sort_value(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.max


def coupon_pink_sort_key(row: list[object] | tuple[object, ...]) -> tuple:
    return (
        coupon_text_sort_value(
            row[COUPON_OUTPUT_HEADER.index("财务大类")]
        ),
        coupon_text_sort_value(row[COUPON_OUTPUT_HEADER.index("品牌")]),
        coupon_text_sort_value(row[COUPON_OUTPUT_HEADER.index("商品名称")]),
        coupon_date_sort_value(row[COUPON_OUTPUT_HEADER.index("单据日期")]),
    )


def coupon_regular_sort_key(row: list[object] | tuple[object, ...]) -> tuple:
    remark = coupon_text_sort_value(
        row[COUPON_OUTPUT_HEADER.index("备注")]
    )
    return (
        COUPON_REMARK_SORT_PRIORITY.get(remark, 2),
        remark,
        coupon_text_sort_value(
            row[COUPON_OUTPUT_HEADER.index("财务大类")]
        ),
        coupon_text_sort_value(row[COUPON_OUTPUT_HEADER.index("品牌")]),
        coupon_text_sort_value(
            row[COUPON_OUTPUT_HEADER.index("详细情况")]
        ),
        coupon_date_sort_value(row[COUPON_OUTPUT_HEADER.index("单据日期")]),
        coupon_text_sort_value(
            row[COUPON_OUTPUT_HEADER.index("商品名称")]
        ),
    )


def coupon_group_regular_sort_key(
    row: list[object] | tuple[object, ...],
) -> tuple:
    remark = coupon_text_sort_value(
        row[COUPON_OUTPUT_HEADER.index("备注")]
    )
    priority = {"已上传": 0, "未上传": 1}.get(remark, 2)
    return (
        priority,
        remark,
        coupon_text_sort_value(
            row[COUPON_OUTPUT_HEADER.index("详细情况")]
        ),
        coupon_date_sort_value(row[COUPON_OUTPUT_HEADER.index("单据日期")]),
        coupon_text_sort_value(
            row[COUPON_OUTPUT_HEADER.index("商品名称")]
        ),
    )


def select_coupon_group_columns(
    row: list[object] | tuple[object, ...],
) -> tuple[object, ...]:
    return tuple(row[index] for index in COUPON_GROUP_COLUMN_INDEXES)


def sort_coupon_detail_rows(
    rows: list[list[object]],
    pink_row_count: int,
) -> None:
    pink_start = len(rows) - pink_row_count
    regular_rows = rows[1:pink_start] if pink_row_count else rows[1:]
    pink_rows = rows[pink_start:] if pink_row_count else []
    regular_rows.sort(key=coupon_regular_sort_key)
    pink_rows.sort(key=coupon_pink_sort_key)
    rows[1:] = [*regular_rows, *pink_rows]


def build_coupon_summary(
    rows: list[list[object]],
    excluded_bottom_rows: int,
) -> tuple[
    list[tuple[object, ...]],
    list[tuple[object, ...]],
    list[tuple[object, ...]],
]:
    category_index = COUPON_OUTPUT_HEADER.index("财务大类")
    brand_index = COUPON_OUTPUT_HEADER.index("品牌")
    remark_index = COUPON_OUTPUT_HEADER.index("备注")
    detail_index = COUPON_OUTPUT_HEADER.index("详细情况")
    subsidy_index = COUPON_OUTPUT_HEADER.index(COUPON_SUBSIDY_HEADER)
    included_rows = (
        rows[1:-excluded_bottom_rows]
        if excluded_bottom_rows > 0
        else rows[1:]
    )
    grouped_counts: Counter[tuple[str, str, str]] = Counter()
    grouped_totals: dict[tuple[str, str, str], Decimal] = {}
    approved_counts: Counter[tuple[str, str]] = Counter()
    approved_totals: dict[tuple[str, str], Decimal] = {}
    remark_counts: Counter[str] = Counter()
    remark_totals: dict[str, Decimal] = {}
    for row in included_rows:
        category = str(row[category_index] or "").strip()
        brand = str(row[brand_index] or "").strip()
        remark = str(row[remark_index] or "").strip()
        detail = str(row[detail_index] or "").strip()
        key = (category, brand, remark)
        grouped_counts[key] += 1
        grouped_totals.setdefault(key, Decimal("0"))
        remark_counts[remark] += 1
        remark_totals.setdefault(remark, Decimal("0"))
        approved_key = (category, brand)
        if "审核通过" in detail:
            approved_counts[approved_key] += 1
            approved_totals.setdefault(approved_key, Decimal("0"))
        subsidy = row[subsidy_index]
        if subsidy not in (None, ""):
            try:
                amount = Decimal(str(subsidy))
                grouped_totals[key] += amount
                remark_totals[remark] += amount
                if "审核通过" in detail:
                    approved_totals[approved_key] += amount
            except InvalidOperation as error:
                raise ValueError(
                    f"{key!r} 的2026家电国补金额无效：{subsidy!r}"
                ) from error

    summary_rows = [
        (
            *key,
            grouped_counts[key],
            float(grouped_totals[key].quantize(Decimal("0.01"))),
        )
        for key in sorted(grouped_counts)
    ]
    summary_rows.append(
        (
            "合计",
            None,
            None,
            sum(grouped_counts.values()),
            float(
                sum(grouped_totals.values(), Decimal("0")).quantize(
                    Decimal("0.01")
                )
            ),
        )
    )
    approved_rows = [
        (
            *key,
            approved_counts[key],
            float(approved_totals[key].quantize(Decimal("0.01"))),
        )
        for key in sorted(approved_counts)
    ]
    approved_rows.append(
        (
            "合计",
            None,
            sum(approved_counts.values()),
            float(
                sum(approved_totals.values(), Decimal("0")).quantize(
                    Decimal("0.01")
                )
            ),
        )
    )
    remark_rows = [
        (
            remark,
            remark_counts[remark],
            float(remark_totals[remark].quantize(Decimal("0.01"))),
        )
        for remark in sorted(remark_counts)
    ]
    remark_rows.append(
        (
            "合计",
            sum(remark_counts.values()),
            float(
                sum(remark_totals.values(), Decimal("0")).quantize(
                    Decimal("0.01")
                )
            ),
        )
    )
    return summary_rows, approved_rows, remark_rows


def coupon_group_sheet_title(
    category: str,
    brand: str,
    used_titles: set[str],
) -> str:
    base_title = f"{category or '未分类'}-{brand or '未品牌'}"
    base_title = INVALID_SHEET_TITLE_RE.sub("_", base_title).strip("'")
    base_title = base_title[:31] or "未分类-未品牌"
    title = base_title
    suffix = 2
    while title in used_titles:
        suffix_text = f"-{suffix}"
        title = f"{base_title[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_titles.add(title)
    return title


def build_coupon_group_sheets(
    rows: list[list[object]],
    excluded_bottom_rows: int,
) -> list[tuple[str, str, str, list[tuple[list[object], bool]]]]:
    category_index = COUPON_OUTPUT_HEADER.index("财务大类")
    brand_index = COUPON_OUTPUT_HEADER.index("品牌")
    first_pink_index = len(rows) - excluded_bottom_rows
    groups: dict[
        tuple[str, str],
        list[tuple[list[object], bool]],
    ] = {}
    for row_index, row in enumerate(rows[1:], start=1):
        category = str(row[category_index] or "").strip()
        brand = str(row[brand_index] or "").strip()
        groups.setdefault((category, brand), []).append(
            (row, row_index >= first_pink_index)
        )

    for grouped_rows in groups.values():
        regular_rows = [
            item for item in grouped_rows if not item[1]
        ]
        pink_rows = [
            item for item in grouped_rows if item[1]
        ]
        regular_rows.sort(key=lambda item: coupon_group_regular_sort_key(item[0]))
        pink_rows.sort(key=lambda item: coupon_pink_sort_key(item[0]))
        grouped_rows[:] = [*regular_rows, *pink_rows]

    used_titles = {"Summary", "Details"}
    return [
        (
            coupon_group_sheet_title(category, brand, used_titles),
            category,
            brand,
            grouped_rows,
        )
        for (category, brand), grouped_rows in sorted(groups.items())
    ]


def merge_coupon_summary_groups(
    sheet,
    start_row: int,
    end_row: int,
    start_column: int = 1,
) -> None:
    centered = Alignment(horizontal="center", vertical="center")
    category_column = start_column
    brand_column = start_column + 1
    for column in (brand_column, category_column):
        group_start = start_row
        while group_start <= end_row:
            category = sheet.cell(group_start, category_column).value
            value = sheet.cell(group_start, column).value
            group_end = group_start
            while (
                group_end + 1 <= end_row
                and sheet.cell(
                    group_end + 1,
                    category_column,
                ).value
                == category
                and sheet.cell(group_end + 1, column).value == value
            ):
                group_end += 1
            if group_end > group_start:
                sheet.merge_cells(
                    start_row=group_start,
                    start_column=column,
                    end_row=group_end,
                    end_column=column,
                )
            sheet.cell(group_start, column).alignment = centered
            group_start = group_end + 1


def merged_coupon_summary_values(
    rows: list[tuple[object, ...]],
) -> list[tuple[object, ...]]:
    merged_rows: list[tuple[object, ...]] = []
    previous_category = None
    previous_brand = None
    for row in rows:
        if row[0] == "合计":
            merged_rows.append(row)
            continue
        category, brand, *remaining = row
        displayed_category = (
            None if category == previous_category else category
        )
        displayed_brand = (
            None
            if category == previous_category and brand == previous_brand
            else brand
        )
        merged_rows.append(
            (displayed_category, displayed_brand, *remaining)
        )
        previous_category = category
        previous_brand = brand
    return merged_rows


def apply_coupon_summary_borders(
    sheet,
    *,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> None:
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_column,
        max_col=max_column,
    ):
        for cell in row:
            cell.border = COUPON_SUMMARY_BORDER


def validate_coupon_output(
    path: Path,
    expected_data_rows: int,
    expected_matched_rows: int,
    remark_lookup: dict[tuple[str, date], str],
    expected_reference_supplement_matches: Counter[
        tuple[str, date, str]
    ],
    expected_matched_subsidy_total: Decimal,
    detail_lookup: dict[str, str],
    expected_uploaded_rows: int,
    reference_universe: set[str],
    expected_unresolved_rows: int,
    expected_unmatched_rows: int,
    expected_reference_decisions: list[tuple[str, str, str, str]],
    expected_summary_rows: list[tuple[object, ...]],
    expected_approved_rows: list[tuple[object, ...]],
    expected_remark_rows: list[tuple[object, ...]],
    expected_group_sheets: list[
        tuple[str, str, str, list[tuple[list[object], bool]]]
    ],
) -> None:
    workbook = load_workbook(path, data_only=True)
    try:
        expected_sheet_names = [
            "Summary",
            "Details",
            *(group[0] for group in expected_group_sheets),
            REFERENCE_REPORT_SHEET,
        ]
        if workbook.sheetnames != expected_sheet_names:
            raise RuntimeError(
                "销售用券工作表校验失败："
                f"预期 {expected_sheet_names}，实际 {workbook.sheetnames}"
            )
        report_sheet = workbook[REFERENCE_REPORT_SHEET]
        report_header = tuple(cell.value for cell in report_sheet[1])
        if report_header != REFERENCE_REPORT_HEADER:
            raise RuntimeError(
                f"参考号处理报告字段标题校验失败：实际为 {report_header}"
            )
        actual_decisions = [
            tuple(
                "" if value is None else str(value)
                for value in row
            )
            for row in report_sheet.iter_rows(min_row=2, values_only=True)
        ]
        if actual_decisions != [
            tuple(decision) for decision in expected_reference_decisions
        ]:
            raise RuntimeError(
                f"参考号处理报告内容校验失败：预期 "
                f"{len(expected_reference_decisions)} 条，"
                f"实际 {len(actual_decisions)} 条"
            )
        sheet = workbook["Details"]
        header = tuple(cell.value for cell in sheet[1])
        if header != COUPON_OUTPUT_HEADER:
            raise RuntimeError(
                f"销售用券字段标题校验失败：实际为 {header}"
            )
        if sheet.max_column != len(COUPON_OUTPUT_HEADER):
            raise RuntimeError(
                f"销售用券列数校验失败：实际为 {sheet.max_column}"
            )
        actual_data_rows = max(sheet.max_row - 1, 0)
        if actual_data_rows != expected_data_rows:
            raise RuntimeError(
                f"销售用券行数校验失败：预期 {expected_data_rows} 条，"
                f"实际 {actual_data_rows} 条"
            )
        detail_column = COUPON_OUTPUT_HEADER.index("详细情况") + 1
        actual_uploaded_rows = sum(
            sheet.cell(row_number, detail_column).value not in (None, "")
            for row_number in range(2, sheet.max_row + 1)
        )
        if actual_uploaded_rows != expected_uploaded_rows:
            raise RuntimeError(
                f"销售用券已上传匹配数校验失败：预期 "
                f"{expected_uploaded_rows} 条，实际 {actual_uploaded_rows} 条"
            )
        remark_column = COUPON_OUTPUT_HEADER.index("备注") + 1
        actual_unuploaded_remark_rows = sum(
            sheet.cell(row_number, remark_column).value == "未上传"
            for row_number in range(2, sheet.max_row + 1)
        )
        expected_unuploaded_remark_rows = expected_unmatched_rows
        if actual_unuploaded_remark_rows != expected_unuploaded_remark_rows:
            raise RuntimeError(
                f"销售用券“未上传”备注数量校验失败：预期 "
                f"{expected_unuploaded_remark_rows} 条，"
                f"实际 {actual_unuploaded_remark_rows} 条"
            )
        matched_start_row = sheet.max_row - expected_matched_rows + 1
        summary_column = COUPON_OUTPUT_HEADER.index("明细摘要") + 1
        actual_unresolved_rows = sum(
            reference not in reference_universe
            for reference in (
                normalize_receipt_identifier(
                    sheet.cell(row_number, summary_column).value
                ).upper()
                for row_number in range(2, matched_start_row)
            )
            if reference
        )
        if actual_unresolved_rows != expected_unresolved_rows:
            raise RuntimeError(
                f"销售用券未解决参考号数量校验失败：预期 "
                f"{expected_unresolved_rows} 条，实际 "
                f"{actual_unresolved_rows} 条"
            )
        actual_regular_rows = [
            row
            for row in sheet.iter_rows(
                min_row=2,
                max_row=matched_start_row - 1,
                min_col=1,
                max_col=len(COUPON_OUTPUT_HEADER),
                values_only=True,
            )
        ]
        if actual_regular_rows != sorted(
            actual_regular_rows,
            key=coupon_regular_sort_key,
        ):
            raise RuntimeError("销售用券明细非粉色区域排序校验失败")
        actual_pink_rows = [
            row
            for row in sheet.iter_rows(
                min_row=matched_start_row,
                max_row=sheet.max_row,
                min_col=1,
                max_col=len(COUPON_OUTPUT_HEADER),
                values_only=True,
            )
        ]
        if actual_pink_rows != sorted(
            actual_pink_rows,
            key=coupon_pink_sort_key,
        ):
            raise RuntimeError("销售用券明细粉色区域排序校验失败")
        product_name_column = COUPON_OUTPUT_HEADER.index("商品名称") + 1
        for column_name, column_number in (
            ("商品名称", product_name_column),
            ("详细情况", detail_column),
        ):
            if any(
                sheet.cell(row_number, column_number).alignment.horizontal
                != "left"
                for row_number in range(2, sheet.max_row + 1)
            ):
                raise RuntimeError(
                    f"销售用券{column_name}列左对齐校验失败"
                )
        brand_column = COUPON_OUTPUT_HEADER.index("品牌") + 1
        remaining_source_brands = {
            str(sheet.cell(row_number, brand_column).value or "").strip()
            for row_number in range(2, sheet.max_row + 1)
        } & COUPON_BRAND_REPLACEMENTS.keys()
        if remaining_source_brands:
            raise RuntimeError(
                "销售用券品牌替换校验失败："
                f"{sorted(remaining_source_brands)}"
            )
        actual_matched_subsidy_total = Decimal("0")
        actual_reference_supplement_matches: Counter[
            tuple[str, date, str]
        ] = Counter()
        subsidy_column = COUPON_OUTPUT_HEADER.index(COUPON_SUBSIDY_HEADER) + 1
        for row_number in range(2, sheet.max_row + 1):
            document_cell = sheet.cell(row_number, 1)
            date_cell = sheet.cell(row_number, 2)
            remark_cell = sheet.cell(row_number, remark_column)
            detail_cell = sheet.cell(row_number, detail_column)
            if "收款" in str(document_cell.value or ""):
                raise RuntimeError(
                    f"销售用券第 {row_number} 行单据号仍包含“收款”"
                )
            if document_cell.number_format != "@":
                raise RuntimeError(
                    f"销售用券第 {row_number} 行单据号不是文本格式"
                )
            if (
                date_cell.value not in (None, "")
                and date_cell.number_format != "yyyy-mm-dd"
            ):
                raise RuntimeError(
                    f"销售用券第 {row_number} 行单据日期格式错误"
                )
            if date_cell.value not in (None, "") and not isinstance(
                date_cell.value,
                (date, datetime),
            ):
                raise RuntimeError(
                    f"销售用券第 {row_number} 行单据日期不是日期值"
                )
            document_date = (
                date_cell.value.date()
                if isinstance(date_cell.value, datetime)
                else date_cell.value
            )
            receipt_remark = remark_lookup.get(
                (
                    normalize_document_number(document_cell.value),
                    document_date,
                ),
                "",
            )
            reference = normalize_receipt_identifier(
                sheet.cell(row_number, summary_column).value
            ).upper()
            expected_pink = (
                expected_matched_rows > 0
                and row_number >= matched_start_row
            )
            supplement_match = (
                normalize_document_number(document_cell.value),
                document_date,
                reference,
            )
            if (
                not expected_pink
                and supplement_match
                in expected_reference_supplement_matches
            ):
                actual_reference_supplement_matches[supplement_match] += 1
            if expected_pink:
                expected_detail = ""
                expected_remark = receipt_remark
            else:
                expected_detail = detail_lookup.get(reference, "")
                if expected_detail:
                    expected_remark = "已上传"
                elif reference not in reference_universe:
                    expected_remark = "未上传"
                else:
                    expected_remark = receipt_remark
            actual_remark = str(remark_cell.value or "")
            if actual_remark != expected_remark:
                raise RuntimeError(
                    f"销售用券第 {row_number} 行备注匹配校验失败"
                )
            if str(detail_cell.value or "") != expected_detail:
                raise RuntimeError(
                    f"销售用券第 {row_number} 行详细情况匹配校验失败"
                )
            is_pink = all(
                sheet.cell(row_number, column).fill.fill_type == "solid"
                and sheet.cell(row_number, column).fill.fgColor.rgb
                in {
                    COUPON_MATCH_FILL_COLOR,
                    f"00{COUPON_MATCH_FILL_COLOR}",
                    f"FF{COUPON_MATCH_FILL_COLOR}",
                }
                for column in range(1, len(COUPON_OUTPUT_HEADER) + 1)
            )
            if is_pink != expected_pink:
                raise RuntimeError(
                    f"销售用券第 {row_number} 行粉色填充或位置校验失败"
                )
            if expected_pink:
                subsidy = sheet.cell(row_number, subsidy_column).value
                if subsidy not in (None, ""):
                    actual_matched_subsidy_total += Decimal(str(subsidy))

        if any(
            actual_reference_supplement_matches[match] < expected_count
            for match, expected_count
            in expected_reference_supplement_matches.items()
        ):
            raise RuntimeError(
                "销售用券补充参考号逐行匹配结果校验失败"
            )
        if as_currency(actual_matched_subsidy_total) != as_currency(
            expected_matched_subsidy_total
        ):
            raise RuntimeError(
                "销售用券匹配行国补合计校验失败："
                f"预期 {expected_matched_subsidy_total}，"
                f"实际 {actual_matched_subsidy_total}"
            )
        summary_sheet = workbook["Summary"]
        summary_header = tuple(
            summary_sheet.cell(1, column).value
            for column in range(1, len(COUPON_SUMMARY_HEADER) + 1)
        )
        if summary_header != COUPON_SUMMARY_HEADER:
            raise RuntimeError(
                f"销售用券汇总字段标题校验失败：实际为 {summary_header}"
            )
        first_total_row = 1 + len(expected_summary_rows)
        actual_summary_rows = [
            tuple(
                summary_sheet.cell(row_number, column_number).value
                for column_number in range(1, 6)
            )
            for row_number in range(2, first_total_row + 1)
        ]
        if actual_summary_rows != merged_coupon_summary_values(
            expected_summary_rows
        ):
            raise RuntimeError("销售用券财务大类、品牌和备注汇总校验失败")
        if not actual_summary_rows or actual_summary_rows[-1][0] != "合计":
            raise RuntimeError("销售用券汇总缺少底部合计行")
        detail_summary_rows = actual_summary_rows[:-1]
        total_summary_row = actual_summary_rows[-1]
        if sum(row[3] for row in detail_summary_rows) != (
            expected_data_rows - expected_matched_rows
        ):
            raise RuntimeError("销售用券汇总包含粉红色数据或数量不完整")
        if total_summary_row[3] != sum(
            row[3] for row in detail_summary_rows
        ):
            raise RuntimeError("销售用券汇总合计数量校验失败")
        if Decimal(str(total_summary_row[4])) != sum(
            (Decimal(str(row[4])) for row in detail_summary_rows),
            Decimal("0"),
        ):
            raise RuntimeError("销售用券汇总合计金额校验失败")
        approved_title_row = 1
        approved_start_column = 7
        if summary_sheet.cell(
            approved_title_row,
            approved_start_column,
        ).value != (
            COUPON_APPROVED_SUMMARY_TITLE
        ):
            raise RuntimeError("销售用券汇总缺少“审核通过明细”表头")
        expected_title_range = (
            f"G{approved_title_row}:J"
            f"{approved_title_row}"
        )
        if expected_title_range not in {
            str(cell_range) for cell_range in summary_sheet.merged_cells.ranges
        }:
            raise RuntimeError("销售用券审核通过明细标题未跨表格全宽合并")
        title_alignment = summary_sheet.cell(
            approved_title_row,
            approved_start_column,
        ).alignment
        if (
            title_alignment.horizontal != "center"
            or title_alignment.vertical != "center"
        ):
            raise RuntimeError("销售用券审核通过明细标题未全局居中")
        approved_header = tuple(
            summary_sheet.cell(
                approved_title_row + 1,
                approved_start_column + column,
            ).value
            for column in range(4)
        )
        if approved_header != COUPON_APPROVED_SUMMARY_HEADER:
            raise RuntimeError("销售用券审核通过明细字段标题校验失败")
        actual_approved_rows = [
            tuple(
                summary_sheet.cell(
                    row_number,
                    approved_start_column + column,
                ).value
                for column in range(4)
            )
            for row_number in range(
                approved_title_row + 2,
                approved_title_row + 2 + len(expected_approved_rows),
            )
        ]
        if actual_approved_rows != merged_coupon_summary_values(
            expected_approved_rows
        ):
            expected_merged_approved_rows = merged_coupon_summary_values(
                expected_approved_rows
            )
            first_difference = next(
                (
                    (index, actual, expected)
                    for index, (actual, expected) in enumerate(
                        zip(
                            actual_approved_rows,
                            expected_merged_approved_rows,
                        ),
                        start=1,
                    )
                    if actual != expected
                ),
                None,
            )
            raise RuntimeError(
                "销售用券审核通过明细汇总校验失败："
                f"{first_difference}"
            )
        if not actual_approved_rows or actual_approved_rows[-1][0] != "合计":
            raise RuntimeError("销售用券审核通过明细缺少底部合计行")
        remark_title_row = (
            approved_title_row + len(expected_approved_rows) + 3
        )
        if any(
            summary_sheet.cell(
                remark_title_row - 1,
                column,
            ).value
            not in (None, "")
            for column in range(7, 11)
        ):
            raise RuntimeError("销售用券审核通过明细与备注汇总之间缺少空白行")
        if summary_sheet.cell(remark_title_row, 7).value != (
            COUPON_REMARK_SUMMARY_TITLE
        ):
            raise RuntimeError("销售用券汇总缺少“备注汇总”标题")
        expected_remark_title_range = (
            f"G{remark_title_row}:I{remark_title_row}"
        )
        if expected_remark_title_range not in {
            str(cell_range) for cell_range in summary_sheet.merged_cells.ranges
        }:
            raise RuntimeError("销售用券备注汇总标题未跨表格全宽合并")
        remark_title_alignment = summary_sheet.cell(
            remark_title_row,
            7,
        ).alignment
        if (
            remark_title_alignment.horizontal != "center"
            or remark_title_alignment.vertical != "center"
        ):
            raise RuntimeError("销售用券备注汇总标题未全局居中")
        actual_remark_header = tuple(
            summary_sheet.cell(remark_title_row + 1, column).value
            for column in range(7, 10)
        )
        if actual_remark_header != COUPON_REMARK_SUMMARY_HEADER:
            raise RuntimeError("销售用券备注汇总字段标题校验失败")
        actual_remark_rows = [
            tuple(
                summary_sheet.cell(row_number, column_number).value
                for column_number in range(7, 10)
            )
            for row_number in range(
                remark_title_row + 2,
                remark_title_row + 2 + len(expected_remark_rows),
            )
        ]
        if actual_remark_rows != expected_remark_rows:
            raise RuntimeError("销售用券备注汇总结果校验失败")
        if not actual_remark_rows or actual_remark_rows[-1][0] != "合计":
            raise RuntimeError("销售用券备注汇总缺少底部合计行")
        remark_detail_rows = actual_remark_rows[:-1]
        remark_total_row = actual_remark_rows[-1]
        if remark_total_row[1] != expected_data_rows - expected_matched_rows:
            raise RuntimeError("销售用券备注汇总数量范围校验失败")
        if remark_total_row[1] != sum(row[1] for row in remark_detail_rows):
            raise RuntimeError("销售用券备注汇总合计数量校验失败")
        if Decimal(str(remark_total_row[2])) != sum(
            (Decimal(str(row[2])) for row in remark_detail_rows),
            Decimal("0"),
        ):
            raise RuntimeError("销售用券备注汇总合计金额校验失败")
        bordered_ranges = (
            (1, len(expected_summary_rows) + 1, 1, 5),
            (
                approved_title_row,
                approved_title_row + len(expected_approved_rows) + 1,
                7,
                10,
            ),
            (
                remark_title_row,
                remark_title_row + len(expected_remark_rows) + 1,
                7,
                9,
            ),
        )
        for min_row, max_row, min_column, max_column in bordered_ranges:
            for row in summary_sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_column,
                max_col=max_column,
            ):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    if any(
                        side.style != "thin"
                        for side in (
                            cell.border.left,
                            cell.border.right,
                            cell.border.top,
                            cell.border.bottom,
                        )
                    ):
                        raise RuntimeError("销售用券汇总表格边框校验失败")
        for (
            sheet_name,
            _,
            _,
            expected_rows,
        ) in expected_group_sheets:
            group_sheet = workbook[sheet_name]
            group_header = tuple(cell.value for cell in group_sheet[1])
            if group_header != COUPON_GROUP_HEADER:
                raise RuntimeError(f"{sheet_name}分类工作表标题校验失败")
            if group_sheet.max_row - 1 != len(expected_rows):
                raise RuntimeError(
                    f"{sheet_name}分类工作表行数校验失败："
                    f"预期 {len(expected_rows)} 条，"
                    f"实际 {group_sheet.max_row - 1} 条"
                )
            for row_number, (_, expected_pink) in enumerate(
                expected_rows,
                start=2,
            ):
                expected_values = select_coupon_group_columns(
                    expected_rows[row_number - 2][0]
                )
                actual_values = tuple(
                    group_sheet.cell(row_number, column).value
                    for column in range(1, len(COUPON_GROUP_HEADER) + 1)
                )
                normalized_actual = tuple(
                    value.date() if isinstance(value, datetime) else value
                    for value in actual_values
                )
                comparable_actual = (
                    coupon_text_sort_value(normalized_actual[0]),
                    normalized_actual[1],
                    coupon_text_sort_value(normalized_actual[2]),
                    (
                        round(float(normalized_actual[3]), 2)
                        if normalized_actual[3] not in (None, "")
                        else normalized_actual[3]
                    ),
                    coupon_text_sort_value(normalized_actual[4]),
                    coupon_text_sort_value(normalized_actual[5]),
                )
                comparable_expected = (
                    coupon_text_sort_value(expected_values[0]),
                    expected_values[1],
                    coupon_text_sort_value(expected_values[2]),
                    (
                        round(float(expected_values[3]), 2)
                        if expected_values[3] not in (None, "")
                        else expected_values[3]
                    ),
                    coupon_text_sort_value(expected_values[4]),
                    coupon_text_sort_value(expected_values[5]),
                )
                if comparable_actual != comparable_expected:
                    raise RuntimeError(
                        f"{sheet_name}分类工作表第 {row_number} 行数据或"
                        f"排序校验失败：实际 {comparable_actual!r}，"
                        f"预期 {comparable_expected!r}"
                    )
                actual_pink = all(
                    group_sheet.cell(
                        row_number,
                        column,
                    ).fill.fill_type == "solid"
                    and group_sheet.cell(
                        row_number,
                        column,
                    ).fill.fgColor.rgb
                    in {
                        COUPON_MATCH_FILL_COLOR,
                        f"00{COUPON_MATCH_FILL_COLOR}",
                        f"FF{COUPON_MATCH_FILL_COLOR}",
                    }
                    for column in range(1, len(COUPON_GROUP_HEADER) + 1)
                )
                if actual_pink != expected_pink:
                    raise RuntimeError(
                        f"{sheet_name}分类工作表粉色标记校验失败"
                    )
    finally:
        workbook.close()


def process_coupon_sales() -> None:
    if not COUPON_SOURCE_FILE.exists():
        raise FileNotFoundError(f"未找到源文件：{COUPON_SOURCE_FILE}")

    rows = read_coupon_rows(COUPON_SOURCE_FILE)
    remark_lookup = load_coupon_remark_lookup(COUPON_REMARK_SOURCE_FILE)
    matched_count, matched_subsidy_total, receipt_remark_count = (
        fill_coupon_remarks(
        rows,
        remark_lookup,
        )
    )
    detail_lookup = load_uploaded_detail_lookup(COUPON_UPLOADED_SOURCE_FILE)
    # Unsubmitted data is no longer supplied, so submitted data is the only
    # source of valid references.
    reference_universe = set(detail_lookup)
    reference_supplement_lookup = load_coupon_reference_supplement(
        COUPON_REFERENCE_SUPPLEMENT_FILE
    )
    (
        reference_supplement_count,
        ambiguous_reference_supplement_count,
        reference_supplement_row_ids,
        reference_supplement_matches,
    ) = fill_coupon_reference_supplement(
        rows,
        reference_supplement_lookup,
        reference_universe,
        matched_count,
    )
    (
        corrected_count,
        unresolved_count,
        correction_collision_count,
        reference_decisions,
    ) = correct_coupon_references(
        rows,
        reference_universe,
        matched_count,
        reference_supplement_row_ids,
    )
    summary_index = COUPON_OUTPUT_HEADER.index("明细摘要")
    regular_rows = rows[1:-matched_count] if matched_count > 0 else rows[1:]
    final_unresolved_reference_count = sum(
        reference not in reference_universe
        for reference in (
            normalize_receipt_identifier(row[summary_index]).upper()
            for row in regular_rows
        )
        if reference
    )
    uploaded_count = fill_uploaded_details(
        rows,
        detail_lookup,
        matched_count,
    )
    unmatched_count = fill_unmatched_remarks(
        rows,
        reference_universe,
        matched_count,
    )
    sort_coupon_detail_rows(rows, matched_count)
    summary_rows, approved_rows, remark_rows = build_coupon_summary(
        rows,
        matched_count,
    )
    group_sheets = build_coupon_group_sheets(rows, matched_count)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Details"
    for row in rows:
        sheet.append(row)

    font_name, font_path = resolve_font()
    measurement_font = load_measurement_font(font_path)
    format_sheet(
        sheet,
        font_name,
        measurement_font,
        ("商品名称", "详细情况"),
    )
    for cell in sheet["A"][1:]:
        cell.number_format = "@"
    for cell in sheet["B"][1:]:
        if cell.value not in (None, ""):
            cell.number_format = "yyyy-mm-dd"
    matched_fill = PatternFill("solid", fgColor=COUPON_MATCH_FILL_COLOR)
    matched_start_row = sheet.max_row - matched_count + 1
    for row in sheet.iter_rows(
        min_row=matched_start_row,
        max_row=sheet.max_row,
    ):
        for cell in row:
            cell.fill = matched_fill

    summary_sheet = workbook.create_sheet("Summary", 0)
    summary_sheet.append(COUPON_SUMMARY_HEADER)
    for summary_row in summary_rows:
        summary_sheet.append(summary_row)
    approved_title_row = 1
    approved_start_column = 7
    summary_sheet.cell(
        approved_title_row,
        approved_start_column,
        COUPON_APPROVED_SUMMARY_TITLE,
    )
    approved_header_row = approved_title_row + 1
    for column, value in enumerate(
        COUPON_APPROVED_SUMMARY_HEADER,
        start=approved_start_column,
    ):
        summary_sheet.cell(approved_header_row, column, value)
    approved_data_start_row = approved_header_row + 1
    for row_number, approved_row in enumerate(
        approved_rows,
        start=approved_data_start_row,
    ):
        for column, value in enumerate(
            approved_row,
            start=approved_start_column,
        ):
            summary_sheet.cell(row_number, column, value)
    remark_title_row = approved_data_start_row + len(approved_rows) + 1
    summary_sheet.cell(
        remark_title_row,
        approved_start_column,
        COUPON_REMARK_SUMMARY_TITLE,
    )
    remark_header_row = remark_title_row + 1
    for column, value in enumerate(
        COUPON_REMARK_SUMMARY_HEADER,
        start=approved_start_column,
    ):
        summary_sheet.cell(remark_header_row, column, value)
    remark_data_start_row = remark_header_row + 1
    for row_number, remark_row in enumerate(
        remark_rows,
        start=remark_data_start_row,
    ):
        for column, value in enumerate(
            remark_row,
            start=approved_start_column,
        ):
            summary_sheet.cell(row_number, column, value)
    format_sheet(summary_sheet, font_name, measurement_font)
    normal_font, header_font, header_fill, centered = create_sheet_styles(
        font_name
    )
    summary_sheet.merge_cells(
        start_row=approved_title_row,
        start_column=approved_start_column,
        end_row=approved_title_row,
        end_column=approved_start_column + 3,
    )
    title_cell = summary_sheet.cell(
        approved_title_row,
        approved_start_column,
    )
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = centered
    for cell in summary_sheet.iter_cols(
        min_col=approved_start_column,
        max_col=approved_start_column + 3,
        min_row=approved_header_row,
        max_row=approved_header_row,
    ):
        cell = cell[0]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
    summary_sheet.merge_cells(
        start_row=remark_title_row,
        start_column=approved_start_column,
        end_row=remark_title_row,
        end_column=approved_start_column + 2,
    )
    remark_title_cell = summary_sheet.cell(
        remark_title_row,
        approved_start_column,
    )
    remark_title_cell.font = header_font
    remark_title_cell.fill = header_fill
    remark_title_cell.alignment = centered
    for cell in summary_sheet.iter_cols(
        min_col=approved_start_column,
        max_col=approved_start_column + 2,
        min_row=remark_header_row,
        max_row=remark_header_row,
    ):
        cell = cell[0]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
    merge_coupon_summary_groups(
        summary_sheet,
        2,
        len(summary_rows),
    )
    merge_coupon_summary_groups(
        summary_sheet,
        approved_data_start_row,
        approved_data_start_row + len(approved_rows) - 2,
        approved_start_column,
    )
    apply_coupon_summary_borders(
        summary_sheet,
        min_row=1,
        max_row=len(summary_rows) + 1,
        min_column=1,
        max_column=5,
    )
    apply_coupon_summary_borders(
        summary_sheet,
        min_row=approved_title_row,
        max_row=approved_data_start_row + len(approved_rows) - 1,
        min_column=approved_start_column,
        max_column=approved_start_column + 3,
    )
    apply_coupon_summary_borders(
        summary_sheet,
        min_row=remark_title_row,
        max_row=remark_data_start_row + len(remark_rows) - 1,
        min_column=approved_start_column,
        max_column=approved_start_column + 2,
    )
    for cell in summary_sheet["E"][1:len(summary_rows) + 1]:
        cell.number_format = "0.00"
    approved_amount_column = approved_start_column + 3
    for row_number in range(
        approved_data_start_row,
        approved_data_start_row + len(approved_rows),
    ):
        summary_sheet.cell(
            row_number,
            approved_amount_column,
        ).number_format = "0.00"
    remark_amount_column = approved_start_column + 2
    for row_number in range(
        remark_data_start_row,
        remark_data_start_row + len(remark_rows),
    ):
        summary_sheet.cell(
            row_number,
            remark_amount_column,
        ).number_format = "0.00"

    for sheet_name, _, _, grouped_rows in group_sheets:
        group_sheet = workbook.create_sheet(sheet_name)
        group_sheet.append(COUPON_GROUP_HEADER)
        for row, _ in grouped_rows:
            group_sheet.append(select_coupon_group_columns(row))
        format_sheet(
            group_sheet,
            font_name,
            measurement_font,
            ("商品名称", "详细情况"),
        )
        for cell in group_sheet["A"][1:]:
            cell.number_format = "@"
        for cell in group_sheet["B"][1:]:
            if cell.value not in (None, ""):
                cell.number_format = "yyyy-mm-dd"
        for row_number, (_, is_pink) in enumerate(grouped_rows, start=2):
            if is_pink:
                for column in range(1, len(COUPON_GROUP_HEADER) + 1):
                    group_sheet.cell(row_number, column).fill = matched_fill

    report_sheet = workbook.create_sheet(REFERENCE_REPORT_SHEET)
    report_sheet.append(REFERENCE_REPORT_HEADER)
    for decision in reference_decisions:
        report_sheet.append(decision)
    format_sheet(report_sheet, font_name, measurement_font, ("说明",))

    data_row_count = max(len(rows) - 1, 0)
    save_workbook_atomically(
        workbook,
        COUPON_OUTPUT_FILE,
        lambda path: validate_coupon_output(
            path,
            data_row_count,
            matched_count,
            remark_lookup,
            reference_supplement_matches,
            matched_subsidy_total,
            detail_lookup,
            uploaded_count,
            reference_universe,
            final_unresolved_reference_count,
            unmatched_count,
            reference_decisions,
            summary_rows,
            approved_rows,
            remark_rows,
            group_sheets,
        ),
    )
    print(f"Subsidy coupon statistics complete: {data_row_count} rows")
    print(f"Receipt remark matches: {receipt_remark_count}")
    print(f"Pink return or exchange rows: {matched_count}")
    print(f"Supplemental reference matches: {reference_supplement_count}")
    print(
        "Ambiguous supplemental reference candidates: "
        f"{ambiguous_reference_supplement_count}"
    )
    print(f"Submitted status matches: {uploaded_count}")
    print(f"Rows not found in submitted data (marked 未上传): {unmatched_count}")
    print(
        f"Automatic reference corrections: {corrected_count}; "
        f"no unique candidate: {unresolved_count}; "
        f"duplicate conflicts: {correction_collision_count}"
    )
    print(
        "Total 2026 appliance subsidy counted as revenue for matched rows: "
        f"{matched_subsidy_total:.2f}"
    )
    print(f"Output file: {COUPON_OUTPUT_FILE}")


def data_processors() -> tuple[tuple[str, Path, Callable[[], None]], ...]:
    """Report the paths configure_data_dir actually resolved, not assumed names.

    Source directories may carry their original Chinese names, so the label and
    the real path are not interchangeable.
    """
    return (
        ("submitted", INPUT_DIR, process_submitted_files),
        ("receipt_statistics", RECEIPTS_SOURCE_FILE, process_receipts),
        ("subsidy_coupon_statistics", COUPON_SOURCE_FILE, process_coupon_sales),
    )


def process_all() -> None:
    for _, source_path, processor in data_processors():
        print(f"Processing: {source_path}")
        processor()


def choose_data_processor() -> Callable[[], None] | None:
    processors = data_processors()

    print("Select a processing mode:")
    for index, (label, _, _) in enumerate(processors, start=1):
        print(f"  {index}. {label}")
    print("  4. all")
    print("  0. Exit")

    while True:
        try:
            choice = input("Enter a number: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nProcessing cancelled")
            return None

        if choice == "0":
            print("Exited")
            return None
        if choice == "4" or choice.lower() == "all":
            print("Processing all data in order: 1, 2, 3")
            return process_all
        if choice.isdigit():
            selected_index = int(choice) - 1
            if 0 <= selected_index < len(processors):
                _, source_path, processor = processors[selected_index]
                print(f"Processing: {source_path}")
                return processor

        print("Invalid input. Enter a menu number or all.")
