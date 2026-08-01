"""家电 side of 销售用券情况统计 processing: the superset report.

家电 is a strict superset of 数码's report: it additionally reads an optional
reference-supplement file, builds 财务大类/品牌 group sheets, and closes
数据汇总 with a five-column (财务大类, 品牌, 备注, 数量, 合计) summary instead
of 数码's three-column one. Those are real differences in what gets built,
not just different constants, so 数码 keeps its own module
(processors/coupons/digital.py) rather than being forced through this one
with a bundle of feature flags.
"""

import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, PatternFill, Side
from python_calamine import CalamineWorkbook

from processors.common.dates import (
    normalize_coupon_date,
    normalize_document_number,
    normalize_receipt_identifier,
)
from processors.common.excel import (
    calamine_rows,
    format_sheet,
    load_measurement_font,
    resolve_font,
)
from processors.receipts import OUTPUT_FILE as RECEIPTS_OUTPUT_FILE
from processors.receipts import RECEIPTS_REMARK_SAME_MODEL_REPLACEMENT
from processors.submitted import PROFILES as SUBMITTED_PROFILES

from . import matching, sources
from .matching import (
    COUPON_REFERENCE_RE,
    as_currency,
    coupon_data_rows,
)
from .report_contract import SUMMARY_HEADER, SUMMARY_SHEET_NAME
from .sources import load_coupon_remark_lookup, load_uploaded_summary
from .validation import (
    is_pink_row,
    validate_detail_sheet_shape,
    validate_document_and_date_cells,
    validate_left_aligned_column,
    validate_matched_subsidy_total,
    validate_pink_position,
    validate_remark_and_detail,
    validate_uploaded_and_unmatched_counts,
)

DETAILS_SHEET_NAME = "家电-明细总表"
# Owned here only to keep this module's own group-sheet titles from
# colliding with processors.coupons.digital's / coupon_report.py's sheets.
DIGITAL_DETAILS_SHEET_NAME = "数码-明细总表"
REFERENCE_REPORT_SHEET_NAME = "Processing Report"

COUPON_SUBSIDY_HEADER = sources.COUPON_FAMILY_SUBSIDY_HEADER
COUPON_REMARK_SOURCE_FILE = RECEIPTS_OUTPUT_FILE
COUPON_UPLOADED_SOURCE_FILE = SUBMITTED_PROFILES["家电"].output_file

# Rows misclassified as digital are excluded from the Summary aggregation
# and never get their own category-brand sheet, though they still appear
# in the Details sheet.
COUPON_EXCLUDED_CATEGORY = "数码"
COUPON_OUTPUT_HEADER = sources.APPLIANCE_PROFILE.output_header
COUPON_GROUP_HEADER = (
    "单据号",
    "单据日期",
    "商品名称",
    COUPON_SUBSIDY_HEADER,
    "备注",
    "详细情况",
)
COUPON_GROUP_COLUMN_INDEXES = tuple(
    COUPON_OUTPUT_HEADER.index(header) for header in COUPON_GROUP_HEADER
)
COUPON_DATE_INDEX = COUPON_OUTPUT_HEADER.index("单据日期")
COUPON_PRODUCT_NAME_INDEX = COUPON_OUTPUT_HEADER.index("商品名称")
COUPON_BRAND_INDEX = COUPON_OUTPUT_HEADER.index("品牌")
COUPON_CATEGORY_INDEX = COUPON_OUTPUT_HEADER.index("财务大类")
COUPON_SUMMARY_COLUMN_INDEX = COUPON_OUTPUT_HEADER.index("明细摘要")
COUPON_SUBSIDY_INDEX = COUPON_OUTPUT_HEADER.index(COUPON_SUBSIDY_HEADER)
COUPON_REMARK_INDEX = COUPON_OUTPUT_HEADER.index("备注")
COUPON_DETAIL_INDEX = COUPON_OUTPUT_HEADER.index("详细情况")
COUPON_MATCH_FILL_COLOR = "FFC7CE"
COUPON_BRAND_REPLACEMENTS = sources.COUPON_BRAND_REPLACEMENTS
# report_contract.py hardcodes this header's text rather than deriving it
# from COUPON_SUBSIDY_HEADER, so this assertion is what actually keeps the
# two in sync — see report_contract.py's module docstring.
assert SUMMARY_HEADER[-1] == f"{COUPON_SUBSIDY_HEADER}合计"
COUPON_SUMMARY_HEADER = SUMMARY_HEADER
# 财务大类 for this project's 已上传/未上传/合计 block at the foot of 数据汇总.
COUPON_SUMMARY_PROJECT_LABEL = "家电"
COUPON_REMARK_SORT_PRIORITY = {
    "未上传": 0,
    "已上传": 1,
}
COUPON_SUMMARY_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
INVALID_SHEET_TITLE_RE = re.compile(r"[\[\]:*?/\\]")
COUPON_REFERENCE_SUPPLEMENT_HEADER = ("参考号", "单据号", "单据日期")


def normalize_coupon_reference_supplement_header(value: object) -> str:
    text = str(value or "").strip()
    if text in COUPON_REFERENCE_SUPPLEMENT_HEADER:
        return text
    try:
        return text.encode("cp949").decode("gb18030")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return text


def load_coupon_reference_supplement(
    source: Path,
) -> dict[tuple[str, date], frozenset[str]]:
    if not source.exists():
        print(f"Optional reference supplement file not found; skipping: {source}")
        return {}

    workbook = CalamineWorkbook.from_path(str(source))
    try:
        sheet = workbook.get_sheet_by_index(0)
        rows_iter = calamine_rows(sheet)
        header = tuple(
            normalize_coupon_reference_supplement_header(value)
            for value in next(rows_iter, [])
        )
        if header != COUPON_REFERENCE_SUPPLEMENT_HEADER:
            raise ValueError(
                f"{source.name} 字段标题不符合要求：实际为 {header}"
            )

        references_by_key: dict[tuple[str, date], set[str]] = {}
        for row_number, row in enumerate(rows_iter, start=2):
            if all(value in (None, "") for value in row):
                continue
            reference = normalize_receipt_identifier(row[0]).upper()
            document_number = normalize_document_number(row[1])
            if not COUPON_REFERENCE_RE.fullmatch(reference):
                raise ValueError(
                    f"{source.name} 第 {row_number} 行参考号格式无效："
                    f"{row[0]!r}"
                )
            if not document_number:
                raise ValueError(
                    f"{source.name} 第 {row_number} 行单据号为空"
                )
            document_date = normalize_coupon_date(row[2], row_number)
            key = (document_number, document_date)
            references_by_key.setdefault(key, set()).add(reference)
        return {
            key: frozenset(references)
            for key, references in references_by_key.items()
        }
    finally:
        workbook.close()


def fill_coupon_reference_supplement(
    rows: list[list[object]],
    reference_lookup: dict[tuple[str, date], frozenset[str]],
    reference_universe: set[str],
    excluded_bottom_rows: int,
) -> tuple[int, int, set[int], Counter[tuple[str, date, str]]]:
    summary_index = COUPON_SUMMARY_COLUMN_INDEX
    included_rows = coupon_data_rows(rows, excluded_bottom_rows)
    matched_count = 0
    ambiguous_count = 0
    matched_row_ids: set[int] = set()
    matched_values: Counter[tuple[str, date, str]] = Counter()
    for row in included_rows:
        current_reference = normalize_receipt_identifier(
            row[summary_index]
        ).upper()
        if current_reference in reference_universe:
            continue
        key = (normalize_document_number(row[0]), row[1])
        references = reference_lookup.get(key)
        if references is None:
            continue
        if len(references) == 1:
            reference = next(iter(references))
        elif current_reference in references:
            reference = current_reference
        else:
            ambiguous_count += 1
            continue
        row[summary_index] = reference
        matched_count += 1
        matched_row_ids.add(id(row))
        matched_values[(key[0], key[1], reference)] += 1
    return (
        matched_count,
        ambiguous_count,
        matched_row_ids,
        matched_values,
    )


def coupon_text_sort_value(value: object) -> str:
    return str(value or "").strip()


def coupon_date_sort_value(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.max


def coupon_pink_sort_key(row: list[object] | tuple[object, ...]) -> tuple:
    return (
        coupon_text_sort_value(row[COUPON_CATEGORY_INDEX]),
        coupon_text_sort_value(row[COUPON_BRAND_INDEX]),
        coupon_text_sort_value(row[COUPON_PRODUCT_NAME_INDEX]),
        coupon_date_sort_value(row[COUPON_DATE_INDEX]),
    )


def coupon_regular_sort_key(row: list[object] | tuple[object, ...]) -> tuple:
    remark = coupon_text_sort_value(row[COUPON_REMARK_INDEX])
    return (
        COUPON_REMARK_SORT_PRIORITY.get(remark, 2),
        remark,
        coupon_text_sort_value(row[COUPON_CATEGORY_INDEX]),
        coupon_text_sort_value(row[COUPON_BRAND_INDEX]),
        coupon_text_sort_value(row[COUPON_DETAIL_INDEX]),
        coupon_date_sort_value(row[COUPON_DATE_INDEX]),
        coupon_text_sort_value(row[COUPON_PRODUCT_NAME_INDEX]),
    )


def coupon_group_regular_sort_key(
    row: list[object] | tuple[object, ...],
) -> tuple:
    remark = coupon_text_sort_value(row[COUPON_REMARK_INDEX])
    priority = {"已上传": 0, "未上传": 1}.get(remark, 2)
    return (
        priority,
        remark,
        coupon_text_sort_value(row[COUPON_DETAIL_INDEX]),
        coupon_date_sort_value(row[COUPON_DATE_INDEX]),
        coupon_text_sort_value(row[COUPON_PRODUCT_NAME_INDEX]),
    )


def select_coupon_group_columns(
    row: list[object] | tuple[object, ...],
) -> tuple[object, ...]:
    return tuple(row[index] for index in COUPON_GROUP_COLUMN_INDEXES)


def sort_coupon_detail_rows(
    rows: list[list[object]],
    pink_row_count: int,
) -> None:
    pink_start = len(rows) - pink_row_count
    regular_rows = rows[1:pink_start] if pink_row_count else rows[1:]
    pink_rows = rows[pink_start:] if pink_row_count else []
    regular_rows.sort(key=coupon_regular_sort_key)
    pink_rows.sort(key=coupon_pink_sort_key)
    rows[1:] = [*regular_rows, *pink_rows]


def coupon_subsidy_count(amount: Decimal) -> int:
    """Count a coupon row as +1, or as -1 when its 国补 is a reversal.

    A return or exchange is recorded as a negative amount that cancels the
    original row, so counting it as -1 keeps the count consistent with the
    total. A zero amount should not occur; it is bad source data and is
    counted as 0 rather than silently inflating the count.
    """
    if amount > 0:
        return 1
    if amount < 0:
        return -1
    return 0


def build_coupon_summary(
    rows: list[list[object]],
    excluded_bottom_rows: int,
    uploaded_subsidy_count: int,
    uploaded_subsidy_total: Decimal,
) -> tuple[list[tuple[object, ...]], int]:
    """Build the 数据汇总 table.

    The table lists 财务大类 / 品牌 / 备注 groups, then closes with 已上传 /
    未上传 / 合计: 已上传 comes from the generated 已上传 workbook's 补贴金额,
    合计 is this coupon file's own 国补 总, and 未上传 is the difference —
    the same three-way split digital already reports.

    Also returns how many rows carried a zero 国补, which is invalid source
    data worth telling the operator about.
    """
    category_index = COUPON_CATEGORY_INDEX
    brand_index = COUPON_BRAND_INDEX
    remark_index = COUPON_REMARK_INDEX
    subsidy_index = COUPON_SUBSIDY_INDEX
    included_rows = coupon_data_rows(rows, excluded_bottom_rows)
    grouped_counts: Counter[tuple[str, str, str]] = Counter()
    grouped_totals: dict[tuple[str, str, str], Decimal] = {}
    coupon_count = 0
    coupon_total = Decimal("0")
    zero_subsidy_count = 0
    for row in included_rows:
        category = str(row[category_index] or "").strip()
        if category == COUPON_EXCLUDED_CATEGORY:
            continue
        brand = str(row[brand_index] or "").strip()
        remark = str(row[remark_index] or "").strip()
        key = (category, brand, remark)
        grouped_counts[key] += 1
        grouped_totals.setdefault(key, Decimal("0"))
        subsidy = row[subsidy_index]
        if subsidy not in (None, ""):
            try:
                amount = Decimal(str(subsidy))
            except InvalidOperation as error:
                raise ValueError(
                    f"{key!r} 的2026家电国补金额无效：{subsidy!r}"
                ) from error
            grouped_totals[key] += amount
            coupon_total += amount
            coupon_count += coupon_subsidy_count(amount)
            if amount == 0:
                zero_subsidy_count += 1

    summary_rows = [
        (
            *key,
            grouped_counts[key],
            float(grouped_totals[key].quantize(Decimal("0.01"))),
        )
        for key in sorted(grouped_counts)
    ]
    # Labelled 财务大类=家电 with the status in 备注, so this block reads the
    # same way as the 数码 block appended after it in the merged 审核明细
    # workbook (see processors/coupon_report.py).
    summary_rows.extend(
        (
            (
                COUPON_SUMMARY_PROJECT_LABEL,
                None,
                "已上传",
                uploaded_subsidy_count,
                float(as_currency(uploaded_subsidy_total)),
            ),
            (
                COUPON_SUMMARY_PROJECT_LABEL,
                None,
                "未上传",
                coupon_count - uploaded_subsidy_count,
                float(as_currency(coupon_total - uploaded_subsidy_total)),
            ),
            (
                COUPON_SUMMARY_PROJECT_LABEL,
                None,
                "合计",
                coupon_count,
                float(as_currency(coupon_total)),
            ),
        )
    )
    return summary_rows, zero_subsidy_count


def coupon_group_sheet_title(
    category: str,
    brand: str,
    used_titles: set[str],
) -> str:
    base_title = f"{category or '未分类'}-{brand or '未品牌'}"
    base_title = INVALID_SHEET_TITLE_RE.sub("_", base_title).strip("'")
    base_title = base_title[:31] or "未分类-未品牌"
    title = base_title
    suffix = 2
    while title in used_titles:
        suffix_text = f"-{suffix}"
        title = f"{base_title[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_titles.add(title)
    return title


def build_coupon_group_sheets(
    rows: list[list[object]],
    excluded_bottom_rows: int,
) -> list[tuple[str, str, str, list[tuple[list[object], bool]]]]:
    category_index = COUPON_CATEGORY_INDEX
    brand_index = COUPON_BRAND_INDEX
    first_pink_index = len(rows) - excluded_bottom_rows
    groups: dict[
        tuple[str, str],
        list[tuple[list[object], bool]],
    ] = {}
    for row_index, row in enumerate(rows[1:], start=1):
        category = str(row[category_index] or "").strip()
        if category == COUPON_EXCLUDED_CATEGORY:
            continue
        brand = str(row[brand_index] or "").strip()
        groups.setdefault((category, brand), []).append(
            (row, row_index >= first_pink_index)
        )

    for grouped_rows in groups.values():
        regular_rows = [item for item in grouped_rows if not item[1]]
        pink_rows = [item for item in grouped_rows if item[1]]
        regular_rows.sort(key=lambda item: coupon_group_regular_sort_key(item[0]))
        pink_rows.sort(key=lambda item: coupon_pink_sort_key(item[0]))
        grouped_rows[:] = [*regular_rows, *pink_rows]

    used_titles = {
        SUMMARY_SHEET_NAME,
        DETAILS_SHEET_NAME,
        DIGITAL_DETAILS_SHEET_NAME,
        REFERENCE_REPORT_SHEET_NAME,
    }
    return [
        (
            coupon_group_sheet_title(category, brand, used_titles),
            category,
            brand,
            grouped_rows,
        )
        for (category, brand), grouped_rows in sorted(groups.items())
    ]


def merge_coupon_summary_groups(
    sheet,
    start_row: int,
    end_row: int,
    start_column: int = 1,
) -> None:
    centered = Alignment(horizontal="center", vertical="center")
    category_column = start_column
    brand_column = start_column + 1
    for column in (brand_column, category_column):
        group_start = start_row
        while group_start <= end_row:
            category = sheet.cell(group_start, category_column).value
            value = sheet.cell(group_start, column).value
            group_end = group_start
            while (
                group_end + 1 <= end_row
                and sheet.cell(
                    group_end + 1,
                    category_column,
                ).value
                == category
                and sheet.cell(group_end + 1, column).value == value
            ):
                group_end += 1
            if group_end > group_start:
                sheet.merge_cells(
                    start_row=group_start,
                    start_column=column,
                    end_row=group_end,
                    end_column=column,
                )
            sheet.cell(group_start, column).alignment = centered
            group_start = group_end + 1


def project_summary_blocks(
    rows: list[tuple[object, ...]],
) -> list[tuple[int, int]]:
    """Locate the trailing per-project 已上传/未上传/合计 blocks.

    They are the rows carrying a 财务大类 but no 品牌 (家电 first, then any
    project appended by processors/coupon_report.py). Returned as inclusive
    0-based (start, end) spans so the caller can merge 财务大类 and 品牌 into
    one cell across each block — the split is meaningless there, and leaving
    an empty 品牌 column beside the label reads as a missing value.
    """
    blocks: list[tuple[int, int]] = []
    start: int | None = None
    for index, row in enumerate(rows):
        if row[1] is None:
            if start is None:
                start = index
            elif rows[start][0] != row[0]:
                blocks.append((start, index - 1))
                start = index
        elif start is not None:
            blocks.append((start, index - 1))
            start = None
    if start is not None:
        blocks.append((start, len(rows) - 1))
    return blocks


def merge_coupon_summary_projects(
    sheet,
    blocks: list[tuple[int, int]],
    header_rows: int = 1,
) -> None:
    centered = Alignment(horizontal="center", vertical="center")
    for start, end in blocks:
        sheet.merge_cells(
            start_row=start + 1 + header_rows,
            start_column=1,
            end_row=end + 1 + header_rows,
            end_column=2,
        )
        sheet.cell(start + 1 + header_rows, 1).alignment = centered


def merged_coupon_summary_values(
    rows: list[tuple[object, ...]],
) -> list[tuple[object, ...]]:
    merged_rows: list[tuple[object, ...]] = []
    previous_category = None
    previous_brand = None
    for row in rows:
        if row[0] == "合计":
            merged_rows.append(row)
            continue
        category, brand, *remaining = row
        displayed_category = (
            None if category == previous_category else category
        )
        displayed_brand = (
            None
            if category == previous_category and brand == previous_brand
            else brand
        )
        merged_rows.append(
            (displayed_category, displayed_brand, *remaining)
        )
        previous_category = category
        previous_brand = brand
    return merged_rows


def apply_coupon_summary_borders(
    sheet,
    *,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> None:
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_column,
        max_col=max_column,
    ):
        for cell in row:
            cell.border = COUPON_SUMMARY_BORDER


@dataclass
class CouponComputation:
    """Everything computed for the 家电 coupon data, before it is written to
    a sheet. Kept separate from workbook assembly so the merged 审核明细
    workbook (built in processors/coupon_report.py) can interleave a 数码
    sheet between this project's own sheets without either project knowing
    about the other at import time."""

    rows: list[list[object]]
    data_row_count: int
    matched_count: int
    matched_subsidy_total: Decimal
    receipt_remark_count: int
    remark_lookup: dict[tuple[str, date], str]
    detail_lookup: dict[str, str]
    reference_universe: set[str]
    reference_supplement_count: int
    ambiguous_reference_supplement_count: int
    reference_supplement_matches: Counter[tuple[str, date, str]]
    corrected_count: int
    unresolved_count: int
    correction_collision_count: int
    reference_decisions: list[tuple[str, str, str, str]]
    final_unresolved_reference_count: int
    uploaded_count: int
    unmatched_count: int
    excluded_category_row_count: int
    uploaded_subsidy_count: int
    uploaded_subsidy_total: Decimal
    zero_subsidy_count: int
    source_total: Decimal | None
    computed_total: Decimal
    summary_rows: list[tuple[object, ...]]
    group_sheets: list[tuple[str, str, str, list[tuple[list[object], bool]]]]


_SOURCE_TOTAL_UNSET = object()


def compute_coupon_data(
    *,
    rows: list[list[object]] | None = None,
    remark_lookup: dict[tuple[str, date], str] | None = None,
    source_total: Decimal | None | object = _SOURCE_TOTAL_UNSET,
) -> CouponComputation:
    if sources.COUPON_SOURCE_FILE is None:
        raise FileNotFoundError(
            f"未在 {sources.DATA_DIR} 中找到文件名包含"
            f"“{sources.COUPON_STATISTICS_KEYWORD}”且表头为"
            f"“{COUPON_SUBSIDY_HEADER}”的 .XLSX 文件"
        )

    if rows is None:
        rows = sources.read_coupon_rows(
            sources.COUPON_SOURCE_FILE, sources.APPLIANCE_PROFILE
        )
    if remark_lookup is None:
        remark_lookup = load_coupon_remark_lookup(COUPON_REMARK_SOURCE_FILE)
    matched_count, matched_subsidy_total, receipt_remark_count = (
        matching.fill_coupon_remarks(
            rows,
            remark_lookup,
            "2026家电国补",
            excluded_remark=RECEIPTS_REMARK_SAME_MODEL_REPLACEMENT,
        )
    )
    detail_lookup, uploaded_subsidy_count, uploaded_subsidy_total = (
        load_uploaded_summary(COUPON_UPLOADED_SOURCE_FILE)
    )
    # Unsubmitted data is no longer supplied, so submitted data is the only
    # source of valid references.
    reference_universe = set(detail_lookup)
    reference_supplement_lookup = load_coupon_reference_supplement(
        sources.COUPON_REFERENCE_SUPPLEMENT_FILE
    )
    (
        reference_supplement_count,
        ambiguous_reference_supplement_count,
        reference_supplement_row_ids,
        reference_supplement_matches,
    ) = fill_coupon_reference_supplement(
        rows,
        reference_supplement_lookup,
        reference_universe,
        matched_count,
    )
    (
        corrected_count,
        unresolved_count,
        correction_collision_count,
        reference_decisions,
    ) = matching.correct_coupon_references(
        rows,
        reference_universe,
        matched_count,
        reference_supplement_row_ids,
    )
    summary_index = COUPON_SUMMARY_COLUMN_INDEX
    regular_rows = coupon_data_rows(rows, matched_count)
    final_unresolved_reference_count = sum(
        reference not in reference_universe
        for reference in (
            normalize_receipt_identifier(row[summary_index]).upper()
            for row in regular_rows
        )
        if reference
    )
    uploaded_count = matching.fill_uploaded_details(
        rows,
        detail_lookup,
        matched_count,
    )
    unmatched_count = matching.fill_unmatched_remarks(
        rows,
        reference_universe,
        matched_count,
    )
    sort_coupon_detail_rows(rows, matched_count)
    excluded_category_row_count = sum(
        1
        for row in coupon_data_rows(rows, matched_count)
        if str(row[COUPON_CATEGORY_INDEX] or "").strip()
        == COUPON_EXCLUDED_CATEGORY
    )
    summary_rows, zero_subsidy_count = build_coupon_summary(
        rows,
        matched_count,
        uploaded_subsidy_count,
        uploaded_subsidy_total,
    )
    group_sheets = build_coupon_group_sheets(rows, matched_count)
    if source_total is _SOURCE_TOTAL_UNSET:
        source_total = sources.read_coupon_source_total(sources.COUPON_SOURCE_FILE)
    computed_total = Decimal(str(summary_rows[-1][4]))

    return CouponComputation(
        rows=rows,
        data_row_count=max(len(rows) - 1, 0),
        matched_count=matched_count,
        matched_subsidy_total=matched_subsidy_total,
        receipt_remark_count=receipt_remark_count,
        remark_lookup=remark_lookup,
        detail_lookup=detail_lookup,
        reference_universe=reference_universe,
        reference_supplement_count=reference_supplement_count,
        ambiguous_reference_supplement_count=(
            ambiguous_reference_supplement_count
        ),
        reference_supplement_matches=reference_supplement_matches,
        corrected_count=corrected_count,
        unresolved_count=unresolved_count,
        correction_collision_count=correction_collision_count,
        reference_decisions=reference_decisions,
        final_unresolved_reference_count=final_unresolved_reference_count,
        uploaded_count=uploaded_count,
        unmatched_count=unmatched_count,
        excluded_category_row_count=excluded_category_row_count,
        uploaded_subsidy_count=uploaded_subsidy_count,
        uploaded_subsidy_total=uploaded_subsidy_total,
        zero_subsidy_count=zero_subsidy_count,
        source_total=source_total,
        computed_total=computed_total,
        summary_rows=summary_rows,
        group_sheets=group_sheets,
    )


def build_summary_and_details_sheets(
    workbook: Workbook,
    computation: CouponComputation,
    extra_summary_rows: list[tuple[object, ...]] = (),
) -> tuple[str, object, PatternFill]:
    """Create 数据汇总 (index 0) and 家电-明细总表. Returns (font_name,
    measurement_font, matched_fill) so the caller can reuse them for sheets
    appended later (digital's detail sheet, this project's group sheets and
    Processing Report) without reloading the font."""
    rows = computation.rows
    matched_count = computation.matched_count
    combined_summary_rows = [*computation.summary_rows, *extra_summary_rows]

    sheet = workbook.create_sheet(DETAILS_SHEET_NAME)
    for row in rows:
        sheet.append(row)

    font_name, font_path = resolve_font()
    measurement_font = load_measurement_font(font_path)
    format_sheet(
        sheet,
        font_name,
        measurement_font,
        ("商品名称", "详细情况"),
    )
    for cell in sheet["A"][1:]:
        cell.number_format = "@"
    for cell in sheet["B"][1:]:
        if cell.value not in (None, ""):
            cell.number_format = "yyyy-mm-dd"
    matched_fill = PatternFill("solid", fgColor=COUPON_MATCH_FILL_COLOR)
    matched_start_row = sheet.max_row - matched_count + 1
    for row in sheet.iter_rows(
        min_row=matched_start_row,
        max_row=sheet.max_row,
    ):
        for cell in row:
            cell.fill = matched_fill

    summary_sheet = workbook.create_sheet(SUMMARY_SHEET_NAME, 0)
    summary_sheet.append(COUPON_SUMMARY_HEADER)
    for summary_row in combined_summary_rows:
        summary_sheet.append(summary_row)
    format_sheet(summary_sheet, font_name, measurement_font)
    # The per-project blocks get one 财务大类+品牌 cell each, so the row-wise
    # 财务大类/品牌 merging has to stop before them or the two would overlap.
    project_blocks = project_summary_blocks(combined_summary_rows)
    brand_rows_end = (
        project_blocks[0][0] if project_blocks else len(combined_summary_rows)
    )
    if brand_rows_end:
        merge_coupon_summary_groups(summary_sheet, 2, brand_rows_end + 1)
    merge_coupon_summary_projects(summary_sheet, project_blocks)
    apply_coupon_summary_borders(
        summary_sheet,
        min_row=1,
        max_row=len(combined_summary_rows) + 1,
        min_column=1,
        max_column=5,
    )
    for cell in summary_sheet["E"][1:len(combined_summary_rows) + 1]:
        cell.number_format = "0.00"

    return font_name, measurement_font, matched_fill


def build_group_sheets(
    workbook: Workbook,
    computation: CouponComputation,
    font_name: str,
    measurement_font: object,
    matched_fill: PatternFill,
) -> None:
    """Append this project's 财务大类-品牌 group sheets after whatever sheets
    already exist in the workbook. The Processing Report is written separately
    by processors/coupon_report.py, which merges both projects' decisions."""
    for sheet_name, _, _, grouped_rows in computation.group_sheets:
        group_sheet = workbook.create_sheet(sheet_name)
        group_sheet.append(COUPON_GROUP_HEADER)
        for row, _ in grouped_rows:
            group_sheet.append(select_coupon_group_columns(row))
        format_sheet(
            group_sheet,
            font_name,
            measurement_font,
            ("商品名称", "详细情况"),
        )
        for cell in group_sheet["A"][1:]:
            cell.number_format = "@"
        for cell in group_sheet["B"][1:]:
            if cell.value not in (None, ""):
                cell.number_format = "yyyy-mm-dd"
        for row_number, (_, is_pink) in enumerate(grouped_rows, start=2):
            if is_pink:
                for column in range(1, len(COUPON_GROUP_HEADER) + 1):
                    group_sheet.cell(row_number, column).fill = matched_fill


def validate_summary_and_details_sheets(
    workbook: Workbook,
    computation: CouponComputation,
    extra_summary_rows: list[tuple[object, ...]] = (),
) -> None:
    expected_data_rows = computation.data_row_count
    expected_matched_rows = computation.matched_count
    remark_lookup = computation.remark_lookup
    expected_reference_supplement_matches = (
        computation.reference_supplement_matches
    )
    expected_matched_subsidy_total = computation.matched_subsidy_total
    detail_lookup = computation.detail_lookup
    expected_uploaded_rows = computation.uploaded_count
    reference_universe = computation.reference_universe
    expected_unresolved_rows = computation.final_unresolved_reference_count
    expected_unmatched_rows = computation.unmatched_count
    expected_excluded_category_rows = computation.excluded_category_row_count
    expected_summary_rows = computation.summary_rows
    combined_summary_rows = [*expected_summary_rows, *extra_summary_rows]

    sheet = workbook[DETAILS_SHEET_NAME]
    validate_detail_sheet_shape(sheet, COUPON_OUTPUT_HEADER, expected_data_rows)
    detail_column, remark_column = validate_uploaded_and_unmatched_counts(
        sheet, COUPON_OUTPUT_HEADER, expected_uploaded_rows, expected_unmatched_rows
    )
    matched_start_row = sheet.max_row - expected_matched_rows + 1
    summary_column = COUPON_OUTPUT_HEADER.index("明细摘要") + 1
    actual_unresolved_rows = sum(
        reference not in reference_universe
        for reference in (
            normalize_receipt_identifier(
                sheet.cell(row_number, summary_column).value
            ).upper()
            for row_number in range(2, matched_start_row)
        )
        if reference
    )
    if actual_unresolved_rows != expected_unresolved_rows:
        raise RuntimeError(
            f"销售用券未解决参考号数量校验失败：预期 "
            f"{expected_unresolved_rows} 条，实际 "
            f"{actual_unresolved_rows} 条"
        )
    actual_regular_rows = [
        row
        for row in sheet.iter_rows(
            min_row=2,
            max_row=matched_start_row - 1,
            min_col=1,
            max_col=len(COUPON_OUTPUT_HEADER),
            values_only=True,
        )
    ]
    if actual_regular_rows != sorted(
        actual_regular_rows,
        key=coupon_regular_sort_key,
    ):
        raise RuntimeError("销售用券明细非粉色区域排序校验失败")
    actual_pink_rows = [
        row
        for row in sheet.iter_rows(
            min_row=matched_start_row,
            max_row=sheet.max_row,
            min_col=1,
            max_col=len(COUPON_OUTPUT_HEADER),
            values_only=True,
        )
    ]
    if actual_pink_rows != sorted(
        actual_pink_rows,
        key=coupon_pink_sort_key,
    ):
        raise RuntimeError("销售用券明细粉色区域排序校验失败")
    product_name_column = COUPON_OUTPUT_HEADER.index("商品名称") + 1
    for column_name, column_number in (
        ("商品名称", product_name_column),
        ("详细情况", detail_column),
    ):
        validate_left_aligned_column(sheet, column_number, column_name)
    brand_column = COUPON_OUTPUT_HEADER.index("品牌") + 1
    remaining_source_brands = {
        str(sheet.cell(row_number, brand_column).value or "").strip()
        for row_number in range(2, sheet.max_row + 1)
    } & COUPON_BRAND_REPLACEMENTS.keys()
    if remaining_source_brands:
        raise RuntimeError(
            "销售用券品牌替换校验失败："
            f"{sorted(remaining_source_brands)}"
        )
    actual_matched_subsidy_total = Decimal("0")
    actual_reference_supplement_matches: Counter[
        tuple[str, date, str]
    ] = Counter()
    subsidy_column = COUPON_OUTPUT_HEADER.index(COUPON_SUBSIDY_HEADER) + 1
    for row_number in range(2, sheet.max_row + 1):
        document_cell = sheet.cell(row_number, 1)
        date_cell = sheet.cell(row_number, 2)
        remark_cell = sheet.cell(row_number, remark_column)
        detail_cell = sheet.cell(row_number, detail_column)
        document_date = validate_document_and_date_cells(
            document_cell, date_cell, row_number
        )
        receipt_remark = remark_lookup.get(
            (
                normalize_document_number(document_cell.value),
                document_date,
            ),
            "",
        )
        reference = normalize_receipt_identifier(
            sheet.cell(row_number, summary_column).value
        ).upper()
        expected_pink = (
            expected_matched_rows > 0
            and row_number >= matched_start_row
        )
        supplement_match = (
            normalize_document_number(document_cell.value),
            document_date,
            reference,
        )
        if (
            not expected_pink
            and supplement_match
            in expected_reference_supplement_matches
        ):
            actual_reference_supplement_matches[supplement_match] += 1
        if expected_pink:
            expected_detail = ""
            expected_remark = receipt_remark
        else:
            expected_detail = detail_lookup.get(reference, "")
            if expected_detail:
                expected_remark = "已上传"
            elif reference not in reference_universe:
                expected_remark = "未上传"
            else:
                expected_remark = receipt_remark
        validate_remark_and_detail(
            remark_cell, detail_cell, expected_remark, expected_detail, row_number
        )
        is_pink = is_pink_row(
            sheet, row_number, len(COUPON_OUTPUT_HEADER), COUPON_MATCH_FILL_COLOR
        )
        validate_pink_position(is_pink, expected_pink, row_number)
        if expected_pink:
            subsidy = sheet.cell(row_number, subsidy_column).value
            if subsidy not in (None, ""):
                actual_matched_subsidy_total += Decimal(str(subsidy))

    if any(
        actual_reference_supplement_matches[match] < expected_count
        for match, expected_count
        in expected_reference_supplement_matches.items()
    ):
        raise RuntimeError(
            "销售用券补充参考号逐行匹配结果校验失败"
        )
    validate_matched_subsidy_total(
        actual_matched_subsidy_total, expected_matched_subsidy_total
    )
    summary_sheet = workbook[SUMMARY_SHEET_NAME]
    summary_header = tuple(
        summary_sheet.cell(1, column).value
        for column in range(1, len(COUPON_SUMMARY_HEADER) + 1)
    )
    if summary_header != COUPON_SUMMARY_HEADER:
        raise RuntimeError(
            f"销售用券汇总字段标题校验失败：实际为 {summary_header}"
        )
    first_total_row = 1 + len(combined_summary_rows)
    actual_summary_rows = [
        tuple(
            summary_sheet.cell(row_number, column_number).value
            for column_number in range(1, 6)
        )
        for row_number in range(2, first_total_row + 1)
    ]
    if actual_summary_rows != merged_coupon_summary_values(
        combined_summary_rows
    ):
        raise RuntimeError("销售用券财务大类、品牌和备注汇总校验失败")
    actual_merges = {
        str(cell_range) for cell_range in summary_sheet.merged_cells.ranges
    }
    for start, end in project_summary_blocks(combined_summary_rows):
        expected_merge = f"A{start + 2}:B{end + 2}"
        if expected_merge not in actual_merges:
            raise RuntimeError(
                "销售用券汇总项目合计块未跨财务大类与品牌两列合并："
                f"缺少 {expected_merge}"
            )
    # The 家电 portion ends in 已上传 / 未上传 / 合计 (in 备注, with 财务大类
    # merged into a single 家电 cell); anything appended after it (digital's
    # rows) is covered by the row-for-row equality check above.
    remark_column = COUPON_SUMMARY_HEADER.index("备注")
    tail_start = len(expected_summary_rows) - 3
    if tail_start < 0 or [
        row[remark_column]
        for row in actual_summary_rows[tail_start:tail_start + 3]
    ] != ["已上传", "未上传", "合计"]:
        raise RuntimeError("销售用券汇总缺少已上传/未上传/合计三行")
    brand_summary_rows = actual_summary_rows[:tail_start]
    uploaded_row, unuploaded_row, total_row = actual_summary_rows[
        tail_start:tail_start + 3
    ]
    if sum(row[3] for row in brand_summary_rows) != (
        expected_data_rows
        - expected_matched_rows
        - expected_excluded_category_rows
    ):
        raise RuntimeError("销售用券汇总包含粉红色数据或数量不完整")
    # 已上传 is measured from the generated 已上传 workbook, so it must equal
    # what that file reports rather than anything derived from the coupon rows.
    if (
        uploaded_row[3] != computation.uploaded_subsidy_count
        or Decimal(str(uploaded_row[4]))
        != as_currency(computation.uploaded_subsidy_total)
    ):
        raise RuntimeError(
            "销售用券汇总“已上传”未与家电_已上传的补贴金额一致："
            f"{uploaded_row[3]} / {uploaded_row[4]}"
        )
    if uploaded_row[3] + unuploaded_row[3] != total_row[3]:
        raise RuntimeError("销售用券汇总合计数量校验失败")
    if Decimal(str(uploaded_row[4])) + Decimal(str(unuploaded_row[4])) != (
        Decimal(str(total_row[4]))
    ):
        raise RuntimeError("销售用券汇总合计金额校验失败")
    # 合计 must be this coupon file's own 国补 total, counted with reversals
    # as -1 so a return and its original cancel out.
    expected_total = Decimal("0")
    expected_count = 0
    for row in coupon_data_rows(computation.rows, expected_matched_rows):
        if str(row[COUPON_CATEGORY_INDEX] or "").strip() == (
            COUPON_EXCLUDED_CATEGORY
        ):
            continue
        subsidy = row[COUPON_SUBSIDY_INDEX]
        if subsidy in (None, ""):
            continue
        amount = Decimal(str(subsidy))
        expected_total += amount
        expected_count += coupon_subsidy_count(amount)
    if total_row[3] != expected_count or Decimal(str(total_row[4])) != (
        as_currency(expected_total)
    ):
        raise RuntimeError(
            "销售用券汇总“合计”未与明细总表的国补合计一致："
            f"预期 {expected_count} / {as_currency(expected_total)}，"
            f"实际 {total_row[3]} / {total_row[4]}"
        )
    for row in summary_sheet.iter_rows(
        min_row=1,
        max_row=len(combined_summary_rows) + 1,
        min_col=1,
        max_col=5,
    ):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if any(
                side.style != "thin"
                for side in (
                    cell.border.left,
                    cell.border.right,
                    cell.border.top,
                    cell.border.bottom,
                )
            ):
                raise RuntimeError("销售用券汇总表格边框校验失败")


def validate_group_sheets(
    workbook: Workbook,
    computation: CouponComputation,
) -> None:
    for (
        sheet_name,
        _,
        _,
        expected_rows,
    ) in computation.group_sheets:
        group_sheet = workbook[sheet_name]
        group_header = tuple(cell.value for cell in group_sheet[1])
        if group_header != COUPON_GROUP_HEADER:
            raise RuntimeError(f"{sheet_name}分类工作表标题校验失败")
        if group_sheet.max_row - 1 != len(expected_rows):
            raise RuntimeError(
                f"{sheet_name}分类工作表行数校验失败："
                f"预期 {len(expected_rows)} 条，"
                f"实际 {group_sheet.max_row - 1} 条"
            )
        for row_number, (_, expected_pink) in enumerate(
            expected_rows,
            start=2,
        ):
            expected_values = select_coupon_group_columns(
                expected_rows[row_number - 2][0]
            )
            actual_values = tuple(
                group_sheet.cell(row_number, column).value
                for column in range(1, len(COUPON_GROUP_HEADER) + 1)
            )
            normalized_actual = tuple(
                value.date() if isinstance(value, datetime) else value
                for value in actual_values
            )
            comparable_actual = (
                coupon_text_sort_value(normalized_actual[0]),
                normalized_actual[1],
                coupon_text_sort_value(normalized_actual[2]),
                (
                    round(float(normalized_actual[3]), 2)
                    if normalized_actual[3] not in (None, "")
                    else normalized_actual[3]
                ),
                coupon_text_sort_value(normalized_actual[4]),
                coupon_text_sort_value(normalized_actual[5]),
            )
            comparable_expected = (
                coupon_text_sort_value(expected_values[0]),
                expected_values[1],
                coupon_text_sort_value(expected_values[2]),
                (
                    round(float(expected_values[3]), 2)
                    if expected_values[3] not in (None, "")
                    else expected_values[3]
                ),
                coupon_text_sort_value(expected_values[4]),
                coupon_text_sort_value(expected_values[5]),
            )
            if comparable_actual != comparable_expected:
                raise RuntimeError(
                    f"{sheet_name}分类工作表第 {row_number} 行数据或"
                    f"排序校验失败：实际 {comparable_actual!r}，"
                    f"预期 {comparable_expected!r}"
                )
            actual_pink = all(
                group_sheet.cell(
                    row_number,
                    column,
                ).fill.fill_type == "solid"
                and group_sheet.cell(
                    row_number,
                    column,
                ).fill.fgColor.rgb
                in {
                    COUPON_MATCH_FILL_COLOR,
                    f"00{COUPON_MATCH_FILL_COLOR}",
                    f"FF{COUPON_MATCH_FILL_COLOR}",
                }
                for column in range(1, len(COUPON_GROUP_HEADER) + 1)
            )
            if actual_pink != expected_pink:
                raise RuntimeError(
                    f"{sheet_name}分类工作表粉色标记校验失败"
                )
