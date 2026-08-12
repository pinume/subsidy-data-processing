"""Locate, read, and classify rows from the merged 销售用券情况统计 export.

家电 and 数码 rows sit in the same sheet, one column of a 国补 pair (columns
26 and 27) populated per row; a single CouponSourceProfile per project
carries the small amount of real per-project variation (which column is
"own" vs "the other side", the output header text, whether brand names get
normalized through config/brand_mapping.yaml) so this module reads the file
exactly once and hands each project its own classified rows, rather than
each project separately locating and re-parsing the same file as the previous
project-specific implementations did.
"""

import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from python_calamine import CalamineWorkbook

from processors.common.config import load_brand_mapping
from processors.common.dates import (
    normalize_coupon_date,
    normalize_document_number,
)
from processors.common.excel import calamine_rows
from processors.common.paths import find_data_files, resolve_unique_file
from processors.common.references import (
    normalize_reference,
    validated_reference,
)

COUPON_STATISTICS_KEYWORD = "销售用券情况统计"
# The coupon export's field header row (row 2) at these columns.
COUPON_FAMILY_SUBSIDY_COLUMN = 26
COUPON_DIGITAL_SUBSIDY_COLUMN = 27
COUPON_FAMILY_SUBSIDY_HEADER = "2026家电国补（计入收入）"
COUPON_DIGITAL_SUBSIDY_HEADER = "2026数码国补（计入收入）"
COUPON_KEPT_SOURCE_COLUMNS_PREFIX = (3, 4, 6, 8, 15, 18)
COUPON_BRAND_REPLACEMENTS = load_brand_mapping()
APPLIANCE_FINANCIAL_CATEGORIES = frozenset(
    {"冰箱", "厨卫", "国产彩电", "空调", "洗衣机"}
)
DIGITAL_FINANCIAL_CATEGORIES = frozenset({"数码", "新业务类"})

DATA_DIR: Path
COUPON_SOURCE_FILE: Path | None
COUPON_REFERENCE_SUPPLEMENT_FILE: Path
COUPON_REFERENCE_SUPPLEMENT_KEYWORD = "新建 Microsoft Excel 工作表"
PAYMENT_DETAIL_SHEETS = {
    "家电": "家电明细",
    "数码": "数码明细",
}


@dataclass(frozen=True)
class CouponSourceProfile:
    name: str
    subsidy_header: str
    normalize_brand: bool

    @property
    def kept_source_columns(self) -> tuple[int, ...]:
        column = (
            COUPON_FAMILY_SUBSIDY_COLUMN
            if self.name == "家电"
            else COUPON_DIGITAL_SUBSIDY_COLUMN
        )
        return (*COUPON_KEPT_SOURCE_COLUMNS_PREFIX, column)

    @property
    def output_header(self) -> tuple[str, ...]:
        return (
            "单据号",
            "单据日期",
            "商品名称",
            "品牌",
            "财务大类",
            "明细摘要",
            self.subsidy_header,
            "备注",
            "详细情况",
            "回款情况",
        )


APPLIANCE_PROFILE = CouponSourceProfile(
    name="家电",
    subsidy_header=COUPON_FAMILY_SUBSIDY_HEADER,
    normalize_brand=True,
)
DIGITAL_PROFILE = CouponSourceProfile(
    name="数码",
    subsidy_header=COUPON_DIGITAL_SUBSIDY_HEADER,
    normalize_brand=False,
)


def classify_coupon_row(
    *,
    appliance_subsidy: object,
    digital_subsidy: object,
    row_number: int,
    source_name: str,
    financial_category: object = None,
    has_receipt_remark: bool = False,
) -> str:
    """Classify a merged 销售用券情况统计 row as "家电" or "数码".

    财务大类 determines the project for known categories unless the receipt
    lookup marks the row as a return/exchange. Other categories retain the
    export's original column-based classification. Both subsidy columns
    populated is source data corruption serious enough to stop the run.
    """
    appliance_nonzero = appliance_subsidy not in (None, "", 0)
    digital_nonzero = digital_subsidy not in (None, "", 0)
    if appliance_nonzero and digital_nonzero:
        raise ValueError(
            f"{source_name} 第 {row_number} 行同时存在家电国补"
            f"（{appliance_subsidy}）与数码国补（{digital_subsidy}），"
            "无法确定该行所属项目"
        )
    category = str(financial_category or "").strip()
    if not has_receipt_remark:
        if category in APPLIANCE_FINANCIAL_CATEGORIES:
            return "家电"
        if category in DIGITAL_FINANCIAL_CATEGORIES:
            return "数码"
    return "数码" if digital_nonzero else "家电"


@dataclass(frozen=True)
class SubsidyCorrection:
    """One row whose subsidy amount moved from one project's column to the
    other's, because 财务大类 said so and no receipt remark overrode it.

    Pure data: read_coupon_export only collects these. What happens to them
    is the caller's choice — processors/coupon_report.py prints them as
    warnings; a future Processing Report entry would consume the same
    records.
    """

    row_number: int
    document_number: str
    financial_category: str
    amount: Decimal
    from_header: str
    to_header: str


@dataclass(frozen=True)
class SubsidyAttribution:
    """What classify_subsidy_attribution decided for one source row."""

    classification: str
    subsidy_value: object
    source_total_adjustment: Decimal
    correction: SubsidyCorrection | None


def classify_subsidy_attribution(
    *,
    appliance_subsidy: object,
    digital_subsidy: object,
    row_number: int,
    source_name: str,
    document_number: str,
    financial_category: object = None,
    has_receipt_remark: bool = False,
) -> SubsidyAttribution:
    """Decide a row's project and how its subsidy amount must move.

    Pure decision: reads no Excel, prints nothing, mutates nothing — so it
    can be tested row by row without a workbook or a mocked print, and the
    caller (read_coupon_export) applies the result exactly once.

    The original project is whichever subsidy column is non-zero (the
    export's own convention); the correct project comes from 财务大类 via
    classify_coupon_row unless a receipt remark marks the row as a
    return/exchange. A mismatch means the amount was recorded under the
    wrong project's column: subsidy_value is the amount the target
    project's column must carry, source_total_adjustment is what the
    export's own 合计 (which counts 家电's column only) needs to stay
    consistent, and correction records the move for reporting. A zero or
    blank amount moves nothing and records nothing.
    """
    original_classification = (
        "数码" if digital_subsidy not in (None, "", 0) else "家电"
    )
    classification = classify_coupon_row(
        appliance_subsidy=appliance_subsidy,
        digital_subsidy=digital_subsidy,
        row_number=row_number,
        source_name=source_name,
        financial_category=financial_category,
        has_receipt_remark=has_receipt_remark,
    )
    if classification == original_classification:
        return SubsidyAttribution(
            classification=classification,
            subsidy_value=(
                digital_subsidy if classification == "数码" else appliance_subsidy
            ),
            source_total_adjustment=Decimal("0"),
            correction=None,
        )

    if classification == "家电":
        moved_subsidy = digital_subsidy
        from_header = COUPON_DIGITAL_SUBSIDY_HEADER
        to_header = COUPON_FAMILY_SUBSIDY_HEADER
    else:
        moved_subsidy = appliance_subsidy
        from_header = COUPON_FAMILY_SUBSIDY_HEADER
        to_header = COUPON_DIGITAL_SUBSIDY_HEADER
    if moved_subsidy in (None, "", 0):
        return SubsidyAttribution(
            classification=classification,
            subsidy_value=moved_subsidy,
            source_total_adjustment=Decimal("0"),
            correction=None,
        )
    try:
        amount = Decimal(str(moved_subsidy))
    except InvalidOperation as error:
        raise ValueError(
            f"{source_name} 第 {row_number} 行补贴金额无效："
            f"{moved_subsidy!r}"
        ) from error
    return SubsidyAttribution(
        classification=classification,
        subsidy_value=moved_subsidy,
        source_total_adjustment=amount if classification == "家电" else -amount,
        correction=SubsidyCorrection(
            row_number=row_number,
            document_number=document_number,
            financial_category=str(financial_category or "").strip(),
            amount=amount,
            from_header=from_header,
            to_header=to_header,
        ),
    )


def load_coupon_remark_lookup(source: Path) -> dict[tuple[str, date], str]:
    if not source.exists():
        raise FileNotFoundError(f"未找到备注匹配文件：{source}")

    workbook = CalamineWorkbook.from_path(str(source))
    try:
        if "Sheet1" not in workbook.sheet_names:
            raise ValueError(f"{source.name} 缺少 Sheet1 工作表")
        sheet = workbook.get_sheet_by_name("Sheet1")
        rows_iter = calamine_rows(sheet)
        header = next(rows_iter, [])
        required_headers = ("单据号", "日期", "备注")
        missing_headers = [
            required_header
            for required_header in required_headers
            if required_header not in header
        ]
        if missing_headers:
            raise ValueError(
                f"{source.name} 缺少字段：{'、'.join(missing_headers)}"
            )

        document_index = header.index("单据号")
        date_index = header.index("日期")
        remark_index = header.index("备注")
        lookup: dict[tuple[str, date], str] = {}
        for row_number, row in enumerate(rows_iter, start=2):
            document_number = normalize_document_number(row[document_index])
            remark = str(row[remark_index] or "").strip()
            if not document_number or not remark:
                continue
            receipt_date = normalize_coupon_date(
                row[date_index],
                row_number,
                source.name,
            )
            key = (document_number, receipt_date)
            existing_remark = lookup.get(key)
            if existing_remark is not None and existing_remark != remark:
                raise ValueError(
                    f"{source.name} 第 {row_number} 行组合键存在冲突备注："
                    f"{document_number} + {receipt_date:%Y-%m-%d}"
                )
            lookup[key] = remark
        return lookup
    finally:
        workbook.close()


def load_uploaded_summary(source: Path) -> tuple[dict[str, str], int, Decimal]:
    """Read a generated 已上传 workbook's Summary sheet once for everything
    both appliance.py and digital.py need from it: the per-reference detail
    lookup and the 补贴金额 count/total, in a single pass over the sheet."""
    if not source.exists():
        raise FileNotFoundError(f"未找到已上传匹配文件：{source}")

    workbook = CalamineWorkbook.from_path(str(source))
    try:
        if "Summary" not in workbook.sheet_names:
            raise ValueError(f"{source.name} 缺少 Summary 工作表")
        sheet = workbook.get_sheet_by_name("Summary")
        rows_iter = calamine_rows(sheet)
        header = next(rows_iter, [])
        required_headers = ("检索参考号", "状态", "描述", "补贴金额")
        missing_headers = [
            required_header
            for required_header in required_headers
            if required_header not in header
        ]
        if missing_headers:
            raise ValueError(
                f"{source.name} 缺少字段：{'、'.join(missing_headers)}"
            )

        reference_index = header.index("检索参考号")
        status_index = header.index("状态")
        description_index = header.index("描述")
        subsidy_index = header.index("补贴金额")
        lookup: dict[str, str] = {}
        subsidy_count = 0
        subsidy_total = Decimal("0")
        for row_number, row in enumerate(rows_iter, start=2):
            reference = normalize_reference(row[reference_index])
            if reference:
                validated_reference(
                    row[reference_index],
                    f"{source.name} 第 {row_number} 行",
                )
                status = str(row[status_index] or "").strip()
                description = str(row[description_index] or "").strip()
                detail = f"{status}：{description}"
                existing_detail = lookup.get(reference)
                if existing_detail is not None and existing_detail != detail:
                    raise ValueError(
                        f"{source.name} 第 {row_number} 行检索参考号存在冲突："
                        f"{reference}"
                    )
                lookup[reference] = detail

            subsidy = row[subsidy_index]
            if subsidy not in (None, ""):
                try:
                    subsidy_total += Decimal(str(subsidy))
                except InvalidOperation as error:
                    raise ValueError(
                        f"{source.name} 第 {row_number} 行补贴金额无效：{subsidy!r}"
                    ) from error
                subsidy_count += 1
        return lookup, subsidy_count, subsidy_total
    finally:
        workbook.close()


def load_payment_reference_locations(
    source: Path,
) -> dict[str, dict[str, str]]:
    """Read authoritative payment references once, separated by project."""
    if not source.exists():
        raise FileNotFoundError(f"未找到回款匹配文件：{source}")

    workbook = CalamineWorkbook.from_path(str(source))
    try:
        result: dict[str, dict[str, str]] = {}
        for project, sheet_name in PAYMENT_DETAIL_SHEETS.items():
            if sheet_name not in workbook.sheet_names:
                raise ValueError(f"{source.name} 缺少 {sheet_name} 工作表")
            sheet = workbook.get_sheet_by_name(sheet_name)
            rows_iter = calamine_rows(sheet)
            header = next(rows_iter, [])
            if "交易参考号" not in header:
                raise ValueError(f"{source.name} 的 {sheet_name} 缺少字段：交易参考号")
            reference_index = header.index("交易参考号")
            locations: dict[str, str] = {}
            for row_number, row in enumerate(rows_iter, start=2):
                location = f"{source.name} 的 {sheet_name} 第 {row_number} 行"
                raw_reference = (
                    row[reference_index]
                    if reference_index < len(row)
                    else None
                )
                if raw_reference in (None, ""):
                    raise ValueError(f"{location}交易参考号为空")
                reference = validated_reference(raw_reference, location)
                locations.setdefault(reference, location)
            result[project] = locations
        return result
    finally:
        workbook.close()


def validate_payment_reference_subset(
    project: str,
    payment_reference_locations: dict[str, str],
    submitted_references: set[str],
) -> None:
    missing = sorted(set(payment_reference_locations) - submitted_references)
    if not missing:
        return
    examples = "；".join(
        f"{reference}（{payment_reference_locations[reference]}）"
        for reference in missing[:10]
    )
    raise ValueError(
        f"{project}回款参考号子集校验失败：共 {len(missing)} 个交易参考号"
        f"未出现在{project}已上传数据中，示例：{examples}"
    )


def _read_header_row(path: Path) -> tuple[object, ...]:
    """Read row 2 (the field header row) via one sequential pass.

    Random sheet.cell(row, column) access on a read_only worksheet re-parses
    the sheet XML from the start on every call — fine for a handful of
    lookups, but O(n) per call adds up to O(n^2) once a caller does that once
    per data row on a 10000+ row export. Every reader in this module goes
    through iter_rows() instead, exactly once per full pass.
    """
    workbook = CalamineWorkbook.from_path(str(path))
    try:
        sheet = workbook.get_sheet_by_index(0)
        rows_iter = calamine_rows(sheet)
        next(rows_iter, None)  # title row
        return tuple(next(rows_iter, None) or ())
    finally:
        workbook.close()


def _header_cell(header_row: tuple[object, ...], column: int) -> str:
    if column - 1 >= len(header_row):
        return ""
    value = header_row[column - 1]
    return str(value).strip() if value is not None else ""


def configure_data_dir(data_dir: Path) -> None:
    global DATA_DIR
    global COUPON_SOURCE_FILE
    global COUPON_REFERENCE_SUPPLEMENT_FILE

    DATA_DIR = data_dir
    candidates = find_data_files(data_dir, COUPON_STATISTICS_KEYWORD, (".xlsx",))
    matches = []
    for candidate in candidates:
        header_row = _read_header_row(candidate)
        if (
            _header_cell(header_row, COUPON_FAMILY_SUBSIDY_COLUMN)
            == COUPON_FAMILY_SUBSIDY_HEADER
            and _header_cell(header_row, COUPON_DIGITAL_SUBSIDY_COLUMN)
            == COUPON_DIGITAL_SUBSIDY_HEADER
        ):
            matches.append(candidate)
    if not matches:
        COUPON_SOURCE_FILE = None
    elif len(matches) > 1:
        raise ValueError(
            f"找到多个表头符合销售用券情况统计导出格式的文件，无法确定使用哪一个：{matches}"
        )
    else:
        COUPON_SOURCE_FILE = matches[0]

    COUPON_REFERENCE_SUPPLEMENT_FILE = resolve_unique_file(
        find_data_files(
            data_dir, COUPON_REFERENCE_SUPPLEMENT_KEYWORD, (".xlsx",)
        )
    ) or data_dir / f"{COUPON_REFERENCE_SUPPLEMENT_KEYWORD}.xlsx"


@dataclass(frozen=True)
class CouponExport:
    appliance_rows: list[list[object]]
    digital_rows: list[list[object]]
    source_total: Decimal | None
    subsidy_corrections: tuple[SubsidyCorrection, ...]


def read_coupon_export(
    source: Path,
    source_workbook=None,
    remark_lookup: dict[tuple[str, date], str] | None = None,
) -> CouponExport:
    """Read the merged export once: classify every row for both projects,
    extract the source's own 合计 total, and collect every subsidy
    attribution correction in the same iter_rows() pass.

    processors/coupon_report.py needs all of these (家电 rows, 数码 rows,
    the source total, the corrections to warn about) every time it runs, so
    the whole read happens here exactly once instead of once per profile
    plus a separate total pass — real waste on a 10000+ row export even
    though each individual read is already O(n).
    """
    owns_workbook = source_workbook is None
    workbook = source_workbook or CalamineWorkbook.from_path(str(source))
    remark_lookup = remark_lookup or {}
    try:
        sheet = workbook.get_sheet_by_index(0)
        all_rows = list(calamine_rows(sheet))
        if len(all_rows) < 3:
            raise ValueError(f"{source.name} 缺少标题行、字段标题行或合计行")
        required_columns = max(
            COUPON_FAMILY_SUBSIDY_COLUMN, COUPON_DIGITAL_SUBSIDY_COLUMN
        )
        header_row = all_rows[1]
        if len(header_row) < required_columns:
            raise ValueError(
                f"{source.name} 列数不足：至少需要 {required_columns} 列"
            )
        last_row = all_rows[-1]
        last_row_marker = last_row[0] if last_row else None
        if str(last_row_marker or "").strip() != "合计":
            raise ValueError(f"{source.name} 最后一行不是合计行")

        for profile in (APPLIANCE_PROFILE, DIGITAL_PROFILE):
            source_header = tuple(
                header_row[column - 1] if column - 1 < len(header_row) else None
                for column in profile.kept_source_columns
            )
            expected_source_header = profile.output_header[
                : len(profile.kept_source_columns)
            ]
            if source_header != expected_source_header:
                raise ValueError(
                    f"{source.name} 保留列字段标题不符合要求："
                    f"预期为 {expected_source_header}，实际为 {source_header}。"
                    f"请检查该文件是否为正确的销售用券情况统计导出文件。"
                )

        total_column = max(APPLIANCE_PROFILE.kept_source_columns)
        total_cell_value = (
            last_row[total_column - 1] if total_column - 1 < len(last_row) else None
        )
        total_raw = "" if total_cell_value is None else str(total_cell_value)
        total_cleaned = re.sub(r"[^0-9.\-]", "", total_raw)
        try:
            source_total = Decimal(total_cleaned) if total_cleaned else None
        except InvalidOperation:
            source_total = None

        family_column = COUPON_FAMILY_SUBSIDY_COLUMN - 1
        digital_column = COUPON_DIGITAL_SUBSIDY_COLUMN - 1
        blank_subsidy_cells = {
            (row_number, column)
            for row_number, values in enumerate(all_rows[2:-1], start=3)
            for column in (family_column, digital_column)
            if column >= len(values) or values[column] in (None, "")
        }
        if blank_subsidy_cells:
            source_type_book = load_workbook(
                source, read_only=True, data_only=False
            )
            try:
                source_type_sheet = source_type_book.worksheets[0]
                if hasattr(source_type_sheet, "reset_dimensions"):
                    source_type_sheet.reset_dimensions()
                for row_number, cells in enumerate(
                    source_type_sheet.iter_rows(
                        min_col=COUPON_FAMILY_SUBSIDY_COLUMN,
                        max_col=COUPON_DIGITAL_SUBSIDY_COLUMN,
                    ),
                    start=1,
                ):
                    for column, cell, header in zip(
                        (family_column, digital_column),
                        cells,
                        (
                            COUPON_FAMILY_SUBSIDY_HEADER,
                            COUPON_DIGITAL_SUBSIDY_HEADER,
                        ),
                        strict=True,
                    ):
                        if (row_number, column) not in blank_subsidy_cells:
                            continue
                        if cell.data_type == "e":
                            raise ValueError(
                                f"{source.name} 第 {row_number} 行字段 {header!r} "
                                f"是 Excel 错误值：{cell.value!r}"
                            )
                        if cell.data_type == "f":
                            raise ValueError(
                                f"{source.name} 第 {row_number} 行字段 {header!r} "
                                "是公式但没有缓存计算结果；请先用 Excel/WPS "
                                "打开并保存，或将公式转换为数值"
                            )
            finally:
                source_type_book.close()
        appliance_rows: list[list[object]] = [list(APPLIANCE_PROFILE.output_header)]
        digital_rows: list[list[object]] = [list(DIGITAL_PROFILE.output_header)]
        subsidy_corrections: list[SubsidyCorrection] = []
        for row_number, values in enumerate(all_rows[2:-1], start=3):
            appliance_values = [
                values[column - 1] if column - 1 < len(values) else None
                for column in APPLIANCE_PROFILE.kept_source_columns
            ]
            digital_values = [
                values[column - 1] if column - 1 < len(values) else None
                for column in DIGITAL_PROFILE.kept_source_columns
            ]
            if not any(value not in (None, "") for value in appliance_values) and (
                not any(value not in (None, "") for value in digital_values)
            ):
                continue
            appliance_subsidy = (
                values[family_column] if family_column < len(values) else None
            )
            digital_subsidy = (
                values[digital_column] if digital_column < len(values) else None
            )
            document_number = (
                ""
                if appliance_values[0] is None
                else str(appliance_values[0]).replace("收款", "")
            )
            document_date = normalize_coupon_date(
                appliance_values[1], row_number, source.name
            )
            remark = remark_lookup.get(
                (normalize_document_number(document_number), document_date),
                "",
            )
            attribution = classify_subsidy_attribution(
                appliance_subsidy=appliance_subsidy,
                digital_subsidy=digital_subsidy,
                row_number=row_number,
                source_name=source.name,
                document_number=document_number,
                financial_category=appliance_values[4],
                has_receipt_remark=bool(remark),
            )
            if attribution.correction is not None:
                subsidy_corrections.append(attribution.correction)
            if source_total is not None:
                source_total += attribution.source_total_adjustment
            if attribution.classification == "家电":
                appliance_values[-1] = attribution.subsidy_value
                profile, row_values, target = (
                    APPLIANCE_PROFILE,
                    appliance_values,
                    appliance_rows,
                )
            else:
                digital_values[-1] = attribution.subsidy_value
                profile, row_values, target = (
                    DIGITAL_PROFILE,
                    digital_values,
                    digital_rows,
                )
            result_row = [
                document_number,
                document_date,
                *row_values[2:],
                None,
                None,
                None,
            ]
            if profile.normalize_brand:
                brand = str(result_row[3] or "").strip()
                result_row[3] = COUPON_BRAND_REPLACEMENTS.get(brand, result_row[3])
            target.append(result_row)
        return CouponExport(
            appliance_rows,
            digital_rows,
            source_total,
            tuple(subsidy_corrections),
        )
    finally:
        if owns_workbook:
            workbook.close()
