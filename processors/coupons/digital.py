"""数码 side of 销售用券情况统计 processing.

数码 has no group sheets, reference-supplement file, or approved-detail
panel, so its summary is a simpler three-column (备注, 数量, 合计) table
instead of 家电's five-column one — see processors/coupons/appliance.py's
module docstring for why this stays a separate module rather than a
feature-flagged variant of that one.
"""

from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook

from processors.common.dates import (
    normalize_document_number,
    normalize_receipt_identifier,
)
from processors.receipts import OUTPUT_FILE as RECEIPTS_OUTPUT_FILE
from processors.submitted import PROFILES as SUBMITTED_PROFILES

from . import matching, sources
from .matching import as_currency
from .sources import load_coupon_remark_lookup, load_uploaded_summary
from .validation import (
    is_pink_row,
    validate_detail_sheet_shape,
    validate_document_and_date_cells,
    validate_left_aligned_column,
    validate_matched_subsidy_total,
    validate_pink_position,
    validate_remark_and_detail,
    validate_uploaded_and_unmatched_counts,
)

COUPON_SUBSIDY_HEADER = sources.COUPON_DIGITAL_SUBSIDY_HEADER
COUPON_REMARK_SOURCE_FILE = RECEIPTS_OUTPUT_FILE
COUPON_UPLOADED_SOURCE_FILE = SUBMITTED_PROFILES["数码"].output_file
COUPON_OUTPUT_HEADER = sources.DIGITAL_PROFILE.output_header
COUPON_MATCH_FILL_COLOR = "FFC7CE"
DETAILS_SHEET_NAME = "数码-明细总表"


def build_coupon_summary(
    rows: list[list[object]],
    uploaded_count: int,
    uploaded_subsidy_total: Decimal,
) -> list[tuple[object, ...]]:
    subsidy_index = matching.SUBSIDY_INDEX
    coupon_count = 0
    coupon_total = Decimal("0")
    for row_number, row in enumerate(rows[1:], start=2):
        subsidy = row[subsidy_index]
        if subsidy in (None, ""):
            continue
        try:
            subsidy_amount = Decimal(str(subsidy))
            coupon_total += subsidy_amount
            if subsidy_amount < 0:
                coupon_count -= 1
            elif subsidy_amount > 0:
                coupon_count += 1
        except InvalidOperation as error:
            raise ValueError(
                f"第 {row_number} 行的2026数码国补金额无效：{subsidy!r}"
            ) from error

    unuploaded_count = coupon_count - uploaded_count
    unuploaded_total = coupon_total - uploaded_subsidy_total
    return [
        (
            "已上传",
            uploaded_count,
            float(as_currency(uploaded_subsidy_total)),
        ),
        (
            "未上传",
            unuploaded_count,
            float(as_currency(unuploaded_total)),
        ),
        ("合计", coupon_count, float(as_currency(coupon_total))),
    ]


@dataclass
class CouponComputation:
    """Everything computed for the 数码 coupon data, before it is written to
    a sheet. Digital has no group sheets, approved-detail panel, or
    reference-supplement file, so this is smaller than the 家电 equivalent
    in processors/coupons/appliance.py."""

    rows: list[list[object]]
    data_row_count: int
    matched_count: int
    matched_subsidy_total: Decimal
    remark_lookup: dict[tuple[str, date], str]
    detail_lookup: dict[str, str]
    reference_universe: set[str]
    uploaded_subsidy_count: int
    uploaded_subsidy_total: Decimal
    uploaded_match_count: int
    unmatched_count: int
    corrected_count: int
    unresolved_count: int
    correction_collision_count: int
    reference_decisions: list[tuple[str, str, str, str]]
    summary_rows: list[tuple[object, ...]]


def compute_coupon_data(
    *,
    rows: list[list[object]] | None = None,
    remark_lookup: dict[tuple[str, date], str] | None = None,
) -> CouponComputation:
    if sources.COUPON_SOURCE_FILE is None:
        raise FileNotFoundError(
            f"未在 {sources.DATA_DIR} 中找到文件名包含"
            f"“{sources.COUPON_STATISTICS_KEYWORD}”且表头为"
            f"“{COUPON_SUBSIDY_HEADER}”的 .XLSX 文件"
        )

    if rows is None:
        rows = sources.read_coupon_rows(
            sources.COUPON_SOURCE_FILE, sources.DIGITAL_PROFILE
        )
    if remark_lookup is None:
        remark_lookup = load_coupon_remark_lookup(COUPON_REMARK_SOURCE_FILE)
    matched_count, matched_subsidy_total, _receipt_remark_count = (
        matching.fill_coupon_remarks(rows, remark_lookup, "2026数码国补")
    )
    detail_lookup, uploaded_subsidy_count, uploaded_subsidy_total = (
        load_uploaded_summary(COUPON_UPLOADED_SOURCE_FILE)
    )
    # Unsubmitted data is no longer supplied, so submitted data is the only
    # source of valid references.
    reference_universe = set(detail_lookup)
    (
        corrected_count,
        unresolved_count,
        correction_collision_count,
        reference_decisions,
    ) = matching.correct_coupon_references(rows, reference_universe)
    uploaded_match_count = matching.fill_uploaded_details(rows, detail_lookup)
    unmatched_count = matching.fill_unmatched_remarks(rows, reference_universe)
    summary_rows = build_coupon_summary(
        rows,
        uploaded_subsidy_count,
        uploaded_subsidy_total,
    )
    if as_currency(matched_subsidy_total) != Decimal("0.00"):
        raise ValueError(
            f"备注匹配行的{COUPON_SUBSIDY_HEADER}合计不为 0："
            f"{matched_subsidy_total}"
        )

    return CouponComputation(
        rows=rows,
        data_row_count=max(len(rows) - 1, 0),
        matched_count=matched_count,
        matched_subsidy_total=matched_subsidy_total,
        remark_lookup=remark_lookup,
        detail_lookup=detail_lookup,
        reference_universe=reference_universe,
        uploaded_subsidy_count=uploaded_subsidy_count,
        uploaded_subsidy_total=uploaded_subsidy_total,
        uploaded_match_count=uploaded_match_count,
        unmatched_count=unmatched_count,
        corrected_count=corrected_count,
        unresolved_count=unresolved_count,
        correction_collision_count=correction_collision_count,
        reference_decisions=reference_decisions,
        summary_rows=summary_rows,
    )


def validate_detail_sheet(
    workbook: Workbook,
    computation: CouponComputation,
) -> None:
    expected_matched_rows = computation.matched_count
    remark_lookup = computation.remark_lookup
    expected_matched_subsidy_total = computation.matched_subsidy_total
    detail_lookup = computation.detail_lookup
    reference_universe = computation.reference_universe

    sheet = workbook[DETAILS_SHEET_NAME]
    validate_detail_sheet_shape(
        sheet, COUPON_OUTPUT_HEADER, computation.data_row_count
    )
    detail_column, remark_column = validate_uploaded_and_unmatched_counts(
        sheet,
        COUPON_OUTPUT_HEADER,
        computation.uploaded_match_count,
        computation.unmatched_count,
    )
    summary_column = COUPON_OUTPUT_HEADER.index("明细摘要") + 1
    product_name_column = COUPON_OUTPUT_HEADER.index("商品名称") + 1
    validate_left_aligned_column(sheet, product_name_column, "商品名称")

    matched_start_row = sheet.max_row - expected_matched_rows + 1
    actual_matched_subsidy_total = Decimal("0")
    subsidy_column = COUPON_OUTPUT_HEADER.index(COUPON_SUBSIDY_HEADER) + 1
    for row_number in range(2, sheet.max_row + 1):
        document_cell = sheet.cell(row_number, 1)
        date_cell = sheet.cell(row_number, 2)
        remark_cell = sheet.cell(row_number, remark_column)
        detail_cell = sheet.cell(row_number, detail_column)
        document_date = validate_document_and_date_cells(
            document_cell, date_cell, row_number
        )
        receipt_remark = remark_lookup.get(
            (
                normalize_document_number(document_cell.value),
                document_date,
            ),
            "",
        )
        reference = normalize_receipt_identifier(
            sheet.cell(row_number, summary_column).value
        ).upper()
        expected_detail = detail_lookup.get(reference, "")
        if expected_detail:
            expected_remark = "已上传"
        elif reference not in reference_universe:
            expected_remark = "未上传"
        else:
            expected_remark = receipt_remark
        validate_remark_and_detail(
            remark_cell, detail_cell, expected_remark, expected_detail, row_number
        )
        expected_pink = (
            expected_matched_rows > 0
            and row_number >= matched_start_row
        )
        is_pink = is_pink_row(
            sheet, row_number, len(COUPON_OUTPUT_HEADER), COUPON_MATCH_FILL_COLOR
        )
        validate_pink_position(is_pink, expected_pink, row_number)
        if expected_pink:
            subsidy = sheet.cell(row_number, subsidy_column).value
            if subsidy not in (None, ""):
                actual_matched_subsidy_total += Decimal(str(subsidy))

    validate_matched_subsidy_total(
        actual_matched_subsidy_total, expected_matched_subsidy_total
    )
    if as_currency(actual_matched_subsidy_total) != Decimal("0.00"):
        raise RuntimeError(
            f"销售用券匹配行的{COUPON_SUBSIDY_HEADER}合计不为 0："
            f"{actual_matched_subsidy_total}"
        )
