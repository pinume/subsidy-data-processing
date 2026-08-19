"""Store subsidy upload/payment status report (门店国补上传及回款情况表).

Unlike the other processors, this one does not read raw exports from the
data directory. It reads two of this project's own outputs — 审核明细.xlsx
(from the 审核明细 processing mode) and 回款明细.xlsx (from the 回款明细
processing mode) — aggregates them by brand and category, and fills a blank
report template that the operator keeps in the data directory.

Current scope is intentionally narrow: one store (益庄店), one year (2026),
one fixed template. ROW_RULES/BRAND_GROUP_RULES and the output filename are
hardcoded rather than configurable — see the module docstring history for why
this is deferred until a second store is actually needed.
"""

from __future__ import annotations

import math
from collections.abc import Sequence
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.workbook.workbook import Workbook
from python_calamine import CalamineWorkbook

from processors.common.console import ConsoleReporter
from processors.common.excel import (
    calamine_rows,
    resolve_font,
    save_workbook_atomically,
)
from processors.common.paths import find_data_files, resolve_unique_file
from processors.common.references import normalize_reference
from processors.coupon_report import OUTPUT_FILE as UPLOAD_FILE
from processors.coupon_report import SUBSIDY_YEAR
from processors.coupon_report import SUMMARY_CORE_HEADER as UPLOAD_HEADER
from processors.coupon_report import SUMMARY_SHEET_NAME as UPLOAD_SHEET_NAME
from processors.coupons import appliance as coupon_appliance
from processors.coupons.matching import as_currency
from processors.payment import APPLIANCE_CATEGORY_MAP as PAYMENT_CATEGORY_MAP
from processors.payment import DERIVED_HEADERS as PAYMENT_DERIVED_HEADERS
from processors.payment import OUTPUT_FILE as PAYMENT_FILE
from processors.payment import PROFILES as PAYMENT_PROFILES
from processors.payment import SUMMARY_HEADERS as PAYMENT_HEADER
from processors.payment import SUMMARY_SHEET_NAME as PAYMENT_SHEET_NAME

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_FILE = (
    OUTPUT_DIR / f"{SUBSIDY_YEAR}年门店国补上传及回款情况表（益庄店）.xlsx"
)
TEMPLATE_FILE_KEYWORD = "门店国补上传及回款情况表"
DATA_DIR: Path

SHANGHAI_TZ = ZoneInfo("Asia/Shanghai")
TEMPLATE_HEADER_PREFIX = (
    f"表1                  {SUBSIDY_YEAR}年（益庄店 ）门店国补上传及回款情况表_"
)
FILLED_ALIGNMENT = Alignment(horizontal="left", vertical="center")
DATA_NUMBER_FORMAT = "General"
PERCENT_NUMBER_FORMAT = "0.00%"
TOTAL_ROW = 33
DETAIL_ROWS = range(3, TOTAL_ROW)
BRAND_GROUP_TOTAL_ROW = 47
BRAND_GROUP_DETAIL_ROWS = range(40, BRAND_GROUP_TOTAL_ROW)
TABLE3_PROJECT_ROWS = {"家电": 50, "数码": 51}
TABLE3_TOTAL_ROW = 52
EXPECTED_SHEET_COUNT = 1
EXPECTED_COLUMN_COUNT = 8  # A..H
EXPECTED_SHEET_TITLE = "益庄"

# 正式模板的版本标识，写在表 3 下方一行。data/ 不进 Git，模板由操作员
# 手工放入各环境，因此这个标记是防止误用旧模板的唯一防线：旧模板
# （含手工修改版）没有它，validate_template 直接给出更换提示。
# 版本标记行必须隐藏，否则会出现在最终报表的打印结果里。
TEMPLATE_VERSION_CELL = "A53"
TEMPLATE_VERSION_ROW = 53
TEMPLATE_VERSION_MARKER = "模板版本：2026-V5"
# 版本标记在第 53 行：模板至少 53 行，与版本契约一致。
MIN_TEMPLATE_ROW_COUNT = TEMPLATE_VERSION_ROW

# 正式模板的全部固定合并区域。校验“必须包含”——模板被改动丢失任一合并
# 都会让标签错位（品类纵向合并、表头分组、表 3 分组），多余合并不拒绝。
# style_id 不校验（Excel 重新保存后可能变化）。
EXPECTED_MERGED_RANGES = frozenset(
    {
        "A1:H1",
        "A3:A13",
        "A14:A18",
        "A19:A25",
        "A26:A31",
        "A33:C33",
        "C38:H38",
    }
)

# A handful of fixed labels that only exist in the right blank template.
# Checked before writing (wrong/stale template must not silently produce a
# wrong report) and again after saving (writing must not have corrupted them).

# 表 1：明细与总计（第 2–33 行）
_TABLE1_CELLS: dict[str, str] = {
    "A2": "品类",
    "B2": "序号",
    "C2": "品牌",
    "D2": "26年发生额",
    "E2": "上传额",
    "F2": "上传率",
    "G2": "回款额",
    "H2": "回款率",

    "C3": "海尔系冰箱",
    "C4": "海尔系洗衣机",
    "C5": "美的系冰箱",
    "C6": "美的系洗衣机",
    "C7": "西门子冰箱",
    "C8": "西门子洗衣机",
    "C9": "博世冰箱",
    "C10": "博世洗衣机",
    "C11": "美菱冰箱",
    "C12": "美菱洗衣机",
    "C13": "小鸭洗衣机",
    "C14": "海信",
    "C15": "创维",
    "C16": "TCL",
    "C17": "海尔",
    "C18": "华为",
    "C19": "格力",
    "C20": "美的",
    "C21": "海尔",
    "C22": "海信",
    "C23": "奥克斯",
    "C24": "科龙",
    "C25": "TCL",
    "C26": "老板",
    "C27": "方太",
    "C28": "AO史密斯",
    "C29": "海尔",
    "C30": "美的",
    "C31": "万家乐",
    "C32": "数码",
    "A33": "费用总计",
}

# 表 2：主要品牌汇总（第 38–47 行）
_TABLE2_CELLS: dict[str, str] = {
    "C38": "表2                            主要品牌国补上传及回款情况",
    "C39": "品牌",
    "D39": "26年国补发生额",
    "E39": "26年国补上传额",
    "F39": "26年国补回款额",
    "G39": "上传率",
    "H39": "回款率",
    "C40": "海尔系",
    "C41": "美的系",
    "C42": "格力",
    "C43": "博西",
    "C44": "海信系",
    "C45": "创维",
    "C46": "TCL",
    "C47": "合计",
}

# 表 3：数量与金额统计（第 49–52 行）
_TABLE3_CELLS: dict[str, str] = {
    "C49": "表3",
    "D49": "审核中（数量）",
    "E49": "审核中（金额）",
    "F49": "未上传（数量）",
    "G49": "未上传（金额）",
    "C50": "家电",
    "C51": "数码",
    "C52": "合计",
}

TEMPLATE_STRUCTURE_CELLS: dict[str, str] = {
    **_TABLE1_CELLS,
    **_TABLE2_CELLS,
    **_TABLE3_CELLS,
    TEMPLATE_VERSION_CELL: TEMPLATE_VERSION_MARKER,
}


def configure_data_dir(data_dir: Path) -> None:
    global DATA_DIR
    DATA_DIR = data_dir


def resolve_template_file() -> Path:
    candidates = find_data_files(DATA_DIR, TEMPLATE_FILE_KEYWORD, (".xlsx",))
    template_file = resolve_unique_file(candidates)
    if template_file is None:
        raise FileNotFoundError(
            f"未在 {DATA_DIR} 找到包含“{TEMPLATE_FILE_KEYWORD}”的空白报表模板"
        )
    return template_file


def _validate_header_title(value: object, source_description: str) -> None:
    text = "" if value is None else str(value)
    if not text.startswith(TEMPLATE_HEADER_PREFIX):
        raise ValueError(
            f"{source_description} A1 标题前缀不符合预期，预期以 {TEMPLATE_HEADER_PREFIX!r} 开头，实际为 {text!r}"
        )
    timestamp_str = text[len(TEMPLATE_HEADER_PREFIX):]
    try:
        parsed = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        raise ValueError(
            f"{source_description} A1 标题时间戳无效，应为合法的 'YYYY-MM-DD HH:MM:SS' 格式，实际为 {timestamp_str!r}"
        ) from None
    if parsed.strftime("%Y-%m-%d %H:%M:%S") != timestamp_str:
        raise ValueError(
            f"{source_description} A1 标题时间戳无效，应为合法的 'YYYY-MM-DD HH:MM:SS' 格式，实际为 {timestamp_str!r}"
        )


def validate_template(workbook: Workbook) -> None:
    """Reject a template that doesn't match the expected layout by content, not just filename.

    Checked before any cell is written, so a wrong/stale template fails fast
    instead of producing a wrong report that only the (more expensive)
    post-save validate_output would catch.
    """
    if len(workbook.sheetnames) != EXPECTED_SHEET_COUNT:
        raise ValueError(
            f"空白模板工作表数量应为 {EXPECTED_SHEET_COUNT}，实际为 {len(workbook.sheetnames)}"
        )
    sheet = workbook[workbook.sheetnames[0]]
    if sheet.title != EXPECTED_SHEET_TITLE:
        raise ValueError(f"空白模板工作表名称应为 {EXPECTED_SHEET_TITLE!r}，实际为 {sheet.title!r}")
    if sheet.max_column != EXPECTED_COLUMN_COUNT:
        raise ValueError(f"空白模板列数应为 {EXPECTED_COLUMN_COUNT}，实际为 {sheet.max_column}")
    if sheet.max_row < MIN_TEMPLATE_ROW_COUNT:
        raise ValueError(f"空白模板行数至少应为 {MIN_TEMPLATE_ROW_COUNT}，实际为 {sheet.max_row}")

    try:
        _validate_header_title(sheet["A1"].value, "空白模板")
    except ValueError as exc:
        raise ValueError(f"{exc}；请更换为新版模板") from None

    # 版本标记先行检查：旧模板（含无标记的手工修改版）在这里被明确拒绝，
    # 而不是混入下面的通用结构错误里让人猜。
    if sheet[TEMPLATE_VERSION_CELL].value != TEMPLATE_VERSION_MARKER:
        raise ValueError(
            f"空白模板版本校验失败：{TEMPLATE_VERSION_CELL} 应为"
            f" {TEMPLATE_VERSION_MARKER!r}，实际为 {sheet[TEMPLATE_VERSION_CELL].value!r}。"
            "请更换为新版模板（含版本标记的正式模板）"
        )
    if not sheet.row_dimensions[TEMPLATE_VERSION_ROW].hidden:
        raise ValueError(
            f"空白模板版本标记行（第 {TEMPLATE_VERSION_ROW} 行）应为隐藏，"
            "否则会出现在报表打印结果中；请更换为新版模板"
        )

    actual_merges = {str(range_) for range_ in sheet.merged_cells.ranges}
    missing_merges = sorted(EXPECTED_MERGED_RANGES - actual_merges)
    if missing_merges:
        raise ValueError(
            "空白模板缺少合并区域：" + "、".join(missing_merges)
            + "；请更换为新版模板"
        )

    # 明细行 3..TOTAL_ROW-1 共 30 行，序号 1..30。
    for expected, row in enumerate(DETAIL_ROWS, start=1):
        actual = sheet[f"B{row}"].value
        if actual != expected:
            raise ValueError(
                f"空白模板序号校验失败：B{row} 应为 {expected}，"
                f"实际为 {actual!r}；请更换为新版模板"
            )

    mismatches = [
        f"{coordinate}（应为 {expected!r}，实际为 {sheet[coordinate].value!r}）"
        for coordinate, expected in TEMPLATE_STRUCTURE_CELLS.items()
        if sheet[coordinate].value != expected
    ]
    if mismatches:
        raise ValueError(
            "空白模板结构校验失败，可能选错了文件或模板已被改动，请更换为新版模板："
            + "；".join(mismatches)
        )


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    replacements = {
        "A.O.史密斯": "AO史密斯",
        "美的系": "美的",
        "东芝JX": "东芝",
        "华为（终端）": "华为",
    }
    return replacements.get(text, text)


@dataclass(frozen=True)
class RowRule:
    row: int
    upload_categories: tuple[str, ...]
    upload_brands: tuple[str, ...]
    payment_categories: tuple[str, ...]
    payment_brands: tuple[str, ...]
    fill_digital: bool = False


ROW_RULES = (
    # 冰洗 3–13
    RowRule(3, ("冰箱",), ("海尔", "卡萨帝"), ("冰箱",), ("海尔", "卡萨帝")),
    RowRule(4, ("洗衣机",), ("海尔", "卡萨帝"), ("洗衣机",), ("海尔", "卡萨帝")),
    RowRule(5, ("冰箱",), ("美的", "COLMO", "东芝"), ("冰箱",), ("美的系", "COLMO", "东芝JX")),
    RowRule(6, ("洗衣机",), ("美的", "小天鹅", "COLMO"), ("洗衣机",), ("美的系", "小天鹅", "COLMO")),
    RowRule(7, ("冰箱",), ("西门子",), ("冰箱",), ("西门子",)),
    RowRule(8, ("洗衣机",), ("西门子",), ("洗衣机",), ("西门子",)),
    RowRule(9, ("冰箱",), ("博世",), ("冰箱",), ("博世",)),
    RowRule(10, ("洗衣机",), ("博世",), ("洗衣机",), ("博世",)),
    RowRule(11, ("冰箱",), ("美菱",), ("冰箱",), ("美菱",)),
    RowRule(12, ("洗衣机",), ("美菱",), ("洗衣机",), ("美菱",)),
    RowRule(13, ("洗衣机",), ("小鸭",), ("洗衣机",), ("小鸭",)),
    # 电视 14–18
    RowRule(14, ("国产彩电",), ("海信",), ("电视",), ("海信",)),
    RowRule(15, ("国产彩电",), ("创维",), ("电视",), ("创维",)),
    RowRule(16, ("国产彩电",), ("TCL",), ("电视",), ("TCL",)),
    RowRule(17, ("国产彩电",), ("海尔", "卡萨帝"), ("电视",), ("海尔", "卡萨帝")),
    RowRule(18, ("国产彩电",), ("华为", "华为（终端）"), ("电视",), ("华为", "华为（终端）")),
    # 空调 19–25
    RowRule(19, ("空调",), ("格力",), ("空调",), ("格力",)),
    RowRule(20, ("空调",), ("美的",), ("空调",), ("美的",)),
    RowRule(21, ("空调",), ("海尔", "卡萨帝"), ("空调",), ("海尔", "卡萨帝")),
    RowRule(22, ("空调",), ("海信",), ("空调",), ("海信",)),
    RowRule(23, ("空调",), ("奥克斯",), ("空调",), ("奥克斯",)),
    RowRule(24, ("空调",), ("科龙",), ("空调",), ("科龙",)),
    RowRule(25, ("空调",), ("TCL",), ("空调",), ("TCL",)),
    # 厨卫 26–31
    RowRule(26, ("厨卫",), ("老板",), ("厨卫",), ("老板",)),
    # 第 27 行：方太（同时汇总厨卫和冰箱两侧方太）
    RowRule(27, ("厨卫", "冰箱"), ("方太",), ("厨卫", "冰箱"), ("方太",)),
    RowRule(28, ("厨卫",), ("AO史密斯", "A.O.史密斯"), ("厨卫",), ("AO史密斯", "A.O.史密斯")),
    RowRule(29, ("厨卫",), ("海尔", "卡萨帝"), ("厨卫",), ("海尔", "卡萨帝")),
    RowRule(30, ("厨卫",), ("美的", "COLMO"), ("厨卫",), ("美的系", "美的", "COLMO")),
    RowRule(31, ("厨卫",), ("万家乐",), ("厨卫",), ("万家乐",)),
    # 3C 数码 32
    RowRule(32, (), (), ("数码",), (), fill_digital=True),
)

# Every payment-file 财务大类 this report knows how to place. A category
# outside both sets raises rather than being silently folded into 数码 — see
# load_payment_data.
HOUSEHOLD_PAYMENT_CATEGORIES = frozenset({"冰箱", "洗衣机", "电视", "空调", "厨卫"})
DIGITAL_PAYMENT_CATEGORIES = frozenset({"手机", "平板", "智能穿戴"})

# 审核侧品类纠正的白名单：只有纠正目标落在 ROW_RULES 的审核侧品类内才
# 执行纠正。回款侧“电视”对应审核侧“国产彩电”——那是两套口径的体系差异，
# 不是错标，纠正会让纠正后的行失去规则匹配，因此这类目标品类不纠正。
UPLOAD_CATEGORIES = frozenset(
    category
    for rule in ROW_RULES
    for category in rule.upload_categories
    if category
)

def _category_correction_exemptions(
    rules: tuple[RowRule, ...],
) -> frozenset[tuple[str, str, str]]:
    """从行规则派生品类纠正的豁免键集合。

    三个分量必须与 _correct_upload_category 的查表方式逐一对齐：
    审核侧品类取 upload_categories，回款侧品类取 payment_categories
    （两侧口径不同，例如 国产彩电/电视），品牌取两侧并集并统一过
    normalize_text——查表用的 record.brand 是归一化后的值，这里不归一化
    会造出永远匹配不上的键，让规则本该压制的纠正静默恢复，且不报错。

    提取成函数是为了能对合成规则做回归测试：常量在导入期求值，
    而 ROW_RULES 今天只有一条规则满足条件，覆盖不到上面这些坑。
    """
    return frozenset(
        (
            normalize_text(upload_category),
            normalize_text(payment_category),
            normalize_text(brand),
        )
        for rule in rules
        if len(rule.upload_categories) > 1
        for brand in (*rule.upload_brands, *rule.payment_brands)
        for upload_category in rule.upload_categories
        for payment_category in rule.payment_categories
        if normalize_text(upload_category) != normalize_text(payment_category)
    )


# 品类纠正的豁免：某些品牌在行规则中横跨多个审核侧品类汇总到同一行，
# 审核侧标注任一品类都是合理的，纠正反而会打破下游的合并逻辑。
# 格式: frozenset of (审核侧品类, 回款侧品类, 品牌)
CATEGORY_CORRECTION_EXEMPTIONS: frozenset[tuple[str, str, str]] = (
    _category_correction_exemptions(ROW_RULES)
)


@dataclass(frozen=True)
class BrandGroupCategory:
    upload_category: str
    payment_category: str
    brands: tuple[str, ...]


@dataclass(frozen=True)
class BrandGroupRule:
    row: int
    name: str
    categories: tuple[BrandGroupCategory, ...]


@dataclass(frozen=True)
class PaymentDetailRecord:
    """回款明细「家电明细」中的一行，按交易参考号索引，用于审核侧品类纠正。

    category 是由编码品类映射出的财务大类；编码品类未配置时为空，此时
    该参考号无法参与纠正（列入人工核对）。raw_category 保留编码品类原文，
    供纠正提示展示依据。
    """

    category: str | None
    raw_category: str
    brand: str
    subsidy: Decimal


@dataclass(frozen=True)
class CategoryCorrection:
    """一条已执行的审核侧品类纠正，用于控制台提示。"""

    document_number: str
    reference: str
    brand: str
    original_category: str
    corrected_category: str
    raw_payment_category: str


BRAND_GROUP_RULES = (
    BrandGroupRule(
        40,
        "海尔系",
        (
            BrandGroupCategory("冰箱", "冰箱", ("海尔", "卡萨帝")),
            BrandGroupCategory("洗衣机", "洗衣机", ("海尔", "卡萨帝")),
            BrandGroupCategory("国产彩电", "电视", ("海尔", "卡萨帝")),
            BrandGroupCategory("空调", "空调", ("海尔", "卡萨帝")),
            BrandGroupCategory("厨卫", "厨卫", ("海尔", "卡萨帝")),
        ),
    ),
    BrandGroupRule(
        41,
        "美的系",
        (
            BrandGroupCategory("冰箱", "冰箱", ("美的", "COLMO", "东芝")),
            BrandGroupCategory("洗衣机", "洗衣机", ("美的", "小天鹅", "COLMO")),
            BrandGroupCategory("空调", "空调", ("美的",)),
            BrandGroupCategory("厨卫", "厨卫", ("美的", "COLMO")),
        ),
    ),
    BrandGroupRule(42, "格力", (BrandGroupCategory("空调", "空调", ("格力",)),)),
    BrandGroupRule(
        43,
        "博西",
        (
            BrandGroupCategory("冰箱", "冰箱", ("西门子", "博世")),
            BrandGroupCategory("洗衣机", "洗衣机", ("西门子", "博世")),
        ),
    ),
    BrandGroupRule(
        44,
        "海信系",
        (
            BrandGroupCategory("国产彩电", "电视", ("海信",)),
            BrandGroupCategory("空调", "空调", ("海信", "科龙")),
        ),
    ),
    BrandGroupRule(45, "创维", (BrandGroupCategory("国产彩电", "电视", ("创维",)),)),
    BrandGroupRule(
        46,
        "TCL",
        (
            BrandGroupCategory("国产彩电", "电视", ("TCL",)),
            BrandGroupCategory("空调", "空调", ("TCL",)),
        ),
    ),
)


def to_decimal(value: object) -> Decimal:
    """Excel 浮点金额统一量化为两位小数。

    源文件里 2280155.7 这类值经 IEEE-754 往返后可能带尾差
    （2280155.70000000000101）；Decimal(str()) 会原样保留它们，导致精确
    相等校验、比例计算和人工核对全部带上尾数。复用 as_currency 的
    ROUND_HALF_UP 规则，与 payment/submitted 的金额口径一致。
    """
    if value is None or value == "":
        return Decimal("0")
    return as_currency(Decimal(str(value)))


def to_count(value: object) -> int:
    """数量必须原样校验：不能用 to_decimal（量化到两位小数会把 1.004
    变成 1.00 而漏过非整数检查），也不能接受 NaN/Infinity。非数值文本
    （Decimal 抛 InvalidOperation）统一转成数量应为整数的 ValueError。"""
    if value is None or value == "":
        return 0
    try:
        count = Decimal(str(value))
    except InvalidOperation:
        raise ValueError(f"数量应为整数，实际为 {value!r}") from None
    if not count.is_finite() or count != count.to_integral_value():
        raise ValueError(f"数量应为整数，实际为 {value!r}")
    return int(count)


def safe_ratio(numerator: Decimal, denominator: Decimal) -> float | None:
    if denominator == 0:
        return None
    ratio = numerator / denominator
    if ratio == 0:
        return None
    return float(ratio)


def decimal_to_cell_value(value: Decimal) -> float | None:
    if value == 0:
        return None
    return float(value)


def apply_filled_style(sheet, coords: tuple[str, ...], font: Font) -> None:
    for coord in coords:
        sheet[coord].alignment = FILLED_ALIGNMENT
        sheet[coord].font = font


def apply_data_number_format(sheet, coords: tuple[str, ...]) -> None:
    for coord in coords:
        sheet[coord].number_format = DATA_NUMBER_FORMAT


def apply_report_font(sheet, font_name: str) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            font = copy(cell.font)
            font.name = font_name
            cell.font = font


def _open_source_workbook(path: Path, business_name: str, processing_mode: str):
    """Open one of this program's own outputs with calamine, not openpyxl.

    Both inputs are read for one small summary sheet each, but openpyxl parses
    the whole shared string table up front no matter which sheet is wanted and
    no matter that read_only was asked for. Since these files are written by
    XlsxWriter, which shares strings rather than inlining them the way openpyxl
    did, that table now holds every string of the thousands of detail rows
    neither loader looks at — 0.17s to reach a 48-row sheet. calamine parses it
    in Rust and the cost disappears.
    """
    if not path.exists():
        raise FileNotFoundError(
            f"未找到{business_name}：{path}，请先运行“{processing_mode}”处理模式生成该文件"
        )
    return CalamineWorkbook.from_path(str(path))


def _sheet_rows_or_raise(
    workbook, sheet_name: str, source_name: str, business_name: str
) -> list[list[object]]:
    if sheet_name not in workbook.sheet_names:
        raise ValueError(
            f"{source_name} 缺少工作表 {sheet_name!r}，文件可能不是本程序生成的{business_name}"
        )
    return list(calamine_rows(workbook.get_sheet_by_name(sheet_name)))


def _validate_header(
    rows: list[list[object]],
    expected_header: Sequence[str],
    source_name: str,
    business_name: str,
) -> None:
    """Check the first row by field name, not just position, so a column reorder or a
    year-label change upstream fails clearly instead of silently misreading data."""
    actual_header = tuple(rows[0][: len(expected_header)]) if rows else ()
    if actual_header != tuple(expected_header):
        raise ValueError(
            f"{source_name} 的{business_name}表头不符合预期：预期 {tuple(expected_header)}，实际 {actual_header}"
        )


def _build_payment_detail_index(
    rows: list[list[object]],
) -> tuple[
    dict[str, PaymentDetailRecord],
    dict[str, tuple[PaymentDetailRecord, ...]],
]:
    """按交易参考号索引回款明细「家电明细」行，供审核侧品类纠正使用。

    - 参考号唯一 → index[ref] = 记录
    - 同参考号多条但 (品类, 品牌) 相同（如一笔销售加一笔冲正，正负成对）→
      记入 ambiguous_refs，纠正阶段仅在回款品类与审核品类不同时提示
    - 同参考号多条且 (品类, 品牌) 不同 → 记入 ambiguous_refs 的完整集合
    - 编码品类不在 PAYMENT_CATEGORY_MAP → 记录 category 为空，纠正阶段
      列入人工核对，不自动纠正
    - 编码品类映射结果与明细财务大类不一致 → 直接报错：口径漂移不能猜测
    """
    detail_header = PAYMENT_PROFILES["家电"].detail_headers + PAYMENT_DERIVED_HEADERS
    positions = {name: index for index, name in enumerate(detail_header)}
    reference_index = positions["交易参考号"]
    raw_category_index = positions["编码品类"]
    category_index = positions["财务大类"]
    brand_index = positions["品牌"]
    subsidy_index = positions["补贴金额"]

    grouped: dict[str, list[PaymentDetailRecord]] = {}
    for row in rows[1:]:
        reference = normalize_reference(row[reference_index])
        if not reference:
            continue  # 空参考号行不参与纠正
        raw_category = normalize_text(row[raw_category_index])
        category = normalize_text(row[category_index])
        brand = normalize_text(row[brand_index])
        mapped = PAYMENT_CATEGORY_MAP.get(raw_category)
        if mapped is not None and mapped != category:
            raise ValueError(
                f"回款明细编码品类 {raw_category!r} 映射为财务大类 {mapped!r}，"
                f"与明细行财务大类 {category!r} 不一致；"
                "请检查 config/payment_brands.yaml 的 categories.appliance"
            )
        grouped.setdefault(reference, []).append(
            PaymentDetailRecord(
                category=mapped,
                raw_category=raw_category,
                brand=brand,
                subsidy=to_decimal(row[subsidy_index]),
            )
        )

    index: dict[str, PaymentDetailRecord] = {}
    ambiguous: dict[str, tuple[PaymentDetailRecord, ...]] = {}
    for reference, records in grouped.items():
        # dict.fromkeys 保留源文件行序去重（set 迭代顺序不稳定，提示候选
        # 会随机排序）。
        distinct = list(dict.fromkeys(records))
        if len(distinct) == 1:
            index[reference] = distinct[0]
        else:
            ambiguous[reference] = tuple(distinct)
    return index, ambiguous


def _correct_upload_category(
    document_number: str,
    category: str,
    brand: str,
    amount: Decimal,
    reference: str,
    payment_index: dict[str, PaymentDetailRecord],
    ambiguous_refs: dict[str, tuple[PaymentDetailRecord, ...]],
) -> tuple[str, PaymentDetailRecord | None, tuple[str, tuple[str, ...]] | None]:
    """按回款明细的交易参考号纠正一条审核记录的品类。

    返回 (纠正后品类, 命中的回款记录, 人工核对提示)。不纠正的情形：
    - 参考号为空或未匹配回款记录：保留原品类（未回款是常态，不提示）
    - 回款品类与审核品类相同：无需纠正
    - 回款品类不在审核侧行规则白名单（电视↔国产彩电是两套口径的差异）：
      不纠正也不提示
    - 品牌不一致 / 补贴金额不一致 / 编码品类未配置 / 参考号对应多条回款
      记录且回款品类唯一：不纠正，列入人工核对
    """
    if not reference:
        return category, None, None

    record = payment_index.get(reference)
    if record is not None:
        if record.category is None:
            return category, record, (
                "审核明细品类纠正未执行：编码品类未配置",
                (f"{document_number}｜{reference}｜品牌 {brand}",),
            )
        if record.category == category:
            return category, record, None
        if record.category not in UPLOAD_CATEGORIES:
            return category, record, None
        if record.brand != brand:
            return category, record, (
                "审核明细品类纠正未执行：品牌不一致",
                (
                    f"{document_number}｜{reference}｜审核侧品牌 {brand}"
                    f" vs 回款侧品牌 {record.brand}",
                ),
            )
        if record.subsidy != amount:
            return category, record, (
                "审核明细品类纠正未执行：补贴金额不一致",
                (
                    f"{document_number}｜{reference}｜审核侧 {amount}"
                    f" vs 回款侧 {record.subsidy}",
                ),
            )
        if (category, record.category, record.brand) in CATEGORY_CORRECTION_EXEMPTIONS:
            return category, record, None
        return record.category, record, None

    records = ambiguous_refs.get(reference)
    if records is not None:
        categories = {record.category for record in records if record.category is not None}
        if len(categories) == 1:
            # 候选品类唯一（如一笔销售加一笔冲正的正负对）：仅当与审核
            # 品类不同且目标在白名单内才提示，避免常态冲正刷屏。
            only = next(iter(categories))
            if category not in categories and only in UPLOAD_CATEGORIES:
                return category, None, (
                    "审核明细品类纠正未执行：参考号对应多条回款记录",
                    (f"{document_number}｜{reference}｜品牌 {brand}",),
                )
        else:
            # 多个候选品类（或全部未配置）：无法唯一确定，必须人工核对，
            # 列出全部候选，绝不自动纠正。
            candidates = "；".join(
                f"{record.raw_category}｜{record.brand}｜{record.subsidy}"
                for record in records
            )
            return category, None, (
                "审核明细品类纠正未执行：参考号对应多个编码品类",
                (
                    f"{document_number}｜{reference}｜品牌 {brand}",
                    f"回款候选：{candidates}",
                ),
            )
    return category, None, None


def _amount_difference_items(
    detail: dict[tuple[str, str], dict[str, Decimal]],
    summary: dict[tuple[str, str], dict[str, Decimal]],
) -> list[str]:
    """两份品牌金额字典的差异描述：缺的键、多的键、金额不同的键。

    全量字典不适合进错误消息（真实数据几十个品牌，每个键带两份子字典
    会让报错刷屏），只产出人类可读的单行差异项。
    """
    items: list[str] = []
    for key in sorted(set(summary) - set(detail)):
        items.append(f"{key[0]}/{key[1]}：仅汇总存在")
    for key in sorted(set(detail) - set(summary)):
        items.append(f"{key[0]}/{key[1]}：仅明细存在")
    for key in sorted(set(detail) & set(summary)):
        for status in ("已上传", "未上传"):
            detail_amount = detail[key].get(status, Decimal("0"))
            summary_amount = summary[key].get(status, Decimal("0"))
            if detail_amount != summary_amount:
                items.append(
                    f"{key[0]}/{key[1]}/{status}：明细 {detail_amount}，汇总 {summary_amount}"
                )
    return items


def _aggregate_upload_detail(
    detail_rows: list[list[object]],
    payment_index: dict[str, PaymentDetailRecord],
    ambiguous_refs: dict[str, tuple[PaymentDetailRecord, ...]],
) -> tuple[
    dict[tuple[str, str], dict[str, Decimal]],  # 纠正前
    dict[tuple[str, str], dict[str, Decimal]],  # 纠正后
    list[CategoryCorrection],
    list[tuple[str, tuple[str, ...]]],
]:
    """从审核明细「家电-明细总表」逐行聚合品牌金额，并按参考号纠正品类。

    与数据汇总品牌行的口径一致：仅备注为已上传/未上传的行计入，错分类到
    数码的行与退换货/倒票行不计入。逐行聚合使品类纠正天然按行生效——
    同 (原品类, 品牌) 下部分行被纠正、部分行保留，各自落入对应报表行。
    """
    header = coupon_appliance.COUPON_OUTPUT_HEADER
    positions = {name: index for index, name in enumerate(header)}
    document_index = positions["单据号"]
    category_index = positions["财务大类"]
    brand_index = positions["品牌"]
    reference_index = positions["明细摘要"]
    subsidy_index = positions[header[6]]
    remark_index = positions["备注"]

    pre_amounts: dict[tuple[str, str], dict[str, Decimal]] = {}
    amounts: dict[tuple[str, str], dict[str, Decimal]] = {}
    corrections: list[CategoryCorrection] = []
    reviews: list[tuple[str, tuple[str, ...]]] = []

    for row in detail_rows[1:]:
        category = normalize_text(row[category_index])
        brand = normalize_text(row[brand_index])
        remark = normalize_text(row[remark_index])
        if category == "数码" or remark not in ("已上传", "未上传"):
            continue
        amount = to_decimal(row[subsidy_index])
        reference = normalize_reference(row[reference_index])
        document_number = str(row[document_index] or "")

        # 纠正前聚合：与数据汇总品牌行逐键比对（load_upload_data 内校验），
        # 任何一行被意外跳过、状态改名或品牌归一化漂移都会在这里暴露。
        pre_bucket = pre_amounts.setdefault(
            (category, brand),
            {"已上传": Decimal("0"), "未上传": Decimal("0")},
        )
        pre_bucket[remark] = pre_bucket.get(remark, Decimal("0")) + amount

        corrected_category, record, review = _correct_upload_category(
            document_number,
            category,
            brand,
            amount,
            reference,
            payment_index,
            ambiguous_refs,
        )
        if review is not None:
            reviews.append(review)
        if record is not None and corrected_category != category:
            corrections.append(
                CategoryCorrection(
                    document_number=document_number,
                    reference=reference,
                    brand=brand,
                    original_category=category,
                    corrected_category=corrected_category,
                    raw_payment_category=record.raw_category,
                )
            )

        key = (corrected_category, brand)
        bucket = amounts.setdefault(
            key,
            {"已上传": Decimal("0"), "未上传": Decimal("0")},
        )
        bucket[remark] = bucket.get(remark, Decimal("0")) + amount

    return pre_amounts, amounts, corrections, reviews


def load_upload_data(
    upload_file: Path = UPLOAD_FILE,
    payment_index: dict[str, PaymentDetailRecord] | None = None,
    ambiguous_refs: dict[str, tuple[PaymentDetailRecord, ...]] | None = None,
) -> tuple[
    dict[tuple[str, str], dict[str, Decimal]],
    dict[str, Decimal],
    dict[str, dict[str, int]],
    dict[str, dict[str, Decimal]],
    list[CategoryCorrection],
    list[tuple[str, tuple[str, ...]]],
]:
    """读审核明细的两个工作表：数据汇总（项目行、数码口径、品牌金额基准）
    与家电-明细总表（品牌金额的行级聚合 + 按回款参考号的品类纠正）。

    品牌行的权威来源是明细总表而不是数据汇总——只有明细行才携带交易参考号，
    品类纠正必须发生在行级。但数据汇总的品牌行保留为校验基准：与明细聚合的
    纠正前金额逐键比对，任何一行被意外跳过、状态改名或品牌归一化漂移都会
    在这里报错而不是静默产出偏差报表。
    """
    payment_index = payment_index or {}
    ambiguous_refs = ambiguous_refs or {}
    workbook = _open_source_workbook(upload_file, "审核明细", "审核明细（销售用券情况统计）")
    try:
        rows = _sheet_rows_or_raise(workbook, UPLOAD_SHEET_NAME, upload_file.name, "审核明细")
        _validate_header(rows, UPLOAD_HEADER, upload_file.name, "审核明细")
        detail_rows = _sheet_rows_or_raise(
            workbook,
            coupon_appliance.DETAILS_SHEET_NAME,
            upload_file.name,
            "审核明细",
        )
        _validate_header(
            detail_rows,
            coupon_appliance.COUPON_OUTPUT_HEADER,
            upload_file.name,
            "审核明细",
        )
    finally:
        workbook.close()

    header_width = len(UPLOAD_HEADER)
    current_category = ""
    current_brand = ""
    digital_uploaded = Decimal("0")
    digital_not_uploaded = Decimal("0")
    digital_total: Decimal | None = None
    upload_counts: dict[str, dict[str, int]] = {
        "家电": {},
        "数码": {},
    }
    upload_amounts: dict[str, dict[str, Decimal]] = {
        "家电": {},
        "数码": {},
    }
    summary_amounts: dict[tuple[str, str], dict[str, Decimal]] = {}

    for row in (r[:header_width] for r in rows[1:]):
        category_raw, brand_raw, status_raw, count_raw, amount_raw = row
        if category_raw:
            current_category = normalize_text(category_raw)
            if not brand_raw:
                current_brand = ""
        if brand_raw:
            current_brand = normalize_text(brand_raw)
        if current_category in {"财务大类", "合计"}:
            current_brand = ""
            continue

        status = normalize_text(status_raw)
        amount = to_decimal(amount_raw)

        if current_category in upload_counts and not current_brand:
            if status in {"已上传", "未上传", "合计"}:
                count = to_count(count_raw)
                if status != "合计":
                    upload_counts[current_category][status] = count
                    upload_amounts[current_category][status] = amount
            if current_category == "数码":
                if status == "已上传":
                    digital_uploaded += amount
                elif status == "未上传":
                    digital_not_uploaded += amount
                elif status == "合计":
                    digital_total = amount
            continue

        if not current_brand or status not in ("已上传", "未上传"):
            continue
        key = (current_category, current_brand)
        bucket = summary_amounts.setdefault(
            key,
            {"已上传": Decimal("0"), "未上传": Decimal("0")},
        )
        bucket[status] = bucket.get(status, Decimal("0")) + amount

    pre_amounts, amounts, corrections, reviews = _aggregate_upload_detail(
        detail_rows,
        payment_index,
        ambiguous_refs,
    )

    if pre_amounts != summary_amounts:
        differences = _amount_difference_items(pre_amounts, summary_amounts)
        shown = differences[:5]
        remainder = len(differences) - len(shown)
        suffix = f"；……其余 {remainder} 项" if remainder else ""
        raise ValueError(
            f"{upload_file.name} 审核明细品牌金额不一致，共 {len(differences)} 项"
            f"（明细状态改名、行被跳过或品牌归一化漂移）："
            + "；".join(shown)
            + suffix
        )

    # 品类纠正只改键不改金额：纠正前后各状态总额必须守恒，防止纠正逻辑
    # 造成金额重复或丢失。
    for status in ("已上传", "未上传"):
        pre_total = sum(
            (bucket.get(status, Decimal("0")) for bucket in pre_amounts.values()),
            Decimal("0"),
        )
        post_total = sum(
            (bucket.get(status, Decimal("0")) for bucket in amounts.values()),
            Decimal("0"),
        )
        if pre_total != post_total:
            raise ValueError(
                f"{upload_file.name} 品类纠正前后 {status} 金额不守恒："
                f"纠正前 {pre_total}，纠正后 {post_total}"
            )

    digital_totals = {
        "发生额": digital_total or digital_uploaded + digital_not_uploaded,
        "上传额": digital_uploaded,
    }

    for project in TABLE3_PROJECT_ROWS:
        missing_statuses = {
            status
            for status in ("已上传", "未上传")
            if status not in upload_counts[project]
        }
        if missing_statuses:
            raise ValueError(
                f"{upload_file.name} 的数据汇总缺少{project}项目汇总状态："
                f"{'、'.join(sorted(missing_statuses))}"
            )

    return amounts, digital_totals, upload_counts, upload_amounts, corrections, reviews


def load_payment_data(
    payment_file: Path = PAYMENT_FILE,
) -> tuple[
    dict[tuple[str, str], Decimal],
    Decimal,
    dict[str, int],
    dict[str, Decimal],
    dict[str, PaymentDetailRecord],
    dict[str, tuple[PaymentDetailRecord, ...]],
]:
    """读回款明细：汇总（品牌金额、项目计数）与家电明细（编码品类索引）。

    编码品类索引供 load_upload_data 按交易参考号纠正审核侧品类——回款明细
    的编码品类（A02-电冰箱 → 冰箱）是权威口径，审核明细的财务大类偶有错标。
    """
    workbook = _open_source_workbook(payment_file, "回款明细", "回款明细（家电+数码）")
    try:
        rows = _sheet_rows_or_raise(workbook, PAYMENT_SHEET_NAME, payment_file.name, "回款明细")
        _validate_header(rows, PAYMENT_HEADER, payment_file.name, "回款明细")
        header_width = len(PAYMENT_HEADER)

        amounts: dict[tuple[str, str], Decimal] = {}
        current_category = ""
        digital_amount = Decimal("0")
        payment_counts: dict[str, int] = {
            "家电": 0,
            "数码": 0,
        }

        for row in (r[:header_width] for r in rows[1:]):
            category_raw, brand_raw, amount_raw, count_raw = row
            if category_raw:
                current_category = normalize_text(category_raw)
            category = current_category

            if not category or category == "合计":
                continue
            if not brand_raw:
                continue

            brand = normalize_text(brand_raw)
            amount = to_decimal(amount_raw)
            count = to_count(count_raw)

            if category in HOUSEHOLD_PAYMENT_CATEGORIES:
                key = (category, brand)
                amounts[key] = amounts.get(key, Decimal("0")) + amount
                payment_counts["家电"] += count
            elif category in DIGITAL_PAYMENT_CATEGORIES:
                digital_amount += amount
                payment_counts["数码"] += count
            else:
                raise ValueError(
                    f"门店报表尚未配置回款品类：{category!r}（品牌 {brand!r}）；"
                    "需先确认业务口径（单独一行/并入数码/不纳入报表），"
                    "再把该品类加入 HOUSEHOLD_PAYMENT_CATEGORIES 或 DIGITAL_PAYMENT_CATEGORIES"
                )

        detail_sheet_name = PAYMENT_PROFILES["家电"].detail_sheet_name
        detail_rows = _sheet_rows_or_raise(
            workbook,
            detail_sheet_name,
            payment_file.name,
            "回款明细",
        )
        _validate_header(
            detail_rows,
            PAYMENT_PROFILES["家电"].detail_headers + PAYMENT_DERIVED_HEADERS,
            payment_file.name,
            "回款明细",
        )
        payment_index, ambiguous_refs = _build_payment_detail_index(detail_rows)
    finally:
        workbook.close()

    payment_amounts = {
        "家电": sum(amounts.values(), Decimal("0")),
        "数码": digital_amount,
    }
    return amounts, digital_amount, payment_counts, payment_amounts, payment_index, ambiguous_refs


def sum_upload_amount(
    upload_data: dict[tuple[str, str], dict[str, Decimal]],
    categories: tuple[str, ...],
    brands: tuple[str, ...],
) -> tuple[Decimal, Decimal]:
    if not categories or not brands:
        return Decimal("0"), Decimal("0")

    occurred = Decimal("0")
    uploaded = Decimal("0")
    normalized_categories = {normalize_text(cat) for cat in categories}
    normalized_brands = {normalize_text(brand) for brand in brands}

    for category in normalized_categories:
        for brand in normalized_brands:
            values = upload_data.get((category, brand))
            if not values:
                continue
            uploaded += values.get("已上传", Decimal("0"))
            occurred += values.get("已上传", Decimal("0")) + values.get("未上传", Decimal("0"))

    return occurred, uploaded


def sum_payment_amount(
    payment_data: dict[tuple[str, str], Decimal],
    categories: tuple[str, ...],
    brands: tuple[str, ...],
) -> Decimal:
    if not categories or not brands:
        return Decimal("0")

    total = Decimal("0")
    normalized_categories = {normalize_text(cat) for cat in categories}
    normalized_brands = {normalize_text(brand) for brand in brands}

    for category in normalized_categories:
        for brand in normalized_brands:
            total += payment_data.get((category, brand), Decimal("0"))

    return total


def _rule_claims(
    categories_attr: str, brands_attr: str, business_name: str
) -> dict[tuple[str, str], int]:
    """Map each (品类, 品牌) ROW_RULES claims to the row that claims it.

    A rule listing brand aliases that normalize to the same value (e.g. both
    "华为" and "华为（终端）") claims that key twice, which is redundant but
    harmless — sum_upload_amount/sum_payment_amount already dedupe brands via
    a set. Only a *different* row claiming an already-claimed key is a real
    ROW_RULES config conflict, not a data problem.
    """
    claims: dict[tuple[str, str], int] = {}
    for rule in ROW_RULES:
        categories = getattr(rule, categories_attr)
        if not categories:
            continue
        for category in categories:
            normalized_category = normalize_text(category)
            for brand in getattr(rule, brands_attr):
                key = (normalized_category, normalize_text(brand))
                if key in claims and claims[key] != rule.row:
                    raise ValueError(
                        f"ROW_RULES 配置冲突：{business_name}品类品牌 {key} 同时被第 {claims[key]} 行"
                        f"和第 {rule.row} 行规则匹配"
                    )
                claims[key] = rule.row
    return claims


def validate_rule_coverage(
    upload_data: dict[tuple[str, str], dict[str, Decimal]],
    payment_data: dict[tuple[str, str], Decimal],
) -> None:
    """Every non-zero (品类, 品牌) group in the source data must hit exactly one ROW_RULES row.

    Guards against a new brand appearing upstream (or a category rename) that
    ROW_RULES hasn't been updated for: without this, that brand's money is
    simply never summed into any row, yet every existing total-vs-detail check
    still balances, because both sides of that check are already missing the
    same money.
    """
    upload_claims = _rule_claims("upload_categories", "upload_brands", "审核明细")
    payment_claims = _rule_claims("payment_categories", "payment_brands", "回款明细")

    unmatched: list[str] = [
        f"审核明细未配置规则：{category}/{brand}，发生额 {occurred}"
        for (category, brand), values in upload_data.items()
        for occurred in (values.get("已上传", Decimal("0")) + values.get("未上传", Decimal("0")),)
        if (category, brand) not in upload_claims and occurred != 0
    ]
    unmatched += [
        f"回款明细未配置规则：{category}/{brand}，回款额 {amount}"
        for (category, brand), amount in payment_data.items()
        if (category, brand) not in payment_claims and amount != 0
    ]
    if unmatched:
        raise ValueError(
            "门店报表规则覆盖率校验失败，以下品类/品牌未命中 ROW_RULES，"
            "请确认是否为新增品牌并同步更新规则：" + "；".join(unmatched)
        )


def write_metrics_row(
    sheet,
    row: int,
    occurred: Decimal,
    uploaded: Decimal,
    paid: Decimal,
    font: Font,
    expected_cells: dict[str, object],
) -> None:
    occurred_value = decimal_to_cell_value(occurred)
    uploaded_value = decimal_to_cell_value(uploaded)
    paid_value = decimal_to_cell_value(paid)
    upload_ratio_value = safe_ratio(uploaded, occurred)
    payment_ratio_value = safe_ratio(paid, occurred)

    values = {
        f"D{row}": occurred_value,
        f"E{row}": uploaded_value,
        f"F{row}": upload_ratio_value,
        f"G{row}": paid_value,
        f"H{row}": payment_ratio_value,
    }

    for coordinate, value in values.items():
        sheet[coordinate] = value
        expected_cells[coordinate] = value

    apply_filled_style(
        sheet,
        (f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}"),
        font,
    )
    apply_data_number_format(
        sheet,
        (f"D{row}", f"E{row}", f"G{row}"),
    )
    sheet[f"F{row}"].number_format = PERCENT_NUMBER_FORMAT
    sheet[f"H{row}"].number_format = PERCENT_NUMBER_FORMAT


def update_totals_row(
    sheet,
    amount_columns: tuple[str, ...],
    ratio_columns: dict[str, tuple[str, str]],
    source_rows: range,
    total_row: int,
    font: Font,
    expected_cells: dict[str, object],
) -> None:
    totals = {
        column: sum(
            (to_decimal(sheet[f"{column}{row}"].value) for row in source_rows),
            Decimal("0"),
        )
        for column in amount_columns
    }

    for column, total in totals.items():
        value = decimal_to_cell_value(total)
        sheet[f"{column}{total_row}"] = value
        expected_cells[f"{column}{total_row}"] = value

    for ratio_column, (numerator_column, denominator_column) in ratio_columns.items():
        value = safe_ratio(totals[numerator_column], totals[denominator_column])
        sheet[f"{ratio_column}{total_row}"] = value
        expected_cells[f"{ratio_column}{total_row}"] = value

    apply_filled_style(
        sheet,
        tuple(f"{column}{total_row}" for column in (*amount_columns, *ratio_columns)),
        font,
    )
    apply_data_number_format(
        sheet,
        tuple(f"{column}{total_row}" for column in amount_columns),
    )
    for ratio_column in ratio_columns:
        sheet[f"{ratio_column}{total_row}"].number_format = PERCENT_NUMBER_FORMAT


def write_row(
    sheet,
    row_rule: RowRule,
    upload_data: dict[tuple[str, str], dict[str, Decimal]],
    payment_data: dict[tuple[str, str], Decimal],
    digital_upload: dict[str, Decimal],
    digital_payment: Decimal,
    font: Font,
    expected_cells: dict[str, object],
) -> None:
    if row_rule.fill_digital:
        occurred = digital_upload["发生额"]
        uploaded = digital_upload["上传额"]
        paid = digital_payment
    else:
        occurred, uploaded = sum_upload_amount(
            upload_data, row_rule.upload_categories, row_rule.upload_brands
        )
        paid = sum_payment_amount(
            payment_data, row_rule.payment_categories, row_rule.payment_brands
        )

    write_metrics_row(
        sheet,
        row_rule.row,
        occurred,
        uploaded,
        paid,
        font,
        expected_cells,
    )


def update_totals(sheet, font: Font, expected_cells: dict[str, object]) -> None:
    update_totals_row(
        sheet,
        amount_columns=("D", "E", "G"),
        ratio_columns={
            "F": ("E", "D"),
            "H": ("G", "D"),
        },
        source_rows=DETAIL_ROWS,
        total_row=TOTAL_ROW,
        font=font,
        expected_cells=expected_cells,
    )


def sum_brand_group(
    upload_data: dict[tuple[str, str], dict[str, Decimal]],
    payment_data: dict[tuple[str, str], Decimal],
    categories: tuple[BrandGroupCategory, ...],
) -> tuple[Decimal, Decimal, Decimal]:
    occurred = Decimal("0")
    uploaded = Decimal("0")
    paid = Decimal("0")

    for category in categories:
        cat_occurred, cat_uploaded = sum_upload_amount(
            upload_data, (category.upload_category,), category.brands
        )
        occurred += cat_occurred
        uploaded += cat_uploaded
        paid += sum_payment_amount(
            payment_data, (category.payment_category,), category.brands
        )

    return occurred, uploaded, paid


def write_brand_group_row(
    sheet,
    brand_group_rule: BrandGroupRule,
    upload_data: dict[tuple[str, str], dict[str, Decimal]],
    payment_data: dict[tuple[str, str], Decimal],
    font: Font,
    expected_cells: dict[str, object],
) -> None:
    occurred, uploaded, paid = sum_brand_group(
        upload_data, payment_data, brand_group_rule.categories
    )
    row = brand_group_rule.row
    values = {
        f"D{row}": decimal_to_cell_value(occurred),
        f"E{row}": decimal_to_cell_value(uploaded),
        f"F{row}": decimal_to_cell_value(paid),
        f"G{row}": safe_ratio(uploaded, occurred),
        f"H{row}": safe_ratio(paid, occurred),
    }
    for coordinate, value in values.items():
        sheet[coordinate] = value
        expected_cells[coordinate] = value

    apply_filled_style(
        sheet,
        (f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}"),
        font,
    )
    apply_data_number_format(
        sheet,
        (f"D{row}", f"E{row}", f"F{row}"),
    )
    sheet[f"G{row}"].number_format = PERCENT_NUMBER_FORMAT
    sheet[f"H{row}"].number_format = PERCENT_NUMBER_FORMAT


def update_brand_group_totals(sheet, font: Font, expected_cells: dict[str, object]) -> None:
    update_totals_row(
        sheet,
        amount_columns=("D", "E", "F"),
        ratio_columns={
            "G": ("E", "D"),
            "H": ("F", "D"),
        },
        source_rows=BRAND_GROUP_DETAIL_ROWS,
        total_row=BRAND_GROUP_TOTAL_ROW,
        font=font,
        expected_cells=expected_cells,
    )


def write_table3(
    sheet,
    upload_counts: dict[str, dict[str, int]],
    upload_amounts: dict[str, dict[str, Decimal]],
    payment_counts: dict[str, int],
    payment_amounts: dict[str, Decimal],
    font: Font,
    expected_cells: dict[str, object],
) -> None:
    pending_count_total = 0
    pending_amount_total = Decimal("0")
    not_uploaded_count_total = 0
    not_uploaded_amount_total = Decimal("0")
    for project, row in TABLE3_PROJECT_ROWS.items():
        pending_count = (
            upload_counts[project]["已上传"]
            - payment_counts[project]
        )
        pending_amount = (
            upload_amounts[project]["已上传"]
            - payment_amounts[project]
        )
        not_uploaded_count = upload_counts[project]["未上传"]
        not_uploaded_amount = upload_amounts[project]["未上传"]

        values = {
            f"D{row}": pending_count or None,
            f"E{row}": decimal_to_cell_value(pending_amount),
            f"F{row}": not_uploaded_count or None,
            f"G{row}": decimal_to_cell_value(not_uploaded_amount),
        }
        for coordinate, value in values.items():
            sheet[coordinate] = value
            expected_cells[coordinate] = value
            sheet[coordinate].font = font

        apply_data_number_format(
            sheet,
            (f"D{row}", f"E{row}", f"F{row}", f"G{row}"),
        )

        pending_count_total += pending_count
        pending_amount_total += pending_amount
        not_uploaded_count_total += not_uploaded_count
        not_uploaded_amount_total += not_uploaded_amount

    total_values = {
        f"D{TABLE3_TOTAL_ROW}": pending_count_total or None,
        f"E{TABLE3_TOTAL_ROW}": decimal_to_cell_value(pending_amount_total),
        f"F{TABLE3_TOTAL_ROW}": not_uploaded_count_total or None,
        f"G{TABLE3_TOTAL_ROW}": decimal_to_cell_value(not_uploaded_amount_total),
    }
    for coordinate, value in total_values.items():
        sheet[coordinate] = value
        expected_cells[coordinate] = value
        sheet[coordinate].font = font

    apply_data_number_format(
        sheet,
        (
            f"D{TABLE3_TOTAL_ROW}",
            f"E{TABLE3_TOTAL_ROW}",
            f"F{TABLE3_TOTAL_ROW}",
            f"G{TABLE3_TOTAL_ROW}",
        ),
    )


def current_timestamp() -> datetime:
    return datetime.now(SHANGHAI_TZ)


def update_header(sheet, timestamp: datetime, expected_cells: dict[str, object]) -> None:
    _validate_header_title(sheet["A1"].value, "空白模板")
    formatted_timestamp = timestamp.strftime("%Y-%m-%d %H:%M:%S")
    value = f"{TEMPLATE_HEADER_PREFIX}{formatted_timestamp}"
    sheet["A1"] = value
    expected_cells["A1"] = value


def _values_match(expected: object, actual: object) -> bool:
    if expected is None or actual is None:
        return expected == actual
    if isinstance(expected, float) or isinstance(actual, float):
        try:
            # 只要有一边是 float 就走数值比较；另一边可能是 str 或别的
            # 类型，靠下面的 TypeError 兜住——这是有意为之的控制流，
            # 不是漏判，所以在这里显式放行 mypy。
            return math.isclose(float(expected), float(actual), rel_tol=1e-9, abs_tol=1e-9)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return False
    return expected == actual


def _validate_totals_match_details(sheet, path_name: str) -> None:
    # 表1
    for column in ("D", "E", "G"):
        recomputed = sum(
            (to_decimal(sheet[f"{column}{row}"].value) for row in DETAIL_ROWS),
            Decimal("0"),
        )
        actual = to_decimal(sheet[f"{column}{TOTAL_ROW}"].value)
        if recomputed != actual:
            raise ValueError(
                f"{path_name} 第 {TOTAL_ROW} 行合计校验失败："
                f"{column}{TOTAL_ROW} 应为 {recomputed}，实际为 {actual}"
            )

    # 表2
    for column in ("D", "E", "F"):
        recomputed = sum(
            (to_decimal(sheet[f"{column}{row}"].value) for row in BRAND_GROUP_DETAIL_ROWS),
            Decimal("0"),
        )
        actual = to_decimal(sheet[f"{column}{BRAND_GROUP_TOTAL_ROW}"].value)
        if recomputed != actual:
            raise ValueError(
                f"{path_name} 第 {BRAND_GROUP_TOTAL_ROW} 行合计校验失败："
                f"{column}{BRAND_GROUP_TOTAL_ROW} 应为 {recomputed}，实际为 {actual}"
            )

    # 表3 独立守恒校验
    d_appliance = to_count(sheet[f"D{TABLE3_PROJECT_ROWS['家电']}"].value)
    d_digital = to_count(sheet[f"D{TABLE3_PROJECT_ROWS['数码']}"].value)
    d_total = to_count(sheet[f"D{TABLE3_TOTAL_ROW}"].value)
    if d_total != d_appliance + d_digital:
        raise ValueError(
            f"{path_name} 第 {TABLE3_TOTAL_ROW} 行审核中数量合计校验失败："
            f"D{TABLE3_TOTAL_ROW} 应为 {d_appliance + d_digital}，实际为 {d_total}"
        )

    f_appliance = to_count(sheet[f"F{TABLE3_PROJECT_ROWS['家电']}"].value)
    f_digital = to_count(sheet[f"F{TABLE3_PROJECT_ROWS['数码']}"].value)
    f_total = to_count(sheet[f"F{TABLE3_TOTAL_ROW}"].value)
    if f_total != f_appliance + f_digital:
        raise ValueError(
            f"{path_name} 第 {TABLE3_TOTAL_ROW} 行未上传数量合计校验失败："
            f"F{TABLE3_TOTAL_ROW} 应为 {f_appliance + f_digital}，实际为 {f_total}"
        )

    e_appliance = to_decimal(sheet[f"E{TABLE3_PROJECT_ROWS['家电']}"].value)
    e_digital = to_decimal(sheet[f"E{TABLE3_PROJECT_ROWS['数码']}"].value)
    e_total = to_decimal(sheet[f"E{TABLE3_TOTAL_ROW}"].value)
    if e_total != e_appliance + e_digital:
        raise ValueError(
            f"{path_name} 第 {TABLE3_TOTAL_ROW} 行审核中金额合计校验失败："
            f"E{TABLE3_TOTAL_ROW} 应为 {e_appliance + e_digital}，实际为 {e_total}"
        )

    g_appliance = to_decimal(sheet[f"G{TABLE3_PROJECT_ROWS['家电']}"].value)
    g_digital = to_decimal(sheet[f"G{TABLE3_PROJECT_ROWS['数码']}"].value)
    g_total = to_decimal(sheet[f"G{TABLE3_TOTAL_ROW}"].value)
    if g_total != g_appliance + g_digital:
        raise ValueError(
            f"{path_name} 第 {TABLE3_TOTAL_ROW} 行未上传金额合计校验失败："
            f"G{TABLE3_TOTAL_ROW} 应为 {g_appliance + g_digital}，实际为 {g_total}"
        )


def _validate_ratios_match_totals(sheet, path_name: str) -> None:
    """Independently recompute each ratio cell from the totals actually saved in
    the file (not from the in-memory value written), catching a numerator/
    denominator mixup that expected-cell comparison alone would miss, since
    that comparison reuses the same safe_ratio() call that wrote the cell."""
    checks = (
        (TOTAL_ROW, {"F": ("E", "D"), "H": ("G", "D")}),
        (BRAND_GROUP_TOTAL_ROW, {"G": ("E", "D"), "H": ("F", "D")}),
    )
    for total_row, ratio_columns in checks:
        for ratio_column, (numerator_column, denominator_column) in ratio_columns.items():
            numerator = to_decimal(sheet[f"{numerator_column}{total_row}"].value)
            denominator = to_decimal(sheet[f"{denominator_column}{total_row}"].value)
            expected_ratio = safe_ratio(numerator, denominator)
            actual_ratio = sheet[f"{ratio_column}{total_row}"].value
            if not _values_match(expected_ratio, actual_ratio):
                raise ValueError(
                    f"{path_name} {ratio_column}{total_row} 与其对应金额之比不符："
                    f"应为 {expected_ratio}，实际为 {actual_ratio}"
                )


def validate_output(
    path: Path,
    expected_cells: dict[str, object],
    expected_sheet_title: str,
) -> None:
    """Re-read the saved workbook and check it matches what process_store_report intended:
    sheet identity, every written cell exactly, and totals/ratios independently recomputed."""
    workbook = load_workbook(path, data_only=True)
    try:
        if len(workbook.sheetnames) != EXPECTED_SHEET_COUNT:
            raise ValueError(
                f"{path.name} 工作表数量应为 {EXPECTED_SHEET_COUNT}，实际为 {len(workbook.sheetnames)}"
            )
        sheet = workbook[workbook.sheetnames[0]]
        if sheet.title != expected_sheet_title:
            raise ValueError(f"{path.name} 工作表名称应为 {expected_sheet_title!r}，实际为 {sheet.title!r}")
        if sheet.max_column != EXPECTED_COLUMN_COUNT:
            raise ValueError(
                f"{path.name} 列数应为 {EXPECTED_COLUMN_COUNT}，实际为 {sheet.max_column}"
            )
        _validate_header_title(sheet["A1"].value, path.name)

        mismatches = [
            f"{coordinate}（应为 {expected!r}，实际为 {sheet[coordinate].value!r}）"
            for coordinate, expected in expected_cells.items()
            if not _values_match(expected, sheet[coordinate].value)
        ]
        if mismatches:
            raise ValueError(
                f"{path.name} 明细单元格校验失败，共 {len(mismatches)} 处不一致，"
                f"示例：{'；'.join(mismatches[:5])}"
            )

        _validate_totals_match_details(sheet, path.name)
        _validate_ratios_match_totals(sheet, path.name)
    finally:
        workbook.close()


def report_category_corrections(
    reporter: ConsoleReporter,
    corrections: list[CategoryCorrection],
) -> None:
    """One consolidated corrected() block: every executed 品类纠正 with its 依据."""
    if not corrections:
        return
    details: list[str] = []
    for correction in corrections:
        details.append(
            f"{correction.document_number}｜{correction.reference}｜"
            f"{correction.brand}｜{correction.original_category}"
            f" → {correction.corrected_category}"
        )
        details.append(
            f"依据：回款明细编码品类 {correction.raw_payment_category}"
        )
    reporter.corrected(f"审核明细品类纠正：{len(corrections)} 条", tuple(details))


def process_store_report(reporter: ConsoleReporter) -> None:
    timestamp = current_timestamp()
    font_name, _ = resolve_font()
    payment_data, digital_payment, payment_counts, payment_amounts, payment_index, ambiguous_refs = (
        load_payment_data(PAYMENT_FILE)
    )
    (
        upload_data,
        digital_upload,
        upload_counts,
        upload_amounts,
        category_corrections,
        review_items,
    ) = load_upload_data(UPLOAD_FILE, payment_index, ambiguous_refs)
    validate_rule_coverage(upload_data, payment_data)
    report_category_corrections(reporter, category_corrections)
    for title, details in review_items:
        reporter.review_required(title, details)

    template_file = resolve_template_file()
    workbook = load_workbook(template_file)
    try:
        validate_template(workbook)
        sheet = workbook[workbook.sheetnames[0]]
        font = Font(name=font_name, size=12)

        expected_cells: dict[str, object] = dict(TEMPLATE_STRUCTURE_CELLS)

        for row_rule in ROW_RULES:
            write_row(
                sheet, row_rule, upload_data, payment_data, digital_upload, digital_payment, font, expected_cells
            )

        update_totals(sheet, font, expected_cells)

        for brand_group_rule in BRAND_GROUP_RULES:
            write_brand_group_row(sheet, brand_group_rule, upload_data, payment_data, font, expected_cells)

        update_brand_group_totals(sheet, font, expected_cells)
        write_table3(
            sheet,
            upload_counts,
            upload_amounts,
            payment_counts,
            payment_amounts,
            font,
            expected_cells,
        )
        update_header(sheet, timestamp, expected_cells)
        apply_report_font(sheet, font_name)
    except BaseException:
        # save_workbook_atomically (below) takes ownership of closing the
        # workbook once writing succeeds; anything raised before that point —
        # a bad template, a write bug — must close it here instead, or the
        # template file stays open until GC.
        workbook.close()
        raise

    save_workbook_atomically(
        workbook,
        OUTPUT_FILE,
        lambda path: validate_output(path, expected_cells, sheet.title),
    )
    reporter.output(OUTPUT_FILE)
