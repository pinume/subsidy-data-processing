from datetime import date, datetime
from pathlib import Path

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from processors.common.config import load_receipt_special_remark_keys
from processors.common.excel import (
    create_sheet_styles,
    load_measurement_font,
    pixels_to_excel_width,
    resolve_font,
    save_workbook_atomically,
    width_measurer,
)
from processors.common.dates import (
    is_valid_original_invoice_number,
    normalize_document_number,
    normalize_receipt_date,
    normalize_receipt_identifier,
    receipt_match_key,
)

from . import _shared
from ._shared import RECEIPTS_OUTPUT_FILE


RECEIPTS_SOURCE_HEADER = ("单据号", "日期", "原票号", "摘要", "商品名称")
RECEIPTS_OUTPUT_HEADER = (*RECEIPTS_SOURCE_HEADER, "备注")
RECEIPTS_REMARK_RETURN = "退换货/倒票（退单）"
RECEIPTS_REMARK_ORIGINAL = "退换货/倒票（原单）"
RECEIPTS_REMARK_BOTH = "退换货/倒票（退单及原单）"
RECEIPTS_REMARK_SAME_MODEL_REPLACEMENT = "已做同型号换货处理"
RECEIPTS_REMARK_SPECIAL = r"退换货\倒票"
RECEIPTS_SPECIAL_REMARK_KEYS = load_receipt_special_remark_keys()
RECEIPTS_ROW_HEIGHT = 20
RECEIPTS_DUPLICATE_FILL_COLOR = "FFC7CE"
RECEIPTS_EXCLUDED_PRODUCT_KEYWORD = "北国"
RECEIPTS_SAME_MODEL_REPLACEMENT_KEYWORD = "同型号换货"


def read_receipt_rows(source: Path) -> list[list[object]]:
    """Read legacy XLS files in pure Python without invoking Excel."""
    source_workbook = xlrd.open_workbook(source)
    try:
        source_sheet = source_workbook.sheet_by_index(0)
        if source_sheet.nrows < 2:
            raise ValueError(f"{source.name} 缺少总标题行或字段标题行")
        source_headers = [
            str(source_sheet.cell_value(1, column_index)).strip()
            for column_index in range(source_sheet.ncols)
        ]
        missing_headers = [
            header for header in RECEIPTS_SOURCE_HEADER
            if header not in source_headers
        ]
        if missing_headers:
            raise ValueError(
                f"{source.name} 缺少必要字段：{tuple(missing_headers)}；"
                f"实际字段 {tuple(source_headers)}"
            )
        source_column_indexes = [
            source_headers.index(header) for header in RECEIPTS_SOURCE_HEADER
        ]

        rows: list[list[object]] = []
        for row_index in range(1, source_sheet.nrows):
            row: list[object] = []
            for column_index in source_column_indexes:
                cell = source_sheet.cell(row_index, column_index)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(
                        value,
                        source_workbook.datemode,
                    )
                elif cell.ctype == xlrd.XL_CELL_NUMBER and value == int(value):
                    value = int(value)
                row.append(value)
            if row_index > 1 and str(row[1]).strip() == "合计":
                continue
            rows.append(row)

        actual_header = tuple(
            str(value).strip() if value is not None else ""
            for value in rows[0]
        )
        if actual_header != RECEIPTS_SOURCE_HEADER:
            raise ValueError(
                f"{source.name} 字段标题不一致："
                f"预期 {RECEIPTS_SOURCE_HEADER}，实际 {actual_header}"
            )
        return rows
    finally:
        source_workbook.release_resources()


def receipt_remark(
    has_original: bool,
    is_referenced: bool,
    is_same_model_replacement: bool = False,
    is_special: bool = False,
) -> str | None:
    if is_special:
        return RECEIPTS_REMARK_SPECIAL
    if has_original and is_referenced:
        return RECEIPTS_REMARK_BOTH
    if is_same_model_replacement:
        return RECEIPTS_REMARK_SAME_MODEL_REPLACEMENT
    if has_original:
        return RECEIPTS_REMARK_RETURN
    if is_referenced:
        return RECEIPTS_REMARK_ORIGINAL
    return None


def prepare_receipt_data(kept_rows: list[list[object]]):
    records: list[dict[str, object]] = []
    key_rows: dict[str, list[int]] = {}
    original_invoice_numbers: set[str] = set()
    referenced_original_invoice_numbers: set[str] = set()
    same_model_replacement_original_invoice_numbers: set[str] = set()
    excluded_product_count = 0

    for source_row, row in enumerate(kept_rows[1:], start=3):
        product_name = normalize_receipt_identifier(row[4])
        if RECEIPTS_EXCLUDED_PRODUCT_KEYWORD in product_name:
            excluded_product_count += 1
            continue

        document_number = normalize_document_number(row[0])
        receipt_date = normalize_receipt_date(row[1], source_row=source_row)
        original_invoice_number = normalize_receipt_identifier(row[2])
        match_key = (
            receipt_match_key(receipt_date, document_number)
            if receipt_date is not None and document_number
            else ""
        )
        if match_key:
            key_rows.setdefault(match_key, []).append(source_row)
        if original_invoice_number:
            original_invoice_numbers.add(original_invoice_number)
            if RECEIPTS_SAME_MODEL_REPLACEMENT_KEYWORD in (
                normalize_receipt_identifier(row[3])
            ):
                same_model_replacement_original_invoice_numbers.add(
                    original_invoice_number
                )
            else:
                referenced_original_invoice_numbers.add(
                    original_invoice_number
                )
        records.append(
            {
                "source_row": source_row,
                "document_number": document_number,
                "receipt_date": receipt_date,
                "original_invoice_number": original_invoice_number,
                "summary": row[3],
                "product_name": product_name or None,
                "match_key": match_key,
            }
        )

    issues: list[tuple[str, str, str, str]] = []
    for match_key, source_rows in key_rows.items():
        if len(source_rows) > 1:
            issues.append(
                (
                    "重复匹配键",
                    "、".join(str(row) for row in source_rows),
                    match_key,
                    "多个数据行生成了相同的日期与单据号组合键",
                )
            )

    only_return_count = 0
    only_original_count = 0
    both_count = 0
    unmatched_original_count = 0
    invalid_original_count = 0
    missing_match_key_count = 0
    output_rows: list[list[object]] = []
    for record in records:
        source_row = int(record["source_row"])
        original_invoice_number = str(record["original_invoice_number"])
        match_key = str(record["match_key"])
        has_original = bool(original_invoice_number)
        is_referenced = bool(
            match_key and match_key in referenced_original_invoice_numbers
        )
        is_same_model_replacement = bool(
            match_key
            and match_key
            in same_model_replacement_original_invoice_numbers
            and not is_referenced
        )
        is_special = match_key in RECEIPTS_SPECIAL_REMARK_KEYS

        if not match_key:
            missing_match_key_count += 1
            issues.append(
                (
                    "缺少匹配键",
                    str(source_row),
                    "",
                    "日期或单据号为空，无法生成匹配键",
                )
            )
        if has_original and not is_valid_original_invoice_number(
            original_invoice_number
        ):
            invalid_original_count += 1
            issues.append(
                (
                    "原票号格式异常",
                    str(source_row),
                    original_invoice_number,
                    "原票号应为6位日期加单据号",
                )
            )
        if has_original and original_invoice_number not in key_rows:
            unmatched_original_count += 1
            issues.append(
                (
                    "原票号未匹配",
                    str(source_row),
                    original_invoice_number,
                    "未找到日期与单据号组合键相同的原单",
                )
            )

        if has_original and (is_referenced or is_same_model_replacement):
            both_count += 1
        elif has_original:
            only_return_count += 1
        elif is_referenced or is_same_model_replacement:
            only_original_count += 1

        output_rows.append(
            [
                record["document_number"],
                record["receipt_date"],
                original_invoice_number or None,
                record["summary"],
                record["product_name"],
                receipt_remark(
                    has_original,
                    is_referenced,
                    is_same_model_replacement,
                    is_special,
                ),
            ]
        )

    stats = {
        "总数据量": len(records),
        "删除北国商品行数": excluded_product_count,
        "仅退单数量": only_return_count,
        "仅原单数量": only_original_count,
        "退单及原单数量": both_count,
        "备注总数": only_return_count + only_original_count + both_count,
        "未匹配原票号数量": unmatched_original_count,
        "重复匹配键数量": sum(
            1 for source_rows in key_rows.values() if len(source_rows) > 1
        ),
        "原票号格式异常数量": invalid_original_count,
        "缺少匹配键数量": missing_match_key_count,
    }
    duplicate_match_keys = {
        match_key
        for match_key, source_rows in key_rows.items()
        if len(source_rows) > 1
    }
    return output_rows, stats, issues, duplicate_match_keys


def validate_receipts_output(
    path: Path,
    expected_data_rows: int,
) -> None:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        if workbook.sheetnames != ["Sheet1"]:
            raise RuntimeError(
                f"收款单工作表校验失败：实际工作表为 {workbook.sheetnames}"
            )

        sheet = workbook["Sheet1"]
        if sheet.max_row - 1 != expected_data_rows:
            raise RuntimeError(
                f"收款单行数校验失败：预期 {expected_data_rows} 条，"
                f"实际 {sheet.max_row - 1} 条"
            )
        if sheet.max_column != len(RECEIPTS_OUTPUT_HEADER):
            raise RuntimeError(
                f"收款单列数校验失败：预期 {len(RECEIPTS_OUTPUT_HEADER)} 列，"
                f"实际 {sheet.max_column} 列"
            )

        header = tuple(cell.value for cell in sheet[1])
        if header != RECEIPTS_OUTPUT_HEADER:
            raise RuntimeError(
                f"收款单表头校验失败：预期 {RECEIPTS_OUTPUT_HEADER}，"
                f"实际 {header}"
            )

        referenced_original_invoice_numbers = {
            normalize_receipt_identifier(sheet.cell(row_number, 3).value)
            for row_number in range(2, sheet.max_row + 1)
            if (
                sheet.cell(row_number, 3).value not in (None, "")
                and normalize_receipt_identifier(
                    sheet.cell(row_number, 4).value
                )
                .find(RECEIPTS_SAME_MODEL_REPLACEMENT_KEYWORD)
                == -1
            )
        }
        same_model_replacement_original_invoice_numbers = {
            normalize_receipt_identifier(sheet.cell(row_number, 3).value)
            for row_number in range(2, sheet.max_row + 1)
            if (
                sheet.cell(row_number, 3).value not in (None, "")
                and normalize_receipt_identifier(
                    sheet.cell(row_number, 4).value
                )
                .find(RECEIPTS_SAME_MODEL_REPLACEMENT_KEYWORD)
                != -1
            )
        }
        match_key_counts: dict[str, int] = {}
        for row_number in range(2, sheet.max_row + 1):
            document_number = normalize_receipt_identifier(
                sheet.cell(row_number, 1).value
            )
            date_value = sheet.cell(row_number, 2).value
            if date_value is None or not document_number:
                continue
            match_key = receipt_match_key(
                date_value.date()
                if isinstance(date_value, datetime)
                else date_value,
                document_number,
            )
            match_key_counts[match_key] = match_key_counts.get(match_key, 0) + 1

        for row_number in range(2, sheet.max_row + 1):
            document_number = sheet.cell(row_number, 1).value
            if document_number is not None and (
                not isinstance(document_number, str)
                or document_number.startswith("收款")
            ):
                raise RuntimeError(
                    f"收款单第 {row_number} 行的单据号格式不正确"
                )

            date_cell = sheet.cell(row_number, 2)
            if date_cell.value is not None:
                if not isinstance(date_cell.value, (date, datetime)):
                    raise RuntimeError(
                        f"收款单第 {row_number} 行的日期不是有效日期"
                    )
                if date_cell.number_format != "yyyymmdd":
                    raise RuntimeError(
                        f"收款单第 {row_number} 行的日期格式不正确"
                    )

            original_invoice_number = normalize_receipt_identifier(
                sheet.cell(row_number, 3).value
            )
            match_key = (
                receipt_match_key(
                    date_cell.value.date()
                    if isinstance(date_cell.value, datetime)
                    else date_cell.value,
                    normalize_receipt_identifier(document_number),
                )
                if date_cell.value is not None and document_number is not None
                else ""
            )
            expected_remark = receipt_remark(
                bool(original_invoice_number),
                bool(
                    match_key
                    and match_key in referenced_original_invoice_numbers
                ),
                bool(
                    match_key
                    and match_key
                    in same_model_replacement_original_invoice_numbers
                    and match_key
                    not in referenced_original_invoice_numbers
                ),
                match_key in RECEIPTS_SPECIAL_REMARK_KEYS,
            )
            if sheet.cell(row_number, 6).value != expected_remark:
                raise RuntimeError(
                    f"收款单第 {row_number} 行的备注校验失败"
                )

            is_duplicate = bool(match_key and match_key_counts[match_key] > 1)
            for cell in sheet[row_number]:
                fill_color = cell.fill.fgColor.rgb
                is_pink = (
                    cell.fill.fill_type == "solid"
                    and fill_color is not None
                    and fill_color[-6:] == RECEIPTS_DUPLICATE_FILL_COLOR[-6:]
                )
                if is_pink != is_duplicate:
                    raise RuntimeError(
                        f"收款单第 {row_number} 行的重复匹配键标记不正确"
                    )
    finally:
        workbook.close()


def process_receipts() -> None:
    if _shared.RECEIPTS_SOURCE_FILE is None:
        raise FileNotFoundError(
            f"未在 {_shared.DATA_DIR} 中找到文件名包含"
            f"“{_shared.RECEIPT_STATISTICS_KEYWORD}”的 .XLS 文件"
        )

    kept_rows = read_receipt_rows(_shared.RECEIPTS_SOURCE_FILE)
    output_rows, stats, _issues, duplicate_match_keys = prepare_receipt_data(
        kept_rows
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(RECEIPTS_OUTPUT_HEADER)
    for row in output_rows:
        sheet.append(row)

    font_name, font_path = resolve_font()
    measurement_font = load_measurement_font(font_path)
    normal_font, header_font, header_fill, centered = create_sheet_styles(
        font_name
    )

    measure = width_measurer(measurement_font)
    maximum_widths = [0.0] * sheet.max_column
    duplicate_fill = PatternFill(
        "solid",
        fgColor=RECEIPTS_DUPLICATE_FILL_COLOR,
    )
    for row_number, row in enumerate(sheet.iter_rows(), start=1):
        sheet.row_dimensions[row_number].height = RECEIPTS_ROW_HEIGHT
        row_match_key = ""
        if row_number >= 2:
            document_number = normalize_receipt_identifier(row[0].value)
            receipt_date = row[1].value
            if receipt_date is not None and document_number:
                row_match_key = receipt_match_key(
                    receipt_date.date()
                    if isinstance(receipt_date, datetime)
                    else receipt_date,
                    document_number,
                )
        for cell in row:
            cell.font = header_font if row_number == 1 else normal_font
            cell.alignment = centered
            if row_number == 1:
                cell.fill = header_fill
            elif row_match_key in duplicate_match_keys:
                cell.fill = duplicate_fill

            if row_number >= 2 and cell.column == 1:
                cell.number_format = "@"
            elif row_number >= 2 and cell.column == 2:
                if cell.value not in (None, ""):
                    cell.number_format = "yyyymmdd"

            width = measure(cell.value)
            if width > maximum_widths[cell.column - 1]:
                maximum_widths[cell.column - 1] = width

    sheet.freeze_panes = "A2"
    for column_index, maximum_pixels in enumerate(maximum_widths, start=1):
        column_letter = sheet.cell(1, column_index).column_letter
        sheet.column_dimensions[column_letter].width = min(
            pixels_to_excel_width(maximum_pixels),
            255,
        )

    row_count = len(output_rows)
    save_workbook_atomically(
        workbook,
        RECEIPTS_OUTPUT_FILE,
        lambda path: validate_receipts_output(path, row_count),
    )
    print(f"Receipt statistics complete: {row_count} rows")
    print(
        f"Remarks: {stats['备注总数']}; "
        f"unmatched original invoices: {stats['未匹配原票号数量']}; "
        f"duplicate match keys: {stats['重复匹配键数量']}"
    )
    print(f"Output file: {RECEIPTS_OUTPUT_FILE}")
