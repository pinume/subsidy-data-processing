"""Merges the 家电 and 数码 coupon pipelines into one 审核明细.xlsx workbook.

Kept as its own module (rather than living inside either project) because
each project's package needs to refer to the other's output file or menu
entry, and importing across processors.coupons.digital <->
processors.coupons.appliance at module load time would be circular. This
module imports both at the top level; main.py only ever reaches it through a
function-local import inside build_processors(), which runs long after both
packages have finished loading.

Re-exports SUMMARY_SHEET_NAME/SUMMARY_HEADER from
processors.coupons.report_contract so store_report.py can depend on this
module's public surface instead of reaching into
processors.coupons.appliance's internals.
"""

from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from python_calamine import CalamineWorkbook

from processors.common.excel import format_sheet, save_workbook_atomically
from processors.coupons import appliance, digital, matching, sources
from processors.coupons.report_contract import SUMMARY_HEADER, SUMMARY_SHEET_NAME
from processors.coupons.sources import load_coupon_remark_lookup

# Not unused despite the lack of a local reference: re-exported for
# store_report.py, which imports SUMMARY_SHEET_NAME/SUMMARY_HEADER from this
# module rather than reaching into processors.coupons.report_contract itself.
# Listing them in __all__ (rather than an "as"-aliased import) tells pyflakes
# the same thing without tripping pylint's redundant-alias rule.
__all__ = ["SUMMARY_HEADER", "SUMMARY_SHEET_NAME"]


BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_FILE = BASE_DIR / "output" / "审核明细.xlsx"

# Defined in appliance.py because its group-sheet titles must avoid colliding
# with it; this module owns what actually goes into the sheet.
REFERENCE_REPORT_SHEET = appliance.REFERENCE_REPORT_SHEET_NAME
REFERENCE_REPORT_HEADER = (
    "项目",
    "处理结果",
    "单据号",
    "单据日期",
    "原参考号",
    "说明",
)
# Corrections first: they are the rows that were actually rewritten and most
# need a human to confirm. Conflicts next, then the leftovers.
REFERENCE_REPORT_ORDER = matching.REFERENCE_REPORT_ORDER
REPORT_PROJECT_ORDER = ("家电", "数码")
# 财务大类 for digital's 已上传/未上传/合计 block at the foot of 数据汇总.
DIGITAL_SUMMARY_PROJECT_LABEL = "数码"


def merged_reference_decisions(
    appliance_computation: appliance.CouponComputation,
    digital_computation: digital.CouponComputation,
) -> list[tuple[object, ...]]:
    """Label each project's decisions and interleave them into one report.

    Digital's decisions used to be computed and then thrown away, leaving its
    automatic corrections with no audit trail at all.
    """
    decisions = [
        ("家电", *decision)
        for decision in appliance_computation.reference_decisions
    ] + [
        ("数码", *decision)
        for decision in digital_computation.reference_decisions
    ]
    decisions.sort(
        key=lambda decision: (
            REFERENCE_REPORT_ORDER[decision[1]],
            REPORT_PROJECT_ORDER.index(decision[0]),
        )
    )
    return decisions


def build_reference_report_sheet(
    workbook: Workbook,
    decisions: list[tuple[object, ...]],
    font_name: str,
    measurement_font: object,
) -> None:
    report_sheet = workbook.create_sheet(REFERENCE_REPORT_SHEET)
    report_sheet.append(REFERENCE_REPORT_HEADER)
    for decision in decisions:
        report_sheet.append(decision)
    format_sheet(report_sheet, font_name, measurement_font, ("说明",))
    document_column = REFERENCE_REPORT_HEADER.index("单据号") + 1
    date_column = REFERENCE_REPORT_HEADER.index("单据日期") + 1
    for row_number in range(2, report_sheet.max_row + 1):
        report_sheet.cell(row_number, document_column).number_format = "@"
        date_cell = report_sheet.cell(row_number, date_column)
        if date_cell.value not in (None, ""):
            date_cell.number_format = "yyyy-mm-dd"


def report_source_total_gap(
    label: str,
    source_total: Decimal | None,
    computed_total: Decimal,
) -> None:
    """Warn when a coupon export's own 合计 row disagrees with its detail rows.

    Both numbers are reported rather than reconciled: the export is a snapshot,
    and a return recorded after the detail rows were written shows up here as a
    gap that closes by itself once the next export includes it. Silently
    trusting either number would hide that.
    """
    if source_total is None or source_total == computed_total:
        return
    print(
        f"[{label}] WARNING: 源文件合计行为 {source_total:,.2f}，"
        f"但其明细行合计为 {computed_total:,.2f}，"
        f"相差 {computed_total - source_total:,.2f}。"
        "报表采用明细行合计以与明细总表保持一致；"
        "该差额通常是导出快照期间产生的退货尚未写入明细，"
        "下次导出补齐后此提示会自动消失"
    )


def digital_extra_summary_rows(
    digital_computation: digital.CouponComputation,
) -> list[tuple[object, ...]]:
    """Recast digital's 3-column summary rows (备注, 数量, 合计) as rows in
    家电's 5-column 数据汇总 table, labeling every row (including digital's
    own "合计" row) with 财务大类="数码" and no 品牌, so the block mirrors the
    家电 one that precedes it (see appliance.COUPON_SUMMARY_PROJECT_LABEL)."""
    return [
        (DIGITAL_SUMMARY_PROJECT_LABEL, None, remark, count, total)
        for remark, count, total in digital_computation.summary_rows
    ]


def process_coupon_sales() -> None:
    coupon_source = sources.COUPON_SOURCE_FILE
    if coupon_source is None:
        # compute_coupon_data() raises the same FileNotFoundError once it
        # sees sources.COUPON_SOURCE_FILE is None; raising it here directly
        # names the file operators are missing without pretending a second
        # (unreachable) code path could still succeed after the first raises.
        raise FileNotFoundError(
            f"未在 {sources.DATA_DIR} 中找到文件名包含"
            f"“{sources.COUPON_STATISTICS_KEYWORD}”且表头符合"
            "家电、数码用券导出格式的 .XLSX 文件"
        )

    source_workbook = CalamineWorkbook.from_path(str(coupon_source))
    try:
        export = sources.read_coupon_export(coupon_source, source_workbook)
    finally:
        source_workbook.close()

    remark_lookup = load_coupon_remark_lookup(appliance.COUPON_REMARK_SOURCE_FILE)
    appliance_computation = appliance.compute_coupon_data(
        rows=export.appliance_rows,
        remark_lookup=remark_lookup,
        source_total=export.source_total,
    )
    digital_computation = digital.compute_coupon_data(
        rows=export.digital_rows,
        remark_lookup=remark_lookup,
    )
    extra_summary_rows = digital_extra_summary_rows(digital_computation)

    workbook = Workbook()
    workbook.remove(workbook.active)

    font_name, measurement_font, matched_fill = (
        appliance.build_summary_and_details_sheets(
            workbook,
            appliance_computation,
            extra_summary_rows,
        )
    )
    digital.build_detail_sheet(workbook, digital_computation)
    appliance.build_group_sheets(
        workbook,
        appliance_computation,
        font_name,
        measurement_font,
        matched_fill,
    )
    decisions = merged_reference_decisions(
        appliance_computation,
        digital_computation,
    )
    build_reference_report_sheet(
        workbook,
        decisions,
        font_name,
        measurement_font,
    )

    save_workbook_atomically(
        workbook,
        OUTPUT_FILE,
        lambda path: validate_merged_coupon_output(
            path,
            appliance_computation,
            digital_computation,
            extra_summary_rows,
            decisions,
        ),
    )

    la = appliance_computation
    print(
        "[家电] Subsidy coupon statistics complete: "
        f"{la.data_row_count} rows"
    )
    print(f"[家电] Receipt remark matches: {la.receipt_remark_count}")
    print(f"[家电] Pink return or exchange rows: {la.matched_count}")
    print(
        "[家电] Supplemental reference matches: "
        f"{la.reference_supplement_count}"
    )
    print(
        "[家电] Ambiguous supplemental reference candidates: "
        f"{la.ambiguous_reference_supplement_count}"
    )
    print(f"[家电] Submitted status matches: {la.uploaded_count}")
    print(
        "[家电] Rows not found in submitted data (marked 未上传): "
        f"{la.unmatched_count}"
    )
    print(
        f"[家电] Automatic reference corrections: {la.corrected_count}; "
        f"no unique candidate: {la.unresolved_count}; "
        f"duplicate conflicts: {la.correction_collision_count}"
    )
    print(
        "[家电] Total 2026 appliance subsidy counted as revenue for "
        f"matched rows: {la.matched_subsidy_total:.2f}"
    )
    if la.zero_subsidy_count:
        print(
            f"[家电] WARNING: {la.zero_subsidy_count} rows have a zero "
            "2026家电国补（计入收入）, which should not occur; they are "
            "counted as 0 — check those source rows"
        )
    report_source_total_gap("家电", la.source_total, la.computed_total)

    dg = digital_computation
    print(
        "[数码] Subsidy coupon statistics complete: "
        f"{dg.data_row_count} rows"
    )
    print(f"[数码] Remark matches: {dg.matched_count}")
    print(f"[数码] Submitted status matches: {dg.uploaded_match_count}")
    print(
        "[数码] Uploaded subsidy rows used in Summary: "
        f"{dg.uploaded_subsidy_count}; total: {dg.uploaded_subsidy_total:.2f}"
    )
    print(
        "[数码] Rows not found in submitted data (marked 未上传): "
        f"{dg.unmatched_count}"
    )
    print(
        f"[数码] Automatic reference corrections: {dg.corrected_count}; "
        f"no unique candidate: {dg.unresolved_count}; "
        f"duplicate conflicts: {dg.correction_collision_count}"
    )
    print(
        "[数码] Total 2026 digital subsidy counted as revenue for "
        f"matched rows: {dg.matched_subsidy_total:.2f}"
    )
    print(f"Output file: {OUTPUT_FILE}")


def validate_reference_report_sheet(
    workbook: Workbook,
    decisions: list[tuple[object, ...]],
) -> None:
    report_sheet = workbook[REFERENCE_REPORT_SHEET]
    report_header = tuple(cell.value for cell in report_sheet[1])
    if report_header != REFERENCE_REPORT_HEADER:
        raise RuntimeError(
            f"参考号处理报告字段标题校验失败：实际为 {report_header}"
        )

    def comparable(values) -> tuple:
        result = []
        for value in values:
            if isinstance(value, datetime):
                value = value.date()
            result.append("" if value is None else value)
        return tuple(result)

    actual = [
        comparable(row)
        for row in report_sheet.iter_rows(
            min_row=2,
            max_col=len(REFERENCE_REPORT_HEADER),
            values_only=True,
        )
    ]
    expected = [comparable(decision) for decision in decisions]
    if actual != expected:
        first_difference = next(
            (
                (index, a, e)
                for index, (a, e) in enumerate(zip(actual, expected), start=2)
                if a != e
            ),
            None,
        )
        raise RuntimeError(
            f"参考号处理报告内容校验失败：预期 {len(expected)} 条，"
            f"实际 {len(actual)} 条；首个差异 {first_difference}"
        )


def validate_merged_coupon_output(
    path: Path,
    appliance_computation: appliance.CouponComputation,
    digital_computation: digital.CouponComputation,
    extra_summary_rows: list[tuple[object, ...]],
    decisions: list[tuple[object, ...]],
) -> None:
    workbook = load_workbook(path, data_only=True)
    try:
        expected_sheet_names = [
            appliance.SUMMARY_SHEET_NAME,
            appliance.DETAILS_SHEET_NAME,
            digital.DETAILS_SHEET_NAME,
            *(
                group[0]
                for group in appliance_computation.group_sheets
            ),
            REFERENCE_REPORT_SHEET,
        ]
        if workbook.sheetnames != expected_sheet_names:
            raise RuntimeError(
                "审核明细工作表校验失败："
                f"预期 {expected_sheet_names}，实际 {workbook.sheetnames}"
            )
        appliance.validate_summary_and_details_sheets(
            workbook,
            appliance_computation,
            extra_summary_rows,
        )
        digital.validate_detail_sheet(workbook, digital_computation)
        appliance.validate_group_sheets(
            workbook,
            appliance_computation,
        )
        validate_reference_report_sheet(workbook, decisions)
    finally:
        workbook.close()
