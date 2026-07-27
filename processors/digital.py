
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from functools import partial
from pathlib import Path

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from processors.common import submitted as common_submitted
from processors.common.config import submitted_file_marker
from processors.common.coupons import (
    as_currency,
    load_coupon_remark_lookup,
    load_uploaded_detail_lookup,
    reference_correction_candidates,
)
from processors.common.excel import (
    format_sheet,
    load_measurement_font,
    load_uploaded_subsidy_stats,
    resolve_font,
)
from processors.common.dates import (
    normalize_coupon_date,
    normalize_document_number,
    normalize_receipt_identifier,
)
from processors.common.paths import (
    find_data_files,
    match_source_file_by_header,
    read_xls_header,
)
from processors.common.submitted import (
    KEPT_SOURCE_COLUMNS,
    KEPT_COLUMN_INDEXES,
    REQUIRED_SUBMITTED_HEADERS,
    STATUS_ORDER,
)
from processors.large_appliances import _shared as large_appliances_shared


BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "output"
DATA_DIR: Path
SUBMITTED_FILES: tuple[Path, ...]
OUTPUT_FILE = OUTPUT_DIR / "数码_已上传.xlsx"
COUPON_SOURCE_FILE: Path | None
# Receipt statistics are shared with large appliances: same source file,
# same output file, processed once using that project's rules (which
# include special cases digital never had, such as 同型号换货).
COUPON_REMARK_SOURCE_FILE = large_appliances_shared.RECEIPTS_OUTPUT_FILE
COUPON_UPLOADED_SOURCE_FILE = OUTPUT_FILE

DATA_TYPE = "数码"
# Files live directly in the flat data directory; the digital submitted
# export is told apart from the large appliances one by filename keyword,
# and the coupon export (whose filename keyword both projects happen to
# share) by header content.
# The submitted marker is derived from config/merchants.yaml in
# configure_data_dir rather than at import time, so a missing or malformed
# config fails the run with a readable error instead of breaking the import.
SUBMITTED_FILE_MARKER: str
COUPON_STATISTICS_KEYWORD = "销售用券情况统计"
# The coupon export's field header row (row 2) at its last kept column
# (column 26); see COUPON_KEPT_SOURCE_COLUMNS below.
COUPON_SUBSIDY_HEADER = "2026数码国补（计入收入）"


def configure_data_dir(data_dir: Path) -> None:
    global DATA_DIR
    global SUBMITTED_FILES
    global COUPON_SOURCE_FILE
    global SUBMITTED_FILE_MARKER

    DATA_DIR = data_dir
    SUBMITTED_FILE_MARKER = submitted_file_marker(DATA_TYPE)
    SUBMITTED_FILES = tuple(
        find_data_files(data_dir, SUBMITTED_FILE_MARKER, (".xlsx",))
    )
    COUPON_SOURCE_FILE = match_source_file_by_header(
        find_data_files(data_dir, COUPON_STATISTICS_KEYWORD, (".xls",)),
        COUPON_SUBSIDY_HEADER,
        read_header=partial(read_xls_header, row=2, column=26),
    )

# Digital: 15% of the transaction, capped at 500 per order. Household
# appliances use the same rate but a 1500 cap — see
# processors/large_appliances/submitted.py.
SUBSIDY_RATE = Decimal("0.15")
SUBSIDY_CAP = Decimal("500")
COUPON_KEPT_SOURCE_COLUMNS = (3, 4, 6, 8, 15, 18, 26)
COUPON_OUTPUT_HEADER = (
    "单据号",
    "单据日期",
    "商品名称",
    "品牌",
    "财务大类",
    "明细摘要",
    COUPON_SUBSIDY_HEADER,
    "备注",
    "详细情况",
)
REFERENCE_REPORT_CORRECTED = "已自动纠正"
REFERENCE_REPORT_UNRESOLVED = "无唯一候选"
REFERENCE_REPORT_COLLISION = "目标冲突"
REFERENCE_REPORT_ORDER = {
    REFERENCE_REPORT_CORRECTED: 0,
    REFERENCE_REPORT_COLLISION: 1,
    REFERENCE_REPORT_UNRESOLVED: 2,
}
COUPON_MATCH_FILL_COLOR = "FFC7CE"
COUPON_REFERENCE_RE = re.compile(r"\d{11}[A-Z]")
COUPON_SUMMARY_HEADER = (
    "备注",
    "数量",
    f"{COUPON_SUBSIDY_HEADER}合计",
)
DETAILS_SHEET_NAME = "数码-明细总表"


def select_columns(row: list[object]) -> list[object]:
    return common_submitted.select_columns(row, KEPT_COLUMN_INDEXES)


def _config() -> common_submitted.SubmittedConfig:
    return common_submitted.SubmittedConfig(
        input_files=SUBMITTED_FILES,
        data_dir=DATA_DIR,
        source_marker=SUBMITTED_FILE_MARKER,
        output_file=OUTPUT_FILE,
        subsidy_rate=SUBSIDY_RATE,
        subsidy_cap=SUBSIDY_CAP,
        kept_columns=KEPT_SOURCE_COLUMNS,
        required_headers=REQUIRED_SUBMITTED_HEADERS,
        status_order=STATUS_ORDER,
    )


def add_subsidy_column(
    row: list[object],
    *,
    is_header: bool = False,
    source_name: str | None = None,
    source_row: int | None = None,
) -> list[object]:
    return common_submitted.add_subsidy_column(
        row,
        subsidy_rate=SUBSIDY_RATE,
        subsidy_cap=SUBSIDY_CAP,
        is_header=is_header,
        source_name=source_name,
        source_row=source_row,
    )


def build_workbook() -> tuple[Workbook, int, int]:
    return common_submitted.build_workbook(_config())


def validate_output(path: Path, expected_data_rows: int) -> None:
    common_submitted.validate_output(path, expected_data_rows, _config())


def process_submitted_files() -> None:
    common_submitted.process_submitted_files(_config())


def read_coupon_rows(source: Path) -> list[list[object]]:
    source_workbook = xlrd.open_workbook(source)
    try:
        source_sheet = source_workbook.sheet_by_index(0)
        if source_sheet.nrows < 3:
            raise ValueError(f"{source.name} 缺少标题行、字段标题行或合计行")
        if source_sheet.ncols < max(COUPON_KEPT_SOURCE_COLUMNS):
            raise ValueError(
                f"{source.name} 列数不足：至少需要 "
                f"{max(COUPON_KEPT_SOURCE_COLUMNS)} 列"
            )
        if str(source_sheet.cell_value(source_sheet.nrows - 1, 0)).strip() != "合计":
            raise ValueError(f"{source.name} 最后一行不是合计行")

        source_header = tuple(
            source_sheet.cell_value(1, column - 1)
            for column in COUPON_KEPT_SOURCE_COLUMNS
        )
        expected_source_header = COUPON_OUTPUT_HEADER[
            :len(COUPON_KEPT_SOURCE_COLUMNS)
        ]
        if source_header != expected_source_header:
            raise ValueError(
                f"{source.name} 保留列字段标题不符合要求："
                f"预期为 {expected_source_header}，实际为 {source_header}。"
                f"如果实际字段包含“2026家电国补（计入收入）”，"
                f"请选择 large appliances 数据类型，或更换为数码销售用券文件。"
            )

        rows: list[list[object]] = [list(COUPON_OUTPUT_HEADER)]
        for row_index in range(2, source_sheet.nrows - 1):
            row: list[object] = []
            for column_index in COUPON_KEPT_SOURCE_COLUMNS:
                cell = source_sheet.cell(row_index, column_index - 1)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(
                        value,
                        source_workbook.datemode,
                    )
                elif (
                    cell.ctype == xlrd.XL_CELL_NUMBER
                    and value == int(value)
                ):
                    value = int(value)
                row.append(value)
            if any(value not in (None, "") for value in row):
                document_number = (
                    ""
                    if row[0] is None
                    else str(row[0]).replace("收款", "")
                )
                document_date = normalize_coupon_date(row[1], row_index + 1)
                rows.append(
                    [document_number, document_date, *row[2:], None, None]
                )
        return rows
    finally:
        source_workbook.release_resources()


def fill_coupon_remarks(
    rows: list[list[object]],
    remark_lookup: dict[tuple[str, date], str],
) -> tuple[int, Decimal]:
    matched_rows: list[list[object]] = []
    unmatched_rows: list[list[object]] = []
    subsidy_index = COUPON_OUTPUT_HEADER.index("2026数码国补（计入收入）")
    matched_subsidy_total = Decimal("0")
    remark_index = COUPON_OUTPUT_HEADER.index("备注")
    for row in rows[1:]:
        key = (normalize_document_number(row[0]), row[1])
        remark = remark_lookup.get(key, "")
        row[remark_index] = remark
        if remark:
            matched_rows.append(row)
            subsidy = row[subsidy_index]
            if subsidy not in (None, ""):
                try:
                    matched_subsidy_total += Decimal(str(subsidy))
                except InvalidOperation as error:
                    raise ValueError(
                        f"组合键 {key} 的2026数码国补金额无效：{subsidy!r}"
                    ) from error
        else:
            unmatched_rows.append(row)
    rows[1:] = [*unmatched_rows, *matched_rows]
    return len(matched_rows), matched_subsidy_total


def fill_uploaded_details(
    rows: list[list[object]],
    detail_lookup: dict[str, str],
) -> int:
    summary_index = COUPON_OUTPUT_HEADER.index("明细摘要")
    remark_index = COUPON_OUTPUT_HEADER.index("备注")
    detail_index = COUPON_OUTPUT_HEADER.index("详细情况")
    matched_count = 0
    for row in rows[1:]:
        reference = normalize_receipt_identifier(
            row[summary_index]
        ).upper()
        detail = detail_lookup.get(reference, "")
        row[detail_index] = detail
        if detail:
            row[remark_index] = "已上传"
            matched_count += 1
    return matched_count


def reference_decision(
    outcome: str,
    row: list[object],
    raw_reference: str,
    note: str,
) -> tuple[str, str, object, str, str]:
    """Identify a decision by 单据号 + 单据日期, not by row position.

    The detail rows get re-sorted after corrections are applied, so a row
    number recorded here would point at the wrong row in the saved sheet.
    """
    return (
        outcome,
        normalize_document_number(row[COUPON_OUTPUT_HEADER.index("单据号")]),
        row[COUPON_OUTPUT_HEADER.index("单据日期")],
        raw_reference,
        note,
    )


def correct_coupon_references(
    rows: list[list[object]],
    reference_universe: set[str],
) -> tuple[int, int, int, list[tuple[str, str, object, str, str]]]:
    """Correct references and record every decision for the processing report.

    The universe is built from submitted data only, so an operator has to be
    able to review each applied correction, not just the counts.

    A well-formed reference with no candidate at all is simply absent from the
    submitted data — the detail row already carries a 未上传 remark, so it is
    counted but kept out of the report. Only malformed references and genuine
    ambiguities are reported, which is what an operator can actually act on.
    """
    summary_index = COUPON_OUTPUT_HEADER.index("明细摘要")
    existing_counts = Counter(
        normalize_receipt_identifier(row[summary_index]).upper()
        for row in rows[1:]
        if normalize_receipt_identifier(row[summary_index])
    )
    proposed: dict[int, str] = {}
    target_counts: Counter[str] = Counter()
    unresolved_count = 0
    decisions: list[tuple[str, str, object, str, str]] = []

    for row_index, row in enumerate(rows[1:], start=1):
        raw_reference = normalize_receipt_identifier(
            row[summary_index]
        ).upper()
        if not raw_reference or raw_reference in reference_universe:
            continue
        candidates = reference_correction_candidates(
            raw_reference,
            reference_universe,
        )
        if len(candidates) != 1:
            unresolved_count += 1
            if candidates or not COUPON_REFERENCE_RE.fullmatch(raw_reference):
                decisions.append(
                    reference_decision(
                        REFERENCE_REPORT_UNRESOLVED,
                        row,
                        raw_reference,
                        f"候选数量 {len(candidates)}，"
                        f"未在已上传数据中找到唯一匹配，保留原值",
                    )
                )
            continue
        target = next(iter(candidates))
        proposed[row_index] = target
        target_counts[target] += 1

    corrected_count = 0
    collision_count = 0
    for row_index, target in proposed.items():
        row = rows[row_index]
        raw_reference = normalize_receipt_identifier(
            row[summary_index]
        ).upper()
        if existing_counts[target] > 0 or target_counts[target] > 1:
            collision_count += 1
            decisions.append(
                reference_decision(
                    REFERENCE_REPORT_COLLISION,
                    row,
                    raw_reference,
                    f"目标参考号 {target} 已被其他行占用，未纠正",
                )
            )
            continue
        row[summary_index] = target
        corrected_count += 1
        decisions.append(
            reference_decision(
                REFERENCE_REPORT_CORRECTED,
                row,
                raw_reference,
                f"已自动纠正为 {target}，请人工复核",
            )
        )

    decisions.sort(key=lambda decision: REFERENCE_REPORT_ORDER[decision[0]])
    return corrected_count, unresolved_count, collision_count, decisions


def fill_unmatched_remarks(
    rows: list[list[object]],
    reference_universe: set[str],
) -> int:
    summary_index = COUPON_OUTPUT_HEADER.index("明细摘要")
    remark_index = COUPON_OUTPUT_HEADER.index("备注")
    unmatched_count = 0
    for row in rows[1:]:
        reference = normalize_receipt_identifier(
            row[summary_index]
        ).upper()
        if reference not in reference_universe:
            row[remark_index] = "未上传"
            unmatched_count += 1
    return unmatched_count


def build_coupon_summary(
    rows: list[list[object]],
    uploaded_count: int,
    uploaded_subsidy_total: Decimal,
) -> list[tuple[object, ...]]:
    subsidy_index = COUPON_OUTPUT_HEADER.index("2026数码国补（计入收入）")
    coupon_count = 0
    coupon_total = Decimal("0")
    for row_number, row in enumerate(rows[1:], start=2):
        subsidy = row[subsidy_index]
        if subsidy in (None, ""):
            continue
        try:
            subsidy_amount = Decimal(str(subsidy))
            coupon_total += subsidy_amount
            if subsidy_amount < 0:
                coupon_count -= 1
            elif subsidy_amount > 0:
                coupon_count += 1
        except InvalidOperation as error:
            raise ValueError(
                f"第 {row_number} 行的2026数码国补金额无效：{subsidy!r}"
            ) from error

    unuploaded_count = coupon_count - uploaded_count
    unuploaded_total = coupon_total - uploaded_subsidy_total
    return [
        (
            "已上传",
            uploaded_count,
            float(as_currency(uploaded_subsidy_total)),
        ),
        (
            "未上传",
            unuploaded_count,
            float(as_currency(unuploaded_total)),
        ),
        ("合计", coupon_count, float(as_currency(coupon_total))),
    ]


@dataclass
class CouponComputation:
    """Everything computed for the 数码 coupon data, before it is written to
    a sheet. Digital has no group sheets, approved-detail panel, or
    reference-supplement file, so this is smaller than the 家电 equivalent
    in processors/large_appliances/coupons.py."""

    rows: list[list[object]]
    data_row_count: int
    matched_count: int
    matched_subsidy_total: Decimal
    remark_lookup: dict[tuple[str, date], str]
    detail_lookup: dict[str, str]
    reference_universe: set[str]
    uploaded_subsidy_count: int
    uploaded_subsidy_total: Decimal
    uploaded_match_count: int
    unmatched_count: int
    corrected_count: int
    unresolved_count: int
    correction_collision_count: int
    reference_decisions: list[tuple[str, str, str, str]]
    summary_rows: list[tuple[object, ...]]


def compute_coupon_data() -> CouponComputation:
    if COUPON_SOURCE_FILE is None:
        raise FileNotFoundError(
            f"未在 {DATA_DIR} 中找到文件名包含"
            f"“{COUPON_STATISTICS_KEYWORD}”且表头为"
            f"“{COUPON_SUBSIDY_HEADER}”的 .XLS 文件"
        )

    rows = read_coupon_rows(COUPON_SOURCE_FILE)
    remark_lookup = load_coupon_remark_lookup(COUPON_REMARK_SOURCE_FILE)
    matched_count, matched_subsidy_total = fill_coupon_remarks(
        rows,
        remark_lookup,
    )
    detail_lookup = load_uploaded_detail_lookup(COUPON_UPLOADED_SOURCE_FILE)
    uploaded_subsidy_count, uploaded_subsidy_total = (
        load_uploaded_subsidy_stats(COUPON_UPLOADED_SOURCE_FILE)
    )
    # Unsubmitted data is no longer supplied, so submitted data is the only
    # source of valid references.
    reference_universe = set(detail_lookup)
    (
        corrected_count,
        unresolved_count,
        correction_collision_count,
        reference_decisions,
    ) = correct_coupon_references(rows, reference_universe)
    uploaded_match_count = fill_uploaded_details(rows, detail_lookup)
    unmatched_count = fill_unmatched_remarks(rows, reference_universe)
    summary_rows = build_coupon_summary(
        rows,
        uploaded_subsidy_count,
        uploaded_subsidy_total,
    )
    if as_currency(matched_subsidy_total) != Decimal("0.00"):
        raise ValueError(
            "备注匹配行的2026数码国补（计入收入）合计不为 0："
            f"{matched_subsidy_total}"
        )

    return CouponComputation(
        rows=rows,
        data_row_count=max(len(rows) - 1, 0),
        matched_count=matched_count,
        matched_subsidy_total=matched_subsidy_total,
        remark_lookup=remark_lookup,
        detail_lookup=detail_lookup,
        reference_universe=reference_universe,
        uploaded_subsidy_count=uploaded_subsidy_count,
        uploaded_subsidy_total=uploaded_subsidy_total,
        uploaded_match_count=uploaded_match_count,
        unmatched_count=unmatched_count,
        corrected_count=corrected_count,
        unresolved_count=unresolved_count,
        correction_collision_count=correction_collision_count,
        reference_decisions=reference_decisions,
        summary_rows=summary_rows,
    )


def build_detail_sheet(
    workbook: Workbook,
    computation: CouponComputation,
) -> tuple[str, object, PatternFill]:
    """Append 数码-明细总表 to the (already partially built) 审核明细
    workbook. Returns (font_name, measurement_font, matched_fill) purely for
    symmetry with the 家电 builder; callers processing only digital's sheet
    can ignore them."""
    sheet = workbook.create_sheet(DETAILS_SHEET_NAME)
    for row in computation.rows:
        sheet.append(row)

    font_name, font_path = resolve_font()
    measurement_font = load_measurement_font(font_path)
    format_sheet(
        sheet,
        font_name,
        measurement_font,
        ("商品名称",),
    )
    for cell in sheet["A"][1:]:
        cell.number_format = "@"
    for cell in sheet["B"][1:]:
        if cell.value not in (None, ""):
            cell.number_format = "yyyy-mm-dd"
    matched_fill = PatternFill("solid", fgColor=COUPON_MATCH_FILL_COLOR)
    matched_start_row = sheet.max_row - computation.matched_count + 1
    for row in sheet.iter_rows(
        min_row=matched_start_row,
        max_row=sheet.max_row,
    ):
        for cell in row:
            cell.fill = matched_fill

    return font_name, measurement_font, matched_fill


def validate_detail_sheet(
    workbook: Workbook,
    computation: CouponComputation,
) -> None:
    expected_data_rows = computation.data_row_count
    expected_matched_rows = computation.matched_count
    remark_lookup = computation.remark_lookup
    expected_matched_subsidy_total = computation.matched_subsidy_total
    detail_lookup = computation.detail_lookup
    expected_uploaded_rows = computation.uploaded_match_count
    reference_universe = computation.reference_universe
    expected_unmatched_rows = computation.unmatched_count

    sheet = workbook[DETAILS_SHEET_NAME]
    header = tuple(cell.value for cell in sheet[1])
    if header != COUPON_OUTPUT_HEADER:
        raise RuntimeError(
            f"销售用券字段标题校验失败：实际为 {header}"
        )
    if sheet.max_column != len(COUPON_OUTPUT_HEADER):
        raise RuntimeError(
            f"销售用券列数校验失败：实际为 {sheet.max_column}"
        )
    actual_data_rows = max(sheet.max_row - 1, 0)
    if actual_data_rows != expected_data_rows:
        raise RuntimeError(
            f"销售用券行数校验失败：预期 {expected_data_rows} 条，"
            f"实际 {actual_data_rows} 条"
        )
    detail_column = COUPON_OUTPUT_HEADER.index("详细情况") + 1
    actual_uploaded_rows = sum(
        sheet.cell(row_number, detail_column).value not in (None, "")
        for row_number in range(2, sheet.max_row + 1)
    )
    if actual_uploaded_rows != expected_uploaded_rows:
        raise RuntimeError(
            f"销售用券已上传匹配数校验失败：预期 "
            f"{expected_uploaded_rows} 条，实际 {actual_uploaded_rows} 条"
        )
    remark_column = COUPON_OUTPUT_HEADER.index("备注") + 1
    actual_unuploaded_remark_rows = sum(
        sheet.cell(row_number, remark_column).value == "未上传"
        for row_number in range(2, sheet.max_row + 1)
    )
    expected_unuploaded_remark_rows = expected_unmatched_rows
    if actual_unuploaded_remark_rows != expected_unuploaded_remark_rows:
        raise RuntimeError(
            f"销售用券“未上传”备注数量校验失败：预期 "
            f"{expected_unuploaded_remark_rows} 条，"
            f"实际 {actual_unuploaded_remark_rows} 条"
        )
    summary_column = COUPON_OUTPUT_HEADER.index("明细摘要") + 1
    product_name_column = COUPON_OUTPUT_HEADER.index("商品名称") + 1
    if any(
        sheet.cell(row_number, product_name_column).alignment.horizontal
        != "left"
        for row_number in range(2, sheet.max_row + 1)
    ):
        raise RuntimeError("销售用券商品名称列左对齐校验失败")
    matched_start_row = sheet.max_row - expected_matched_rows + 1
    actual_matched_subsidy_total = Decimal("0")
    subsidy_column = COUPON_OUTPUT_HEADER.index(COUPON_SUBSIDY_HEADER) + 1
    for row_number in range(2, sheet.max_row + 1):
        document_cell = sheet.cell(row_number, 1)
        date_cell = sheet.cell(row_number, 2)
        remark_cell = sheet.cell(row_number, remark_column)
        detail_cell = sheet.cell(row_number, detail_column)
        if "收款" in str(document_cell.value or ""):
            raise RuntimeError(
                f"销售用券第 {row_number} 行单据号仍包含“收款”"
            )
        if document_cell.number_format != "@":
            raise RuntimeError(
                f"销售用券第 {row_number} 行单据号不是文本格式"
            )
        if (
            date_cell.value not in (None, "")
            and date_cell.number_format != "yyyy-mm-dd"
        ):
            raise RuntimeError(
                f"销售用券第 {row_number} 行单据日期格式错误"
            )
        if date_cell.value not in (None, "") and not isinstance(
            date_cell.value,
            (date, datetime),
        ):
            raise RuntimeError(
                f"销售用券第 {row_number} 行单据日期不是日期值"
            )
        document_date = (
            date_cell.value.date()
            if isinstance(date_cell.value, datetime)
            else date_cell.value
        )
        receipt_remark = remark_lookup.get(
            (
                normalize_document_number(document_cell.value),
                document_date,
            ),
            "",
        )
        reference = normalize_receipt_identifier(
            sheet.cell(row_number, summary_column).value
        ).upper()
        expected_detail = detail_lookup.get(reference, "")
        if expected_detail:
            expected_remark = "已上传"
        elif reference not in reference_universe:
            expected_remark = "未上传"
        else:
            expected_remark = receipt_remark
        actual_remark = str(remark_cell.value or "")
        if actual_remark != expected_remark:
            raise RuntimeError(
                f"销售用券第 {row_number} 行备注匹配校验失败"
            )
        if str(detail_cell.value or "") != expected_detail:
            raise RuntimeError(
                f"销售用券第 {row_number} 行详细情况匹配校验失败"
            )
        expected_pink = (
            expected_matched_rows > 0
            and row_number >= matched_start_row
        )
        is_pink = all(
            cell.fill.fill_type == "solid"
            and cell.fill.fgColor.rgb
            in {
                COUPON_MATCH_FILL_COLOR,
                f"00{COUPON_MATCH_FILL_COLOR}",
                f"FF{COUPON_MATCH_FILL_COLOR}",
            }
            for cell in sheet[row_number]
        )
        if is_pink != expected_pink:
            raise RuntimeError(
                f"销售用券第 {row_number} 行粉色填充或位置校验失败"
            )
        if expected_pink:
            subsidy = sheet.cell(row_number, subsidy_column).value
            if subsidy not in (None, ""):
                actual_matched_subsidy_total += Decimal(str(subsidy))

    if as_currency(actual_matched_subsidy_total) != as_currency(
        expected_matched_subsidy_total
    ):
        raise RuntimeError(
            "销售用券匹配行国补合计校验失败："
            f"预期 {expected_matched_subsidy_total}，"
            f"实际 {actual_matched_subsidy_total}"
        )
    if as_currency(actual_matched_subsidy_total) != Decimal("0.00"):
        raise RuntimeError(
            "销售用券匹配行的2026数码国补（计入收入）合计不为 0："
            f"{actual_matched_subsidy_total}"
        )

