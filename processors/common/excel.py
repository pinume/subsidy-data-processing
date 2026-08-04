import os
import shutil
import stat
import time
from collections.abc import Callable, Iterator
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from PIL import ImageFont
from python_calamine import CalamineSheet

FONT_NAME = "Maple Mono NF CN"
FALLBACK_FONT_NAME = "微软雅黑"
FONT_SIZE = 11
ROW_HEIGHT = 18
FONT_PATH = Path(r"C:\Windows\Fonts\MapleMono-NF-CN-Regular.ttf")
FALLBACK_FONT_PATH = Path(r"C:\Windows\Fonts\msyh.ttc")
FONT_CANDIDATES = (
    (
        FONT_NAME,
        (
            FONT_PATH,
            Path.home()
            / ".local/share/fonts/MapleMono-NF-CN/MapleMono-NF-CN-Regular.ttf",
        ),
        (
            "MapleMono-NF-CN-Regular.ttf",
            "MapleMonoNormal-NF-CN-Regular.ttf",
        ),
    ),
    (
        FALLBACK_FONT_NAME,
        (
            FALLBACK_FONT_PATH,
            Path("/usr/local/share/fonts/github-fonts/MSYH.TTC"),
            Path("/usr/local/share/fonts/github-fonts/MSYHL.TTC"),
        ),
        (
            "msyh.ttc",
            "msyh.ttf",
            "MSYH.TTC",
            "MSYHL.TTC",
        ),
    ),
    (
        "Noto Sans CJK SC",
        (
            Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
            Path("/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc"),
        ),
        (
            "NotoSansCJK-Regular.ttc",
            "NotoSansSC-Regular.otf",
            "NotoSansCJKsc-Regular.otf",
        ),
    ),
    (
        "Source Han Sans SC",
        (),
        (
            "SourceHanSansSC-Regular.otf",
            "SourceHanSansCN-Regular.otf",
        ),
    ),
    (
        "PingFang SC",
        (Path("/System/Library/Fonts/PingFang.ttc"),),
        ("PingFang.ttc",),
    ),
    (
        "WenQuanYi Micro Hei",
        (),
        (
            "wqy-microhei.ttc",
            "WenQuanYi Micro Hei.ttf",
        ),
    ),
)
FONT_SEARCH_ROOTS = (
    Path.home() / ".local/share/fonts",
    Path.home() / ".fonts",
    Path.home() / "Library/Fonts",
    Path("/usr/share/fonts"),
    Path("/usr/local/share/fonts"),
    Path("/Library/Fonts"),
    Path("/System/Library/Fonts"),
    Path(r"C:\Windows\Fonts"),
)


def normalize_calamine_value(value: object) -> object:
    """Bridge calamine's cell typing to openpyxl's read_only, data_only output.

    calamine represents a blank cell as "" where openpyxl gives None, and a
    whole number as float where openpyxl gives int when the stored value has
    no fractional part. Normalizing here keeps every downstream consumer —
    written against openpyxl's original output — working unchanged.

    Two differences are deliberately left alone, so this is a close bridge
    rather than an equivalent one:

    A date-formatted cell with no time part arrives as datetime.date, where
    openpyxl always gave datetime.datetime. Every reader of these values
    narrows a datetime to its date anyway, so the plain date is already the
    wanted form; converting it back up would only add a round trip.

    An error cell (#DIV/0!, #N/A) is indistinguishable from a blank one:
    calamine has no error type at all and yields "" for both, so nothing
    here can recover the "#DIV/0!" string openpyxl produced. Callers of
    financially significant fields must therefore inspect the original
    cell type when a calamine value unexpectedly appears blank.
    """
    if value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def calamine_rows(sheet: CalamineSheet) -> Iterator[list[object]]:
    """Yield every row of a calamine sheet, cell values normalized to match
    openpyxl's read_only, data_only output (see normalize_calamine_value).

    Two shape differences from openpyxl are corrected here as well. A sheet
    whose used range starts below row 1 still gets its leading blank rows
    from iter_rows(), so row numbers stay absolute, but leading blank
    *columns* are dropped — every row would start at the used range's first
    column. They are padded back so index 0 always means column A, which
    submitted's fixed D/E/F… column letters depend on, and which payment
    needs in order to address the same column in a sheet read by calamine
    and its openpyxl-read formula twin.

    A sheet with no cells at all reports start None and makes iter_rows()
    panic inside the Rust extension, so it yields nothing instead.
    """
    if sheet.start is None:
        return
    leading_blanks = [None] * sheet.start[1]
    for row in sheet.iter_rows():
        yield leading_blanks + [normalize_calamine_value(value) for value in row]


def _find_font_file(file_names: tuple[str, ...]) -> Path | None:
    for root in FONT_SEARCH_ROOTS:
        if not root.exists():
            continue

        for file_name in file_names:
            direct_path = root / file_name
            if direct_path.exists():
                return direct_path

        for file_name in file_names:
            try:
                return next(root.rglob(file_name))
            except (OSError, StopIteration):
                continue
    return None


def resolve_font() -> tuple[str, Path]:
    """Resolve the Excel font name and a local font file used for measuring.

    Excel output uses Maple Mono NF CN whenever its font file is available;
    every other locally available measurement font maps to 微软雅黑 so that
    machines without Maple produce the requested, consistent fallback name.
    """
    configured_path = os.environ.get("UPLOAD_DATA_FONT_PATH")
    if configured_path:
        font_path = Path(configured_path).expanduser()
        if not font_path.exists():
            raise FileNotFoundError(f"Configured font file does not exist: {font_path}")
        return os.environ.get("UPLOAD_DATA_FONT_NAME", font_path.stem), font_path

    for font_name, paths, file_names in FONT_CANDIDATES:
        for path in paths:
            if path.exists():
                return (
                    FONT_NAME if font_name == FONT_NAME else FALLBACK_FONT_NAME,
                    path,
                )

        font_path = _find_font_file(file_names)
        if font_path is not None:
            return (
                FONT_NAME if font_name == FONT_NAME else FALLBACK_FONT_NAME,
                font_path,
            )

    raise FileNotFoundError(
        "No supported font file was found. Install Maple Mono NF CN, Microsoft "
        "YaHei, Noto Sans CJK SC, or PingFang SC, or set UPLOAD_DATA_FONT_PATH."
    )


def load_measurement_font(font_path: Path):
    return ImageFont.truetype(str(font_path), size=15)


def measurement_text(value: object) -> str:
    if value is None:
        return ""
    return value.strftime("%Y%m%d") if isinstance(value, (date, datetime)) else str(value)


# Strings whose width would stop being the sum of their characters' widths the
# moment a font kerned, ligated, or shaped them: classic kerning pairs, an "ffi"
# ligature candidate, and the character classes this program actually writes
# (digits and uppercase in SN/IMEI/invoice numbers, CJK punctuation, runs of
# Chinese). A font passes only if every one of them measures additively.
WIDTH_ADDITIVITY_PROBES = (
    "AV",
    "To",
    "Ta",
    "ffi",
    "0123456789",
    "SN2026ABCD",
    "，。、；：",
    "连续中文字符",
    "中文A1，混排",
)


def widths_are_additive(value_font) -> bool:
    """Whether a string's width always equals the sum of its characters'.

    True for the monospace font this project prefers, false for proportional
    fonts such as 微软雅黑, which kern. Probing the font itself rather than
    trusting its name keeps any unrecognised font on the exact, slower path.
    """
    for probe in WIDTH_ADDITIVITY_PROBES:
        summed = sum(value_font.getlength(character) for character in probe)
        if value_font.getlength(probe) != summed:
            return False
    return True


def width_measurer(value_font) -> Callable[[object], float]:
    """Measure text width, caching by rendered text.

    Sheets repeat dates, statuses, and remarks heavily, but SN/IMEI/invoice
    columns are effectively all-distinct, so caching by whole string cannot
    bound the number of font calls. When the font measures additively (see
    widths_are_additive) a string's width can be summed from its characters
    instead, which replaces that unbounded set of distinct strings with the
    bounded set of distinct characters. Fonts that kern keep the whole-string
    measurement, so their output is unchanged.
    """
    cache: dict[str, float] = {}

    if not widths_are_additive(value_font):

        def measure(value: object) -> float:
            text = measurement_text(value)
            if not text:
                return 0
            width = cache.get(text)
            if width is None:
                width = value_font.getlength(text)
                cache[text] = width
            return width

        return measure

    character_widths: dict[str, float] = {}

    def measure(value: object) -> float:
        text = measurement_text(value)
        if not text:
            return 0
        width = cache.get(text)
        if width is None:
            width = 0.0
            for character in text:
                character_width = character_widths.get(character)
                if character_width is None:
                    character_width = value_font.getlength(character)
                    character_widths[character] = character_width
                width += character_width
            cache[text] = width
        return width

    return measure


# 255 character units is Excel's widest column, at 7 pixels per unit.
MAX_COLUMN_PIXELS = 255 * 7


def pixels_to_column_pixels(pixels: float) -> int:
    """Padded pixel width for XlsxWriter's set_column_pixels().

    The padding calculation reproduces the widths from the former openpyxl
    writer to within a rounding step without XlsxWriter adding extra padding.
    """
    return min(round((pixels * 1.1) + 16), MAX_COLUMN_PIXELS)


# What openpyxl assigned to a cell on its own when given a date or datetime.
DATE_NUMBER_FORMAT = "yyyy-mm-dd"
DATETIME_NUMBER_FORMAT = "yyyy-mm-dd h:mm:ss"


def sheet_format_set(workbook, font_name: str) -> dict[str, object]:
    """The four cell formats every plain output table uses.

    XlsxWriter deduplicates identical formats into one style-table entry, so
    building this per sheet costs nothing and keeps each sheet's writer
    self-contained.
    """
    base = {
        "font_name": font_name,
        "font_size": FONT_SIZE,
        "font_color": "#000000",
        "align": "center",
        "valign": "vcenter",
    }
    return {
        "header": workbook.add_format(
            {**base, "bold": True, "font_color": "#FFFFFF", "bg_color": "#000000"}
        ),
        "center": workbook.add_format(base),
        "left": workbook.add_format({**base, "align": "left"}),
        # openpyxl stamped these two formats onto a cell automatically the
        # moment a date or datetime was assigned to it; XlsxWriter applies
        # nothing and would leave the value showing as its serial number
        # (2026-01-02 as 46024). Same formats, so a native Excel date looks
        # identical whichever writer produced the file.
        "date": workbook.add_format({**base, "num_format": DATE_NUMBER_FORMAT}),
        "datetime": workbook.add_format(
            {**base, "num_format": DATETIME_NUMBER_FORMAT}
        ),
    }


def write_formatted_sheet(
    workbook,
    sheet_name: str,
    header,
    rows,
    font_name: str,
    measurement_font,
    *,
    left_aligned_headers: tuple[str, ...] = ("描述",),
    number_formats: dict[str, str] | None = None,
    autofilter: bool = True,
) -> None:
    """XlsxWriter equivalent of format_sheet for a plain header-plus-body table.

    Same result as building the sheet with openpyxl and running format_sheet
    over it: black header row, centered body, per-column widths measured from
    the widest value, frozen header, and an autofilter over the whole table.
    number_formats maps a header name to the number format its column takes,
    defaulting to 补贴金额 as "0.00" exactly as format_sheet applied it.
    """
    if number_formats is None:
        number_formats = {"补贴金额": "0.00"}
    sheet = workbook.add_worksheet(sheet_name)
    formats = sheet_format_set(workbook, font_name)
    numbered_formats = {
        index: workbook.add_format(
            {
                "font_name": font_name,
                "font_size": FONT_SIZE,
                "font_color": "#000000",
                "align": "center",
                "valign": "vcenter",
                "num_format": number_format,
            }
        )
        for index, name in enumerate(header)
        if (number_format := number_formats.get(name)) is not None
    }
    left_columns = {
        index for index, name in enumerate(header) if name in left_aligned_headers
    }

    measure = width_measurer(measurement_font)
    maximum_widths = [measure(value) for value in header]
    sheet.set_row(0, ROW_HEIGHT)
    for column, value in enumerate(header):
        sheet.write(0, column, value, formats["header"])

    for row_number, row in enumerate(rows, start=1):
        sheet.set_row(row_number, ROW_HEIGHT)
        for column, value in enumerate(row):
            # datetime is a subclass of date, so it has to be tested first.
            if isinstance(value, datetime):
                cell_format = formats["datetime"]
            elif isinstance(value, date):
                cell_format = formats["date"]
            elif column in left_columns:
                cell_format = formats["left"]
            elif column in numbered_formats and value is not None:
                cell_format = numbered_formats[column]
            else:
                cell_format = formats["center"]
            sheet.write(row_number, column, value, cell_format)
            if column < len(maximum_widths):
                maximum_widths[column] = max(maximum_widths[column], measure(value))

    sheet.freeze_panes(1, 0)
    if autofilter:
        sheet.autofilter(0, 0, len(rows), len(header) - 1)
    for column, maximum_pixels in enumerate(maximum_widths):
        sheet.set_column_pixels(
            column, column, pixels_to_column_pixels(maximum_pixels)
        )
    return sheet


STALE_TEMPORARY_FILE_AGE_SECONDS = 180


@dataclass(frozen=True)
class StaleFileCleanup:
    """What startup cleanup did, reported by the caller rather than printed.

    removed lists the dot-prefixed leftovers deleted; failed lists
    (file name, reason) pairs for files that could not be removed.
    """

    removed: tuple[str, ...]
    failed: tuple[tuple[str, str], ...]


def remove_stale_temporary_files(
    output_dir: Path,
    minimum_age_seconds: float = STALE_TEMPORARY_FILE_AGE_SECONDS,
) -> StaleFileCleanup:
    """Delete leftover intermediate files from an interrupted earlier run.

    Everything this program writes into the output directory as an
    intermediate is dot-prefixed: save_workbook_atomically's ".<名字>-<随机>"
    temporary file. It is removed on a normal run; a crash or a killed
    process leaves it behind, where it accumulates invisibly and still holds
    business data. Excel's own lock files are named "~$...", so they are
    never touched.

    Only files older than minimum_age_seconds are removed. This program was
    never designed for two instances to run against the same output
    directory at once, but without an age check this cleanup would make that
    actively unsafe instead of merely unsupported: a second instance
    starting up would delete the first instance's temporary file while it is
    still being written, and the first instance's own save would then fail
    trying to rename a file that no longer exists. The age check does not
    make concurrent runs supported — it only keeps a second instance's
    startup from corrupting a first instance's in-flight save. The default
    (180s) is comfortably longer than any single sheet this program writes
    has been observed to take to save.
    """
    if not output_dir.is_dir():
        return StaleFileCleanup((), ())

    now = time.time()
    removed: list[str] = []
    failed: list[tuple[str, str]] = []
    for path in sorted(output_dir.iterdir()):
        if not path.is_file() or not path.name.startswith("."):
            continue
        try:
            age_seconds = now - path.stat().st_mtime
        except OSError:
            continue
        if age_seconds < minimum_age_seconds:
            continue
        try:
            path.chmod(path.stat().st_mode | stat.S_IWRITE)
            path.unlink()
        except OSError as error:
            failed.append((path.name, str(error)))
            continue
        removed.append(path.name)
    return StaleFileCleanup(tuple(removed), tuple(failed))


def save_workbook_atomically(
    workbook: Workbook,
    output_path: Path,
    validator: Callable[[Path], None],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with NamedTemporaryFile(
            prefix=f".{output_path.stem}-",
            suffix=output_path.suffix,
            dir=output_path.parent,
            delete=False,
        ) as temporary_file:
            temporary_path = Path(temporary_file.name)

        workbook.save(temporary_path)
        validator(temporary_path)
        os.replace(temporary_path, output_path)
        temporary_path = None
    finally:
        workbook.close()
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def write_xlsx_atomically(
    output_path: Path,
    writer: Callable[[Path], None],
    validator: Callable[[Path], None],
) -> None:
    """Write a path-based XLSX producer with the same atomic contract.

    XlsxWriter takes its destination path when the workbook is created, unlike
    openpyxl's save-at-the-end API. Keeping that difference here lets either
    writer use the same write, validate, and replace transaction.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with NamedTemporaryFile(
            prefix=f".{output_path.stem}-",
            suffix=output_path.suffix,
            dir=output_path.parent,
            delete=False,
        ) as temporary_file:
            temporary_path = Path(temporary_file.name)

        writer(temporary_path)
        validator(temporary_path)
        os.replace(temporary_path, output_path)
        temporary_path = None
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def run_with_output_rollback(
    output_paths: tuple[Path, ...],
    operation: Callable[[], object],
) -> object:
    """Restore every managed output if a multi-file operation fails.

    Individual workbooks are already written atomically. This adds the missing
    transaction boundary around commands that intentionally produce several
    files: an error in a later step must not leave earlier outputs from the new
    run beside older outputs from the previous run.
    """
    unique_paths = tuple(dict.fromkeys(output_paths))
    backups: dict[Path, Path | None] = {}
    temporary_backups: set[Path] = set()
    try:
        for output_path in unique_paths:
            if not output_path.exists():
                backups[output_path] = None
                continue
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with NamedTemporaryFile(
                prefix=f".{output_path.stem}-rollback-",
                suffix=output_path.suffix,
                dir=output_path.parent,
                delete=False,
            ) as backup_file:
                backup_path = Path(backup_file.name)
            temporary_backups.add(backup_path)
            shutil.copy2(output_path, backup_path)
            backups[output_path] = backup_path

        return operation()
    except BaseException as operation_error:
        rollback_errors: list[str] = []
        for output_path, backup_path in backups.items():
            try:
                if backup_path is None:
                    output_path.unlink(missing_ok=True)
                else:
                    os.replace(backup_path, output_path)
                    backups[output_path] = None
                    temporary_backups.discard(backup_path)
            except OSError as rollback_error:
                rollback_errors.append(f"{output_path.name}: {rollback_error}")
        if rollback_errors:
            raise RuntimeError(
                "处理失败且输出回滚不完整："
                + "；".join(rollback_errors)
            ) from operation_error
        raise
    finally:
        for backup_path in temporary_backups:
            backup_path.unlink(missing_ok=True)
