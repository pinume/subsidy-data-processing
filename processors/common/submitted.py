"""Shared "submitted data" pipeline for both projects.

Household appliances and digital read the same source column layout, apply
the same 15% subsidy rate, and validate the same shape of output; only the
subsidy cap, the source files, and the output file differ between them. Each
project's own module still declares its own SUBSIDY_RATE / SUBSIDY_CAP
constants (see README.md's Subsidy Rules section for why the caps are not
merged into one shared default) and passes them in here.
"""

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string

from processors.common.excel import (
    format_sheet,
    load_measurement_font,
    read_rows,
    resolve_font,
)


# Keep the required source columns in their original order.
KEPT_SOURCE_COLUMNS = ("D", "E", "F", "G", "I", "J", "Q", "S", "U", "W", "X")
KEPT_COLUMN_INDEXES = tuple(
    column_index_from_string(column) for column in KEPT_SOURCE_COLUMNS
)
# "补贴金额" is inserted by add_subsidy_column, so only source fields are listed.
REQUIRED_SUBMITTED_HEADERS = ("状态", "描述", "交易金额")
STATUS_ORDER = (
    "核销失败",
    "审核失败",
    "暂存",
    "同步(已上送)",
    "待审核",
    "审核通过",
)
STATUS_PRIORITY = {status: index for index, status in enumerate(STATUS_ORDER)}


def select_columns(row: list[object]) -> list[object]:
    return [
        row[column - 1] if column <= len(row) else None
        for column in KEPT_COLUMN_INDEXES
    ]


def add_subsidy_column(
    row: list[object],
    *,
    subsidy_rate: Decimal,
    subsidy_cap: Decimal,
    is_header: bool = False,
    source_name: str | None = None,
    source_row: int | None = None,
) -> list[object]:
    result = list(row)
    if is_header:
        result.insert(3, "补贴金额")
        return result

    amount = result[2]
    if amount in (None, ""):
        subsidy = None
    else:
        try:
            calculated = Decimal(str(amount)) * subsidy_rate
            subsidy = float(
                min(calculated, subsidy_cap).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP,
                )
            )
        except (InvalidOperation, ValueError) as error:
            location = (
                f"{source_name} 第 {source_row} 行"
                if source_name is not None and source_row is not None
                else "数据"
            )
            raise ValueError(f"{location}的交易金额无效：{amount!r}") from error

    result.insert(3, subsidy)
    return result


def build_workbook(
    files: list[Path],
    *,
    data_dir: Path,
    file_marker: str,
    subsidy_rate: Decimal,
    subsidy_cap: Decimal,
) -> tuple[Workbook, int, int]:
    if not files:
        raise FileNotFoundError(
            f"未在 {data_dir} 中找到文件名包含"
            f"“{file_marker}”的 .xlsx 文件"
        )

    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "Summary"

    expected_header: list[object] | None = None
    output_header: list[object] | None = None
    data_row_count = 0
    data_rows: list[list[object]] = []

    for path in files:
        rows = read_rows(path)
        title = next(rows, None)
        header = next(rows, None)
        if title is None or header is None:
            raise ValueError(f"{path.name} 缺少标题行或表头行")

        if expected_header is None:
            expected_header = header
            # Drop the source title row and use the selected headers as row 1.
            output_header = add_subsidy_column(
                select_columns(header),
                subsidy_rate=subsidy_rate,
                subsidy_cap=subsidy_cap,
                is_header=True,
            )
            # Reject a wrong export before parsing any rows, so the operator sees
            # the missing fields instead of a downstream value error.
            missing_headers = [
                required
                for required in REQUIRED_SUBMITTED_HEADERS
                if required not in output_header
            ]
            if missing_headers:
                raise ValueError(
                    f"{path.name} 不是已上传数据的导出格式，缺少字段："
                    f"{'、'.join(missing_headers)}；"
                    f"实际字段 {tuple(output_header)}"
                )
            sheet.append(output_header)
        elif header != expected_header:
            raise ValueError(f"{path.name} 的表头与第一个文件不一致")

        for source_row, row in enumerate(rows, start=3):
            if any(value not in (None, "") for value in row):
                data_rows.append(
                    add_subsidy_column(
                        select_columns(row),
                        subsidy_rate=subsidy_rate,
                        subsidy_cap=subsidy_cap,
                        source_name=path.name,
                        source_row=source_row,
                    )
                )
                data_row_count += 1

    if output_header is None:
        raise RuntimeError("未能生成输出表头")

    status_column_index = output_header.index("状态")
    data_rows.sort(
        key=lambda row: STATUS_PRIORITY.get(
            str(row[status_column_index]) if row[status_column_index] is not None else "",
            len(STATUS_ORDER),
        )
    )
    for row in data_rows:
        sheet.append(row)

    description_column_index = output_header.index("描述")
    rows_by_status: dict[str, list[list[object]]] = {
        status: [] for status in STATUS_ORDER
    }
    for row in data_rows:
        status = str(row[status_column_index] or "")
        if status in rows_by_status:
            rows_by_status[status].append(row)

    font_name, font_path = resolve_font()
    measurement_font = load_measurement_font(font_path)
    format_sheet(sheet, font_name, measurement_font)
    for status in STATUS_ORDER:
        status_sheet = workbook.create_sheet(title=status)
        status_sheet.append(output_header)
        status_rows = rows_by_status[status]
        status_rows.sort(
            key=lambda row: (
                row[description_column_index] not in (None, ""),
                str(row[description_column_index] or ""),
            ),
            reverse=True,
        )
        for row in status_rows:
            status_sheet.append(row)
        format_sheet(status_sheet, font_name, measurement_font)

    return workbook, len(files), data_row_count


def validate_output(
    path: Path,
    expected_data_rows: int,
    *,
    subsidy_rate: Decimal,
    subsidy_cap: Decimal,
) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        expected_sheet_names = ["Summary", *STATUS_ORDER]
        if workbook.sheetnames != expected_sheet_names:
            raise RuntimeError(
                f"工作表校验失败：预期 {expected_sheet_names}，"
                f"实际 {workbook.sheetnames}"
            )

        sheet = workbook["Summary"]
        actual_data_rows = max(sheet.max_row - 1, 0)
        if actual_data_rows != expected_data_rows:
            raise RuntimeError(
                f"输出校验失败：预期 {expected_data_rows} 条，实际 {actual_data_rows} 条"
            )
        expected_columns = len(KEPT_SOURCE_COLUMNS) + 1
        if sheet.max_column != expected_columns:
            raise RuntimeError(
                f"输出校验失败：预期 {expected_columns} 列，"
                f"实际 {sheet.max_column} 列"
            )

        header = tuple(cell.value for cell in next(sheet.iter_rows(max_row=1)))
        status_column = header.index("状态")
        description_column = header.index("描述")
        amount_column = header.index("交易金额")
        subsidy_column = header.index("补贴金额")

        status_total = sum(
            max(workbook[status].max_row - 1, 0)
            for status in STATUS_ORDER
        )
        known_status_total = sum(
            1
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[status_column] in STATUS_ORDER
        )
        if status_total != known_status_total:
            raise RuntimeError(
                f"状态工作表校验失败：预期共 {known_status_total} 条，"
                f"实际共 {status_total} 条"
            )

        for status in STATUS_ORDER:
            status_sheet = workbook[status]
            status_header = tuple(
                cell.value for cell in next(status_sheet.iter_rows(max_row=1))
            )
            if status_header != header:
                raise RuntimeError(f"{status}工作表的标题行与汇总表不一致")

            descriptions: list[str] = []
            blank_description_found = False
            for row in status_sheet.iter_rows(min_row=2, values_only=True):
                if row[status_column] != status:
                    raise RuntimeError(f"{status}工作表中存在其他状态的数据")

                description = row[description_column]
                if description in (None, ""):
                    blank_description_found = True
                else:
                    if blank_description_found:
                        raise RuntimeError(
                            f"{status}工作表的空白描述未全部排在末尾"
                        )
                    descriptions.append(str(description))

                amount = row[amount_column]
                subsidy = row[subsidy_column]
                if amount not in (None, ""):
                    expected_subsidy = min(
                        Decimal(str(amount)) * subsidy_rate,
                        subsidy_cap,
                    ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    if Decimal(str(subsidy)) != expected_subsidy:
                        raise RuntimeError(
                            f"{status}工作表存在补贴金额计算错误"
                        )

            if descriptions != sorted(descriptions, reverse=True):
                raise RuntimeError(f"{status}工作表的描述列未按降序排列")
    finally:
        workbook.close()
