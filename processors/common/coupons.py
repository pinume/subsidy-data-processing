"""Shared helpers for the household-appliance and digital coupon pipelines."""

import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

from processors.common.dates import (
    normalize_coupon_date,
    normalize_document_number,
    normalize_receipt_identifier,
)

COUPON_REFERENCE_RE = re.compile(r"\d{11}[A-Z]")


@dataclass(frozen=True)
class ReferenceCorrectionIndex:
    references: frozenset[str]
    digit_prefixes: dict[str, frozenset[str]]
    deletion_variants: dict[str, frozenset[str]]
    substitution_variants: dict[tuple[int, str], frozenset[str]]


def build_reference_correction_index(
    reference_universe: set[str],
) -> ReferenceCorrectionIndex:
    references = frozenset(reference.upper() for reference in reference_universe)
    invalid_references = sorted(
        reference
        for reference in references
        if not COUPON_REFERENCE_RE.fullmatch(reference)
    )
    if invalid_references:
        raise ValueError(
            "参考号格式应为11位数字后跟一个大写字母，"
            f"实际存在无效参考号：{invalid_references[:5]}"
        )
    digit_prefixes: dict[str, set[str]] = {}
    deletion_variants: dict[str, set[str]] = {}
    substitution_variants: dict[tuple[int, str], set[str]] = {}
    for reference in references:
        digit_prefixes.setdefault(reference[:11], set()).add(reference)
        for index in range(len(reference)):
            without_character = reference[:index] + reference[index + 1:]
            deletion_variants.setdefault(without_character, set()).add(
                reference
            )
            substitution_variants.setdefault(
                (index, without_character),
                set(),
            ).add(reference)
    return ReferenceCorrectionIndex(
        references=references,
        digit_prefixes={
            prefix: frozenset(candidates)
            for prefix, candidates in digit_prefixes.items()
        },
        deletion_variants={
            variant: frozenset(candidates)
            for variant, candidates in deletion_variants.items()
        },
        substitution_variants={
            variant: frozenset(candidates)
            for variant, candidates in substitution_variants.items()
        },
    )


def as_currency(amount: Decimal) -> Decimal:
    """Round monetary comparisons to cents."""
    return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def classify_coupon_row(
    *,
    appliance_subsidy: object,
    digital_subsidy: object,
    row_number: int,
    source_name: str,
) -> str:
    """Classify a merged 销售用券情况统计 row as "家电" or "数码".

    The merged export is documented (README) to carry exactly one of the two
    国补 columns populated per row. Both populated is source data corruption
    serious enough to stop the run rather than silently pick a side; neither
    populated defaults to 家电, where the existing zero-国补 warning already
    surfaces it to the operator as bad data.
    """
    appliance_nonzero = appliance_subsidy not in (None, "", 0)
    digital_nonzero = digital_subsidy not in (None, "", 0)
    if appliance_nonzero and digital_nonzero:
        raise ValueError(
            f"{source_name} 第 {row_number} 行同时存在家电国补"
            f"（{appliance_subsidy}）与数码国补（{digital_subsidy}），"
            "无法确定该行所属项目"
        )
    return "数码" if digital_nonzero else "家电"


def load_coupon_remark_lookup(source: Path) -> dict[tuple[str, date], str]:
    if not source.exists():
        raise FileNotFoundError(f"未找到备注匹配文件：{source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if "Sheet1" not in workbook.sheetnames:
            raise ValueError(f"{source.name} 缺少 Sheet1 工作表")
        sheet = workbook["Sheet1"]
        header = [cell.value for cell in sheet[1]]
        required_headers = ("单据号", "日期", "备注")
        missing_headers = [
            required_header
            for required_header in required_headers
            if required_header not in header
        ]
        if missing_headers:
            raise ValueError(
                f"{source.name} 缺少字段：{'、'.join(missing_headers)}"
            )

        document_index = header.index("单据号")
        date_index = header.index("日期")
        remark_index = header.index("备注")
        lookup: dict[tuple[str, date], str] = {}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            document_number = normalize_document_number(row[document_index])
            remark = str(row[remark_index] or "").strip()
            if not document_number or not remark:
                continue
            receipt_date = normalize_coupon_date(
                row[date_index],
                row_number,
            )
            key = (document_number, receipt_date)
            existing_remark = lookup.get(key)
            if existing_remark is not None and existing_remark != remark:
                raise ValueError(
                    f"{source.name} 第 {row_number} 行组合键存在冲突备注："
                    f"{document_number} + {receipt_date:%Y-%m-%d}"
                )
            lookup[key] = remark
        return lookup
    finally:
        workbook.close()


def load_uploaded_detail_lookup(source: Path) -> dict[str, str]:
    if not source.exists():
        raise FileNotFoundError(f"未找到已上传匹配文件：{source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if "Summary" not in workbook.sheetnames:
            raise ValueError(f"{source.name} 缺少 Summary 工作表")
        sheet = workbook["Summary"]
        header = [cell.value for cell in sheet[1]]
        required_headers = ("检索参考号", "状态", "描述")
        missing_headers = [
            required_header
            for required_header in required_headers
            if required_header not in header
        ]
        if missing_headers:
            raise ValueError(
                f"{source.name} 缺少字段：{'、'.join(missing_headers)}"
            )

        reference_index = header.index("检索参考号")
        status_index = header.index("状态")
        description_index = header.index("描述")
        lookup: dict[str, str] = {}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            reference = normalize_receipt_identifier(
                row[reference_index]
            ).upper()
            if not reference:
                continue
            if not COUPON_REFERENCE_RE.fullmatch(reference):
                raise ValueError(
                    f"{source.name} 第 {row_number} 行检索参考号格式无效："
                    f"{row[reference_index]!r}；"
                    "正确格式应为11位数字后跟一个大写字母"
                )
            status = str(row[status_index] or "").strip()
            description = str(row[description_index] or "").strip()
            detail = f"{status}：{description}"
            existing_detail = lookup.get(reference)
            if existing_detail is not None and existing_detail != detail:
                raise ValueError(
                    f"{source.name} 第 {row_number} 行检索参考号存在冲突："
                    f"{reference}"
                )
            lookup[reference] = detail
        return lookup
    finally:
        workbook.close()


def reference_correction_candidates(
    raw_reference: str,
    reference_universe: set[str] | ReferenceCorrectionIndex,
) -> set[str]:
    index = (
        reference_universe
        if isinstance(reference_universe, ReferenceCorrectionIndex)
        else build_reference_correction_index(reference_universe)
    )
    references = index.references
    candidates: set[str] = set()
    upper_reference = raw_reference.upper()
    compact = re.sub(r"\s+", "", upper_reference)
    cleaned = re.sub(r"[^0-9A-Z]", "", upper_reference)

    for token in re.findall(
        r"(?<!\d)(\d{11}[A-Z])(?![A-Z0-9])",
        upper_reference,
    ):
        if token in references:
            candidates.add(token)
    if cleaned in references:
        candidates.add(cleaned)
    if re.fullmatch(r"\d{11}", compact):
        candidates.update(index.digit_prefixes.get(compact, ()))
    if len(compact) == 11:
        candidates.update(index.deletion_variants.get(compact, ()))
    elif len(compact) == 13:
        for character_index in range(13):
            candidate = (
                compact[:character_index] + compact[character_index + 1:]
            )
            if candidate in references:
                candidates.add(candidate)
    elif len(compact) == 12:
        for character_index in range(12):
            without_character = (
                compact[:character_index] + compact[character_index + 1:]
            )
            candidates.update(
                index.substitution_variants.get(
                    (character_index, without_character),
                    (),
                )
            )
    return candidates
