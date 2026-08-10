"""Payment collection (回款明细) processing for household appliances and digital.

Both subsidy programs export a "补贴明细" workbook per store. This module
normalizes those exports into one workbook holding a detail sheet per data
type plus a shared 汇总 sheet. The data type is detected per source file, so
the two programs never have to be selected by hand.
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from python_calamine import CalamineWorkbook
from xlsxwriter import Workbook

from processors import submitted
from processors.common.config import load_payment_brand_config, merchant_id
from processors.common.console import ConsoleReporter, format_count
from processors.common.excel import (
    FONT_SIZE,
    ROW_HEIGHT,
    calamine_rows,
    load_measurement_font,
    pixels_to_column_pixels,
    resolve_font,
    sheet_format_set,
    width_measurer,
    write_formatted_sheet,
    write_xlsx_atomically,
)
from processors.common.paths import find_data_files
from processors.common.references import validated_reference

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "output"
DATA_DIR: Path
SOURCE_FILES: tuple[Path, ...]
OUTPUT_FILE = OUTPUT_DIR / "回款明细.xlsx"

# Both programs name their export "...补贴明细"; which program a file belongs
# to is decided per file by detect_profile, not by the search keyword.
SOURCE_FILE_KEYWORD = "补贴明细"
SUPPORTED_SUFFIXES = (".xlsx",)

HEADER_ALIASES = {
    "交易完成时间": "交易时间",
    "销方名称": "销售企业",
    "销售企业名称": "销售企业",
    "核销商编": "商户编号",
    "经销商编号": "商户编号",
    "应收销售金额": "销售金额",
    "参考号": "交易参考号",
    "订单号": "交易订单号",
    "补贴销售金额": "补贴金额",
    "品类": "编码品类",
    "商品明细": "商品名称",
    "发票号码": "发票号",
}
APPLIANCE_DETAIL_HEADERS = (
    "拨付批次",
    "交易时间",
    "交易参考号",
    "交易订单号",
    "销售企业",
    "商户编号",
    "销售金额",
    "其他支付",
    "实收销售金额",
    "补贴金额",
    "补贴比例",
    "SN码",
    "所在地区",
    "商品编码",
    "能耗等级",
    "编码品类",
    "商品名称",
    "发票金额",
    "发票号",
    "ID",
)
DIGITAL_DETAIL_HEADERS = (
    "拨付批次",
    "交易时间",
    "交易参考号",
    "交易订单号",
    "销售企业",
    "商户编号",
    "销售金额",
    "其他支付",
    "实收销售金额",
    "补贴金额",
    "编码品类",
    "商品名称",
    "SN码",
    "IMEI1码",
    "IMEI2码",
    "商品编码",
    "所在地区",
    "发票号",
    "备注",
)
DERIVED_HEADERS = ("财务大类", "品牌")
# Category maps, brand keyword lists, brand normalization, the 美的系 group,
# and model->brand aliases all live in config/payment_brands.yaml — see that
# file for what each section means and why (list order is match priority for
# brand_keywords, and category order there drives sort/summary order).
_BRAND_CONFIG = load_payment_brand_config()
APPLIANCE_CATEGORY_MAP = _BRAND_CONFIG.appliance_categories
DIGITAL_CATEGORY_MAP = _BRAND_CONFIG.digital_categories
APPLIANCE_BRAND_KEYWORDS = _BRAND_CONFIG.appliance_brand_keywords
DIGITAL_BRAND_KEYWORDS = _BRAND_CONFIG.digital_brand_keywords
APPLIANCE_BRAND_NORMALIZATION_MAP = _BRAND_CONFIG.appliance_brand_normalization
MIDEA_GROUP_CATEGORIES = _BRAND_CONFIG.midea_group_categories
MIDEA_GROUP_BRANDS = _BRAND_CONFIG.midea_group_brands
APPLIANCE_BRAND_MODEL_ALIASES = _BRAND_CONFIG.appliance_brand_model_aliases
MODEL_TOKEN_PATTERN = re.compile(
    r"(?=[A-Z0-9._/+\-]*\d)[A-Z0-9]+(?:[._/+\-][A-Z0-9]+)*",
    re.IGNORECASE,
)
SUMMARY_SHEET_NAME = "汇总"
SUMMARY_HEADERS = ["财务大类", "品牌", "补贴金额合计", "补贴金额计数"]
DETAIL_SORT_HEADERS = ("财务大类", "品牌", "交易时间", "商品名称")


@dataclass(frozen=True)
class ProcessingProfile:
    name: str
    detail_sheet_name: str
    detail_headers: tuple[str, ...]
    optional_headers: frozenset[str]
    category_map: dict[str, str]
    brand_keywords: tuple[tuple[str, tuple[str, ...]], ...]
    brand_normalization_map: dict[str, str]
    brand_model_aliases: dict[str, str]


@dataclass(frozen=True)
class PaymentOutputExpectation:
    profile: ProcessingProfile
    row_count: int
    merchant_id: str


@dataclass(frozen=True)
class DetailSection:
    """One data type's merged detail rows, before they become a worksheet."""

    name: str
    header: tuple[str, ...]
    rows: list[list[object]]
    # Rows whose brand could not be identified from keywords or model
    # aliases; reported as a warning only when non-zero.
    unidentified_brands: int = 0
    supplemented_references: int = 0


@dataclass(frozen=True)
class PaymentReport:
    summary_rows: list[list[object]]
    summary_bold_rows: list[int]
    summary_merge_ranges: list[tuple[int, int]]
    details: list[DetailSection]
    section_profiles: list[ProcessingProfile]
    summary_groups: int


PROFILES = {
    "家电": ProcessingProfile(
        name="家电",
        detail_sheet_name="家电明细",
        detail_headers=APPLIANCE_DETAIL_HEADERS,
        optional_headers=frozenset({"发票金额"}),
        category_map=APPLIANCE_CATEGORY_MAP,
        brand_keywords=APPLIANCE_BRAND_KEYWORDS,
        brand_normalization_map=APPLIANCE_BRAND_NORMALIZATION_MAP,
        brand_model_aliases=APPLIANCE_BRAND_MODEL_ALIASES,
    ),
    "数码": ProcessingProfile(
        name="数码",
        detail_sheet_name="数码明细",
        detail_headers=DIGITAL_DETAIL_HEADERS,
        optional_headers=frozenset(),
        category_map=DIGITAL_CATEGORY_MAP,
        brand_keywords=DIGITAL_BRAND_KEYWORDS,
        brand_normalization_map={},
        brand_model_aliases={},
    ),
}
# 汇总 keeps this order, so 家电 always precedes 数码 no matter which source
# file the data directory happens to list first.
PROFILE_ORDER = ("家电", "数码")
FILENAME_KEYWORDS = (
    ("数码补贴明细", "数码"),
    ("以旧换新补贴明细", "家电"),
)


def configure_data_dir(data_dir: Path) -> None:
    global DATA_DIR
    global SOURCE_FILES

    DATA_DIR = data_dir
    SOURCE_FILES = tuple(
        find_data_files(data_dir, SOURCE_FILE_KEYWORD, SUPPORTED_SUFFIXES)
    )


def _iter_actual_cell_rows_with_numbers(worksheet):
    """Cell-preserving twin used only to recover types calamine discards."""
    if hasattr(worksheet, "reset_dimensions"):
        worksheet.reset_dimensions()
    for row_number, row in enumerate(worksheet.iter_rows(min_col=1), 1):
        if any(cell.value not in (None, "") for cell in row):
            yield row_number, row


def _iter_actual_calamine_rows_with_numbers(sheet):
    """Yield numbered non-empty rows from calamine's actual used range."""
    for row_number, row in enumerate(calamine_rows(sheet), 1):
        values = tuple(row)
        if any(value not in (None, "") for value in values):
            yield row_number, values


def _cell_value(row: tuple, index: int):
    return row[index] if index < len(row) else None


def _normalize_header_names(row: tuple) -> list[str]:
    headers = []
    for value in row:
        normalized = value.strip() if isinstance(value, str) else value
        headers.append(
            HEADER_ALIASES.get(normalized, normalized)
            if normalized is not None
            else ""
        )
    return headers


def _is_header_row(row: tuple) -> bool:
    return any(
        isinstance(value, str) and value.strip() == "拨付批次" for value in row
    )


def _missing_required_headers(
    headers: list[str], profile: ProcessingProfile
) -> list[str]:
    present = set(headers)
    return [
        name
        for name in profile.detail_headers
        if name not in profile.optional_headers and name not in present
    ]


def _detect_profile_from_filename(source_name: str) -> ProcessingProfile | None:
    matches = {
        profile_name
        for keyword, profile_name in FILENAME_KEYWORDS
        if keyword in source_name
    }
    if len(matches) == 1:
        return PROFILES[next(iter(matches))]
    return None


def detect_profile(path: Path, source_name: str | None = None) -> ProcessingProfile:
    """Pick the processing profile from the file name, else the header row."""
    if source_name:
        profile = _detect_profile_from_filename(source_name)
        if profile is not None:
            return profile
    book = CalamineWorkbook.from_path(str(path))
    try:
        for sheet_name in book.sheet_names:
            if sheet_name == SUMMARY_SHEET_NAME:
                continue
            sheet = book.get_sheet_by_name(sheet_name)
            for _, row in _iter_actual_calamine_rows_with_numbers(sheet):
                if not _is_header_row(row):
                    continue
                headers = _normalize_header_names(row)
                matches = {
                    name: profile
                    for name, profile in PROFILES.items()
                    if not _missing_required_headers(headers, profile)
                }
                if len(matches) == 1:
                    return next(iter(matches.values()))
                if matches:
                    raise ValueError(
                        "无法确定数据类型，表头同时符合："
                        f"{sorted(matches)}，请手动拆分原始数据"
                    )
                details = "；".join(
                    f"{profile.name}缺少 {_missing_required_headers(headers, profile)}"
                    for profile in PROFILES.values()
                )
                raise ValueError(f"无法识别数据类型：{details}")
    finally:
        book.close()
    raise ValueError(f"工作簿 {path.name!r} 中没有找到明细表头")


def _get_source_positions(
    row: tuple, sheet_name: str, profile: ProcessingProfile
) -> dict[str, int]:
    """Normalize and validate one source header row."""
    headers = _normalize_header_names(row)
    header_counts = Counter(name for name in headers if name)
    duplicates = sorted(name for name, count in header_counts.items() if count > 1)
    if duplicates:
        raise ValueError(f"工作表 {sheet_name!r} 存在重复表头：{duplicates}")

    positions = {name: index for index, name in enumerate(headers) if name}
    missing = _missing_required_headers(headers, profile)
    if missing:
        raise ValueError(f"工作表 {sheet_name!r} 缺少字段：{missing}")
    return positions


def _collect_normalized_detail(
    rows_with_numbers,
    sheet_name: str,
    *,
    profile: ProcessingProfile,
    merchant_id: str,
    formula_workbook=None,
) -> tuple[list[list[object]], int]:
    """Normalize one source sheet without creating output Cell objects yet.

    rows_with_numbers takes (row_number, values) pairs rather than a
    worksheet object so this same logic serves both the calamine-based read
    of the source workbook and — for a cell needing its original type —
    _iter_actual_cell_rows_with_numbers's openpyxl read of the source workbook.
    """
    source_positions: dict[str, int] = {}
    normalized_rows: list[list[object]] = []
    unidentified_brands = 0
    category_index = profile.detail_headers.index("编码品类")
    product_index = profile.detail_headers.index("商品名称")
    subsidy_index_in_headers = profile.detail_headers.index("补贴金额")
    subsidy_header = "补贴金额"
    formula_rows_iterator = None

    def _find_source_cell(row_number: int, column_index: int):
        # Advances the cell-preserving iterator forward only as far as needed;
        # opens the (expensive) openpyxl workbook lazily on first use.
        nonlocal formula_rows_iterator
        if formula_rows_iterator is None:
            formula_sheet = formula_workbook()[sheet_name]
            formula_rows_iterator = _iter_actual_cell_rows_with_numbers(formula_sheet)
        for formula_row_number, formula_row in formula_rows_iterator:
            if formula_row_number == row_number:
                return _cell_value(formula_row, column_index)
            if formula_row_number > row_number:
                break
        return None

    for row_number, row in rows_with_numbers:
        if not source_positions and _is_header_row(row):
            source_positions = _get_source_positions(row, sheet_name, profile)
            continue

        if not source_positions:
            continue

        row_merchant_id = _cell_value(row, source_positions["商户编号"])
        # Any row that is not this merchant's is skipped without comment — the
        # source carries every merchant's sales, so most rows are someone
        # else's. That deliberately includes a 商户编号 holding an Excel error
        # value: the 数码 source has one (#N/A at F460, a 404.85 subsidy row),
        # and it was confirmed as data that legitimately does not belong to
        # this merchant. Raising on it, as the 补贴金额 column does, would stop
        # every run on a file that is fine.
        if str(row_merchant_id).strip() != merchant_id:
            continue

        normalized = [
            _cell_value(row, source_positions[name]) if name in source_positions else None
            for name in profile.detail_headers
        ]
        subsidy_value = normalized[subsidy_index_in_headers]
        if (
            subsidy_value is None
            and formula_workbook is not None
            and subsidy_header in source_positions
        ):
            source_cell = _find_source_cell(
                row_number, source_positions[subsidy_header]
            )
            if source_cell is not None and source_cell.data_type == "e":
                raise ValueError(
                    f"工作表 {sheet_name!r} 第 {row_number} 行字段 "
                    f"{subsidy_header!r} 是 Excel 错误值：{source_cell.value!r}"
                )
            if source_cell is not None and source_cell.data_type == "f":
                raise ValueError(
                    f"工作表 {sheet_name!r} 第 {row_number} 行字段 "
                    f"{subsidy_header!r} 是公式但没有缓存计算结果；"
                    "请先用 Excel/WPS 打开并保存，或将公式转换为数值"
                )
        encoded_category = normalized[category_index]
        financial_category = profile.category_map.get(encoded_category)
        if financial_category is None:
            raise ValueError(
                f"工作表 {sheet_name!r} 第 {row_number} 行存在未配置的编码品类："
                f"{encoded_category!r}"
            )
        product_name = normalized[product_index]
        brand = _extract_brand(product_name, profile)
        brand = _normalize_financial_brand(brand, financial_category)
        if brand is None:
            unidentified_brands += 1
        normalized_rows.append(normalized + [financial_category, brand])

    if not source_positions:
        raise ValueError(f"工作表 {sheet_name!r} 未找到明细表头")
    return normalized_rows, unidentified_brands


def _extract_brand(product_name, profile: ProcessingProfile) -> str | None:
    if product_name in (None, ""):
        return None
    normalized_name = str(product_name).strip().casefold()
    for brand, keywords in profile.brand_keywords:
        if any(keyword.casefold() in normalized_name for keyword in keywords):
            return profile.brand_normalization_map.get(brand, brand)
    for model, brand in profile.brand_model_aliases.items():
        if model.casefold() in normalized_name:
            return profile.brand_normalization_map.get(brand, brand)
    return None


def _normalize_financial_brand(
    brand: str | None, financial_category: str
) -> str | None:
    if (
        financial_category in MIDEA_GROUP_CATEGORIES
        and brand in MIDEA_GROUP_BRANDS
    ):
        return "美的系"
    return brand


def _extract_model_tokens(product_name) -> set[str]:
    if product_name in (None, ""):
        return set()
    return {
        token.upper().strip("._/+-")
        for token in MODEL_TOKEN_PATTERN.findall(str(product_name).upper())
        if len(token.strip("._/+-")) >= 6
    }


def _infer_missing_brands(
    rows: list[list[object]], headers: tuple[str, ...]
) -> int:
    """Infer brands in plain records before they become worksheet cells."""
    product_column = headers.index("商品名称")
    brand_column = headers.index("品牌")
    token_brands: dict[str, Counter] = defaultdict(Counter)

    for row in rows:
        brand = row[brand_column]
        if brand in (None, ""):
            continue
        for token in _extract_model_tokens(row[product_column]):
            token_brands[token][brand] += 1

    unique_token_brand = {
        token: next(iter(counts))
        for token, counts in token_brands.items()
        if len(counts) == 1
    }
    inferred = 0
    for row in rows:
        if row[brand_column] not in (None, ""):
            continue
        candidates = {
            unique_token_brand[token]
            for token in _extract_model_tokens(row[product_column])
            if token in unique_token_brand
        }
        if len(candidates) == 1:
            row[brand_column] = candidates.pop()
            inferred += 1
    return inferred


def _category_order(category_map: dict[str, str]) -> dict[str, int]:
    return {
        category: index
        for index, category in enumerate(dict.fromkeys(category_map.values()))
    }


def _sort_scalar(value) -> tuple[int, str]:
    """Return a type-stable scalar sort key for worksheet values."""
    if value in (None, ""):
        return (1, "")
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return (0, value.isoformat())
    return (0, str(value).strip())


def _sort_detail_rows(
    rows: list[list[object]],
    headers: tuple[str, ...],
    category_map: dict[str, str],
) -> int:
    """Sort plain detail records before writing the output worksheet."""
    data_row_count = len(rows)
    if data_row_count <= 1:
        return data_row_count

    header_positions = {header: index for index, header in enumerate(headers)}
    missing = [header for header in DETAIL_SORT_HEADERS if header not in header_positions]
    if missing:
        raise ValueError(f"明细缺少排序字段：{missing}")

    category_column = header_positions["财务大类"]
    brand_column = header_positions["品牌"]
    transaction_time_column = header_positions["交易时间"]
    product_column = header_positions["商品名称"]
    category_order = _category_order(category_map)

    rows.sort(
        key=lambda item: (
            category_order.get(
                str(item[category_column]),
                len(category_order),
            ),
            _sort_scalar(item[brand_column]),
            _sort_scalar(item[transaction_time_column]),
            _sort_scalar(item[product_column]),
        ),
    )
    return data_row_count


def _sum_detail_groups(detail_sections) -> dict[tuple[str, str], list]:
    """Total 补贴金额 by (财务大类, 品牌) across detail sheets.

    detail_sections takes (sheet_name, rows) pairs rather than worksheets so
    the same totals can be computed from the workbook under construction and
    from the saved file, whichever library read it.
    """
    groups: dict[tuple[str, str], list] = {}
    for sheet_name, sheet_rows in detail_sections:
        rows = iter(sheet_rows)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"{sheet_name}缺少表头")
        header_positions = {
            value: index
            for index, value in enumerate(header)
            if value not in (None, "")
        }
        category_column = header_positions["财务大类"]
        brand_column = header_positions["品牌"]
        subsidy_column = header_positions["补贴金额"]

        for row_number, row in enumerate(rows, start=2):
            category = row[category_column]
            brand = row[brand_column]
            subsidy = row[subsidy_column]
            if category in (None, ""):
                raise ValueError(
                    f"{sheet_name}第 {row_number} 行缺少财务大类"
                )
            key = (str(category), "" if brand in (None, "") else str(brand))
            if key not in groups:
                groups[key] = [Decimal("0"), 0]
            if subsidy not in (None, ""):
                try:
                    subsidy_amount = Decimal(str(subsidy))
                except (InvalidOperation, ValueError) as error:
                    raise ValueError(
                        f"{sheet_name}第 {row_number} 行补贴金额不是有效数值"
                    ) from error
                groups[key][0] += subsidy_amount
                groups[key][1] += -1 if subsidy_amount < 0 else 1
    return groups


def _append_summary_row(rows, category, brand, amount: Decimal, count: int) -> None:
    rounded_amount = amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    rows.append([category, brand, float(rounded_amount), count])


def _build_summary_rows(
    sections: list[tuple[str, list, dict[str, str]]],
) -> tuple[list[list[object]], list[int], int]:
    """Build 汇总 as plain rows: one subtotaled block per data type.

    Sections keep their given order so 家电 and 数码 never interleave, and are
    separated by one blank row instead of a labeled title row. Returns the
    rows, the 1-based sheet row numbers of the bold 合计 rows, and the group
    count; the writer turns those into cells.
    """
    rows: list[list[object]] = [list(SUMMARY_HEADERS)]
    total_groups = 0
    bold_rows: list[int] = []
    grand_amount = Decimal("0")
    grand_count = 0
    written_sections = 0

    for _label, detail_sections, category_map in sections:
        groups = _sum_detail_groups(detail_sections)
        if not groups:
            continue
        if written_sections:
            rows.append([None, None, None, None])
        written_sections += 1

        category_order = _category_order(category_map)
        for (category, brand), (amount, count) in sorted(
            groups.items(),
            key=lambda item: (
                category_order.get(item[0][0], len(category_order)),
                item[0][1],
            ),
        ):
            _append_summary_row(rows, category, brand, amount, count)
        total_groups += len(groups)

        section_amount = sum((values[0] for values in groups.values()), Decimal("0"))
        section_count = sum(values[1] for values in groups.values())
        _append_summary_row(rows, "合计", None, section_amount, section_count)
        bold_rows.append(len(rows))
        grand_amount += section_amount
        grand_count += section_count

    _append_summary_row(rows, "合计", None, grand_amount, grand_count)
    bold_rows.append(len(rows))
    return rows, bold_rows, total_groups


def _summary_merge_ranges(rows: list[list[object]]) -> list[tuple[int, int]]:
    """Row spans of 财务大类 runs worth merging, as 1-based sheet rows.

    Mirrors what _merge_summary_categories did against a built worksheet: the
    header and the bottom grand-total row are excluded, and a run of one row
    merges nothing but still gets centered by the writer.
    """
    first_data_row = 2
    last_data_row = len(rows) - 1  # Exclude the bottom total row.
    if last_data_row < first_data_row:
        return []

    ranges: list[tuple[int, int]] = []
    group_start = first_data_row
    current_value = rows[group_start - 1][0]
    for row_number in range(first_data_row + 1, last_data_row + 2):
        next_value = (
            rows[row_number - 1][0] if row_number <= last_data_row else object()
        )
        if next_value == current_value:
            continue
        ranges.append((group_start, row_number - 1))
        group_start = row_number
        current_value = next_value
    return ranges


def _process_sources(
    sources: list[Path],
    profile: ProcessingProfile,
    merchant_id: str,
) -> DetailSection:
    """Merge every source file of one data type into a single detail section.

    Returns rows rather than a worksheet: the caller aggregates 汇总 from them
    and only then hands everything to the writer, so nothing is turned into
    cells until the data is final.
    """
    merged_rows = 0
    merged_header_written = False
    normalized_rows: list[list[object]] = []

    unidentified_brands = 0
    for path in sources:
        source_book = CalamineWorkbook.from_path(str(path))
        # Formula and error cell types aren't available from calamine's cached
        # values, so a blank subsidy needs a lazy openpyxl check of that cell.
        formula_book: object | None = None

        def get_formula_book(path=path):
            nonlocal formula_book
            if formula_book is None:
                formula_book = load_workbook(path, read_only=True, data_only=False)
            return formula_book

        try:
            for sheet_name in source_book.sheet_names:
                if sheet_name == SUMMARY_SHEET_NAME:
                    continue
                source_sheet = source_book.get_sheet_by_name(sheet_name)
                sheet_rows, missing_brands = _collect_normalized_detail(
                    _iter_actual_calamine_rows_with_numbers(source_sheet),
                    sheet_name,
                    profile=profile,
                    merchant_id=merchant_id,
                    formula_workbook=get_formula_book,
                )
                normalized_rows.extend(sheet_rows)
                merged_rows += len(sheet_rows)
                unidentified_brands += missing_brands
                merged_header_written = True
        finally:
            source_book.close()
            if formula_book is not None:
                formula_book.close()

    if not merged_header_written:
        raise ValueError(f"{profile.name}的原始数据中没有可整合的明细 Sheet")
    if merged_rows == 0:
        source_names = "、".join(path.name for path in sources)
        raise ValueError(
            f"{profile.name}来源文件中未找到商户 {merchant_id} 的数据："
            f"{source_names}；请检查 config/merchants.yaml"
        )
    output_headers = profile.detail_headers + DERIVED_HEADERS
    inferred_brands = _infer_missing_brands(normalized_rows, output_headers)
    unidentified_brands -= inferred_brands
    _sort_detail_rows(
        normalized_rows,
        output_headers,
        profile.category_map,
    )
    return DetailSection(
        name=profile.detail_sheet_name,
        header=output_headers,
        rows=normalized_rows,
        unidentified_brands=unidentified_brands,
    )


def _fill_missing_references(
    detail: DetailSection,
    submitted_report: submitted.SubmittedReport,
    profile_name: str,
) -> int:
    """Fill blank payment references from one uniquely matching upload order."""
    reference_index = detail.header.index("交易参考号")
    order_index = detail.header.index("交易订单号")
    amount_index = detail.header.index("销售金额")
    submitted_order_index = submitted_report.header.index("订单号")
    submitted_reference_index = submitted_report.header.index("检索参考号")
    submitted_amount_index = submitted_report.header.index("交易金额")

    rows_by_order: dict[str, list[list[object]]] = defaultdict(list)
    for row in submitted_report.summary_rows:
        order = str(row[submitted_order_index] or "").strip()
        if order:
            rows_by_order[order].append(row)

    supplemented = 0
    for row in detail.rows:
        if str(row[reference_index] or "").strip():
            continue
        order = str(row[order_index] or "").strip()
        matches = rows_by_order.get(order, [])
        if len(matches) != 1:
            raise ValueError(
                f"{profile_name}回款订单 {order or '<空>'} 的交易参考号为空，"
                f"按交易订单号在已上传数据中找到 {len(matches)} 条记录，无法唯一补全"
            )
        match = matches[0]
        location = f"{profile_name}已上传数据订单 {order} 的"
        reference = validated_reference(
            match[submitted_reference_index],
            location,
        )
        try:
            payment_amount = Decimal(str(row[amount_index]))
            submitted_amount = Decimal(str(match[submitted_amount_index]))
        except (InvalidOperation, ValueError) as error:
            raise ValueError(f"{location}交易金额无效，无法补全交易参考号") from error
        if payment_amount != submitted_amount:
            raise ValueError(
                f"{profile_name}回款订单 {order} 的销售金额 {payment_amount} "
                f"与已上传交易金额 {submitted_amount} 不一致，无法补全交易参考号"
            )
        row[reference_index] = reference
        supplemented += 1
    return supplemented


def _classify_sources() -> dict[str, list[Path]]:
    """Group the data directory's subsidy exports by data type."""
    if not SOURCE_FILES:
        raise FileNotFoundError(
            f"未在 {DATA_DIR} 找到包含“{SOURCE_FILE_KEYWORD}”的回款原始数据文件"
        )

    classified: dict[str, list[Path]] = {}
    for source in SOURCE_FILES:
        profile = detect_profile(source, source.name)
        classified.setdefault(profile.name, []).append(source)
    return classified


def build_report() -> PaymentReport:
    """Assemble every sheet's rows in memory, before any of them become cells."""
    classified = _classify_sources()
    sections: list[tuple[ProcessingProfile, DetailSection]] = []
    for profile_name in PROFILE_ORDER:
        sources = classified.get(profile_name)
        if not sources:
            continue
        profile = PROFILES[profile_name]
        detail = _process_sources(
            sources,
            profile,
            merchant_id(profile_name),
        )
        reference_index = detail.header.index("交易参考号")
        if any(
            not str(row[reference_index] or "").strip()
            for row in detail.rows
        ):
            submitted.configure_data_dir(DATA_DIR)
            supplemented = _fill_missing_references(
                detail,
                submitted.build_report(profile_name),
                profile_name,
            )
            detail = DetailSection(
                name=detail.name,
                header=detail.header,
                rows=detail.rows,
                unidentified_brands=detail.unidentified_brands,
                supplemented_references=supplemented,
            )
        sections.append((profile, detail))

    if not sections:
        raise ValueError("没有任何明细数据可输出")

    summary_rows, bold_rows, summary_groups = _build_summary_rows(
        [
            (
                profile.name,
                [(detail.name, [detail.header, *detail.rows])],
                profile.category_map,
            )
            for profile, detail in sections
        ]
    )
    # Merging a run of 财务大类 leaves only the top cell holding the value, so
    # the rows are blanked here rather than in the writer: validate_output
    # compares these same rows against the saved file, and would otherwise see
    # a category the file no longer carries.
    merge_ranges = _summary_merge_ranges(summary_rows)
    for first_row, last_row in merge_ranges:
        for row_number in range(first_row + 1, last_row + 1):
            summary_rows[row_number - 1][0] = None

    return PaymentReport(
        summary_rows=summary_rows,
        summary_bold_rows=bold_rows,
        summary_merge_ranges=merge_ranges,
        details=[detail for _profile, detail in sections],
        section_profiles=[profile for profile, _detail in sections],
        summary_groups=summary_groups,
    )


def write_workbook(path: Path, report: PaymentReport) -> None:
    font_name, font_path = resolve_font()
    measurement_font = load_measurement_font(font_path)
    with Workbook(
        str(path),
        {
            # Deliberately not constant_memory: 汇总 merges runs of 财务大类
            # after its rows are written, and constant_memory flushes each row
            # on the next one, so merge_range would silently do nothing. The
            # largest sheet here is a few thousand rows, well within memory.
            "strings_to_urls": False,
            # 商品名称 is free text from the source export; a leading "=" is
            # data, not a formula to evaluate.
            "strings_to_formulas": False,
        },
    ) as workbook:
        _write_summary_sheet(workbook, report, font_name, measurement_font)
        for detail in report.details:
            write_formatted_sheet(
                workbook,
                detail.name,
                detail.header,
                detail.rows,
                font_name,
                measurement_font,
            )


def _write_summary_sheet(
    workbook, report: PaymentReport, font_name: str, measurement_font
) -> None:
    """汇总 is a report, not a filterable table: merged 财务大类 runs, a blank
    separator row between data types, bold 合计 rows, and no autofilter."""
    sheet = workbook.add_worksheet(SUMMARY_SHEET_NAME)
    base = {
        "font_name": font_name,
        "font_size": FONT_SIZE,
        "font_color": "#000000",
        "align": "center",
        "valign": "vcenter",
    }
    formats = sheet_format_set(workbook, font_name)
    cell_formats = {
        (bold, number_format): workbook.add_format(
            {**base, **({"bold": True} if bold else {})}
            | ({"num_format": number_format} if number_format else {})
        )
        for bold in (False, True)
        for number_format in (None, "0.00", "0")
    }
    # Matches what format_sheet applied to 汇总 before: the amount column
    # carries two decimals, the count column none.
    column_number_formats = {2: "0.00", 3: "0"}

    measure = width_measurer(measurement_font)
    header = report.summary_rows[0]
    maximum_widths = [measure(value) for value in header]
    sheet.set_row(0, ROW_HEIGHT)
    for column, value in enumerate(header):
        sheet.write(0, column, value, formats["header"])

    bold_rows = set(report.summary_bold_rows)
    for row_number, row in enumerate(report.summary_rows[1:], start=2):
        sheet.set_row(row_number - 1, ROW_HEIGHT)
        is_bold = row_number in bold_rows
        for column, value in enumerate(row):
            number_format = (
                column_number_formats.get(column) if value is not None else None
            )
            sheet.write(
                row_number - 1, column, value, cell_formats[(is_bold, number_format)]
            )
            if column < len(maximum_widths):
                maximum_widths[column] = max(maximum_widths[column], measure(value))

    for first_row, last_row in report.summary_merge_ranges:
        if last_row > first_row:
            sheet.merge_range(
                first_row - 1,
                0,
                last_row - 1,
                0,
                report.summary_rows[first_row - 1][0],
                cell_formats[(first_row in bold_rows, None)],
            )

    sheet.freeze_panes(1, 0)
    for column, maximum_pixels in enumerate(maximum_widths):
        sheet.set_column_pixels(
            column, column, pixels_to_column_pixels(maximum_pixels)
        )


def _summary_snapshot(rows) -> list[tuple[object, ...]]:
    """Truncate 汇总 to its declared columns so the two sides compare equal.

    calamine pads every row out to the sheet's used width, and openpyxl's
    max_col did the truncating before; doing it here keeps the snapshot taken
    from the workbook being built comparable with the one read back.
    """
    return [tuple(row[: len(SUMMARY_HEADERS)]) for row in rows]


def validate_output(
    path: Path,
    expectations: dict[str, PaymentOutputExpectation],
    expected_summary: list[tuple[object, ...]],
) -> None:
    """Re-read the saved workbook and cross-check details against the summary.

    Values only — no font or fill checks — so this reads through calamine,
    which parses the file several times faster than openpyxl did.
    """
    workbook = CalamineWorkbook.from_path(str(path))
    try:
        expected_sheet_names = [SUMMARY_SHEET_NAME, *expectations]
        if workbook.sheet_names != expected_sheet_names:
            raise ValueError(
                f"{path.name} 工作表应为 {expected_sheet_names}，"
                f"实际为 {workbook.sheet_names}"
            )

        detail_sections = []
        for sheet_name, expectation in expectations.items():
            rows = list(calamine_rows(workbook.get_sheet_by_name(sheet_name)))
            expected_header = expectation.profile.detail_headers + DERIVED_HEADERS
            actual_header = tuple(rows[0][: len(expected_header)]) if rows else ()
            if actual_header != expected_header:
                raise ValueError(
                    f"{path.name} 的“{sheet_name}”表头不符合要求："
                    f"预期 {expected_header}，实际 {actual_header}"
                )

            written = max(len(rows) - 1, 0)
            if written != expectation.row_count:
                raise ValueError(
                    f"{path.name} 的“{sheet_name}”应有 "
                    f"{expectation.row_count} 行数据，"
                    f"实际 {written} 行"
                )

            merchant_column = expected_header.index("商户编号")
            wrong_merchants = [
                row_number
                for row_number, row in enumerate(rows[1:], start=2)
                if str(row[merchant_column]).strip() != expectation.merchant_id
            ]
            if wrong_merchants:
                raise ValueError(
                    f"{path.name} 的“{sheet_name}”存在非目标商户数据，"
                    f"首个异常行：{wrong_merchants[0]}"
                )
            detail_sections.append((sheet_name, rows))

        actual_summary = _summary_snapshot(
            calamine_rows(workbook.get_sheet_by_name(SUMMARY_SHEET_NAME))
        )
        if actual_summary != expected_summary:
            raise ValueError(f"{path.name} 的“{SUMMARY_SHEET_NAME}”内容校验失败")

        detail_groups = _sum_detail_groups(detail_sections)
        detail_total = sum(
            (values[0] for values in detail_groups.values()),
            Decimal("0"),
        ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        detail_count = sum(values[1] for values in detail_groups.values())
        grand_total = actual_summary[-1]
        if (
            grand_total[0] != "合计"
            or Decimal(str(grand_total[2])).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )
            != detail_total
            or grand_total[3] != detail_count
        ):
            raise ValueError(
                f"{path.name} 的汇总总计与明细不一致："
                f"明细为 {detail_total} / {detail_count}"
            )
    finally:
        workbook.close()


def process_payment_files(reporter: ConsoleReporter) -> None:
    report = build_report()
    expectations = {
        detail.name: PaymentOutputExpectation(
            profile=profile,
            row_count=len(detail.rows),
            merchant_id=merchant_id(profile.name),
        )
        for profile, detail in zip(report.section_profiles, report.details)
    }
    expected_summary = _summary_snapshot(report.summary_rows)
    write_xlsx_atomically(
        OUTPUT_FILE,
        lambda path: write_workbook(path, report),
        lambda path: validate_output(path, expectations, expected_summary),
    )
    for profile, detail in zip(report.section_profiles, report.details):
        reporter.metric(
            profile.name,
            f"{format_count(len(detail.rows))} 条",
        )
        if detail.unidentified_brands:
            reporter.review_required(
                f"{profile.name}有 {format_count(detail.unidentified_brands)} 行"
                "品牌未识别",
                ("处理：品牌列为空，请在 config/payment_brands.yaml 中补充关键词",),
            )
        if detail.supplemented_references:
            reporter.metric(
                f"{profile.name}补全交易参考号",
                f"{format_count(detail.supplemented_references)} 条",
            )
    reporter.metric("汇总", f"{format_count(report.summary_groups)} 组")
    reporter.output(OUTPUT_FILE)
