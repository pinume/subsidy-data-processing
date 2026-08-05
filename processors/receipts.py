"""收款单统计 (receipt statistics) processing, shared by both projects.

Household appliances and digital both read this same source file and are
processed with the same rules; there has never been a digital-specific
variant, so this lives at the top level rather than under either project's
package.
"""

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from python_calamine import CalamineWorkbook
from xlsxwriter import Workbook

from processors.common.console import ConsoleReporter, format_count
from processors.common.dates import (
    is_valid_original_invoice_number,
    normalize_document_number,
    normalize_receipt_date,
    normalize_receipt_identifier,
    receipt_match_key,
)
from processors.common.excel import (
    FONT_SIZE,
    ROW_HEIGHT,
    load_measurement_font,
    normalize_calamine_value,
    pixels_to_column_pixels,
    resolve_font,
    width_measurer,
    write_xlsx_atomically,
)
from processors.common.paths import find_data_files, resolve_unique_file

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_FILE = OUTPUT_DIR / "收款单统计.xlsx"

DATA_DIR: Path
RECEIPTS_SOURCE_FILE: Path | None
RECEIPT_STATISTICS_KEYWORD = "收款单统计"

# Located in the 60-column export by header name, but read positionally from
# here on, so this order is the contract: 单据号 and 日期 lead because
# is_empty_receipt_row and is_receipt_total_row index them directly, and
# prepare_receipt_data reads 原票号, 商品名称 and 销售类别 at 2, 3 and 4.
# 摘要 and 销售金额 were selected once and are no longer read by anything.
RECEIPTS_SOURCE_HEADER = (
    "单据号",
    "日期",
    "原票号",
    "商品名称",
    "销售类别",
)
# 原票号, 商品名称 and 销售类别 are read from the source and drive the remark,
# the 北国 exclusion and the sort, but nothing downstream consumes them: the
# coupon remark lookup reads exactly these three columns out of Sheet1.
RECEIPTS_OUTPUT_HEADER = ("单据号", "日期", "备注")
RECEIPTS_DOCUMENT_INDEX = RECEIPTS_OUTPUT_HEADER.index("单据号")
RECEIPTS_DATE_INDEX = RECEIPTS_OUTPUT_HEADER.index("日期")
RECEIPTS_REMARK_INDEX = RECEIPTS_OUTPUT_HEADER.index("备注")
RECEIPTS_REMARK_RETURN = "退换货/倒票（退单）"
RECEIPTS_REMARK_ORIGINAL = "退换货/倒票（原单）"
RECEIPTS_REMARK_BOTH = "退换货/倒票（退单及原单）"
# 销售类别 alone decides this, and nothing else is checked. A 零售补差 or
# 同型号换货 is booked as the original sale being refunded and immediately
# re-sold — in this export as two rows sharing one 单据号/日期 with opposite
# 销售金额 — but the pairing is not verified: a category naming one of these
# is taken at its word even if its counterpart row is missing.
#
# Neither row is itself a return to report, and the 原票号 they carry normally
# points at a sale that was replaced rather than returned, so they stay out of
# the remark rules. Both categories can nevertheless be intermediate links:
# when a later ordinary 退货 names one of them, follow its 原票号 back to the
# terminal original sale. This keeps the bridge rows blank while pairing the
# actual return with the subsidy-bearing original in 审核明细.
RECEIPTS_UNREMARKED_SALE_CATEGORIES = frozenset({"零售补差", "同型号换货"})
RECEIPTS_REFERENCE_BRIDGE_CATEGORIES = frozenset({"零售补差", "同型号换货"})
RECEIPTS_REMARKS = frozenset(
    {
        RECEIPTS_REMARK_RETURN,
        RECEIPTS_REMARK_ORIGINAL,
        RECEIPTS_REMARK_BOTH,
    }
)
RECEIPTS_REMARK_ORDER: dict[str | None, int] = {
    None: 0,
    RECEIPTS_REMARK_ORIGINAL: 1,
    RECEIPTS_REMARK_RETURN: 2,
    RECEIPTS_REMARK_BOTH: 3,
}
RECEIPTS_ROW_HEIGHT = 20
RECEIPTS_REMARK_FILL_COLOR = "FFC7CE"
# The same colour as openpyxl may report it: bare, alpha-prefixed by the reader,
# or alpha-prefixed by XlsxWriter on the way out.
RECEIPTS_REMARK_FILL_SPELLINGS = frozenset(
    {
        RECEIPTS_REMARK_FILL_COLOR,
        f"00{RECEIPTS_REMARK_FILL_COLOR}",
        f"FF{RECEIPTS_REMARK_FILL_COLOR}",
    }
)
RECEIPTS_EXCLUDED_PRODUCT_KEYWORD = "北国"

ISSUES_SHEET_NAME = "问题明细"
ISSUES_HEADER = ("问题类型", "行号", "内容", "说明")


@dataclass(frozen=True)
class ExcludedProductRecord:
    """One row excluded by the 北国 product rule, for the console warning.

    Deliberately not an issue: the exclusion is business policy, not a data
    problem, so these records never reach 问题明细. They only let the
    operator warning name actual rows instead of a bare count.
    """

    source_row: int
    document_number: str
    product_name: str


def configure_data_dir(data_dir: Path) -> None:
    global DATA_DIR
    global RECEIPTS_SOURCE_FILE

    DATA_DIR = data_dir
    RECEIPTS_SOURCE_FILE = resolve_unique_file(
        find_data_files(data_dir, RECEIPT_STATISTICS_KEYWORD, (".xlsx",))
    )


def read_receipt_rows(source: Path) -> list[list[object]]:
    """Read the 收款单统计 .xlsx export: title row, header row, data, 合计.

    Returns every source row at its original position, including blank rows
    and the 合计 row. Dropping them here would renumber everything after
    them, and prepare_receipt_data derives each row's Excel row number from
    its position in this list — so an operator chasing a reported problem
    row would be sent to the wrong line. Filtering happens there instead,
    inside the loop that already knows the true row number.

    Reads via python-calamine rather than openpyxl: the source export always
    carries far more columns than the five kept here (60 in a real file), and
    openpyxl's read_only parser still parses every column of every row before
    select() discards the rest, making it the dominant cost of this whole
    processing step. calamine parses the same file roughly 5x faster.
    """
    workbook = CalamineWorkbook.from_path(str(source))
    try:
        sheet = workbook.get_sheet_by_index(0)
        if sheet.start is None:
            # iter_rows() panics inside the Rust extension on a sheet with no
            # cells at all, so a blank export has to be rejected before it.
            raise ValueError(f"{source.name} 缺少总标题行或字段标题行")
        # Raw iter_rows(), not calamine_rows(): normalizing all 60 source
        # columns to keep only 5 of them is what this reader exists to avoid.
        # Column positions are only ever resolved against this same row shape
        # by header name, so calamine's dropped leading blank columns — which
        # calamine_rows pads back — cannot misalign anything here.
        rows_iter = sheet.iter_rows()
        title_row = next(rows_iter, None)
        header_row = next(rows_iter, None)
        if title_row is None or header_row is None:
            raise ValueError(f"{source.name} 缺少总标题行或字段标题行")
        header_row = [normalize_calamine_value(value) for value in header_row]

        source_headers = [
            str(value).strip() if value is not None else ""
            for value in header_row
        ]
        missing_headers = [
            header for header in RECEIPTS_SOURCE_HEADER
            if header not in source_headers
        ]
        if missing_headers:
            raise ValueError(
                f"{source.name} 缺少必要字段：{tuple(missing_headers)}；"
                f"实际字段 {tuple(source_headers)}"
            )
        source_column_indexes = [
            source_headers.index(header) for header in RECEIPTS_SOURCE_HEADER
        ]

        def select(values: list[object]) -> list[object]:
            return [
                normalize_calamine_value(values[index])
                if index < len(values) else None
                for index in source_column_indexes
            ]

        rows: list[list[object]] = [select(header_row)]
        for source_row in rows_iter:
            rows.append(select(source_row))

        actual_header = tuple(
            str(value).strip() if value is not None else ""
            for value in rows[0]
        )
        if actual_header != RECEIPTS_SOURCE_HEADER:
            raise ValueError(
                f"{source.name} 字段标题不一致："
                f"预期 {RECEIPTS_SOURCE_HEADER}，实际 {actual_header}"
            )
        return rows
    finally:
        workbook.close()


def is_empty_receipt_row(row: list[object]) -> bool:
    """Whether every kept field is blank.

    A row is only blank when all of them are: a row carrying just one value is
    real, incomplete data that still belongs in the output and in 问题明细.
    The comparison is against the stripped string so a formatting-only row
    Excel left behind counts as blank, while a numeric 0 does not.
    """
    return all(value is None or str(value).strip() == "" for value in row)


def is_receipt_total_row(row: list[object]) -> bool:
    """Whether this is the export's own 合计 row.

    Only 单据号 and 日期 are examined, and only for an exact match: "合计"
    appearing inside a 商品名称 is ordinary business text, and
    treating it as a total row would silently delete a real sale. Which of
    the two columns carries the word has varied between exports, so both
    are accepted.
    """
    document_number = str(row[0] or "").strip()
    receipt_date = str(row[1] or "").strip()
    return document_number == "合计" or receipt_date == "合计"


def receipt_remark(
    has_original: bool,
    is_referenced: bool,
) -> str | None:
    if has_original and is_referenced:
        return RECEIPTS_REMARK_BOTH
    if has_original:
        return RECEIPTS_REMARK_RETURN
    if is_referenced:
        return RECEIPTS_REMARK_ORIGINAL
    return None


def _receipt_remark_flags(
    record: dict[str, object],
    referenced_original_invoice_numbers: set[str],
) -> tuple[bool, bool]:
    match_key = str(record["match_key"])
    has_original = bool(record["original_invoice_number"])
    is_referenced = bool(
        match_key and match_key in referenced_original_invoice_numbers
    )
    return has_original, is_referenced


def _expand_referenced_original_invoice_numbers(
    referenced_original_invoice_numbers: set[str],
    bridge_originals_by_match_key: dict[str, set[str]],
) -> None:
    """Follow referenced bridge documents back to their original sales."""
    pending = list(referenced_original_invoice_numbers)
    while pending:
        referenced_match_key = pending.pop()
        for original_invoice_number in bridge_originals_by_match_key.get(
            referenced_match_key, ()
        ):
            if original_invoice_number in referenced_original_invoice_numbers:
                continue
            referenced_original_invoice_numbers.add(original_invoice_number)
            pending.append(original_invoice_number)


def receipt_output_sort_key(
    remark: object,
    receipt_date: object,
    document_number: object,
    product_name: object,
) -> tuple[int, str, bool, date, str, str]:
    """Sort by 备注、日期、单据号、商品名称; missing dates sort last.

    An empty remark sorts before every actual remark, so ordinary sales stay
    together at the top and the highlighted return/exchange rows form the
    trailing section of the output. Remark groups use the explicit business
    order above rather than their display text, so copy changes cannot
    silently reorder the output. Unknown remarks sort after all known groups.
    """
    normalized_remark = normalize_receipt_identifier(remark) or None
    known_remark = normalized_remark in RECEIPTS_REMARK_ORDER
    normalized_date = (
        receipt_date.date() if isinstance(receipt_date, datetime) else receipt_date
    )
    has_no_date = not isinstance(normalized_date, date)
    return (
        RECEIPTS_REMARK_ORDER.get(normalized_remark, len(RECEIPTS_REMARK_ORDER)),
        "" if known_remark else normalized_remark,
        has_no_date,
        date.max if has_no_date else normalized_date,
        normalize_receipt_identifier(document_number),
        normalize_receipt_identifier(product_name),
    )


def prepare_receipt_data(kept_rows: list[list[object]], source_name: str):
    records: list[dict[str, object]] = []
    key_rows: dict[str, list[int]] = {}
    referenced_original_invoice_numbers: set[str] = set()
    bridge_originals_by_match_key: dict[str, set[str]] = {}
    excluded_product_count = 0
    excluded_product_records: list[ExcludedProductRecord] = []
    blank_row_count = 0
    total_row_count = 0

    # start=3 because kept_rows[0] is the field header (Excel row 2) and the
    # export opens with a title row; every row read stays at its original
    # index, so source_row is the real Excel row number even when rows above
    # it are skipped below.
    for source_row, row in enumerate(kept_rows[1:], start=3):
        if is_empty_receipt_row(row):
            blank_row_count += 1
            continue
        if is_receipt_total_row(row):
            total_row_count += 1
            continue

        product_name = normalize_receipt_identifier(row[3])
        if RECEIPTS_EXCLUDED_PRODUCT_KEYWORD in product_name:
            excluded_product_count += 1
            excluded_product_records.append(
                ExcludedProductRecord(
                    source_row=source_row,
                    document_number=normalize_document_number(row[0]),
                    product_name=product_name,
                )
            )
            continue

        document_number = normalize_document_number(row[0])
        receipt_date = normalize_receipt_date(
            row[1], source_row=source_row, source_name=source_name
        )
        original_invoice_number = normalize_receipt_identifier(row[2])
        sale_category = normalize_receipt_identifier(row[4])
        is_unremarked_sale_category = (
            sale_category in RECEIPTS_UNREMARKED_SALE_CATEGORIES
        )
        match_key = (
            receipt_match_key(receipt_date, document_number)
            if receipt_date is not None and document_number
            else ""
        )
        if match_key:
            key_rows.setdefault(match_key, []).append(source_row)
        if original_invoice_number and not is_unremarked_sale_category:
            referenced_original_invoice_numbers.add(original_invoice_number)
        if (
            match_key
            and original_invoice_number
            and sale_category in RECEIPTS_REFERENCE_BRIDGE_CATEGORIES
        ):
            bridge_originals_by_match_key.setdefault(match_key, set()).add(
                original_invoice_number
            )
        records.append(
            {
                "document_number": document_number,
                "receipt_date": receipt_date,
                "original_invoice_number": original_invoice_number,
                "product_name": product_name or None,
                "match_key": match_key,
                "sale_category": sale_category,
            }
        )

    # A later return can reference an exchange or price-adjustment document
    # rather than the subsidy-bearing original. Traverse those intermediate
    # documents only after all rows have been indexed. The visited set also
    # makes malformed cyclic references harmless.
    _expand_referenced_original_invoice_numbers(
        referenced_original_invoice_numbers,
        bridge_originals_by_match_key,
    )

    # Remarks depend on the complete, expanded reference set, so compute them
    # only after every source row has been scanned. Sorting then uses the final
    # remark as its primary key.
    for record in records:
        if str(record["sale_category"]) in RECEIPTS_UNREMARKED_SALE_CATEGORIES:
            record["remark"] = None
            continue
        (
            has_original,
            is_referenced,
        ) = _receipt_remark_flags(
            record,
            referenced_original_invoice_numbers,
        )
        record["remark"] = receipt_remark(
            has_original,
            is_referenced,
        )

    records.sort(
        key=lambda record: receipt_output_sort_key(
            record["remark"],
            record["receipt_date"],
            record["document_number"],
            record["product_name"],
        )
    )

    # Every record ends up in output_rows in this sorted order, one row per
    # record. Assign output coordinates only now so 问题明细 points at the row
    # an operator will find in the generated Sheet1, not the raw import.
    for position, record in enumerate(records, start=2):
        record["output_row"] = position

    # Same-document 烟机+灶具 (and similar kitchen-suite) sales are entered as
    # multiple line items sharing one 单据号/日期, which is expected here, not
    # a data problem — so it is neither reported in 问题明细 nor highlighted.
    issues: list[tuple[str, str, str, str]] = []

    # 原票号 referencing a sale from a year earlier than anything in this
    # file can never resolve to a match_key here by construction (the file
    # only carries one year's receipts), so it is not a data problem either.
    receipt_years = {
        record["receipt_date"].year
        for record in records
        if record["receipt_date"] is not None
    }
    min_receipt_year = min(receipt_years) if receipt_years else None

    only_return_count = 0
    only_original_count = 0
    both_count = 0
    unmatched_original_count = 0
    invalid_original_count = 0
    missing_match_key_count = 0
    remark_count = 0
    output_rows: list[list[object]] = []
    for record in records:
        output_row = int(record["output_row"])
        original_invoice_number = str(record["original_invoice_number"])
        match_key = str(record["match_key"])
        # The same exclusion the remark went through: 原票号 is not acted on
        # for these rows, so neither the 退单/原单 tallies nor the issues
        # derived from it may claim them, or the counts stop describing the
        # sheet — every one of them carries an 原票号 and would land in
        # 仅退单数量 while its 备注 stays blank.
        is_unremarked = (
            str(record["sale_category"]) in RECEIPTS_UNREMARKED_SALE_CATEGORIES
        )
        has_original, is_referenced = (
            (False, False)
            if is_unremarked
            else _receipt_remark_flags(
                record,
                referenced_original_invoice_numbers,
            )
        )

        if not match_key:
            missing_match_key_count += 1
            issues.append(
                (
                    "缺少匹配键",
                    str(output_row),
                    "",
                    "日期或单据号为空，无法生成匹配键",
                )
            )
        if has_original and not is_valid_original_invoice_number(
            original_invoice_number
        ):
            invalid_original_count += 1
            issues.append(
                (
                    "原票号格式异常",
                    str(output_row),
                    original_invoice_number,
                    "原票号应为6位日期加单据号",
                )
            )
        if has_original and original_invoice_number not in key_rows:
            is_prior_period_reference = (
                min_receipt_year is not None
                and is_valid_original_invoice_number(original_invoice_number)
                and datetime.strptime(
                    original_invoice_number[:6], "%y%m%d"
                ).year
                < min_receipt_year
            )
            if not is_prior_period_reference:
                unmatched_original_count += 1
                issues.append(
                    (
                        "原票号未匹配",
                        str(output_row),
                        original_invoice_number,
                        "未找到日期与单据号组合键相同的原单",
                    )
                )

        if has_original and is_referenced:
            both_count += 1
        elif has_original:
            only_return_count += 1
        elif is_referenced:
            only_original_count += 1

        remark = record["remark"]
        if remark:
            remark_count += 1

        output_rows.append(
            [
                record["document_number"],
                record["receipt_date"],
                remark,
            ]
        )

    stats = {
        "总数据量": len(records),
        "删除北国商品行数": excluded_product_count,
        "跳过空白行数": blank_row_count,
        "跳过合计行数": total_row_count,
        "仅退单数量": only_return_count,
        "仅原单数量": only_original_count,
        "退单及原单数量": both_count,
        "备注总数": remark_count,
        "未匹配原票号数量": unmatched_original_count,
        "重复匹配键数量": sum(
            1 for source_rows in key_rows.values() if len(source_rows) > 1
        ),
        "原票号格式异常数量": invalid_original_count,
        "缺少匹配键数量": missing_match_key_count,
        # Console-warning only; never written to 问题明细 (the exclusion is
        # business policy, not a data problem).
        "北国剔除明细": tuple(excluded_product_records),
    }
    return output_rows, stats, issues


def _write_issues_sheet(
    workbook: Workbook,
    issues: list[tuple[str, str, str, str]],
    font_name: str,
    measurement_font,
) -> None:
    sheet = workbook.add_worksheet(ISSUES_SHEET_NAME)
    header_format = workbook.add_format(
        {
            "font_name": font_name,
            "font_size": FONT_SIZE,
            "bold": True,
            "font_color": "#FFFFFF",
            "bg_color": "#000000",
            "align": "center",
            "valign": "vcenter",
        }
    )
    centered_format = workbook.add_format(
        {
            "font_name": font_name,
            "font_size": FONT_SIZE,
            "font_color": "#000000",
            "align": "center",
            "valign": "vcenter",
        }
    )
    left_format = workbook.add_format(
        {
            "font_name": font_name,
            "font_size": FONT_SIZE,
            "font_color": "#000000",
            "align": "left",
            "valign": "vcenter",
        }
    )
    left_columns = {
        ISSUES_HEADER.index("内容"),
        ISSUES_HEADER.index("说明"),
    }
    measure = width_measurer(measurement_font)
    maximum_widths = [measure(value) for value in ISSUES_HEADER]

    sheet.set_row(0, ROW_HEIGHT)
    for column, value in enumerate(ISSUES_HEADER):
        sheet.write(0, column, value, header_format)
    for row_number, row in enumerate(issues, start=1):
        sheet.set_row(row_number, ROW_HEIGHT)
        for column, value in enumerate(row):
            cell_format = left_format if column in left_columns else centered_format
            sheet.write(row_number, column, value, cell_format)
            maximum_widths[column] = max(maximum_widths[column], measure(value))

    sheet.freeze_panes(1, 0)
    sheet.autofilter(0, 0, len(issues), len(ISSUES_HEADER) - 1)
    for column, maximum_pixels in enumerate(maximum_widths):
        sheet.set_column_pixels(
            column,
            column,
            pixels_to_column_pixels(maximum_pixels),
        )


def _comparable_receipt_row(row) -> tuple[object, ...]:
    """One output row reduced to what survives a round trip through .xlsx.

    openpyxl hands back a datetime for a date cell and None for a blank one,
    so both sides go through here before being compared.
    """
    document_number, receipt_date, remark = row
    return (
        normalize_receipt_identifier(document_number),
        receipt_date.date() if isinstance(receipt_date, datetime) else receipt_date,
        remark if remark not in (None, "") else None,
    )


def validate_receipts_output(
    path: Path,
    expected_rows: list[list[object]],
    expected_issues: list[tuple[str, str, str, str]],
) -> None:
    """Re-read the workbook just written and check it against what was built.

    The remark used to be recomputed here from 原票号 and 摘要, which the output
    no longer carries. That independent derivation is covered by
    prepare_receipt_data's own tests; what a read-back can still establish —
    and what only a read-back can — is that the writer put those rows on the
    sheet unaltered, with the right number formats, fills and ordering.
    """
    expected_data_rows = len(expected_rows)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        expected_sheet_names = (
            ["Sheet1"] if not expected_issues else ["Sheet1", ISSUES_SHEET_NAME]
        )
        if workbook.sheetnames != expected_sheet_names:
            raise RuntimeError(
                f"收款单工作表校验失败：实际工作表为 {workbook.sheetnames}"
            )

        sheet = workbook["Sheet1"]
        if sheet.max_row - 1 != expected_data_rows:
            raise RuntimeError(
                f"收款单行数校验失败：预期 {expected_data_rows} 条，"
                f"实际 {sheet.max_row - 1} 条"
            )
        if sheet.max_column != len(RECEIPTS_OUTPUT_HEADER):
            raise RuntimeError(
                f"收款单列数校验失败：预期 {len(RECEIPTS_OUTPUT_HEADER)} 列，"
                f"实际 {sheet.max_column} 列"
            )

        rows = sheet.iter_rows(
            min_row=1,
            min_col=1,
            max_col=len(RECEIPTS_OUTPUT_HEADER),
        )
        header_cells = next(rows, ())
        header = tuple(cell.value for cell in header_cells)
        if header != RECEIPTS_OUTPUT_HEADER:
            raise RuntimeError(
                f"收款单表头校验失败：预期 {RECEIPTS_OUTPUT_HEADER}，"
                f"实际 {header}"
            )

        snapshots = []
        for row_number, row in enumerate(rows, start=2):
            document_cell, date_cell, remark_cell = row
            # openpyxl does not promise a str here: a fill this program wrote
            # yields one, but a theme-coloured cell yields an RGB object that
            # cannot be sliced. Comparing against the three spellings of the
            # colour works for either, and matches how the coupon validators
            # already do it.
            pink_cells = tuple(
                cell.fill.fill_type == "solid"
                and cell.fill.fgColor.rgb in RECEIPTS_REMARK_FILL_SPELLINGS
                for cell in row
            )
            snapshots.append(
                (
                    row_number,
                    document_cell.value,
                    date_cell.value,
                    remark_cell.value,
                    date_cell.number_format,
                    pink_cells,
                )
            )

        actual_rows = [
            _comparable_receipt_row(snapshot[1:4]) for snapshot in snapshots
        ]
        if actual_rows != [_comparable_receipt_row(row) for row in expected_rows]:
            raise RuntimeError("收款单数据行与生成结果不一致")

        for (
            row_number,
            document_number,
            date_value,
            actual_remark,
            date_number_format,
            pink_cells,
        ) in snapshots:
            if document_number is not None and (
                not isinstance(document_number, str)
                or document_number.startswith("收款")
            ):
                raise RuntimeError(
                    f"收款单第 {row_number} 行的单据号格式不正确"
                )

            if date_value is not None:
                if not isinstance(date_value, (date, datetime)):
                    raise RuntimeError(
                        f"收款单第 {row_number} 行的日期不是有效日期"
                    )
                if date_number_format != "yyyymmdd":
                    raise RuntimeError(
                        f"收款单第 {row_number} 行的日期格式不正确"
                    )

            if actual_remark is not None and actual_remark not in RECEIPTS_REMARKS:
                raise RuntimeError(
                    f"收款单第 {row_number} 行的备注不是已知备注：{actual_remark!r}"
                )

            should_be_pink = actual_remark is not None
            for is_pink in pink_cells:
                if is_pink != should_be_pink:
                    raise RuntimeError(
                        f"收款单第 {row_number} 行的退换货备注填充不正确"
                    )

        # 商品名称 is the last sort key but is no longer written out, so only
        # the three leading keys can be re-derived here. Sorting by four keys
        # still leaves the output ordered by the first three, so a break in
        # this weaker check is still a real ordering fault.
        actual_sort_keys = [
            receipt_output_sort_key(
                actual_remark,
                date_value,
                document_number,
                None,
            )
            for (
                _row_number,
                document_number,
                date_value,
                actual_remark,
                _date_number_format,
                _pink_cells,
            ) in snapshots
        ]
        if actual_sort_keys != sorted(actual_sort_keys):
            raise RuntimeError(
                "收款单排序校验失败：应按备注、日期、单据号升序排列"
            )

        if expected_issues:
            issues_sheet = workbook[ISSUES_SHEET_NAME]
            issues_header = tuple(
                cell.value for cell in next(issues_sheet.iter_rows(max_row=1))
            )
            if issues_header != ISSUES_HEADER:
                raise RuntimeError(
                    f"{ISSUES_SHEET_NAME}工作表字段标题校验失败：实际为 {issues_header}"
                )

            def comparable(row: tuple[object, ...]) -> tuple[object, ...]:
                # A cell saved with "" round-trips through openpyxl as None,
                # so blanks on either side must compare equal.
                return tuple("" if value is None else value for value in row)

            actual_issues = [
                comparable(row)
                for row in issues_sheet.iter_rows(min_row=2, values_only=True)
            ]
            expected = [comparable(tuple(issue)) for issue in expected_issues]
            if actual_issues != expected:
                raise RuntimeError(
                    f"{ISSUES_SHEET_NAME}工作表内容校验失败：预期 {len(expected)} 条，"
                    f"实际 {len(actual_issues)} 条"
                )
    finally:
        workbook.close()


def _write_receipts_workbook(
    path: Path,
    output_rows: list[list[object]],
    issues: list[tuple[str, str, str, str]],
) -> None:
    font_name, font_path = resolve_font()
    measurement_font = load_measurement_font(font_path)
    with Workbook(
        str(path),
        {
            "constant_memory": True,
            "strings_to_urls": False,
            # 商品名称 and 备注 are free text copied from the source export; a
            # value starting with "=" is data, not a formula to evaluate.
            "strings_to_formulas": False,
        },
    ) as workbook:
        sheet = workbook.add_worksheet("Sheet1")
        header_format = workbook.add_format(
            {
                "font_name": font_name,
                "font_size": FONT_SIZE,
                "bold": True,
                "font_color": "#FFFFFF",
                "bg_color": "#000000",
                "align": "center",
                "valign": "vcenter",
            }
        )
        body_formats = {}
        for has_remark in (False, True):
            for number_format_kind in ("general", "text", "date"):
                properties = {
                    "font_name": font_name,
                    "font_size": FONT_SIZE,
                    "font_color": "#000000",
                    "align": "center",
                    "valign": "vcenter",
                }
                if has_remark:
                    properties["bg_color"] = f"#{RECEIPTS_REMARK_FILL_COLOR}"
                if number_format_kind == "text":
                    properties["num_format"] = "@"
                elif number_format_kind == "date":
                    properties["num_format"] = "yyyymmdd"
                body_formats[(has_remark, number_format_kind)] = (
                    workbook.add_format(properties)
                )

        measure = width_measurer(measurement_font)
        maximum_widths = [measure(value) for value in RECEIPTS_OUTPUT_HEADER]
        sheet.set_row(0, RECEIPTS_ROW_HEIGHT)
        for column, value in enumerate(RECEIPTS_OUTPUT_HEADER):
            sheet.write(0, column, value, header_format)

        for row_number, row in enumerate(output_rows, start=1):
            sheet.set_row(row_number, RECEIPTS_ROW_HEIGHT)
            has_remark = row[RECEIPTS_REMARK_INDEX] is not None
            for column, value in enumerate(row):
                number_format_kind = (
                    "text"
                    if column == RECEIPTS_DOCUMENT_INDEX
                    else "date"
                    if column == RECEIPTS_DATE_INDEX and value not in (None, "")
                    else "general"
                )
                sheet.write(
                    row_number,
                    column,
                    value,
                    body_formats[(has_remark, number_format_kind)],
                )
                maximum_widths[column] = max(maximum_widths[column], measure(value))

        sheet.freeze_panes(1, 0)
        for column, maximum_pixels in enumerate(maximum_widths):
            sheet.set_column_pixels(
                column,
                column,
                pixels_to_column_pixels(maximum_pixels),
            )

        if issues:
            _write_issues_sheet(
                workbook,
                issues,
                font_name,
                measurement_font,
            )


def process_receipts(reporter: ConsoleReporter) -> None:
    if RECEIPTS_SOURCE_FILE is None:
        raise FileNotFoundError(
            f"未在 {DATA_DIR} 中找到文件名包含"
            f"“{RECEIPT_STATISTICS_KEYWORD}”的 .XLSX 文件"
        )

    kept_rows = read_receipt_rows(RECEIPTS_SOURCE_FILE)
    output_rows, stats, issues = prepare_receipt_data(
        kept_rows, source_name=RECEIPTS_SOURCE_FILE.name
    )
    row_count = len(output_rows)
    write_xlsx_atomically(
        OUTPUT_FILE,
        lambda path: _write_receipts_workbook(
            path,
            output_rows,
            issues,
        ),
        lambda path: validate_receipts_output(path, output_rows, issues),
    )

    excluded_records = stats["北国剔除明细"]
    reporter.metric(
        "原始数据",
        f"{format_count(row_count + len(excluded_records))} 行",
    )
    reporter.metric(
        "北国商品",
        f"按业务规则剔除 {format_count(len(excluded_records))} 行",
    )
    reporter.metric("有效数据", f"{format_count(row_count)} 行")
    if excluded_records:
        examples = tuple(
            f"源第 {record.source_row} 行，单据号 "
            f"{record.document_number or '(空)'}，"
            f"商品名称 {record.product_name}"
            for record in excluded_records
        )
        reporter.detail(
            f"按业务规则剔除“{RECEIPTS_EXCLUDED_PRODUCT_KEYWORD}”商品："
            f"{format_count(len(excluded_records))} 行",
            examples,
        )
    if issues:
        reporter.review_required(
            f"存在 {format_count(len(issues))} 条数据问题",
            (f"详见“{ISSUES_SHEET_NAME}”工作表",),
        )
    reporter.output(OUTPUT_FILE)
